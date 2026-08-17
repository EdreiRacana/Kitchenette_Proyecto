"""Export XLSX de respuestas del asistente.

Diseño:
  - Cada tool declara qué claves del `data` son "tabulares" (arrays de
    dicts) y qué claves son escalares. build_xlsx() genera una hoja con:
      * Encabezado: título de la consulta + fecha de emisión.
      * Bloque de escalares (si aplica): cuadro clave/valor.
      * Tabla principal: headers + filas.
  - Para tools cuyo `data` no tiene tablas, se exporta un XLSX con el
    texto plano formateado (fallback). Nunca falla.
  - openpyxl ya está en requirements.txt (3.1.2) — sin nuevas
    dependencias. Estilos consistentes con el resto del ERP.
"""
from __future__ import annotations
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# Estilos base — el visual del ERP: azul acero (nova), gris oscuro texto,
# borde tenue. Consistente con los XLSX del módulo BI / accounting.
_TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="0B2740")
_SUBTITLE_FONT = Font(name="Calibri", size=10, color="586880")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
_TOTAL_FONT = Font(name="Calibri", size=11, bold=True)
_CELL_FONT = Font(name="Calibri", size=10.5)
_BORDER = Border(
    left=Side(style="thin", color="D6DEE7"),
    right=Side(style="thin", color="D6DEE7"),
    top=Side(style="thin", color="D6DEE7"),
    bottom=Side(style="thin", color="D6DEE7"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")


# ─── Especificación por tool ──────────────────────────────────────────
# Cada spec tiene:
#   title    : título humano del reporte.
#   scalars  : [(key_en_data, label_visible)] — cuadro clave/valor arriba.
#   table    : {
#       key      : nombre del array dentro de data,
#       headers  : [(campo, label, tipo)]  tipo: str|int|money|pct
#     } o lista de tablas.
_SPEC: Dict[str, Dict[str, Any]] = {
    "ventas_periodo": {
        "title": "Ventas del periodo",
        "scalars": [("periodo", "Periodo"), ("count", "Pedidos"),
                     ("total", "Total ventas"), ("ticket_promedio", "Ticket promedio")],
    },
    "top_productos": {
        "title": "Top productos",
        "scalars": [("periodo", "Periodo"), ("por", "Ordenado por")],
        "table": {"key": "items", "headers": [
            ("product_name", "Producto", "str"),
            ("sku", "SKU", "str"),
            ("quantity", "Unidades", "int"),
            ("revenue", "Ingreso", "money"),
        ]},
    },
    "top_clientes": {
        "title": "Top clientes",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "items", "headers": [
            ("name", "Cliente", "str"), ("pedidos", "Pedidos", "int"),
            ("revenue", "Ingreso", "money"),
        ]},
    },
    "ticket_promedio_ventas": {
        "title": "Ticket promedio de ventas",
        "scalars": [("periodo", "Periodo"), ("count", "Pedidos"),
                     ("ticket", "Ticket promedio")],
    },
    "pedidos_pendientes": {
        "title": "Pedidos pendientes",
        "scalars": [("count", "Total"), ("total_saldo", "Saldo pendiente")],
        "table": {"key": "items", "headers": [
            ("folio", "Folio", "str"), ("status", "Estado", "str"),
            ("saldo", "Saldo", "money"), ("dias", "Días", "int"),
        ]},
    },
    "cxc_resumen": {
        "title": "Cuentas por cobrar",
        "scalars": [("total", "Total por cobrar")],
        "table": {"key": "top_debtors", "headers": [
            ("name", "Cliente", "str"), ("saldo", "Saldo", "money"),
        ]},
    },
    "cxp_resumen": {
        "title": "Cuentas por pagar",
        "scalars": [("total", "Total por pagar")],
        "table": {"key": "top_creditors", "headers": [
            ("name", "Proveedor", "str"), ("saldo", "Saldo", "money"),
        ]},
    },
    "top_deudores": {
        "title": "Top deudores",
        "table": {"key": "items", "headers": [
            ("name", "Cliente", "str"), ("saldo", "Saldo", "money"),
        ]},
    },
    "top_acreedores": {
        "title": "Top acreedores",
        "table": {"key": "items", "headers": [
            ("name", "Proveedor", "str"), ("saldo", "Saldo", "money"),
        ]},
    },
    "saldo_bancos": {
        "title": "Saldos bancarios",
        "scalars": [("total_mxn", "Total MXN")],
        "table": {"key": "accounts", "headers": [
            ("name", "Cuenta", "str"), ("balance", "Saldo", "money"),
            ("currency", "Moneda", "str"),
        ]},
    },
    "stock_critico": {
        "title": "Stock crítico",
        "scalars": [("count", "Total"), ("agotados", "Agotados")],
        "table": {"key": "items", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("warehouse", "Almacén", "str"),
            ("stock", "Existencia", "int"), ("reorder", "Reorden", "int"),
        ]},
    },
    "caducidades_proximas": {
        "title": "Caducidades próximas",
        "scalars": [("dias", "Ventana (días)"), ("count", "Total lotes")],
        "table": {"key": "items", "headers": [
            ("product_name", "Producto", "str"),
            ("quantity_remaining", "Unidades", "int"),
            ("days_left", "Días restantes", "int"),
        ]},
    },
    "desempeno_cadena": {
        "title": "Desempeño por cadena",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "items", "headers": [
            ("name", "Cadena", "str"), ("pedidos", "Pedidos", "int"),
            ("revenue", "Ingreso", "money"),
        ]},
    },
    "desempeno_tienda": {
        "title": "Desempeño por tienda",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "top", "headers": [
            ("name", "Tienda", "str"), ("cadena", "Cadena", "str"),
            ("unidades", "Unidades", "int"), ("revenue", "Ingreso", "money"),
        ]},
    },
    "sell_through_por_tienda": {
        "title": "Sell-out por tienda",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "items", "headers": [
            ("name", "Tienda", "str"), ("units_out", "Unidades vendidas", "int"),
        ]},
    },
    "ventas_pos_dia": {
        "title": "Ventas POS del día",
        "scalars": [("fecha", "Fecha"), ("tickets", "Tickets"),
                     ("total", "Total"), ("ticket_promedio", "Ticket promedio"),
                     ("sesiones_cerradas", "Sesiones cerradas"),
                     ("diferencia_arqueo_total", "Diferencia arqueo")],
    },
    "ventas_pos_hora": {
        "title": "Ventas POS por hora",
        "scalars": [("fecha", "Fecha")],
        "table": {"key": "horas", "headers": [
            ("hora", "Hora", "int"), ("tickets", "Tickets", "int"),
            ("monto", "Monto", "money"),
        ]},
    },
    "concentracion_clientes": {
        "title": "Concentración de clientes (Pareto)",
        "scalars": [("periodo", "Periodo"), ("n_clientes", "Total clientes"),
                     ("clientes_hasta_80", "Clientes al 80%"),
                     ("top3_pct", "Top 3 (%)"), ("total", "Ingreso total")],
    },
    "sin_movimiento": {
        "title": "SKUs sin movimiento",
        "scalars": [("dias", "Ventana (días)"), ("count", "Total")],
        "table": {"key": "items", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("stock", "Stock parado", "int"),
        ]},
    },
    "rotacion_producto": {
        "title": "Rotación (WoS)",
        "scalars": [("n_evaluados", "SKUs evaluados"),
                     ("wos_promedio", "WoS promedio (sem)")],
        "table": {"key": "rapidos", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("stock", "Stock", "int"), ("vel_sem", "Velocidad/sem", "int"),
            ("wos", "WoS (sem)", "int"),
        ]},
    },
    "utilidad_bruta": {
        "title": "Utilidad bruta",
        "scalars": [("periodo", "Periodo"), ("ingreso", "Ingreso"),
                     ("costo", "Costo"), ("utilidad", "Utilidad"),
                     ("margen_pct", "Margen %")],
    },
    "cotizaciones_abiertas": {
        "title": "Cotizaciones abiertas",
        "scalars": [("count", "Total"), ("monto_total", "Monto total")],
        "table": {"key": "recientes", "headers": [
            ("folio", "Folio", "str"), ("monto", "Monto", "money"),
        ]},
    },
    "clientes_inactivos": {
        "title": "Clientes inactivos",
        "scalars": [("dias", "Umbral (días)"), ("count", "Total")],
        "table": {"key": "items", "headers": [
            ("name", "Cliente", "str"),
            ("dias_sin_comprar", "Días sin comprar", "int"),
        ]},
    },
    "devoluciones_periodo": {
        "title": "Devoluciones del periodo",
        "scalars": [("periodo", "Periodo"), ("count", "Total"),
                     ("monto", "Monto reembolsado")],
    },
    "cxc_vencen_semana": {
        "title": "CxC que vencen esta semana",
        "scalars": [("count", "Total"), ("total", "Monto total")],
        "table": {"key": "items", "headers": [
            ("folio", "Folio", "str"), ("cliente", "Cliente", "str"),
            ("monto", "Monto", "money"), ("vence", "Vence", "str"),
        ]},
    },
    "cxp_vencen_semana": {
        "title": "CxP que vencen esta semana",
        "scalars": [("count", "Total"), ("total", "Monto total")],
        "table": {"key": "items", "headers": [
            ("folio", "Folio", "str"), ("proveedor", "Proveedor", "str"),
            ("monto", "Monto", "money"), ("vence", "Vence", "str"),
        ]},
    },
    "flujo_neto_30d": {
        "title": "Flujo neto proyectado 30 días",
        "scalars": [("cxc", "CxC esperada"), ("cxp", "CxP esperada"),
                     ("neto", "Neto proyectado")],
    },
    "oc_abiertas": {
        "title": "Órdenes de compra abiertas",
        "scalars": [("count", "Total"), ("monto_total", "Monto total")],
        "table": {"key": "items", "headers": [
            ("folio", "Folio", "str"), ("proveedor", "Proveedor", "str"),
            ("status", "Estado", "str"), ("monto", "Monto", "money"),
            ("vence", "Vence", "str"),
        ]},
    },
    "oc_atrasadas": {
        "title": "Órdenes de compra atrasadas",
        "scalars": [("count", "Total")],
        "table": {"key": "items", "headers": [
            ("folio", "Folio", "str"), ("proveedor", "Proveedor", "str"),
            ("monto", "Monto", "money"),
            ("dias_retraso", "Días retraso", "int"),
        ]},
    },
    "top_proveedores": {
        "title": "Top proveedores",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "items", "headers": [
            ("name", "Proveedor", "str"),
            ("gasto", "Gasto", "money"), ("facturas", "Facturas", "int"),
        ]},
    },
    "valor_inventario": {
        "title": "Valor de inventario",
        "scalars": [("valor_total", "Valor total"),
                     ("skus_con_stock", "SKUs con stock")],
    },
    "merma_mes": {
        "title": "Merma del periodo",
        "scalars": [("periodo", "Periodo"), ("movimientos", "Movimientos"),
                     ("unidades", "Unidades"), ("valor", "Valor")],
    },
    "ingresos_vs_egresos": {
        "title": "Ingresos vs egresos",
        "scalars": [("periodo", "Periodo"), ("ingresos", "Ingresos"),
                     ("egresos", "Egresos"), ("neto", "Neto")],
    },
    "gastos_por_categoria": {
        "title": "Gastos por categoría",
        "scalars": [("periodo", "Periodo"), ("total", "Total")],
        "table": {"key": "items", "headers": [
            ("categoria", "Categoría", "str"), ("monto", "Monto", "money"),
        ]},
    },
    "movimientos_no_conciliados": {
        "title": "Movimientos bancarios sin conciliar",
        "scalars": [("count", "Total"), ("monto", "Monto")],
    },
    "nomina_periodo": {
        "title": "Nómina del periodo",
        "scalars": [("periodo", "Periodo"), ("status", "Estado"),
                     ("kind", "Tipo"), ("empleados", "Empleados"),
                     ("bruto", "Bruto"), ("neto", "Neto"),
                     ("imss_patronal", "IMSS patronal")],
    },
    "empleados_activos": {
        "title": "Plantilla",
        "scalars": [("activos", "Activos"), ("altas_mes", "Altas del mes")],
    },
    "incapacidades_mes": {
        "title": "Incapacidades",
        "scalars": [("periodo", "Periodo"), ("total", "Total")],
    },
    "contratos_por_vencer": {
        "title": "Contratos por vencer",
        "scalars": [("dias", "Ventana (días)"), ("count", "Total")],
        "table": {"key": "items", "headers": [
            ("empleado", "Empleado", "str"), ("tipo", "Tipo", "str"),
            ("vence", "Vence", "str"), ("dias", "Días restantes", "int"),
        ]},
    },
    "cumpleanos_mes": {
        "title": "Cumpleaños del mes",
        "scalars": [("count", "Total")],
        "table": {"key": "items", "headers": [
            ("nombre", "Empleado", "str"), ("dia", "Fecha", "str"),
        ]},
    },
    "isr_nomina_mes": {
        "title": "ISR retenido en nómina",
        "scalars": [("periodo", "Periodo"), ("isr_retenido", "ISR retenido")],
    },
    "corte_caja_actual": {
        "title": "Corte de caja actual",
        "scalars": [("abiertas", "Sesiones abiertas")],
        "table": {"key": "sesiones", "headers": [
            ("terminal", "Terminal", "str"), ("abierta", "Abierta desde", "str"),
            ("esperado", "Esperado", "money"),
        ]},
    },
    "formas_pago_pos": {
        "title": "Formas de pago POS",
        "scalars": [("fecha", "Fecha"), ("total", "Total")],
        "table": {"key": "items", "headers": [
            ("metodo", "Método", "str"), ("tickets", "Tickets", "int"),
            ("monto", "Monto", "money"),
        ]},
    },
    "top_cajeros_dia": {
        "title": "Top cajeros del día",
        "scalars": [("fecha", "Fecha")],
        "table": {"key": "items", "headers": [
            ("cajero", "Cajero", "str"), ("tickets", "Tickets", "int"),
            ("monto", "Monto", "money"),
        ]},
    },
    "flujo_efectivo_proyectado": {
        "title": "Flujo de efectivo proyectado",
        "scalars": [("saldo_actual", "Saldo actual"),
                     ("cobranza_esperada", "Cobranza esperada"),
                     ("pagos_esperados", "Pagos esperados"),
                     ("proyeccion_30d", "Proyección 30 días")],
    },
    "nomina_vs_ventas": {
        "title": "Nómina vs ventas",
        "scalars": [("ventas", "Ventas del mes"), ("nomina", "Nómina bruta"),
                     ("pct_costo_laboral", "% costo laboral")],
    },
    "tiendas_wos_critico": {
        "title": "Tiendas WoS crítico",
        "scalars": [("count", "Total")],
        "table": {"key": "items", "headers": [
            ("name", "Tienda", "str"), ("cadena", "Cadena", "str"),
            ("on_hand", "Stock", "int"), ("wos", "WoS (sem)", "int"),
            ("umbral", "Umbral", "int"),
        ]},
    },
    "tiendas_sobrestock": {
        "title": "Tiendas con sobre-stock",
        "scalars": [("count", "Total")],
        "table": {"key": "items", "headers": [
            ("name", "Tienda", "str"), ("cadena", "Cadena", "str"),
            ("on_hand", "Stock", "int"), ("wos", "WoS (sem)", "int"),
            ("umbral", "Umbral", "int"),
        ]},
    },
    "fill_rate_cadena": {
        "title": "Fill rate por cadena",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "items", "headers": [
            ("cadena", "Cadena", "str"), ("vendido", "Vendido", "int"),
            ("devuelto", "Devuelto", "int"),
            ("fill_rate_pct", "Fill rate %", "int"),
        ]},
    },
    "return_rate_cadena": {
        "title": "Return rate por cadena",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "items", "headers": [
            ("cadena", "Cadena", "str"),
            ("return_rate_pct", "Return rate %", "int"),
            ("umbral", "Umbral %", "int"),
        ]},
    },
    "aging_cxc": {
        "title": "Aging de cuentas por cobrar",
        "scalars": [("total", "Total")],
    },
    "dso_dpo": {
        "title": "DSO / DPO",
        "scalars": [("dso_dias", "DSO (días de cobro)"),
                     ("dpo_dias", "DPO (días de pago)"),
                     ("cxc", "CxC"), ("cxp", "CxP")],
    },
    "pagos_programados": {
        "title": "Pagos programados",
        "scalars": [("count", "Total"), ("total", "Monto total")],
        "table": {"key": "items", "headers": [
            ("fecha", "Fecha", "str"), ("tipo", "Tipo", "str"),
            ("concepto", "Concepto", "str"),
            ("metodo", "Método", "str"), ("monto", "Monto", "money"),
        ]},
    },
    "aguinaldo_devengado": {
        "title": "Aguinaldo devengado",
        "scalars": [("empleados", "Empleados"), ("total", "Total")],
        "table": {"key": "top", "headers": [
            ("empleado", "Empleado", "str"),
            ("aguinaldo", "Aguinaldo", "money"),
        ]},
    },
    "vacaciones_pendientes": {
        "title": "Vacaciones pendientes",
        "scalars": [("empleados_con_saldo", "Empleados con saldo"),
                     ("total_dias", "Total días")],
        "table": {"key": "top", "headers": [
            ("empleado", "Empleado", "str"),
            ("pendientes", "Días pendientes", "int"),
        ]},
    },
    "imss_a_pagar": {
        "title": "IMSS a pagar",
        "scalars": [("periodo", "Periodo"), ("obrero", "Cuota obrero"),
                     ("patronal", "Cuota patronal"),
                     ("infonavit_patronal", "INFONAVIT patronal"),
                     ("total", "Total")],
    },
    "ptu_estimado": {
        "title": "PTU estimado",
        "scalars": [("anio", "Año"), ("utilidad_repartible", "Utilidad repartible"),
                     ("ptu_pagado", "PTU pagado"), ("excluidos", "Excluidos"),
                     ("status", "Estado")],
    },
    "iva_mes": {
        "title": "IVA del periodo",
        "scalars": [("periodo", "Periodo"), ("trasladado", "IVA trasladado"),
                     ("acreditable", "IVA acreditable"), ("saldo", "Saldo")],
    },
    "lead_time_proveedor": {
        "title": "Lead time por proveedor",
        "scalars": [("promedio_dias", "Promedio (días)"), ("count", "Proveedores")],
        "table": {"key": "mas_lentos", "headers": [
            ("name", "Proveedor", "str"), ("dias", "Días", "int"),
        ]},
    },
    "reordenar_sin_oc": {
        "title": "SKUs por reordenar sin OC abierta",
        "scalars": [("count", "Total")],
        "table": {"key": "items", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("stock", "Stock", "int"), ("reorder", "Reorden", "int"),
        ]},
    },
    "variacion_costo": {
        "title": "Variaciones de costo",
        "scalars": [("count", "Total")],
        "table": {"key": "items", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("anterior", "Costo anterior", "money"),
            ("actual", "Costo actual", "money"),
            ("variacion_pct", "Variación %", "int"),
        ]},
    },
    "top_valor_inmovilizado": {
        "title": "Top valor inmovilizado",
        "table": {"key": "items", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("unidades", "Unidades", "int"), ("valor", "Valor", "money"),
        ]},
    },
    "faltantes_para_pedidos": {
        "title": "Faltantes para pedidos abiertos",
        "scalars": [("count", "Total")],
        "table": {"key": "items", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("requerido", "Requerido", "int"), ("stock", "Stock", "int"),
            ("faltan", "Faltan", "int"),
        ]},
    },
    "descuentos_pos_dia": {
        "title": "Descuentos POS del día",
        "scalars": [("fecha", "Fecha"),
                     ("tickets_con_descuento", "Tickets con descuento"),
                     ("monto_descontado", "Monto descontado")],
    },
    "devoluciones_pos_dia": {
        "title": "Devoluciones POS del día",
        "scalars": [("fecha", "Fecha"), ("count", "Total"),
                     ("monto", "Monto")],
    },
    "cancelaciones_pos_dia": {
        "title": "Cancelaciones del día",
        "scalars": [("fecha", "Fecha"), ("count", "Total"),
                     ("monto", "Monto")],
    },
    "top_producto_pos_dia": {
        "title": "Top productos POS del día",
        "scalars": [("fecha", "Fecha")],
        "table": {"key": "items", "headers": [
            ("name", "Producto", "str"), ("sku", "SKU", "str"),
            ("unidades", "Unidades", "int"), ("revenue", "Ingreso", "money"),
        ]},
    },
    "top_vendedores": {
        "title": "Top vendedores",
        "scalars": [("periodo", "Periodo")],
        "table": {"key": "items", "headers": [
            ("vendedor", "Vendedor", "str"),
            ("pedidos", "Pedidos", "int"),
            ("revenue", "Ingreso", "money"),
        ]},
    },
    "ventas_pos_periodo": {
        "title": "Ventas POS del periodo",
        "scalars": [("periodo", "Periodo"), ("tickets", "Tickets"),
                     ("total", "Total"), ("ticket_promedio", "Ticket promedio")],
    },
    "ventas_cliente": {
        "title": "Ficha por nombre",
        "scalars": [("nombre_busqueda", "Nombre buscado")],
        "tables": [
            {"key": "vendedores", "title": "Como vendedor", "headers": [
                ("nombre", "Vendedor", "str"),
                ("pedidos", "Pedidos", "int"),
                ("total_vendido", "Total vendido", "money"),
                ("ultima_venta", "Última venta", "str"),
            ]},
            {"key": "clientes", "title": "Como cliente", "headers": [
                ("nombre", "Cliente", "str"),
                ("pedidos", "Pedidos", "int"),
                ("total_comprado", "Total comprado", "money"),
                ("saldo_pendiente", "Saldo pendiente", "money"),
                ("ultima_compra", "Última compra", "str"),
            ]},
            {"key": "empleados", "title": "Como empleado", "headers": [
                ("nombre", "Empleado", "str"),
                ("numero", "N° empleado", "str"),
                ("puesto", "Puesto", "str"),
                ("departamento", "Departamento", "str"),
                ("ingreso", "Fecha ingreso", "str"),
                ("estado", "Estado", "str"),
            ]},
        ],
    },
    "ventas_persona": {
        "title": "Ficha por nombre (vendedor / cliente)",
        "scalars": [("nombre_busqueda", "Nombre buscado")],
        # Dos tablas: una por rol. Cada una se imprime solo si tiene
        # filas — así cuando el nombre solo aparece como cliente (ej.
        # Argelia), la tabla de vendedores se omite silenciosamente y
        # el XLSX igual sale con datos.
        "tables": [
            {"key": "vendedores", "title": "Como vendedor", "headers": [
                ("nombre", "Vendedor", "str"),
                ("pedidos", "Pedidos", "int"),
                ("total_vendido", "Total vendido", "money"),
                ("ultima_venta", "Última venta", "str"),
            ]},
            {"key": "clientes", "title": "Como cliente", "headers": [
                ("nombre", "Cliente", "str"),
                ("pedidos", "Pedidos", "int"),
                ("total_comprado", "Total comprado", "money"),
                ("saldo_pendiente", "Saldo pendiente", "money"),
                ("ultima_compra", "Última compra", "str"),
            ]},
            {"key": "empleados", "title": "Como empleado", "headers": [
                ("nombre", "Empleado", "str"),
                ("numero", "N° empleado", "str"),
                ("puesto", "Puesto", "str"),
                ("departamento", "Departamento", "str"),
                ("ingreso", "Fecha ingreso", "str"),
                ("estado", "Estado", "str"),
                # Los siguientes solo tienen valor cuando el usuario tenía
                # permiso HR al pedir la consulta (include_hr_details=True).
                # Si no, quedan vacíos y openpyxl los deja en blanco.
                ("salario_base", "Salario", "money"),
                ("sbc", "SBC", "money"),
                ("frecuencia", "Frecuencia", "str"),
                ("vacaciones_pendientes", "Vacaciones pend.", "int"),
                ("telefono", "Teléfono", "str"),
                ("banco", "Banco", "str"),
                ("clabe", "CLABE", "str"),
                ("deducciones", "Deducciones", "str"),
            ]},
        ],
    },
}


def _tables_of(spec: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Normaliza spec.table (dict, único) o spec.tables (list) → siempre
    lista de dicts. Vacía si no hay ninguna."""
    if spec.get("tables"):
        return list(spec["tables"])
    tbl = spec.get("table")
    return [tbl] if tbl else []


def is_exportable(data: Optional[Dict[str, Any]]) -> bool:
    """True si la respuesta tiene contenido que valga la pena exportar
    (al menos una tabla con al menos una fila, o suficientes escalares
    para armar una hoja significativa). Usado por el frontend para
    decidir mostrar el botón 'Descargar Excel'."""
    if not data or not isinstance(data, dict):
        return False
    tool = data.get("tool")
    spec = _SPEC.get(tool)
    if not spec:
        return False
    for tbl in _tables_of(spec):
        rows = data.get(tbl["key"]) or []
        if isinstance(rows, list) and len(rows) > 0:
            return True
    # Si no hay tabla con filas pero sí escalares, dejamos exportar solo si
    # hay ≥2 escalares con valor no-nulo — evita exportar una sola cifra.
    scalars = spec.get("scalars", [])
    non_null = sum(1 for k, _ in scalars if data.get(k) is not None)
    return non_null >= 2


def _fmt(value: Any, tipo: str) -> Any:
    """Formatea un valor según su tipo. Para money e int devuelve numérico
    para que Excel lo entienda como número (permite sumas y filtros)."""
    if value is None or value == "":
        return ""
    if tipo == "money":
        try:
            return round(float(value), 2)
        except Exception:
            return value
    if tipo == "int":
        try:
            return int(float(value))
        except Exception:
            return value
    if tipo == "pct":
        try:
            return round(float(value), 1)
        except Exception:
            return value
    return str(value)


def _apply_number_format(cell, tipo: str) -> None:
    if tipo == "money":
        cell.number_format = "$#,##0.00"
        cell.alignment = _RIGHT
    elif tipo == "int":
        cell.number_format = "#,##0"
        cell.alignment = _RIGHT
    elif tipo == "pct":
        cell.number_format = "0.0"
        cell.alignment = _RIGHT
    else:
        cell.alignment = _LEFT


def build_xlsx(tool_name: str, data: Dict[str, Any],
                question: Optional[str] = None) -> bytes:
    """Genera un XLSX en memoria y devuelve bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Asistente"
    spec = _SPEC.get(tool_name, {"title": tool_name})

    # Fila 1: título
    ws.cell(row=1, column=1, value=spec.get("title", tool_name)).font = _TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    # Fila 2: pregunta original (si viene) + fecha
    subtitle = f"Emitido: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    if question:
        subtitle = f"Consulta: “{question[:120]}”   ·   {subtitle}"
    ws.cell(row=2, column=1, value=subtitle).font = _SUBTITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)

    row = 4
    # Bloque de escalares (si aplica)
    scalars = spec.get("scalars", [])
    if scalars:
        for key, label in scalars:
            val = data.get(key)
            if val is None:
                continue
            c1 = ws.cell(row=row, column=1, value=label)
            c1.font = Font(bold=True, color="586880")
            c1.alignment = _LEFT
            c2 = ws.cell(row=row, column=2, value=val)
            # Heurística: si el label habla de porcentaje o el valor es
            # numérico con "pct" en el nombre → formato apropiado.
            if isinstance(val, (int, float)):
                if "%" in label.lower() or "pct" in key.lower():
                    c2.number_format = "0.0"
                elif any(w in label.lower() for w in ("total", "monto", "ingreso",
                                                       "saldo", "valor", "bruto",
                                                       "neto", "gasto", "costo",
                                                       "utilidad", "cuota", "esperada",
                                                       "esperados", "reembolsado",
                                                       "descontado", "actual",
                                                       "retenido", "proyección")):
                    c2.number_format = "$#,##0.00"
                else:
                    c2.number_format = "#,##0"
                c2.alignment = _RIGHT
            else:
                c2.alignment = _LEFT
            row += 1
        row += 1  # línea en blanco

    # Bloque(s) de tabla — soporta spec.table (dict, único) y spec.tables (lista).
    for tbl in _tables_of(spec):
        rows_data = data.get(tbl["key"]) or []
        if not rows_data:
            continue
        # Sub-título opcional por tabla (útil cuando hay varias).
        if tbl.get("title"):
            c = ws.cell(row=row, column=1, value=tbl["title"])
            c.font = Font(bold=True, size=11, color="1E3A5F")
            row += 1
        headers = tbl["headers"]
        for i, (_field, label, _tipo) in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = _HEADER_FONT
            c.fill = _HEADER_FILL
            c.alignment = _CENTER
            c.border = _BORDER
        row += 1
        for item in rows_data:
            if not isinstance(item, dict):
                continue
            for i, (field, _label, tipo) in enumerate(headers, 1):
                v = _fmt(item.get(field), tipo)
                c = ws.cell(row=row, column=i, value=v)
                c.font = _CELL_FONT
                c.border = _BORDER
                _apply_number_format(c, tipo)
            row += 1
        row += 1  # separador entre tablas

    # Ajuste de anchos: mide contenido de cada columna, cap 45.
    for col_idx in range(1, ws.max_column + 1):
        max_len = 10
        letter = get_column_letter(col_idx)
        for cell in ws[letter]:
            try:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)) + 2, 45))
            except Exception:
                pass
        ws.column_dimensions[letter].width = max_len

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def safe_filename(tool_name: str) -> str:
    """Nombre de archivo estable: asistente_<tool>_<YYYYMMDD>.xlsx"""
    return f"asistente_{tool_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
