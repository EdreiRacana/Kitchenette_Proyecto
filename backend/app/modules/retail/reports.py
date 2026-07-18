"""Generadores de reportes Excel + PDF del módulo Retail.

Todos los archivos comparten un header con branding (color, logo, empresa,
periodo) y estilo consistente (Helvetica-Bold para totales, backgrounds
por status, freeze panes).
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


BRAND_BLUE = "1E3A8A"
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor=BRAND_BLUE)
TOTAL_FONT = Font(bold=True, size=11)
GREY_ROW = PatternFill("solid", fgColor="F1F5F9")
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")

# Fills por status WOS
STATUS_FILL = {
    "critical": PatternFill("solid", fgColor="FEE2E2"),
    "replenish": PatternFill("solid", fgColor="FEF3C7"),
    "healthy": PatternFill("solid", fgColor="D1FAE5"),
    "overstock": PatternFill("solid", fgColor="DBEAFE"),
    "no_data": PatternFill("solid", fgColor="F3F4F6"),
}
STATUS_LABEL = {
    "critical": "Crítico",
    "replenish": "Resurtir",
    "healthy": "Sano",
    "overstock": "Sobreinventario",
    "no_data": "Sin datos",
}
SEV_FILL = {
    "urgent": PatternFill("solid", fgColor="FEE2E2"),
    "high": PatternFill("solid", fgColor="FEF3C7"),
    "medium": PatternFill("solid", fgColor="DBEAFE"),
    "low": PatternFill("solid", fgColor="F3F4F6"),
}


def _style_header(ws, ncols: int) -> None:
    for i in range(1, ncols + 1):
        c = ws.cell(row=1, column=i)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    ws.freeze_panes = "A2"


def _autosize(ws, min_w: int = 10, max_w: int = 42) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = max(min_w, min(length + 2, max_w))


def _company_header(ws, title: str, subtitle: str, ncols: int) -> None:
    """Fila 1 con nombre del reporte + fecha de emisión (grey)."""
    ws.insert_rows(1)
    ws.insert_rows(1)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14, color=BRAND_BLUE)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    cell.alignment = LEFT
    sub = ws.cell(row=2, column=1,
                    value=f"{subtitle} · Emitido {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    sub.font = Font(size=9, color="64748B")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws.freeze_panes = "A4"


def _to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Reporte 1: Sell-out crudo ────────────────────────────────────────────

def build_sellout_report(rows: List[Any]) -> bytes:
    """rows son SellOutReportOut (o similares) — se usa como filas de export."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sell-out"
    headers = [
        "Cadena", "Tienda", "SKU", "Producto",
        "Tipo periodo", "Inicio", "Fin",
        "Unidades vendidas", "Devueltas", "Netas",
        "Stock final", "Ingreso", "Devoluciones $", "Ingreso neto",
        "% Devoluciones",
        "Fuente", "Notas",
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    total_units = 0
    total_returned = 0
    total_stock = 0
    total_rev = 0.0
    total_ret_amt = 0.0
    for r in rows:
        u = int(r.units_sold or 0)
        ru = int(getattr(r, "units_returned", 0) or 0)
        net_u = max(u - ru, 0)
        rev = float(r.revenue or 0.0)
        ret_amt = float(getattr(r, "returns_amount", 0.0) or 0.0)
        net_rev = round(max(rev - ret_amt, 0.0), 2)
        ret_pct = round((ru / u * 100.0), 2) if u > 0 else 0.0
        total_units += u
        total_returned += ru
        total_stock += int(r.units_on_hand or 0)
        total_rev += rev
        total_ret_amt += ret_amt
        ws.append([
            r.channel_name or "", r.store_name or "",
            r.sku or "", r.product_name or "",
            r.period_type or "",
            r.period_start.strftime("%Y-%m-%d") if r.period_start else "",
            r.period_end.strftime("%Y-%m-%d") if r.period_end else "",
            u, ru, net_u,
            int(r.units_on_hand or 0),
            rev, ret_amt, net_rev,
            ret_pct,
            r.source or "",
            r.notes or "",
        ])

    # Fila de totales
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    net_units_total = max(total_units - total_returned, 0)
    net_rev_total = round(max(total_rev - total_ret_amt, 0.0), 2)
    ret_pct_total = round((total_returned / total_units * 100.0), 2) if total_units > 0 else 0.0
    totals_by_col = {
        8: total_units, 9: total_returned, 10: net_units_total,
        11: total_stock, 12: round(total_rev, 2),
        13: round(total_ret_amt, 2), 14: net_rev_total,
        15: ret_pct_total,
    }
    for col, val in totals_by_col.items():
        c = ws.cell(row=total_row, column=col, value=val)
        c.font = TOTAL_FONT
        c.fill = GREY_ROW

    _autosize(ws)
    _company_header(
        ws, "Reporte de Sell-out",
        f"{len(rows)} filas · {total_units:,} u vendidas · {total_returned:,} devueltas "
        f"({ret_pct_total:.1f}%) · $ {total_rev:,.2f} bruto · $ {net_rev_total:,.2f} neto",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 2: Dashboard KPIs + Stores + SKUs ────────────────────────────

def build_dashboard_report(kpis: Any, stores: List[Any], skus: List[Any]) -> bytes:
    wb = Workbook()

    # Hoja 1: KPIs
    ws = wb.active
    ws.title = "KPIs"
    ws.append(["Métrica", "Valor"])
    _style_header(ws, 2)
    ws.append(["Sell-out unidades", kpis.total_sell_out_units])
    ws.append(["Sell-out ingreso", round(kpis.total_sell_out_revenue, 2)])
    ws.append(["Devoluciones unidades", getattr(kpis, "total_returns_units", 0)])
    ws.append(["Devoluciones importe", round(getattr(kpis, "total_returns_amount", 0.0), 2)])
    ws.append(["Tasa devoluciones (%)", getattr(kpis, "return_rate_pct", 0.0)])
    ws.append(["Unidades netas", getattr(kpis, "net_units", 0)])
    ws.append(["Ingreso neto", round(getattr(kpis, "net_revenue", 0.0), 2)])
    ws.append(["Sell-in unidades", kpis.total_sell_in_units])
    ws.append(["Sell-in ingreso", round(kpis.total_sell_in_revenue, 2)])
    ws.append(["Sell-through (%)", kpis.sell_through_pct])
    ws.append(["Stock on-hand total", kpis.total_on_hand])
    ws.append(["WOS promedio (sem)", kpis.avg_wos_weeks])
    ws.append(["Tiendas críticas", kpis.critical_stores_count])
    ws.append(["Tiendas sobreinventario", kpis.overstock_stores_count])
    ws.append(["Tiendas activas", kpis.stores_active_count])
    ws.append(["SKUs activos", kpis.skus_active_count])
    _autosize(ws)
    _company_header(
        ws, "Dashboard Retail",
        f"Cadena: {kpis.channel_name or 'Todas'} · Periodo {kpis.period_start.strftime('%d/%m/%Y')} → {kpis.period_end.strftime('%d/%m/%Y')}",
        2,
    )

    # Hoja 2: Tiendas por WOS
    ws2 = wb.create_sheet("Tiendas por WOS")
    ws2.append(["Tienda", "Cadena", "Vendidas (4 sem)", "Vel semanal",
                 "Stock", "WOS", "Status"])
    _style_header(ws2, 7)
    for s in stores:
        row = [
            s.store_name, s.channel_name or "",
            int(s.total_units_sold or 0), float(s.avg_weekly_units or 0),
            int(s.total_on_hand or 0), float(s.wos_weeks or 0),
            STATUS_LABEL.get(s.status, s.status),
        ]
        ws2.append(row)
        fill = STATUS_FILL.get(s.status)
        if fill:
            for col in range(1, len(row) + 1):
                ws2.cell(row=ws2.max_row, column=col).fill = fill
    _autosize(ws2)

    # Hoja 3: SKUs velocidad
    ws3 = wb.create_sheet("SKUs velocidad")
    ws3.append(["SKU", "Producto", "Tiendas",
                 "Vendidas (4 sem)", "Vel semanal", "Stock", "WOS", "Status"])
    _style_header(ws3, 8)
    for k in skus:
        row = [
            k.sku or "", k.product_name or "", k.stores_count,
            int(k.total_units_sold or 0), float(k.avg_weekly_units or 0),
            int(k.total_on_hand or 0), float(k.wos_weeks or 0),
            STATUS_LABEL.get(k.status, k.status),
        ]
        ws3.append(row)
        fill = STATUS_FILL.get(k.status)
        if fill:
            for col in range(1, len(row) + 1):
                ws3.cell(row=ws3.max_row, column=col).fill = fill
    _autosize(ws3)

    return _to_bytes(wb)


# ── Reporte 3: Reabasto (sugerencias) ────────────────────────────────────

PRIO_FILL = {
    "urgent": PatternFill("solid", fgColor="FEE2E2"),
    "high":   PatternFill("solid", fgColor="FEF3C7"),
    "normal": PatternFill("solid", fgColor="DBEAFE"),
}
PRIO_LABEL = {"urgent": "Urgente", "high": "Alta", "normal": "Normal"}


def build_replenishment_report(resp: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reabasto"
    headers = ["Prioridad", "Cadena", "Tienda", "SKU", "Producto",
                "Stock", "Vel. semanal", "WOS", "Sugerido", "Motivo"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for s in resp.suggestions:
        row = [
            PRIO_LABEL.get(s.priority, s.priority),
            s.channel_name, s.store_name,
            s.sku or "", s.product_name or "",
            int(s.current_on_hand or 0),
            float(s.avg_weekly_units or 0),
            float(s.wos_weeks or 0),
            int(s.suggested_units or 0),
            s.reason,
        ]
        ws.append(row)
        fill = PRIO_FILL.get(s.priority)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    _autosize(ws)
    _company_header(
        ws, "Sugerencias de reabasto",
        f"Urgentes {resp.urgent_count} · Alta {resp.high_count} · Normal {resp.normal_count} · Meta {resp.target_wos_weeks} sem",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 4: Heatmap tiendas × SKUs ────────────────────────────────────

def build_heatmap_report(hm: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Heatmap"

    # Fila header: tienda + una col por variant
    headers = ["Tienda", "Cadena"] + [
        (v.sku or v.product_name or "—")[:20] for v in hm.variants
    ]
    ws.append(headers)
    _style_header(ws, len(headers))

    # Index cells por store x variant
    cells_by = {}
    for c in hm.cells:
        cells_by[(c.store_id, c.variant_id)] = c

    metric = hm.metric

    for s in hm.stores:
        row = [s.name, s.channel_name or ""]
        for v in hm.variants:
            c = cells_by.get((s.id, v.id))
            if c is None:
                row.append("")
                continue
            if metric == "units_sold":
                row.append(int(c.units_sold or 0))
            elif metric == "on_hand":
                row.append(int(c.on_hand or 0))
            else:
                row.append(float(c.value) if c.value is not None else "∞")
        ws.append(row)
        r_idx = ws.max_row
        for i, v in enumerate(hm.variants, start=3):
            c = cells_by.get((s.id, v.id))
            if c is None:
                continue
            fill = STATUS_FILL.get(c.status)
            if fill:
                ws.cell(row=r_idx, column=i).fill = fill

    _autosize(ws, min_w=6, max_w=16)
    _company_header(
        ws, f"Heatmap Retail ({metric})",
        f"{len(hm.stores)} tiendas × {len(hm.variants)} SKUs",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 5: Clasificación ABC ─────────────────────────────────────────

ABC_FILL = {
    "A": PatternFill("solid", fgColor="D1FAE5"),
    "B": PatternFill("solid", fgColor="DBEAFE"),
    "C": PatternFill("solid", fgColor="F3F4F6"),
}


def build_abc_report(abc: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ABC"
    headers = ["Rank", "SKU", "Producto", "Tiendas",
                "Unidades", "Ingreso", "%", "% Acumulado", "Clase"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in abc.rows:
        row = [
            r.rank, r.sku or "", r.product_name or "",
            r.stores_count, r.total_units, r.total_revenue,
            r.revenue_pct, r.cumulative_pct, r.abc_class,
        ]
        ws.append(row)
        fill = ABC_FILL.get(r.abc_class)
        if fill:
            ws.cell(row=ws.max_row, column=9).fill = fill

    _autosize(ws)
    _company_header(
        ws, "Clasificación ABC (Pareto)",
        f"A: {abc.class_a_count} · B: {abc.class_b_count} · C: {abc.class_c_count} · Total $ {abc.total_revenue:,.2f}",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 6: Alertas ───────────────────────────────────────────────────

SEV_LABEL = {"urgent": "Urgente", "high": "Alta", "medium": "Media", "low": "Baja"}
ALERT_TYPE_LABEL = {
    "stockout_imminent": "Stock crítico",
    "stockout": "Sin stock",
    "overstock": "Sobreinventario",
    "no_movement": "Sin movimiento",
    "sell_through_low": "Sell-through bajo",
    "high_return_rate": "Devoluciones altas",
}


def build_alerts_report(alerts: List[Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Alertas"
    headers = ["Severidad", "Tipo", "Cadena", "Tienda", "SKU", "Producto",
                "Stock", "Vel sem", "WOS", "Mensaje", "Estado", "Creada"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for a in alerts:
        row = [
            SEV_LABEL.get(a.severity, a.severity),
            ALERT_TYPE_LABEL.get(a.alert_type, a.alert_type),
            a.channel_name or "", a.store_name or "",
            a.sku or "", a.product_name or "",
            int(a.on_hand_snapshot) if a.on_hand_snapshot is not None else "",
            float(a.weekly_velocity_snapshot) if a.weekly_velocity_snapshot is not None else "",
            float(a.wos_snapshot) if a.wos_snapshot is not None else "",
            a.message,
            a.status,
            a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "",
        ]
        ws.append(row)
        fill = SEV_FILL.get(a.severity)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    _autosize(ws)
    _company_header(
        ws, "Alertas de Retail",
        f"{len(alerts)} alertas en el filtro solicitado",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 7: Reconciliación de consignación ───────────────────────────

RECON_FILL = {
    "match":              PatternFill("solid", fgColor="D1FAE5"),
    "short_at_warehouse": PatternFill("solid", fgColor="FEE2E2"),
    "over_at_warehouse":  PatternFill("solid", fgColor="DBEAFE"),
    "no_data":            PatternFill("solid", fgColor="F3F4F6"),
}
RECON_LABEL = {
    "match": "Cuadra",
    "short_at_warehouse": "Faltante en tu almacén",
    "over_at_warehouse": "Sobrante en tu almacén",
    "no_data": "Sin datos",
}


def build_consignment_report(recon: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Consignación"
    headers = ["Tienda", "Cadena", "Almacén", "SKU", "Producto",
                "Reportado tienda", "Fecha reporte",
                "En almacén", "Diferencia", "Status"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in recon.rows:
        row = [
            r.store_name, r.channel_name or "", r.warehouse_name,
            r.sku or "", r.product_name or "",
            int(r.reported_on_hand or 0),
            r.reported_at.strftime("%Y-%m-%d") if r.reported_at else "",
            int(r.warehouse_stock or 0),
            int(r.difference or 0),
            RECON_LABEL.get(r.status, r.status),
        ]
        ws.append(row)
        fill = RECON_FILL.get(r.status)
        if fill:
            ws.cell(row=ws.max_row, column=10).fill = fill

    _autosize(ws)
    _company_header(
        ws, "Reconciliación de consignación",
        f"Total {recon.total_rows} · Cuadran {recon.matched} · Con descuadre {recon.with_diff}",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 8: Tendencia (time-series) ──────────────────────────────────

def build_trend_report(tr: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Tendencia"
    headers = ["Periodo", "Inicio", "Fin", "Vendidas", "Devueltas", "Netas",
                "Ingreso", "Devoluciones $", "Ingreso neto",
                "On-hand", "Tiendas"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for p in tr.points:
        ws.append([
            p.label,
            p.period_start.strftime("%Y-%m-%d") if p.period_start else "",
            p.period_end.strftime("%Y-%m-%d") if p.period_end else "",
            int(p.units_sold or 0), int(p.units_returned or 0), int(p.net_units or 0),
            float(p.revenue or 0.0), float(p.returns_amount or 0.0), float(p.net_revenue or 0.0),
            int(p.on_hand or 0), int(p.stores_reporting or 0),
        ])

    wow = ""
    if tr.wow_units_pct is not None:
        wow = f" · Δ unidades últ. periodo {tr.wow_units_pct:+.1f}%"
    _autosize(ws)
    _company_header(
        ws, "Tendencia de Sell-out",
        f"{len(tr.points)} periodos ({tr.period_type}) · {tr.total_units:,} u · $ {tr.total_revenue:,.2f}{wow}",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 9: Distribución numérica (voids) ────────────────────────────

DIST_FILL = {
    "excellent": PatternFill("solid", fgColor="D1FAE5"),
    "good":      PatternFill("solid", fgColor="DBEAFE"),
    "low":       PatternFill("solid", fgColor="FEF3C7"),
    "critical":  PatternFill("solid", fgColor="FEE2E2"),
}
DIST_LABEL = {
    "excellent": "Excelente", "good": "Buena", "low": "Baja", "critical": "Crítica",
}


def build_distribution_report(dist: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Distribución"
    headers = ["SKU", "Producto", "Tiendas que venden", "Con stock",
                "Total tiendas", "Distribución %", "Voids (huecos)",
                "Unidades", "Prom u/tienda", "Nivel"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in dist.rows:
        ws.append([
            r.sku or "", r.product_name or "",
            int(r.stores_selling or 0), int(r.stores_stocking or 0),
            int(r.total_stores or 0), float(r.distribution_pct or 0.0),
            int(r.void_stores or 0), int(r.total_units or 0),
            float(r.avg_units_per_store or 0.0),
            DIST_LABEL.get(r.status, r.status),
        ])
        fill = DIST_FILL.get(r.status)
        if fill:
            ws.cell(row=ws.max_row, column=10).fill = fill

    _autosize(ws)
    _company_header(
        ws, "Distribución numérica por SKU",
        f"{len(dist.rows)} SKUs · {dist.total_stores} tiendas activas · "
        f"un 'void' es una tienda que aún no vende el producto",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 10: Venta perdida por stockout ──────────────────────────────

def build_lost_sales_report(ls: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Venta perdida"
    headers = ["Severidad", "Cadena", "Tienda", "SKU", "Producto",
                "Vel. semanal", "Sem. sin stock", "Unidades perdidas",
                "Precio", "Venta perdida $"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in ls.rows:
        ws.append([
            SEV_LABEL.get(r.severity, r.severity),
            r.channel_name or "", r.store_name or "",
            r.sku or "", r.product_name or "",
            float(r.avg_weekly_units or 0.0), float(r.weeks_out_of_stock or 0.0),
            int(r.lost_units or 0), float(r.unit_price or 0.0),
            float(r.lost_revenue or 0.0),
        ])
        fill = SEV_FILL.get(r.severity)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    # Fila total
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="TOTAL").font = TOTAL_FONT
    ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
    c8 = ws.cell(row=total_row, column=8, value=int(ls.total_lost_units or 0))
    c8.font = TOTAL_FONT; c8.fill = GREY_ROW
    c10 = ws.cell(row=total_row, column=10, value=round(float(ls.total_lost_revenue or 0.0), 2))
    c10.font = TOTAL_FONT; c10.fill = GREY_ROW

    _autosize(ws)
    _company_header(
        ws, "Venta perdida por agotados",
        f"{ls.affected_combos} combos agotados · {ls.total_lost_units:,} u perdidas · "
        f"$ {ls.total_lost_revenue:,.2f} sin vender",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 11: Rentabilidad (márgenes + GMROI) ─────────────────────────

def _margin_fill(margin_pct: float) -> Optional[PatternFill]:
    if margin_pct >= 35:
        return PatternFill("solid", fgColor="D1FAE5")   # verde
    if margin_pct >= 20:
        return PatternFill("solid", fgColor="DBEAFE")   # azul
    if margin_pct >= 8:
        return PatternFill("solid", fgColor="FEF3C7")   # ámbar
    return PatternFill("solid", fgColor="FEE2E2")       # rojo


GROUP_BY_LABEL = {
    "sku": "SKU", "category": "Categoría", "store": "Tienda", "channel": "Cadena",
}


def build_profitability_report(prof: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rentabilidad"
    dim = GROUP_BY_LABEL.get(prof.group_by, "Dimensión")
    is_sku = prof.group_by == "sku"
    headers = ([dim] + (["Producto"] if is_sku else []) +
               ["Unidades", "Ingreso", "COGS", "Margen bruto", "Margen %",
                "Inv. a costo", "GMROI"])
    ws.append(headers)
    _style_header(ws, len(headers))

    margin_col = len(headers) - 2  # columna "Margen %"
    for r in prof.rows:
        row = [r.dimension_label]
        if is_sku:
            row.append(r.product_name or "")
        row += [
            int(r.units_sold or 0), float(r.revenue or 0.0),
            float(r.cogs or 0.0), float(r.gross_margin or 0.0),
            float(r.margin_pct or 0.0), float(r.inventory_cost or 0.0),
            float(r.gmroi) if r.gmroi is not None else "",
        ]
        ws.append(row)
        fill = _margin_fill(float(r.margin_pct or 0.0))
        if fill:
            ws.cell(row=ws.max_row, column=margin_col).fill = fill

    # Fila total
    tr = ws.max_row + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = TOTAL_FONT
    span_end = 2 if is_sku else 1
    if is_sku:
        ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=2)
    base = 2 if is_sku else 1
    totals = [
        int(prof.total_units or 0), float(prof.total_revenue or 0.0),
        float(prof.total_cogs or 0.0), float(prof.total_gross_margin or 0.0),
        float(prof.total_margin_pct or 0.0), float(prof.total_inventory_cost or 0.0),
        float(prof.total_gmroi) if prof.total_gmroi is not None else "",
    ]
    for i, val in enumerate(totals):
        c = ws.cell(row=tr, column=base + 1 + i, value=val)
        c.font = TOTAL_FONT
        c.fill = GREY_ROW

    _autosize(ws)
    warn = ""
    if prof.variants_without_cost > 0:
        warn = f" · ⚠ {prof.variants_without_cost} SKUs sin costo (margen subestimado)"
    _company_header(
        ws, f"Rentabilidad por {dim}",
        f"Margen bruto $ {prof.total_gross_margin:,.2f} ({prof.total_margin_pct:.1f}%) · "
        f"GMROI {prof.total_gmroi if prof.total_gmroi is not None else '—'} · "
        f"Últimos {prof.days} días{warn}",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 12: Exceso de inventario + rotación ─────────────────────────

def build_excess_inventory_report(exc: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Exceso inventario"
    headers = ["Severidad", "Cadena", "Tienda", "SKU", "Producto",
                "On-hand", "Vel. sem", "WOS", "DOH (días)",
                "Exceso u.", "Costo u.", "Exceso $", "Dead stock"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in exc.rows:
        ws.append([
            SEV_LABEL.get(r.severity, r.severity),
            r.channel_name or "", r.store_name or "",
            r.sku or "", r.product_name or "",
            int(r.on_hand or 0), float(r.avg_weekly_units or 0.0),
            float(r.wos_weeks) if r.wos_weeks is not None else "∞",
            float(r.doh_days) if r.doh_days is not None else "∞",
            int(r.excess_units or 0), float(r.unit_cost or 0.0),
            float(r.excess_cost or 0.0),
            "Sí" if r.is_dead_stock else "",
        ])
        fill = SEV_FILL.get(r.severity)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill

    tr = ws.max_row + 1
    ws.cell(row=tr, column=1, value="TOTAL").font = TOTAL_FONT
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=9)
    c10 = ws.cell(row=tr, column=10, value=int(exc.total_excess_units or 0))
    c10.font = TOTAL_FONT; c10.fill = GREY_ROW
    c12 = ws.cell(row=tr, column=12, value=round(float(exc.total_excess_cost or 0.0), 2))
    c12.font = TOTAL_FONT; c12.fill = GREY_ROW

    _autosize(ws)
    turn = f"{exc.inventory_turnover:.2f}" if exc.inventory_turnover is not None else "—"
    doh = f"{exc.days_of_inventory:.0f}" if exc.days_of_inventory is not None else "—"
    _company_header(
        ws, "Exceso de inventario y rotación",
        f"$ {exc.total_excess_cost:,.2f} detenido en exceso "
        f"(dead stock $ {exc.dead_stock_cost:,.2f}) · "
        f"Inventario a costo $ {exc.total_inventory_cost:,.2f} · "
        f"Rotación {turn}x/año · Días de inventario {doh}",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 13: Antigüedad de inventario (aging) ────────────────────────

AGING_FILL = {
    "0-30":  PatternFill("solid", fgColor="D1FAE5"),
    "31-60": PatternFill("solid", fgColor="DBEAFE"),
    "61-90": PatternFill("solid", fgColor="FEF3C7"),
    "90+":   PatternFill("solid", fgColor="FEE2E2"),
    "never": PatternFill("solid", fgColor="FECACA"),
}


def build_aging_report(aging: Any) -> bytes:
    wb = Workbook()
    # Hoja 1: resumen por cubeta
    ws = wb.active
    ws.title = "Resumen"
    ws.append(["Antigüedad", "Unidades", "Valor a costo", "% del valor"])
    _style_header(ws, 4)
    for b in aging.buckets:
        ws.append([b.label, int(b.units or 0), float(b.value or 0.0), float(b.pct_of_value or 0.0)])
        fill = AGING_FILL.get(b.bucket)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill
    _autosize(ws)
    _company_header(
        ws, "Antigüedad de inventario",
        f"Inventario a costo $ {aging.total_stock_value:,.2f} · "
        f"En riesgo de obsolescencia $ {aging.obsolete_value:,.2f} ({aging.obsolete_pct:.1f}%)",
        4,
    )

    # Hoja 2: detalle por SKU
    ws2 = wb.create_sheet("Detalle")
    headers = ["Antigüedad", "Cadena", "Tienda", "SKU", "Producto",
                "On-hand", "Última venta", "Días sin vender", "Costo u.", "Valor"]
    ws2.append(headers)
    _style_header(ws2, len(headers))
    bucket_label = {b.bucket: b.label for b in aging.buckets}
    for r in aging.rows:
        ws2.append([
            bucket_label.get(r.bucket, r.bucket),
            r.channel_name or "", r.store_name or "",
            r.sku or "", r.product_name or "",
            int(r.on_hand or 0),
            r.last_sale_date.strftime("%Y-%m-%d") if r.last_sale_date else "Nunca",
            r.days_since_last_sale if r.days_since_last_sale is not None else "—",
            float(r.unit_cost or 0.0), float(r.stock_value or 0.0),
        ])
        fill = AGING_FILL.get(r.bucket)
        if fill:
            ws2.cell(row=ws2.max_row, column=1).fill = fill
    _autosize(ws2)
    return _to_bytes(wb)


# ── Reporte 14: Nivel de servicio / fill rate ───────────────────────────

SL_FILL = {
    "excellent": PatternFill("solid", fgColor="D1FAE5"),
    "good":      PatternFill("solid", fgColor="DBEAFE"),
    "low":       PatternFill("solid", fgColor="FEF3C7"),
    "critical":  PatternFill("solid", fgColor="FEE2E2"),
}
SL_LABEL = {"excellent": "Excelente", "good": "Bueno", "low": "Bajo", "critical": "Crítico"}
SL_DIM = {"store": "Tienda", "sku": "SKU", "channel": "Cadena"}


def build_service_level_report(sl: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Nivel de servicio"
    dim = SL_DIM.get(sl.group_by, "Dimensión")
    is_sku = sl.group_by == "sku"
    headers = ([dim] + (["Producto"] if is_sku else []) +
               ["Observaciones", "Con stock", "In-stock %",
                "Vendidas", "Perdidas est.", "Fill rate %", "Nivel"])
    ws.append(headers)
    _style_header(ws, len(headers))

    status_col = len(headers)
    for r in sl.rows:
        row = [r.dimension_label]
        if is_sku:
            row.append(r.product_name or "")
        row += [
            int(r.total_periods or 0), int(r.in_stock_periods or 0),
            float(r.in_stock_rate_pct or 0.0),
            int(r.units_sold or 0), int(r.estimated_lost_units or 0),
            float(r.fill_rate_pct or 0.0),
            SL_LABEL.get(r.status, r.status),
        ]
        ws.append(row)
        fill = SL_FILL.get(r.status)
        if fill:
            ws.cell(row=ws.max_row, column=status_col).fill = fill

    _autosize(ws)
    _company_header(
        ws, f"Nivel de servicio por {dim}",
        f"In-stock (OSA) {sl.overall_in_stock_rate_pct:.1f}% · "
        f"Quiebre {sl.overall_stockout_rate_pct:.1f}% · "
        f"Fill rate {sl.overall_fill_rate_pct:.1f}% · "
        f"{sl.total_estimated_lost:,} u perdidas est. · Últimas {sl.weeks_back} sem",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte 15: ABC-XYZ (segmentación de surtido) ───────────────────────

_ABCXYZ_CELL_FILL = {
    "A": PatternFill("solid", fgColor="D1FAE5"),
    "B": PatternFill("solid", fgColor="DBEAFE"),
    "C": PatternFill("solid", fgColor="F3F4F6"),
}


def build_abc_xyz_report(axz: Any) -> bytes:
    wb = Workbook()
    # Hoja 1: matriz 3×3 (conteo + % de ingreso por celda)
    ws = wb.active
    ws.title = "Matriz ABC-XYZ"
    ws.append(["", "X · estable", "Y · variable", "Z · errático"])
    _style_header(ws, 4)
    cells = {c.combined: c for c in axz.matrix}
    for a in ("A", "B", "C"):
        row = [f"Clase {a}"]
        for x in ("X", "Y", "Z"):
            c = cells.get(a + x)
            row.append(f"{c.count} SKUs · {c.revenue_pct:.1f}% ingreso" if c else "—")
        ws.append(row)
        fill = _ABCXYZ_CELL_FILL.get(a)
        if fill:
            ws.cell(row=ws.max_row, column=1).fill = fill
    _autosize(ws)
    _company_header(
        ws, "Matriz ABC-XYZ",
        f"ABC por facturación × XYZ por variabilidad de demanda · "
        f"{len(axz.rows)} SKUs · {axz.weeks} semanas · Total $ {axz.total_revenue:,.2f}",
        4,
    )

    # Hoja 2: detalle por SKU
    ws2 = wb.create_sheet("Detalle")
    headers = ["Clase", "SKU", "Producto", "Unidades", "Ingreso", "% Ingreso",
                "% Acum", "Prom sem", "CV", "ABC", "XYZ", "Estrategia"]
    ws2.append(headers)
    _style_header(ws2, len(headers))
    for r in axz.rows:
        ws2.append([
            r.combined_class, r.sku or "", r.product_name or "",
            int(r.total_units or 0), float(r.total_revenue or 0.0),
            float(r.revenue_pct or 0.0), float(r.cumulative_pct or 0.0),
            float(r.avg_weekly_units or 0.0),
            float(r.cv) if r.cv is not None else "∞",
            r.abc_class, r.xyz_class, r.strategy,
        ])
        fill = _ABCXYZ_CELL_FILL.get(r.abc_class)
        if fill:
            ws2.cell(row=ws2.max_row, column=1).fill = fill
    _autosize(ws2)
    return _to_bytes(wb)


# ── Reporte 16: Inteligencia de precios ─────────────────────────────────

_ELASTICITY_LABEL = {
    "elastic": "Elástico", "inelastic": "Inelástico", "unit": "Unitario", "n/a": "—",
}


def build_pricing_report(pr: Any) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Precios"
    headers = ["SKU", "Producto", "Unidades", "Precio prom", "Mín", "Máx",
                "Volatilidad %", "Cambio %", "Precio lista", "Elasticidad", "Tipo"]
    ws.append(headers)
    _style_header(ws, len(headers))

    for r in pr.rows:
        ws.append([
            r.sku or "", r.product_name or "",
            int(r.units_sold or 0), float(r.avg_price or 0.0),
            float(r.min_price or 0.0), float(r.max_price or 0.0),
            float(r.price_volatility_pct or 0.0), float(r.price_change_pct or 0.0),
            float(r.list_price) if r.list_price is not None else "",
            float(r.elasticity) if r.elasticity is not None else "",
            _ELASTICITY_LABEL.get(r.elasticity_label, r.elasticity_label),
        ])

    _autosize(ws)
    _company_header(
        ws, "Inteligencia de precios",
        f"{len(pr.rows)} SKUs · precio implícito (ingreso ÷ unidades) · "
        f"elasticidad estimada · últimos {pr.days} días",
        len(headers),
    )
    return _to_bytes(wb)


# ── Reporte ejecutivo PDF semanal ────────────────────────────────────────

def build_executive_pdf(
    company: dict, kpis: Any, stores: List[Any], skus: List[Any],
    alerts: List[Any], repl: Any,
) -> bytes:
    """Reporte ejecutivo semanal para el gerente comercial.

    Estructura:
      Header con logo y branding.
      Bloque KPIs: sell-in, sell-out, sell-through, WOS, tiendas críticas.
      Top 5 tiendas y bottom 5 (por WOS).
      Top 10 SKUs por unidades vendidas.
      Alertas urgentes activas.
      Top sugerencias de reabasto.
      Firmas.
    """
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.units import mm
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        KeepTogether,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=LETTER,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"Reporte ejecutivo Retail",
    )
    styles = getSampleStyleSheet()
    story = []

    brand_hex = company.get("brand_color") or "#1E3A8A"
    brand = colors.HexColor(brand_hex)
    grey_bg = colors.HexColor("#F0F3F8")
    grey_border = colors.HexColor("#CCCCCC")
    grey_grid = colors.HexColor("#EEEEEE")
    label_grey = colors.HexColor("#666666")

    h1 = ParagraphStyle("h1", parent=styles["Normal"], fontSize=10.5,
                        textColor=brand, spaceAfter=2, fontName="Helvetica-Bold")

    # Header
    empresa_lines = []
    if company.get("commercial_name"):
        empresa_lines.append(f"<b>{company['commercial_name']}</b>")
    if company.get("legal_name"):
        empresa_lines.append(company["legal_name"])
    p_empresa = Paragraph("<br/>".join(empresa_lines), ParagraphStyle(
        "e", parent=styles["Normal"], fontSize=8.5, leading=11,
        textColor=colors.HexColor("#333")))
    p_titulo = Paragraph(
        f"<font size='18' color='{brand_hex}'><b>REPORTE EJECUTIVO</b></font><br/>"
        f"<font size='9' color='#666'>Retail Analytics</font>",
        ParagraphStyle("t", parent=styles["Normal"], alignment=TA_RIGHT))
    header = Table([[p_empresa, p_titulo]], colWidths=[110 * mm, 70 * mm])
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LINEBELOW", (0, 0), (-1, 0), 1.5, brand),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
    ]))
    story.append(header)
    story.append(Spacer(1, 4 * mm))

    # Meta
    meta = [
        ["Cadena:", kpis.channel_name or "Todas",
         "Periodo:", f"{kpis.period_start.strftime('%d/%m/%Y')} → {kpis.period_end.strftime('%d/%m/%Y')}"],
        ["Emitido:", datetime.now().strftime("%d/%m/%Y %H:%M"),
         "Ventana KPI:", "Últimas 4 semanas"],
    ]
    t_meta = Table(meta, colWidths=[22 * mm, 65 * mm, 22 * mm, 65 * mm])
    t_meta.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("TEXTCOLOR", (0, 0), (0, -1), label_grey),
        ("TEXTCOLOR", (2, 0), (2, -1), label_grey),
        ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (3, -1), "Helvetica-Bold"),
    ]))
    story.append(t_meta)
    story.append(Spacer(1, 5 * mm))

    # KPIs
    def _mxn(n: float) -> str:
        return "$" + f"{n:,.2f}"

    ret_units = int(getattr(kpis, "total_returns_units", 0) or 0)
    ret_amt = float(getattr(kpis, "total_returns_amount", 0.0) or 0.0)
    ret_pct = float(getattr(kpis, "return_rate_pct", 0.0) or 0.0)
    net_units = int(getattr(kpis, "net_units", 0) or 0)
    net_rev = float(getattr(kpis, "net_revenue", 0.0) or 0.0)
    kpi_rows = [
        ["Indicador", "Valor"],
        ["Sell-out (unidades)", f"{kpis.total_sell_out_units:,}"],
        ["Sell-out (ingreso)", _mxn(kpis.total_sell_out_revenue)],
        ["Devoluciones (unidades)", f"{ret_units:,}"],
        ["Devoluciones (importe)", _mxn(ret_amt)],
        ["Tasa de devoluciones", f"{ret_pct:.1f}%"],
        ["Unidades netas", f"{net_units:,}"],
        ["Ingreso neto", _mxn(net_rev)],
        ["Sell-in (unidades)", f"{kpis.total_sell_in_units:,}"],
        ["Sell-in (ingreso)", _mxn(kpis.total_sell_in_revenue)],
        ["Sell-through %", f"{kpis.sell_through_pct:.1f}%"],
        ["Stock on-hand total", f"{kpis.total_on_hand:,}"],
        ["WOS promedio", f"{kpis.avg_wos_weeks:.1f} sem"],
        ["Tiendas en crítico", f"{kpis.critical_stores_count}"],
        ["Tiendas en sobreinventario", f"{kpis.overstock_stores_count}"],
        ["Tiendas activas", f"{kpis.stores_active_count}"],
        ["SKUs activos", f"{kpis.skus_active_count}"],
    ]
    story.append(Paragraph("KPIs DEL PERIODO", h1))
    t_kpi = Table(kpi_rows, colWidths=[110 * mm, 60 * mm])
    t_kpi.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), grey_bg),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.4, grey_border),
        ("INNERGRID", (0, 0), (-1, -1), 0.3, grey_grid),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(KeepTogether(t_kpi))
    story.append(Spacer(1, 4 * mm))

    # Top 5 tiendas críticas
    critical_stores = [s for s in stores if s.status in ("critical", "replenish")][:8]
    if critical_stores:
        story.append(Paragraph("TIENDAS QUE REQUIEREN ATENCIÓN", h1))
        rows = [["Tienda", "Cadena", "Stock", "Vel sem", "WOS", "Status"]]
        for s in critical_stores:
            rows.append([
                s.store_name[:35], s.channel_name or "",
                str(int(s.total_on_hand or 0)),
                f"{float(s.avg_weekly_units or 0):.1f}",
                f"{float(s.wos_weeks or 0):.1f}",
                STATUS_LABEL.get(s.status, s.status),
            ])
        t = Table(rows, colWidths=[60 * mm, 40 * mm, 18 * mm, 20 * mm, 15 * mm, 27 * mm])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0), grey_bg),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ALIGN", (2, 1), (5, -1), "RIGHT"),
            ("BOX", (0, 0), (-1, -1), 0.4, grey_border),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, grey_grid),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(KeepTogether(t))
        story.append(Spacer(1, 4 * mm))

    # Top 10 SKUs
    if skus:
        story.append(Paragraph("TOP SKUs POR VENTAS (4 SEMANAS)", h1))
        rows = [["SKU", "Producto", "Tiendas", "Unidades", "Vel sem"]]
        for k in skus[:10]:
            rows.append([
                (k.sku or "")[:20], (k.product_name or "")[:38],
                str(k.stores_count),
                str(int(k.total_units_sold or 0)),
                f"{float(k.avg_weekly_units or 0):.1f}",
            ])
        t = Table(rows, colWidths=[35 * mm, 75 * mm, 20 * mm, 22 * mm, 22 * mm])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8.5),
            ("BACKGROUND", (0, 0), (-1, 0), grey_bg),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTNAME", (0, 1), (0, -1), "Courier"),
            ("ALIGN", (2, 1), (4, -1), "RIGHT"),
            ("BOX", (0, 0), (-1, -1), 0.4, grey_border),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, grey_grid),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(KeepTogether(t))
        story.append(Spacer(1, 4 * mm))

    # Alertas urgentes
    urgent_alerts = [a for a in alerts if a.severity in ("urgent", "high") and a.status == "open"][:10]
    if urgent_alerts:
        story.append(Paragraph("ALERTAS ACTIVAS URGENTES / ALTA", h1))
        rows = [["Severidad", "Tipo", "Tienda / Producto", "Mensaje"]]
        for a in urgent_alerts:
            product_line = f"{a.store_name or ''} · {a.product_name or a.sku or '—'}"
            rows.append([
                SEV_LABEL.get(a.severity, a.severity),
                ALERT_TYPE_LABEL.get(a.alert_type, a.alert_type),
                product_line[:45],
                (a.message or "")[:60],
            ])
        t = Table(rows, colWidths=[22 * mm, 30 * mm, 60 * mm, 62 * mm])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), grey_bg),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOX", (0, 0), (-1, -1), 0.4, grey_border),
            ("INNERGRID", (0, 0), (-1, -1), 0.3, grey_grid),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        story.append(KeepTogether(t))
        story.append(Spacer(1, 4 * mm))

    # Reabasto urgente
    if repl and repl.urgent_count + repl.high_count > 0:
        story.append(Paragraph("SUGERENCIAS DE REABASTO URGENTE / ALTA", h1))
        rows = [["Prio", "Tienda", "SKU", "Stock", "Sugerido", "Motivo"]]
        for s in repl.suggestions[:15]:
            if s.priority not in ("urgent", "high"):
                continue
            rows.append([
                PRIO_LABEL.get(s.priority, s.priority),
                (s.store_name or "")[:22],
                (s.sku or "")[:18],
                str(int(s.current_on_hand or 0)),
                str(int(s.suggested_units or 0)),
                (s.reason or "")[:40],
            ])
        if len(rows) > 1:
            t = Table(rows, colWidths=[18 * mm, 40 * mm, 32 * mm, 15 * mm, 22 * mm, 47 * mm])
            t.setStyle(TableStyle([
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BACKGROUND", (0, 0), (-1, 0), grey_bg),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTNAME", (2, 1), (2, -1), "Courier"),
                ("ALIGN", (3, 1), (4, -1), "RIGHT"),
                ("BOX", (0, 0), (-1, -1), 0.4, grey_border),
                ("INNERGRID", (0, 0), (-1, -1), 0.3, grey_grid),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]))
            story.append(KeepTogether(t))

    story.append(Spacer(1, 6 * mm))
    story.append(Paragraph(
        f"<font size='7' color='#999'>STHENOVA ERP · Retail Analytics · "
        f"Emitido {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}</font>",
        ParagraphStyle("f", parent=styles["Normal"], alignment=TA_CENTER)))

    doc.build(story)
    return buf.getvalue()
