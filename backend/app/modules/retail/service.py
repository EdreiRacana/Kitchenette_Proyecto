"""Business logic del módulo Retail Sell-out Analytics.

Diseño:
- CRUD de channels, stores y sellout reports.
- KPIs: sell-through %, WOS ponderado, tiendas críticas/sobreinventario.
- Replenishment engine: sugiere unidades a mandar para llegar al WOS target
  de la cadena. Prioridad urgent/high/normal según qué tan rojo esté.

WOS (Weeks of Supply) = on_hand / velocidad_semanal_promedio
- Se calcula con las últimas 4 semanas (ventana estándar en retail analytics).
- Si no hay ventas → WOS = ∞ (marcamos 999 y status="no_data").
- Se compara vs los umbrales de la cadena para dar el status:
    WOS < critical → critical
    critical <= WOS < target → replenish
    target <= WOS <= overstock → healthy
    WOS > overstock → overstock
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime, timedelta, timezone
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import and_, delete, func, select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging import get_logger
from app.modules.customers import models as cust_models
from app.modules.inventory import models as inv_models

from . import models, schemas

log = get_logger(__name__)


# ── Constantes de análisis ───────────────────────────────────────────────

VELOCITY_WINDOW_DAYS = 28              # 4 semanas rolling
WOS_INFINITY = 999.0                    # marcador para "sin ventas"


def _clamp(x: float, lo: float = 0.0, hi: float = WOS_INFINITY) -> float:
    return max(lo, min(hi, x))


def _wos_status(wos: float, critical: float, target: float, overstock: float,
                has_sales: bool) -> str:
    if not has_sales:
        return "no_data"
    if wos < critical:
        return "critical"
    if wos < target:
        return "replenish"
    if wos <= overstock:
        return "healthy"
    return "overstock"


# ── Channels ─────────────────────────────────────────────────────────────

async def list_channels(db: AsyncSession) -> List[schemas.RetailChannelOut]:
    stmt = select(models.RetailChannel).order_by(models.RetailChannel.name.asc())
    rows = list((await db.execute(stmt)).scalars().all())
    # Enriquecer con nombre de customer y count de stores
    cust_ids = {r.customer_id for r in rows if r.customer_id}
    cust_names: Dict[int, str] = {}
    if cust_ids:
        cres = await db.execute(
            select(cust_models.Customer.id, cust_models.Customer.name)
            .where(cust_models.Customer.id.in_(cust_ids))
        )
        cust_names = {r.id: r.name for r in cres}
    counts_stmt = (
        select(models.RetailStore.channel_id, func.count(models.RetailStore.id))
        .group_by(models.RetailStore.channel_id)
    )
    counts = dict((await db.execute(counts_stmt)).all())
    out = []
    for r in rows:
        out.append(schemas.RetailChannelOut(
            id=r.id, name=r.name, code=r.code, customer_id=r.customer_id,
            target_wos_weeks=r.target_wos_weeks,
            critical_wos_weeks=r.critical_wos_weeks,
            overstock_wos_weeks=r.overstock_wos_weeks,
            is_active=r.is_active, notes=r.notes,
            customer_name=cust_names.get(r.customer_id or 0),
            stores_count=int(counts.get(r.id, 0)),
            created_at=r.created_at,
        ))
    return out


async def get_channel(db: AsyncSession, channel_id: int) -> Optional[models.RetailChannel]:
    return await db.get(models.RetailChannel, channel_id)


async def create_channel(db: AsyncSession, data: schemas.RetailChannelCreate) -> models.RetailChannel:
    ch = models.RetailChannel(**data.model_dump(exclude_unset=True))
    db.add(ch); await db.commit(); await db.refresh(ch)
    return ch


async def update_channel(db: AsyncSession, channel_id: int,
                          data: schemas.RetailChannelUpdate) -> Optional[models.RetailChannel]:
    ch = await db.get(models.RetailChannel, channel_id)
    if ch is None:
        return None
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(ch, k, v)
    await db.commit(); await db.refresh(ch)
    return ch


async def delete_channel(db: AsyncSession, channel_id: int) -> bool:
    ch = await db.get(models.RetailChannel, channel_id)
    if ch is None:
        return False
    await db.delete(ch); await db.commit()
    return True


# ── Stores ───────────────────────────────────────────────────────────────

async def list_stores(db: AsyncSession, channel_id: Optional[int] = None,
                       active_only: bool = False) -> List[schemas.RetailStoreOut]:
    stmt = select(models.RetailStore, models.RetailChannel.name).join(
        models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id
    )
    if channel_id is not None:
        stmt = stmt.where(models.RetailStore.channel_id == channel_id)
    if active_only:
        stmt = stmt.where(models.RetailStore.is_active.is_(True))
    stmt = stmt.order_by(
        models.RetailChannel.name.asc(), models.RetailStore.name.asc()
    )
    rows = (await db.execute(stmt)).all()

    # Cargar nombres de warehouses referenciados
    wh_ids = {s.consignment_warehouse_id for s, _ in rows if s.consignment_warehouse_id}
    wh_names: Dict[int, str] = {}
    if wh_ids:
        wres = await db.execute(
            select(inv_models.Warehouse.id, inv_models.Warehouse.name)
            .where(inv_models.Warehouse.id.in_(wh_ids))
        )
        wh_names = {r.id: r.name for r in wres}

    return [
        schemas.RetailStoreOut(
            id=s.id, channel_id=s.channel_id, channel_name=cname,
            name=s.name, code=s.code, external_code=s.external_code,
            city=s.city, state=s.state, region=s.region,
            store_format=s.store_format, address=s.address,
            contact_name=s.contact_name, contact_phone=s.contact_phone,
            consignment_warehouse_id=s.consignment_warehouse_id,
            consignment_warehouse_name=wh_names.get(s.consignment_warehouse_id or -1),
            is_active=s.is_active, notes=s.notes, created_at=s.created_at,
        ) for s, cname in rows
    ]


async def get_store(db: AsyncSession, store_id: int) -> Optional[models.RetailStore]:
    return await db.get(models.RetailStore, store_id)


async def create_store(db: AsyncSession, data: schemas.RetailStoreCreate) -> models.RetailStore:
    s = models.RetailStore(**data.model_dump(exclude_unset=True))
    db.add(s); await db.commit(); await db.refresh(s)
    return s


async def bulk_create_stores(db: AsyncSession, data: schemas.BulkStoresRequest
                              ) -> schemas.BulkStoresResponse:
    ch = await db.get(models.RetailChannel, data.channel_id)
    if ch is None:
        raise ValueError("Cadena no encontrada")

    # Skip duplicados por external_code + channel_id (o name)
    ext_codes = {s.external_code for s in data.stores if s.external_code}
    existing_ext: set = set()
    if ext_codes:
        res = await db.execute(
            select(models.RetailStore.external_code).where(
                models.RetailStore.channel_id == data.channel_id,
                models.RetailStore.external_code.in_(ext_codes),
            )
        )
        existing_ext = {r[0] for r in res}

    created: List[models.RetailStore] = []
    skipped = 0
    for st in data.stores:
        if st.external_code and st.external_code in existing_ext:
            skipped += 1
            continue
        obj = models.RetailStore(channel_id=data.channel_id, **st.model_dump(exclude_unset=True))
        db.add(obj); created.append(obj)
    await db.commit()
    for c in created:
        await db.refresh(c)
    return schemas.BulkStoresResponse(
        created=len(created), skipped=skipped,
        stores=[
            schemas.RetailStoreOut(
                id=c.id, channel_id=c.channel_id, channel_name=ch.name,
                name=c.name, code=c.code, external_code=c.external_code,
                city=c.city, state=c.state, region=c.region,
                store_format=c.store_format, address=c.address,
                contact_name=c.contact_name, contact_phone=c.contact_phone,
                is_active=c.is_active, notes=c.notes, created_at=c.created_at,
            ) for c in created
        ],
    )


async def update_store(db: AsyncSession, store_id: int,
                        data: schemas.RetailStoreUpdate) -> Optional[models.RetailStore]:
    s = await db.get(models.RetailStore, store_id)
    if s is None:
        return None
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(s, k, v)
    await db.commit(); await db.refresh(s)
    return s


async def delete_store(db: AsyncSession, store_id: int) -> bool:
    s = await db.get(models.RetailStore, store_id)
    if s is None:
        return False
    await db.delete(s); await db.commit()
    return True


# ── Sell-out Reports ─────────────────────────────────────────────────────

async def _fill_variant_snapshot(db: AsyncSession, report: models.SellOutReport) -> None:
    if report.variant_id and (not report.product_name or not report.sku):
        v = await db.get(inv_models.ProductVariant, report.variant_id)
        if v:
            if not report.sku:
                report.sku = v.sku
            if not report.product_name and v.product_id:
                p = await db.get(inv_models.Product, v.product_id)
                if p:
                    report.product_name = p.name


async def list_sellout(db: AsyncSession,
                        channel_id: Optional[int] = None,
                        store_id: Optional[int] = None,
                        variant_id: Optional[int] = None,
                        period_start_gte: Optional[datetime] = None,
                        period_start_lt: Optional[datetime] = None,
                        limit: int = 500) -> List[schemas.SellOutReportOut]:
    stmt = (
        select(
            models.SellOutReport,
            models.RetailStore.name.label("store_name"),
            models.RetailStore.channel_id.label("channel_id"),
            models.RetailChannel.name.label("channel_name"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
    )
    conds = []
    if channel_id is not None:
        conds.append(models.RetailStore.channel_id == channel_id)
    if store_id is not None:
        conds.append(models.SellOutReport.store_id == store_id)
    if variant_id is not None:
        conds.append(models.SellOutReport.variant_id == variant_id)
    if period_start_gte is not None:
        conds.append(models.SellOutReport.period_start >= period_start_gte)
    if period_start_lt is not None:
        conds.append(models.SellOutReport.period_start < period_start_lt)
    if conds:
        stmt = stmt.where(and_(*conds))
    stmt = stmt.order_by(models.SellOutReport.period_start.desc(),
                          models.SellOutReport.id.desc()).limit(limit)
    rows = (await db.execute(stmt)).all()
    out = []
    for r, store_name, ch_id, ch_name in rows:
        out.append(schemas.SellOutReportOut(
            id=r.id, store_id=r.store_id, store_name=store_name,
            channel_id=ch_id, channel_name=ch_name,
            variant_id=r.variant_id, product_name=r.product_name, sku=r.sku,
            period_start=r.period_start, period_end=r.period_end,
            period_type=r.period_type,
            units_sold=r.units_sold,
            units_returned=int(r.units_returned or 0),
            units_on_hand=r.units_on_hand,
            revenue=r.revenue,
            returns_amount=float(r.returns_amount or 0.0),
            source=r.source, notes=r.notes,
            created_at=r.created_at,
        ))
    return out


async def create_sellout(db: AsyncSession, data: schemas.SellOutReportCreate,
                          user_id: Optional[int] = None) -> models.SellOutReport:
    # Upsert por (store, variant, period_start, period_type)
    existing_stmt = select(models.SellOutReport).where(
        models.SellOutReport.store_id == data.store_id,
        models.SellOutReport.variant_id == data.variant_id,
        models.SellOutReport.period_start == data.period_start,
        models.SellOutReport.period_type == data.period_type,
    )
    existing = (await db.execute(existing_stmt)).scalars().first()
    if existing:
        for k, v in data.model_dump(exclude_unset=True).items():
            if k in ("store_id", "variant_id", "period_start", "period_type"):
                continue
            setattr(existing, k, v)
        await _fill_variant_snapshot(db, existing)
        existing.uploaded_by_user_id = user_id or existing.uploaded_by_user_id
        await db.commit(); await db.refresh(existing)
        await _apply_consignment_movement(db, existing, user_id)
        return existing
    r = models.SellOutReport(
        **data.model_dump(exclude_unset=True),
        uploaded_by_user_id=user_id,
    )
    await _fill_variant_snapshot(db, r)
    db.add(r); await db.commit(); await db.refresh(r)
    await _apply_consignment_movement(db, r, user_id)
    try:
        await evaluate_alerts(db)
    except Exception as e:
        log.warning("evaluate_alerts falló tras create_sellout %s: %s", r.id, e)
    return r


async def _apply_consignment_movement(
    db: AsyncSession, report: models.SellOutReport, user_id: Optional[int],
) -> None:
    """Si la tienda tiene warehouse de consignación asignado y hay variant en
    catálogo, ajusta el stock: descuenta el DELTA entre units_sold reportadas
    y stock_consumed previo. Idempotente: reimportar el mismo reporte no dobla
    el descuento; incrementar el reporte descuenta la diferencia; disminuirlo
    reingresa (adjustment +).
    """
    if not report.variant_id:
        return
    store = await db.get(models.RetailStore, report.store_id)
    if store is None or store.consignment_warehouse_id is None:
        return
    already = int(report.stock_consumed or 0)
    target = int(report.units_sold or 0)
    delta = target - already
    if delta == 0:
        return

    from app.modules.inventory import schemas as inv_schemas, service as inv_service
    try:
        if delta > 0:
            mov = inv_schemas.StockMovementCreate(
                variant_id=report.variant_id,
                warehouse_id=store.consignment_warehouse_id,
                quantity=delta,
                movement_type="out",
                reference=f"retail_sellout:{report.id}",
                notes=(f"Venta consignación — tienda {store.name} · "
                        f"{report.product_name or report.sku or ''}"),
            )
        else:
            # Reversión: si el reporte bajó, reingresa stock como adjustment +
            mov = inv_schemas.StockMovementCreate(
                variant_id=report.variant_id,
                warehouse_id=store.consignment_warehouse_id,
                quantity=abs(delta),
                movement_type="adjustment",
                reference=f"retail_sellout:{report.id}:reverse",
                notes=("Reversión de sell-out reportado a la baja"),
            )
        await inv_service.adjust_stock(db, mov, user_id=user_id)
        report.stock_consumed = target
        await db.commit()
    except Exception as e:
        # No revertir el sell-out si el ajuste de stock falla — deja constancia
        # y sigue. La reconciliación reportará el descuadre.
        log.warning(
            "consignment_movement falló para report=%s store=%s variant=%s: %s",
            report.id, report.store_id, report.variant_id, e,
        )


async def list_consignment_warehouses(db: AsyncSession) -> List[schemas.ConsignmentWarehouseOption]:
    res = await db.execute(
        select(inv_models.Warehouse)
        .where(inv_models.Warehouse.type == "consignment")
        .order_by(inv_models.Warehouse.name)
    )
    return [
        schemas.ConsignmentWarehouseOption(
            id=w.id, name=w.name, location=w.location, is_active=w.is_active,
        ) for w in res.scalars().all()
    ]


async def consignment_reconciliation(
    db: AsyncSession, channel_id: Optional[int] = None,
) -> schemas.ConsignmentReconResponse:
    """Compara on_hand reportado por la tienda vs stock actual del almacén
    de consignación asignado, por (store, variant). Sólo tiendas con warehouse."""
    from datetime import datetime, timezone

    now = datetime.now(timezone.utc)

    stores_stmt = (
        select(models.RetailStore, models.RetailChannel.name)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
        .where(models.RetailStore.consignment_warehouse_id.isnot(None),
                models.RetailStore.is_active.is_(True))
    )
    if channel_id is not None:
        stores_stmt = stores_stmt.where(models.RetailStore.channel_id == channel_id)
    stores = (await db.execute(stores_stmt)).all()
    if not stores:
        return schemas.ConsignmentReconResponse(
            generated_at=now, channel_id=channel_id, total_rows=0,
            matched=0, with_diff=0, rows=[],
        )

    warehouse_ids = {s.consignment_warehouse_id for s, _ in stores}
    wh_names = {
        w.id: w.name for w in (await db.execute(
            select(inv_models.Warehouse).where(inv_models.Warehouse.id.in_(warehouse_ids))
        )).scalars().all()
    }

    rows: List[schemas.ConsignmentReconRow] = []
    matched = 0
    with_diff = 0

    for store, channel_name in stores:
        # Últimos on_hand reportados por (variant) — el reporte más reciente
        last_stmt = (
            select(
                models.SellOutReport.variant_id,
                models.SellOutReport.product_name,
                models.SellOutReport.sku,
                func.max(models.SellOutReport.period_start).label("last_period"),
            )
            .where(models.SellOutReport.store_id == store.id,
                    models.SellOutReport.variant_id.isnot(None))
            .group_by(
                models.SellOutReport.variant_id,
                models.SellOutReport.product_name,
                models.SellOutReport.sku,
            )
        )
        last_rows = (await db.execute(last_stmt)).all()

        # Stock actual del warehouse por variant
        stock_rows = (await db.execute(
            select(inv_models.StockLevel.variant_id, inv_models.StockLevel.quantity)
            .where(inv_models.StockLevel.warehouse_id == store.consignment_warehouse_id)
        )).all()
        stock_by_variant: Dict[int, int] = {int(vid): int(q) for vid, q in stock_rows}

        for lr in last_rows:
            vid = int(lr.variant_id) if lr.variant_id else None
            if vid is None:
                continue
            oh_report = int((await db.execute(
                select(func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0))
                .where(models.SellOutReport.store_id == store.id,
                        models.SellOutReport.variant_id == vid,
                        models.SellOutReport.period_start == lr.last_period)
            )).scalar() or 0)
            wh_stock = stock_by_variant.get(vid, 0)
            diff = wh_stock - oh_report
            if abs(diff) < 1:
                status = "match"; matched += 1
            elif diff > 0:
                status = "over_at_warehouse"; with_diff += 1
            else:
                status = "short_at_warehouse"; with_diff += 1
            rows.append(schemas.ConsignmentReconRow(
                store_id=store.id, store_name=store.name,
                channel_name=channel_name,
                warehouse_id=store.consignment_warehouse_id,
                warehouse_name=wh_names.get(store.consignment_warehouse_id, "—"),
                variant_id=vid,
                product_name=lr.product_name, sku=lr.sku,
                reported_on_hand=oh_report,
                reported_at=lr.last_period,
                warehouse_stock=wh_stock,
                difference=diff,
                status=status,
            ))
        # Además, cualquier variant con stock en warehouse pero sin reporte
        for vid, qty in stock_by_variant.items():
            if any(r.variant_id == vid for r in rows if r.store_id == store.id):
                continue
            v = await db.get(inv_models.ProductVariant, vid)
            pname = None
            if v and v.product_id:
                p = await db.get(inv_models.Product, v.product_id)
                if p:
                    pname = p.name
            rows.append(schemas.ConsignmentReconRow(
                store_id=store.id, store_name=store.name,
                channel_name=channel_name,
                warehouse_id=store.consignment_warehouse_id,
                warehouse_name=wh_names.get(store.consignment_warehouse_id, "—"),
                variant_id=vid, product_name=pname, sku=v.sku if v else None,
                reported_on_hand=0, reported_at=None,
                warehouse_stock=qty, difference=qty,
                status="no_data" if qty == 0 else "over_at_warehouse",
            ))
            if qty > 0:
                with_diff += 1

    rows.sort(key=lambda r: (
        0 if r.status == "short_at_warehouse" else 1 if r.status == "over_at_warehouse"
        else 2 if r.status == "match" else 3,
        -abs(r.difference),
    ))
    return schemas.ConsignmentReconResponse(
        generated_at=now, channel_id=channel_id,
        total_rows=len(rows), matched=matched, with_diff=with_diff,
        rows=rows,
    )


async def update_sellout(db: AsyncSession, report_id: int,
                          data: schemas.SellOutReportUpdate) -> Optional[models.SellOutReport]:
    r = await db.get(models.SellOutReport, report_id)
    if r is None:
        return None
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(r, k, v)
    await db.commit(); await db.refresh(r)
    return r


async def delete_sellout(db: AsyncSession, report_id: int) -> bool:
    r = await db.get(models.SellOutReport, report_id)
    if r is None:
        return False
    await db.delete(r); await db.commit()
    return True


# ── Dashboard KPIs ───────────────────────────────────────────────────────

async def dashboard_kpis(db: AsyncSession, channel_id: Optional[int] = None,
                          days: int = 30) -> schemas.RetailKPIs:
    now = datetime.now(timezone.utc)
    period_start = now - timedelta(days=days)

    # Sell-out (incluye devoluciones para poder calcular netos y tasa)
    so_stmt = (
        select(
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units"),
            func.coalesce(func.sum(models.SellOutReport.revenue), 0.0).label("revenue"),
            func.coalesce(func.sum(models.SellOutReport.units_returned), 0).label("units_returned"),
            func.coalesce(func.sum(models.SellOutReport.returns_amount), 0.0).label("returns_amount"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(models.SellOutReport.period_start >= period_start)
    )
    if channel_id is not None:
        so_stmt = so_stmt.where(models.RetailStore.channel_id == channel_id)
    so = (await db.execute(so_stmt)).one()

    # Sell-in (facturación al customer vinculado a la cadena)
    from app.modules.sales import models as sales_models

    sell_in_units = 0
    sell_in_revenue = 0.0
    channel_ids: List[int] = []
    if channel_id is not None:
        channel_ids = [channel_id]
    else:
        cids = (await db.execute(
            select(models.RetailChannel.id).where(models.RetailChannel.is_active.is_(True))
        )).all()
        channel_ids = [c[0] for c in cids]

    if channel_ids:
        cust_map_stmt = select(models.RetailChannel.id, models.RetailChannel.customer_id).where(
            models.RetailChannel.id.in_(channel_ids)
        )
        cust_ids = [c for _, c in (await db.execute(cust_map_stmt)).all() if c]
        if cust_ids:
            si_stmt = (
                select(
                    func.coalesce(func.sum(sales_models.OrderItem.quantity), 0).label("units"),
                    func.coalesce(func.sum(sales_models.OrderItem.total), 0.0).label("revenue"),
                )
                .join(sales_models.Order, sales_models.OrderItem.order_id == sales_models.Order.id)
                .where(
                    sales_models.Order.customer_id.in_(cust_ids),
                    sales_models.Order.kind == "order",
                    sales_models.Order.status != "cancelled",
                    sales_models.Order.created_at >= period_start,
                )
            )
            si = (await db.execute(si_stmt)).one()
            sell_in_units = int(si.units or 0)
            sell_in_revenue = float(si.revenue or 0.0)

    # WOS por store (para agregado)
    stores_stmt = (
        select(models.RetailStore.id, models.RetailStore.channel_id,
                models.RetailChannel.critical_wos_weeks,
                models.RetailChannel.target_wos_weeks,
                models.RetailChannel.overstock_wos_weeks)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
        .where(models.RetailStore.is_active.is_(True))
    )
    if channel_id is not None:
        stores_stmt = stores_stmt.where(models.RetailStore.channel_id == channel_id)
    stores_rows = (await db.execute(stores_stmt)).all()

    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)
    # On-hand por store: sumar units_on_hand del último periodo reportado por store.
    on_hand_by_store: Dict[int, int] = {}
    last_periods = (await db.execute(
        select(
            models.SellOutReport.store_id,
            func.max(models.SellOutReport.period_start).label("last"),
        ).group_by(models.SellOutReport.store_id)
    )).all()
    for sid, last in last_periods:
        if last is None:
            continue
        oh = (await db.execute(
            select(func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0))
            .where(models.SellOutReport.store_id == sid,
                    models.SellOutReport.period_start == last)
        )).scalar() or 0
        on_hand_by_store[int(sid)] = int(oh)

    # Velocidad semanal por store: units_sold en la ventana / (días/7)
    vel_rows = (await db.execute(
        select(
            models.SellOutReport.store_id,
            func.sum(models.SellOutReport.units_sold).label("units"),
        ).where(models.SellOutReport.period_start >= velocity_from)
         .group_by(models.SellOutReport.store_id)
    )).all()
    vel_by_store: Dict[int, float] = {
        int(sid): (float(units or 0) / (VELOCITY_WINDOW_DAYS / 7.0))
        for sid, units in vel_rows
    }

    critical_stores = 0
    overstock_stores = 0
    total_on_hand = 0
    weighted_wos_num = 0.0
    weighted_wos_den = 0.0

    for sid, _cid, crit_w, tgt_w, over_w in stores_rows:
        on_hand = on_hand_by_store.get(int(sid), 0)
        vel = vel_by_store.get(int(sid), 0.0)
        total_on_hand += on_hand
        if vel <= 0:
            continue
        wos = _clamp(on_hand / vel)
        if wos < float(crit_w or 2.0):
            critical_stores += 1
        if wos > float(over_w or 12.0):
            overstock_stores += 1
        # Ponderación por on_hand para el promedio
        weighted_wos_num += wos * on_hand
        weighted_wos_den += on_hand

    avg_wos = round(weighted_wos_num / weighted_wos_den, 2) if weighted_wos_den > 0 else 0.0

    # Skus activos en el periodo
    skus_count = int((await db.execute(
        select(func.count(func.distinct(models.SellOutReport.variant_id)))
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(models.SellOutReport.period_start >= period_start,
                *([models.RetailStore.channel_id == channel_id] if channel_id else []))
    )).scalar() or 0)

    stores_active = len(stores_rows)

    total_so = int(so.units or 0)
    sell_through = round((total_so / sell_in_units * 100.0), 2) if sell_in_units > 0 else 0.0

    channel_name = None
    if channel_id is not None:
        ch = await db.get(models.RetailChannel, channel_id)
        channel_name = ch.name if ch else None

    total_returns_units = int(so.units_returned or 0)
    total_returns_amount = round(float(so.returns_amount or 0.0), 2)
    gross_revenue = float(so.revenue or 0.0)
    return_rate = round((total_returns_units / total_so * 100.0), 2) if total_so > 0 else 0.0
    net_units = max(total_so - total_returns_units, 0)
    net_revenue = round(max(gross_revenue - total_returns_amount, 0.0), 2)

    return schemas.RetailKPIs(
        channel_id=channel_id, channel_name=channel_name,
        period_start=period_start, period_end=now,
        total_sell_out_units=total_so,
        total_sell_out_revenue=round(gross_revenue, 2),
        total_sell_in_units=sell_in_units,
        total_sell_in_revenue=round(sell_in_revenue, 2),
        sell_through_pct=sell_through,
        total_returns_units=total_returns_units,
        total_returns_amount=total_returns_amount,
        return_rate_pct=return_rate,
        net_units=net_units,
        net_revenue=net_revenue,
        total_on_hand=total_on_hand,
        avg_wos_weeks=avg_wos,
        critical_stores_count=critical_stores,
        overstock_stores_count=overstock_stores,
        stores_active_count=stores_active,
        skus_active_count=skus_count,
    )


# ── Store velocity list ─────────────────────────────────────────────────

async def stores_velocity(db: AsyncSession, channel_id: Optional[int] = None
                            ) -> List[schemas.StoreVelocityRow]:
    now = datetime.now(timezone.utc)
    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)

    stores_stmt = (
        select(models.RetailStore, models.RetailChannel)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
        .where(models.RetailStore.is_active.is_(True))
    )
    if channel_id is not None:
        stores_stmt = stores_stmt.where(models.RetailStore.channel_id == channel_id)
    stores = (await db.execute(stores_stmt)).all()

    # Sums por store en la ventana
    vel_map: Dict[int, Tuple[int, int]] = {}
    vel_rows = (await db.execute(
        select(
            models.SellOutReport.store_id,
            func.sum(models.SellOutReport.units_sold).label("units"),
        ).where(models.SellOutReport.period_start >= velocity_from)
         .group_by(models.SellOutReport.store_id)
    )).all()
    for sid, u in vel_rows:
        vel_map[int(sid)] = (int(u or 0), 0)

    # On-hand del último reporte de cada store (sum on-hand de la fecha más reciente por store)
    onhand_map: Dict[int, int] = {}
    last_periods = (await db.execute(
        select(
            models.SellOutReport.store_id,
            func.max(models.SellOutReport.period_start).label("last"),
        ).group_by(models.SellOutReport.store_id)
    )).all()
    for sid, last in last_periods:
        if last is None:
            continue
        s = (await db.execute(
            select(func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0))
            .where(models.SellOutReport.store_id == sid,
                    models.SellOutReport.period_start == last)
        )).scalar() or 0
        onhand_map[int(sid)] = int(s)

    out: List[schemas.StoreVelocityRow] = []
    for store, ch in stores:
        units = vel_map.get(store.id, (0, 0))[0]
        avg_weekly = round(units / (VELOCITY_WINDOW_DAYS / 7.0), 2)
        on_hand = onhand_map.get(store.id, 0)
        wos = _clamp(on_hand / avg_weekly) if avg_weekly > 0 else WOS_INFINITY
        status = _wos_status(
            wos, ch.critical_wos_weeks, ch.target_wos_weeks,
            ch.overstock_wos_weeks, has_sales=avg_weekly > 0,
        )
        out.append(schemas.StoreVelocityRow(
            store_id=store.id, store_name=store.name,
            channel_name=ch.name,
            total_units_sold=units, avg_weekly_units=avg_weekly,
            total_on_hand=on_hand, wos_weeks=round(wos, 2),
            status=status,
        ))
    out.sort(key=lambda r: (
        0 if r.status == "critical" else 1 if r.status == "replenish"
        else 3 if r.status == "overstock" else 2 if r.status == "healthy" else 4,
        r.wos_weeks,
    ))
    return out


# ── SKU velocity ────────────────────────────────────────────────────────

async def skus_velocity(db: AsyncSession, channel_id: Optional[int] = None,
                          limit: int = 100) -> List[schemas.SKUVelocityRow]:
    now = datetime.now(timezone.utc)
    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)

    base_join = (
        select(
            models.SellOutReport.variant_id,
            func.max(models.SellOutReport.sku).label("sku"),
            func.max(models.SellOutReport.product_name).label("product_name"),
            func.count(func.distinct(models.SellOutReport.store_id)).label("stores"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(models.SellOutReport.period_start >= velocity_from)
        .group_by(models.SellOutReport.variant_id)
        .order_by(func.sum(models.SellOutReport.units_sold).desc())
        .limit(limit)
    )
    if channel_id is not None:
        base_join = base_join.where(models.RetailStore.channel_id == channel_id)

    vel_rows = (await db.execute(base_join)).all()

    # On-hand por variant (último periodo por store, sumado)
    on_hand_by_variant: Dict[int, int] = {}
    for vid, *_ in vel_rows:
        if vid is None:
            continue
        oh_stmt = (
            select(
                models.SellOutReport.store_id,
                func.max(models.SellOutReport.period_start).label("last"),
            ).where(models.SellOutReport.variant_id == vid)
             .group_by(models.SellOutReport.store_id)
        )
        rows = (await db.execute(oh_stmt)).all()
        total = 0
        for sid, last in rows:
            oh = (await db.execute(
                select(func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0))
                .where(models.SellOutReport.variant_id == vid,
                        models.SellOutReport.store_id == sid,
                        models.SellOutReport.period_start == last)
            )).scalar() or 0
            total += int(oh)
        on_hand_by_variant[int(vid)] = total

    # Umbrales globales (promedio) para status por SKU
    thresholds = (await db.execute(
        select(
            func.avg(models.RetailChannel.critical_wos_weeks),
            func.avg(models.RetailChannel.target_wos_weeks),
            func.avg(models.RetailChannel.overstock_wos_weeks),
        )
    )).one()
    crit = float(thresholds[0] or 2.0)
    tgt = float(thresholds[1] or 4.0)
    over = float(thresholds[2] or 12.0)

    out: List[schemas.SKUVelocityRow] = []
    for vid, sku, pname, stores, units in vel_rows:
        avg_weekly = round(float(units or 0) / (VELOCITY_WINDOW_DAYS / 7.0), 2)
        on_hand = on_hand_by_variant.get(int(vid) if vid else -1, 0)
        wos = _clamp(on_hand / avg_weekly) if avg_weekly > 0 else WOS_INFINITY
        status = _wos_status(wos, crit, tgt, over, has_sales=avg_weekly > 0)
        out.append(schemas.SKUVelocityRow(
            variant_id=int(vid) if vid else None,
            sku=sku, product_name=pname,
            stores_count=int(stores or 0),
            total_units_sold=int(units or 0),
            avg_weekly_units=avg_weekly,
            total_on_hand=on_hand,
            wos_weeks=round(wos, 2),
            status=status,
        ))
    return out


# ── Replenishment engine ────────────────────────────────────────────────

async def replenishment(db: AsyncSession, channel_id: Optional[int] = None
                          ) -> schemas.ReplenishmentResponse:
    now = datetime.now(timezone.utc)
    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)

    # Combos (store, variant) con último periodo y sums en la ventana
    stmt = (
        select(
            models.RetailStore.id.label("store_id"),
            models.RetailStore.name.label("store_name"),
            models.RetailChannel.id.label("channel_id"),
            models.RetailChannel.name.label("channel_name"),
            models.RetailChannel.target_wos_weeks,
            models.RetailChannel.critical_wos_weeks,
            models.SellOutReport.variant_id,
            models.SellOutReport.product_name,
            models.SellOutReport.sku,
            func.max(models.SellOutReport.period_start).label("last_period"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units_sold"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
        .where(models.SellOutReport.period_start >= velocity_from,
                models.RetailStore.is_active.is_(True))
        .group_by(
            models.RetailStore.id, models.RetailStore.name,
            models.RetailChannel.id, models.RetailChannel.name,
            models.RetailChannel.target_wos_weeks,
            models.RetailChannel.critical_wos_weeks,
            models.SellOutReport.variant_id, models.SellOutReport.product_name,
            models.SellOutReport.sku,
        )
    )
    if channel_id is not None:
        stmt = stmt.where(models.RetailStore.channel_id == channel_id)

    rows = (await db.execute(stmt)).all()

    suggestions: List[schemas.ReplenishmentSuggestion] = []
    urgent = high = normal = 0

    for r in rows:
        # On-hand actual: último reporte de ese (store, variant)
        oh = (await db.execute(
            select(func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0))
            .where(models.SellOutReport.store_id == r.store_id,
                    models.SellOutReport.variant_id == r.variant_id,
                    models.SellOutReport.period_start == r.last_period)
        )).scalar() or 0
        on_hand = int(oh)

        avg_weekly = float(r.units_sold or 0) / (VELOCITY_WINDOW_DAYS / 7.0)
        if avg_weekly <= 0:
            continue
        wos = on_hand / avg_weekly
        tgt = float(r.target_wos_weeks or 4.0)
        crit = float(r.critical_wos_weeks or 2.0)
        if wos >= tgt:
            continue

        # Cuánto mandar para llegar a target
        needed = int(round(avg_weekly * tgt - on_hand))
        if needed <= 0:
            continue

        if wos < crit:
            prio, reason = "urgent", f"Stock crítico: {wos:.1f} sem. < mín {crit:.0f}"
            urgent += 1
        elif wos < tgt * 0.7:
            prio, reason = "high", f"Debajo del objetivo: {wos:.1f} sem. de {tgt:.0f}"
            high += 1
        else:
            prio, reason = "normal", f"Reabasto rutinario para llegar a {tgt:.0f} sem."
            normal += 1

        suggestions.append(schemas.ReplenishmentSuggestion(
            store_id=r.store_id, store_name=r.store_name,
            channel_id=r.channel_id, channel_name=r.channel_name,
            variant_id=r.variant_id, product_name=r.product_name, sku=r.sku,
            current_on_hand=on_hand, avg_weekly_units=round(avg_weekly, 2),
            wos_weeks=round(wos, 2), suggested_units=needed,
            priority=prio, reason=reason,
        ))

    priority_order = {"urgent": 0, "high": 1, "normal": 2}
    suggestions.sort(key=lambda s: (priority_order[s.priority], s.wos_weeks))

    tgt_default = 4.0
    crit_default = 2.0
    if channel_id is not None:
        ch = await db.get(models.RetailChannel, channel_id)
        if ch:
            tgt_default = ch.target_wos_weeks
            crit_default = ch.critical_wos_weeks

    return schemas.ReplenishmentResponse(
        channel_id=channel_id, generated_at=now,
        target_wos_weeks=tgt_default, critical_wos_weeks=crit_default,
        suggestions=suggestions,
        urgent_count=urgent, high_count=high, normal_count=normal,
    )


# ── Store performance ───────────────────────────────────────────────────

async def store_performance(db: AsyncSession, store_id: int, weeks_back: int = 12
                              ) -> Optional[schemas.StorePerformanceOut]:
    store = await db.get(models.RetailStore, store_id)
    if store is None:
        return None
    channel = await db.get(models.RetailChannel, store.channel_id)

    now = datetime.now(timezone.utc)
    since = now - timedelta(weeks=weeks_back)

    rows = (await db.execute(
        select(
            models.SellOutReport.period_start,
            models.SellOutReport.period_end,
            func.sum(models.SellOutReport.units_sold).label("units"),
            func.sum(models.SellOutReport.units_on_hand).label("on_hand"),
            func.sum(models.SellOutReport.revenue).label("revenue"),
        )
        .where(models.SellOutReport.store_id == store_id,
                models.SellOutReport.period_start >= since)
        .group_by(models.SellOutReport.period_start, models.SellOutReport.period_end)
        .order_by(models.SellOutReport.period_start.asc())
    )).all()

    periods = [
        schemas.StorePerformancePeriod(
            period_start=ps, period_end=pe,
            units_sold=int(u or 0), on_hand_end=int(oh or 0),
            revenue=round(float(rev or 0.0), 2),
        ) for ps, pe, u, oh, rev in rows
    ]
    total_units = sum(p.units_sold for p in periods)
    total_rev = round(sum(p.revenue for p in periods), 2)
    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)

    def _ge(a: datetime, b: datetime) -> bool:
        # SQLite regresa datetimes naive; alinea antes de comparar.
        if a.tzinfo is None and b.tzinfo is not None:
            a = a.replace(tzinfo=b.tzinfo)
        elif b.tzinfo is None and a.tzinfo is not None:
            b = b.replace(tzinfo=a.tzinfo)
        return a >= b

    recent_units = sum(p.units_sold for p in periods if _ge(p.period_start, velocity_from))
    avg_weekly = round(recent_units / (VELOCITY_WINDOW_DAYS / 7.0), 2)
    latest_on_hand = periods[-1].on_hand_end if periods else 0
    wos = _clamp(latest_on_hand / avg_weekly) if avg_weekly > 0 else WOS_INFINITY
    status = _wos_status(
        wos,
        channel.critical_wos_weeks if channel else 2.0,
        channel.target_wos_weeks if channel else 4.0,
        channel.overstock_wos_weeks if channel else 12.0,
        has_sales=avg_weekly > 0,
    )
    return schemas.StorePerformanceOut(
        store_id=store.id, store_name=store.name,
        channel_name=channel.name if channel else "",
        periods=periods,
        total_units_sold=total_units, total_revenue=total_rev,
        avg_weekly_units=avg_weekly, latest_on_hand=latest_on_hand,
        wos_weeks=round(wos, 2), status=status,
    )


# ── Bulk import: plantilla + parser ─────────────────────────────────────

TEMPLATE_HEADERS = [
    "cadena_codigo", "cadena_nombre",
    "tienda_codigo", "tienda_nombre",
    "sku", "producto_nombre",
    "periodo_tipo", "periodo_inicio", "periodo_fin",
    "unidades_vendidas", "unidades_devueltas", "unidades_stock",
    "ingreso", "importe_devoluciones",
    "notas",
]


async def _template_hints(db: AsyncSession) -> Dict[str, List[dict]]:
    """Cadenas, tiendas y SKUs activos para las hojas de referencia."""
    chans = (await db.execute(
        select(models.RetailChannel).where(models.RetailChannel.is_active.is_(True))
        .order_by(models.RetailChannel.name).limit(500)
    )).scalars().all()

    stores = (await db.execute(
        select(models.RetailStore, models.RetailChannel.name)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
        .where(models.RetailStore.is_active.is_(True))
        .order_by(models.RetailChannel.name, models.RetailStore.name)
        .limit(5000)
    )).all()

    variants = (await db.execute(
        select(inv_models.ProductVariant.sku, inv_models.ProductVariant.price,
                inv_models.Product.name)
        .join(inv_models.Product, inv_models.ProductVariant.product_id == inv_models.Product.id)
        .where(inv_models.ProductVariant.is_active.is_(True))
        .limit(3000)
    )).all()

    return {
        "cadenas": [{"nombre": c.name, "codigo": c.code or ""} for c in chans],
        "tiendas": [
            {
                "cadena": chn, "nombre": s.name, "codigo_interno": s.code or "",
                "codigo_externo": s.external_code or "",
                "ciudad": s.city or "", "estado": s.state or "",
            } for s, chn in stores
        ],
        "productos": [
            {"sku": v.sku, "nombre": v.name, "precio_sugerido": float(v.price or 0.0)}
            for v in variants
        ],
    }


async def build_sellout_template_xlsx(db: AsyncSession) -> bytes:
    """Plantilla profesional multi-hoja con instrucciones + referencias."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hints = await _template_hints(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "SellOut"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    center = Alignment(horizontal="center")

    ws.append(TEMPLATE_HEADERS)
    for col_idx, _h in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    widths = [14, 26, 14, 26, 16, 30, 12, 14, 14, 14, 14, 14, 14, 14, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Fila-guía
    example_channel = hints["cadenas"][0]["nombre"] if hints["cadenas"] else "Walmart"
    example_channel_code = hints["cadenas"][0]["codigo"] if hints["cadenas"] and hints["cadenas"][0]["codigo"] else "WMT"
    example_store = hints["tiendas"][0]["nombre"] if hints["tiendas"] else "Sucursal Centro"
    example_store_code = hints["tiendas"][0]["codigo_externo"] if hints["tiendas"] and hints["tiendas"][0]["codigo_externo"] else "1001"
    example_sku = hints["productos"][0]["sku"] if hints["productos"] else "SKU-001"
    example_prod = hints["productos"][0]["nombre"] if hints["productos"] else "Producto de ejemplo"

    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)

    example_row = [
        example_channel_code, example_channel,
        example_store_code, example_store,
        example_sku, example_prod,
        "week", week_start.isoformat(), week_end.isoformat(),
        12, 1, 45, 12 * 1290.0, 1290.0,
        "Ejemplo — puedes borrar esta fila",
    ]
    ws.append(example_row)
    grey = PatternFill("solid", fgColor="F1F5F9")
    italic = Font(italic=True, color="64748B")
    for col_idx in range(1, len(example_row) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = grey
        cell.font = italic
    ws.freeze_panes = "A2"

    def _add_ref_sheet(name: str, rows: List[dict], headers: List[str]):
        s = wb.create_sheet(name)
        s.append(headers)
        for c in s[1]:
            c.font = header_font
            c.fill = header_fill
        for r in rows:
            s.append([r.get(h.lower().replace(" ", "_"), "") for h in headers])
        for i, h in enumerate(headers, start=1):
            s.column_dimensions[get_column_letter(i)].width = max(14, len(h) + 4)

    _add_ref_sheet("Cadenas", hints["cadenas"], ["nombre", "codigo"])
    _add_ref_sheet(
        "Tiendas", hints["tiendas"],
        ["cadena", "nombre", "codigo_interno", "codigo_externo", "ciudad", "estado"],
    )
    _add_ref_sheet("Productos", hints["productos"], ["sku", "nombre", "precio_sugerido"])

    ws_i = wb.create_sheet("Instrucciones", 0)
    lines = [
        "Plantilla de Sell-out Retail",
        "",
        "Cómo llenarla:",
        "  1) Una fila = una venta reportada de UN SKU en UNA tienda en UN periodo.",
        "  2) Matcheo de cadena: por 'cadena_codigo' (recomendado); si no, por 'cadena_nombre' exacto.",
        "  3) Matcheo de tienda: por 'tienda_codigo' (el nº de tienda del cliente); si no, por nombre.",
        "  4) SKU: se busca en tu catálogo. Si no existe, se guarda como snapshot con el nombre que pongas.",
        "  5) 'periodo_tipo' acepta: day, week o month. Default: week.",
        "  6) Fechas en formato YYYY-MM-DD. Si dejas 'periodo_fin' vacío, se calcula según el tipo.",
        "  7) Si vuelves a subir una fila con el mismo (tienda, sku, periodo_inicio, periodo_tipo), se actualiza (no se duplica).",
        "  8) 'unidades_devueltas' e 'importe_devoluciones' son opcionales — si el reporte de la cadena las trae, cárgalas: el sistema calcula tasa de devoluciones y ventas netas, y alerta cuando pasan del umbral.",
        "  9) Consulta las hojas 'Cadenas', 'Tiendas' y 'Productos' para copiar los códigos exactos.",
    ]
    for row_ix, line in enumerate(lines, start=1):
        c = ws_i.cell(row=row_ix, column=1, value=line)
        if row_ix == 1:
            c.font = Font(bold=True, size=14, color="1E3A8A")
    ws_i.column_dimensions["A"].width = 110

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_sellout_template_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    today = date.today()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    writer.writerow([
        "WMT", "Walmart",
        "1001", "Sucursal Centro",
        "SKU-001", "Producto de ejemplo",
        "week", week_start.isoformat(), week_end.isoformat(),
        12, 1, 45, 12 * 1290.0, 1290.0,
        "Ejemplo — puedes borrar esta fila",
    ])
    return buf.getvalue().encode("utf-8-sig")


def _parse_int(v: Any) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _parse_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _parse_iso_date(v: Any) -> Optional[datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.replace(tzinfo=v.tzinfo or timezone.utc)
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day, tzinfo=timezone.utc)
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError:
        return None


def _default_period_end(start: datetime, period_type: str) -> datetime:
    if period_type == "day":
        return start.replace(hour=23, minute=59, second=59)
    if period_type == "month":
        # último instante del mes calendario
        y, m = start.year, start.month
        if m == 12:
            nxt = datetime(y + 1, 1, 1, tzinfo=start.tzinfo)
        else:
            nxt = datetime(y, m + 1, 1, tzinfo=start.tzinfo)
        return nxt - timedelta(seconds=1)
    # week (default): 6 días completos
    return (start + timedelta(days=6)).replace(hour=23, minute=59, second=59)


async def _resolve_import_hints(
    db: AsyncSession,
) -> Tuple[Dict[str, int], Dict[Tuple[int, str], int], Dict[Tuple[int, str], int],
            Dict[str, Tuple[int, str, float]]]:
    """Índices para matcheo rápido:
      - chan_by_key: 'CODE' → channel_id, 'NAME::name lower' → channel_id
      - store_by_ext: (channel_id, external_code upper) → store_id
      - store_by_name: (channel_id, name lower) → store_id
      - variant_by_sku: sku upper → (variant_id, product_name, price)
    """
    chans = (await db.execute(select(models.RetailChannel))).scalars().all()
    chan_by_key: Dict[str, int] = {}
    for c in chans:
        if c.code:
            chan_by_key[c.code.strip().upper()] = c.id
        if c.name:
            chan_by_key[f"NAME::{c.name.strip().lower()}"] = c.id

    stores = (await db.execute(select(models.RetailStore))).scalars().all()
    store_by_ext: Dict[Tuple[int, str], int] = {}
    store_by_name: Dict[Tuple[int, str], int] = {}
    for s in stores:
        if s.external_code:
            store_by_ext[(s.channel_id, s.external_code.strip().upper())] = s.id
        if s.code:
            store_by_ext[(s.channel_id, s.code.strip().upper())] = s.id
        if s.name:
            store_by_name[(s.channel_id, s.name.strip().lower())] = s.id

    variants = (await db.execute(
        select(inv_models.ProductVariant.id, inv_models.ProductVariant.sku,
                inv_models.ProductVariant.price, inv_models.Product.name)
        .join(inv_models.Product, inv_models.ProductVariant.product_id == inv_models.Product.id)
    )).all()
    variant_by_sku: Dict[str, Tuple[int, str, float]] = {}
    for vid, sku, price, pname in variants:
        if sku:
            variant_by_sku[sku.strip().upper()] = (vid, pname or "", float(price or 0.0))

    return chan_by_key, store_by_ext, store_by_name, variant_by_sku


async def import_sellout(
    db: AsyncSession, file_bytes: bytes, filename: str,
    user_id: Optional[int] = None,
) -> schemas.ImportSellOutResponse:
    """Ingesta bulk de sell-out. Detecta xlsx/csv por extensión.

    - Matchea cadena por código, luego por nombre.
    - Matchea tienda por código externo/interno, luego por nombre.
    - Vincula SKU si existe en catálogo; si no, guarda snapshot con el nombre.
    - Upsert por (store, variant, period_start, period_type).
    - Fila con datos vacíos → skipped.
    - Fila con error → guarda en `errors`, sigue con las demás.
    """
    name_lower = (filename or "").lower()
    is_xlsx = name_lower.endswith(".xlsx") or name_lower.endswith(".xlsm")

    rows: List[Dict[str, Any]] = []
    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        target = "SellOut" if "SellOut" in wb.sheetnames else wb.active.title
        ws = wb[target]
        header: Optional[List[str]] = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [str(c).strip().lower() if c is not None else "" for c in row]
                continue
            values = [c for c in row]
            if all(v is None or v == "" for v in values):
                continue
            rows.append(dict(zip(header, values)))
    else:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            data = {k.strip().lower(): v for k, v in r.items() if k}
            if all((v is None or str(v).strip() == "") for v in data.values()):
                continue
            rows.append(data)

    chan_by_key, store_by_ext, store_by_name, variant_by_sku = await _resolve_import_hints(db)

    created = 0
    updated = 0
    skipped = 0
    errors: List[schemas.ImportRowError] = []
    touched_reports: List[models.SellOutReport] = []

    for row_idx, row in enumerate(rows, start=2):
        chan_code = str(row.get("cadena_codigo") or "").strip().upper()
        chan_name = str(row.get("cadena_nombre") or "").strip()
        store_code = str(row.get("tienda_codigo") or "").strip().upper()
        store_name = str(row.get("tienda_nombre") or "").strip()
        sku = str(row.get("sku") or "").strip().upper()
        product_name = str(row.get("producto_nombre") or "").strip()
        period_type = str(row.get("periodo_tipo") or "week").strip().lower()
        if period_type not in ("day", "week", "month"):
            period_type = "week"

        # Fila totalmente vacía → skip
        if not (chan_code or chan_name or store_code or store_name or sku or product_name):
            skipped += 1
            continue

        # Cadena
        chan_id: Optional[int] = chan_by_key.get(chan_code) if chan_code else None
        if chan_id is None and chan_name:
            chan_id = chan_by_key.get(f"NAME::{chan_name.lower()}")
        if chan_id is None:
            errors.append(schemas.ImportRowError(
                row=row_idx,
                reason=f"Cadena no encontrada (código='{chan_code}', nombre='{chan_name}')",
            ))
            continue

        # Tienda
        store_id: Optional[int] = None
        if store_code:
            store_id = store_by_ext.get((chan_id, store_code))
        if store_id is None and store_name:
            store_id = store_by_name.get((chan_id, store_name.lower()))
        if store_id is None:
            errors.append(schemas.ImportRowError(
                row=row_idx,
                reason=f"Tienda no encontrada en cadena {chan_id} (código='{store_code}', nombre='{store_name}')",
            ))
            continue

        # SKU (opcional en catálogo)
        variant_id: Optional[int] = None
        matched_pname: Optional[str] = None
        matched_price: float = 0.0
        if sku and sku in variant_by_sku:
            variant_id, matched_pname, matched_price = variant_by_sku[sku]

        final_name = matched_pname or product_name or f"SKU {sku}" if sku else product_name
        if not final_name:
            errors.append(schemas.ImportRowError(
                row=row_idx, reason="Sin producto: pon SKU válido o producto_nombre.",
            ))
            continue

        # Fechas
        period_start = _parse_iso_date(row.get("periodo_inicio"))
        if period_start is None:
            errors.append(schemas.ImportRowError(
                row=row_idx, reason="periodo_inicio inválido o vacío (formato YYYY-MM-DD).",
            ))
            continue
        period_end = _parse_iso_date(row.get("periodo_fin"))
        if period_end is None:
            period_end = _default_period_end(period_start, period_type)

        units_sold = _parse_int(row.get("unidades_vendidas"))
        units_returned = _parse_int(row.get("unidades_devueltas"))
        units_on_hand = _parse_int(row.get("unidades_stock"))
        revenue = _parse_float(row.get("ingreso"))
        returns_amount = _parse_float(row.get("importe_devoluciones"))
        notes = (str(row.get("notas") or "").strip() or None)

        # Muchas cadenas reportan sólo unidades (sin pesos). Si el archivo no
        # trae ingreso pero el SKU matchea catálogo con precio de lista,
        # estimamos ingreso = unidades × precio para que el dashboard no se
        # quede en ceros. Igual para el importe de devoluciones.
        estimated_revenue = False
        if revenue <= 0 and units_sold > 0 and matched_price > 0:
            revenue = round(units_sold * matched_price, 2)
            estimated_revenue = True
        if returns_amount <= 0 and units_returned > 0 and matched_price > 0:
            returns_amount = round(units_returned * matched_price, 2)
            estimated_revenue = True
        if estimated_revenue:
            tag = "ingreso estimado por precio de lista"
            notes = f"{notes} · {tag}" if notes else tag.capitalize()

        # Upsert
        existing = (await db.execute(
            select(models.SellOutReport).where(
                models.SellOutReport.store_id == store_id,
                models.SellOutReport.variant_id == variant_id,
                models.SellOutReport.period_start == period_start,
                models.SellOutReport.period_type == period_type,
            )
        )).scalars().first()

        if existing:
            existing.period_end = period_end
            existing.units_sold = units_sold
            existing.units_returned = units_returned
            existing.units_on_hand = units_on_hand
            existing.revenue = revenue
            existing.returns_amount = returns_amount
            existing.notes = notes
            existing.product_name = final_name
            existing.sku = sku or existing.sku
            existing.source = "xlsx" if is_xlsx else "csv"
            existing.uploaded_by_user_id = user_id or existing.uploaded_by_user_id
            updated += 1
            touched_reports.append(existing)
        else:
            r = models.SellOutReport(
                store_id=store_id, variant_id=variant_id,
                product_name=final_name, sku=sku or None,
                period_start=period_start, period_end=period_end,
                period_type=period_type,
                units_sold=units_sold, units_returned=units_returned,
                units_on_hand=units_on_hand,
                revenue=revenue, returns_amount=returns_amount,
                source="xlsx" if is_xlsx else "csv",
                uploaded_by_user_id=user_id, notes=notes,
            )
            db.add(r)
            created += 1
            touched_reports.append(r)

    await db.commit()

    # Aplicar movimientos de consignación SOLO a los reports afectados por
    # este import (los que quedaron actualizados o creados con store en
    # consignación). Idempotente por stock_consumed, así que reprocesar no
    # duplica descuentos; el filtro es para no barrer 2000 reports en cada
    # import cuando la base crece.
    touched_ids = [r.id for r in touched_reports if r is not None]
    if touched_ids:
        fresh = (await db.execute(
            select(models.SellOutReport)
            .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
            .where(
                models.SellOutReport.id.in_(touched_ids),
                models.RetailStore.consignment_warehouse_id.isnot(None),
            )
        )).scalars().all()
        for r in fresh:
            try:
                await _apply_consignment_movement(db, r, user_id)
            except Exception as e:
                log.warning("consignment hook falló en import row report=%s: %s", r.id, e)
                continue

    # Auto-evalúa alertas para las cadenas tocadas
    try:
        await evaluate_alerts(db)
    except Exception as e:
        log.warning("evaluate_alerts falló tras import_sellout: %s", e)

    return schemas.ImportSellOutResponse(
        total_rows=len(rows), created=created, updated=updated,
        skipped=skipped, errors=errors,
    )


# ── Alerts engine ───────────────────────────────────────────────────────

# Cuánto tiempo debe pasar para poder crear una alerta nueva del mismo
# tipo+store+variant tras cerrarse la anterior. Evita ruido de reapertura.
_ALERT_REOPEN_COOLDOWN = timedelta(hours=6)


def _severity_for_wos(wos: float, critical: float, target: float) -> str:
    if wos <= 0:
        return "urgent"
    if wos < critical:
        return "urgent"
    if wos < critical * 1.5 or wos < target * 0.5:
        return "high"
    return "medium"


def _is_wos_healthy(wos: float, critical: float, target: float, overstock: float) -> bool:
    return critical <= wos <= overstock


async def _summarize_alerts(db: AsyncSession, channel_id: Optional[int] = None
                              ) -> Tuple[int, int]:
    q = select(func.count(models.RetailAlert.id)).where(
        models.RetailAlert.status.in_(("open", "acknowledged"))
    )
    if channel_id is not None:
        q = q.where(models.RetailAlert.channel_id == channel_id)
    total_open = int((await db.execute(q)).scalar() or 0)

    q2 = q.where(models.RetailAlert.severity == "urgent")
    urgent = int((await db.execute(q2)).scalar() or 0)
    return total_open, urgent


async def evaluate_alerts(
    db: AsyncSession, channel_id: Optional[int] = None,
) -> schemas.EvaluateAlertsResponse:
    """Recorre stores × variants con sell-out reciente y genera / auto-resuelve
    alertas según las reglas de la cadena. Idempotente.
    """
    now = datetime.now(timezone.utc)
    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)

    # Cadenas a considerar
    ch_stmt = select(models.RetailChannel).where(
        models.RetailChannel.alerts_enabled.is_(True),
        models.RetailChannel.is_active.is_(True),
    )
    if channel_id is not None:
        ch_stmt = ch_stmt.where(models.RetailChannel.id == channel_id)
    channels = list((await db.execute(ch_stmt)).scalars().all())
    if not channels:
        total_open, urgent = await _summarize_alerts(db, channel_id)
        return schemas.EvaluateAlertsResponse(
            created=0, auto_resolved=0, total_open=total_open, urgent_open=urgent,
        )

    channel_ids = [c.id for c in channels]

    # Combos (store, variant) con sell-out en ventana + último período por combo
    combos_stmt = (
        select(
            models.RetailStore.id.label("store_id"),
            models.RetailStore.name.label("store_name"),
            models.RetailStore.channel_id.label("channel_id"),
            models.SellOutReport.variant_id,
            models.SellOutReport.product_name,
            models.SellOutReport.sku,
            func.max(models.SellOutReport.period_start).label("last_period"),
            func.max(models.SellOutReport.period_end).label("last_period_end"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units_window"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(
            models.RetailStore.channel_id.in_(channel_ids),
            models.RetailStore.is_active.is_(True),
        )
        .group_by(
            models.RetailStore.id, models.RetailStore.name,
            models.RetailStore.channel_id,
            models.SellOutReport.variant_id,
            models.SellOutReport.product_name,
            models.SellOutReport.sku,
        )
    )
    combos = (await db.execute(combos_stmt)).all()

    # Para el sell-through: sell-out por cadena en ventana
    sell_out_by_channel: Dict[int, int] = {}
    for r in combos:
        # Únicamente sumamos ventas en la ventana
        # (Se recalcula abajo por precisión sobre period_start)
        pass

    from app.modules.sales import models as sales_models

    sellout_agg = (await db.execute(
        select(
            models.RetailStore.channel_id,
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(models.SellOutReport.period_start >= velocity_from,
                models.RetailStore.channel_id.in_(channel_ids))
        .group_by(models.RetailStore.channel_id)
    )).all()
    for cid, u in sellout_agg:
        sell_out_by_channel[int(cid)] = int(u or 0)

    # Sell-in por cadena (con customer vinculado)
    sell_in_by_channel: Dict[int, int] = {ch.id: 0 for ch in channels}
    cust_map = {ch.id: ch.customer_id for ch in channels if ch.customer_id}
    if cust_map:
        cust_ids = list(cust_map.values())
        si_rows = (await db.execute(
            select(
                sales_models.Order.customer_id,
                func.coalesce(func.sum(sales_models.OrderItem.quantity), 0),
            )
            .join(sales_models.OrderItem, sales_models.OrderItem.order_id == sales_models.Order.id)
            .where(
                sales_models.Order.customer_id.in_(cust_ids),
                sales_models.Order.kind == "order",
                sales_models.Order.status != "cancelled",
                sales_models.Order.created_at >= velocity_from,
            )
            .group_by(sales_models.Order.customer_id)
        )).all()
        cust_to_ch = {v: k for k, v in cust_map.items()}
        for cust_id, u in si_rows:
            ch_id = cust_to_ch.get(int(cust_id))
            if ch_id:
                sell_in_by_channel[ch_id] = int(u or 0)

    # Alertas abiertas actuales — para dedupe y auto-resolve
    open_alerts_rows = (await db.execute(
        select(models.RetailAlert).where(
            models.RetailAlert.channel_id.in_(channel_ids),
            models.RetailAlert.status.in_(("open", "acknowledged")),
        )
    )).scalars().all()
    open_by_key: Dict[Tuple[str, int, Optional[int]], models.RetailAlert] = {}
    for a in open_alerts_rows:
        open_by_key[(a.alert_type, a.store_id, a.variant_id)] = a

    # Índice por canal (config)
    ch_by_id: Dict[int, models.RetailChannel] = {c.id: c for c in channels}

    created = 0
    auto_resolved = 0
    seen_keys: set = set()
    now_dt = now

    async def _upsert(
        alert_type: str, store_id: int, variant_id: Optional[int],
        channel_id_val: int, message: str, severity: str,
        wos: Optional[float], on_hand: Optional[int], velocity: Optional[float],
        store_name: Optional[str], product_name: Optional[str], sku: Optional[str],
    ):
        nonlocal created
        key = (alert_type, store_id, variant_id)
        seen_keys.add(key)
        existing = open_by_key.get(key)
        if existing:
            # Update snapshot para reflejar el estado más reciente
            existing.severity = severity
            existing.message = message
            existing.wos_snapshot = wos
            existing.on_hand_snapshot = on_hand
            existing.weekly_velocity_snapshot = velocity
            existing.store_name = store_name
            existing.product_name = product_name
            existing.sku = sku
            return
        # Verifica cooldown: no reabrir si acabamos de cerrar la misma
        prev_stmt = (
            select(models.RetailAlert)
            .where(
                models.RetailAlert.alert_type == alert_type,
                models.RetailAlert.store_id == store_id,
                models.RetailAlert.variant_id == variant_id,
                models.RetailAlert.status.in_(("resolved", "dismissed")),
                models.RetailAlert.resolved_at.isnot(None),
            )
            .order_by(models.RetailAlert.resolved_at.desc())
            .limit(1)
        )
        prev = (await db.execute(prev_stmt)).scalars().first()
        if prev and prev.resolved_at and (now_dt - prev.resolved_at) < _ALERT_REOPEN_COOLDOWN:
            return
        a = models.RetailAlert(
            channel_id=channel_id_val, store_id=store_id, variant_id=variant_id,
            alert_type=alert_type, severity=severity, message=message,
            wos_snapshot=wos, on_hand_snapshot=on_hand,
            weekly_velocity_snapshot=velocity,
            store_name=store_name, product_name=product_name, sku=sku,
            status="open",
        )
        db.add(a)
        created += 1

    # Evaluación combo por combo
    for r in combos:
        ch = ch_by_id.get(int(r.channel_id))
        if ch is None:
            continue
        # On-hand actual: suma del último período de ese (store, variant)
        on_hand = int((await db.execute(
            select(func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0))
            .where(
                models.SellOutReport.store_id == r.store_id,
                models.SellOutReport.variant_id == r.variant_id,
                models.SellOutReport.period_start == r.last_period,
            )
        )).scalar() or 0)

        # Solo la ventana de velocidad
        units_win_row = (await db.execute(
            select(func.coalesce(func.sum(models.SellOutReport.units_sold), 0))
            .where(
                models.SellOutReport.store_id == r.store_id,
                models.SellOutReport.variant_id == r.variant_id,
                models.SellOutReport.period_start >= velocity_from,
            )
        )).scalar() or 0
        units_win = int(units_win_row)
        velocity = units_win / (VELOCITY_WINDOW_DAYS / 7.0) if units_win > 0 else 0.0

        wos = (on_hand / velocity) if velocity > 0 else WOS_INFINITY

        product_name = r.product_name
        sku = r.sku
        store_name = r.store_name

        # Regla stockout puro
        if on_hand == 0 and velocity > 0:
            await _upsert(
                "stockout", r.store_id, r.variant_id, ch.id,
                message=(f"Sin stock en {store_name} · {product_name or sku or 'SKU'}. "
                        f"Velocidad {velocity:.1f} u/sem — venta detenida."),
                severity="urgent",
                wos=0.0, on_hand=0, velocity=velocity,
                store_name=store_name, product_name=product_name, sku=sku,
            )
        # Regla stockout_imminent
        elif velocity > 0 and wos < ch.critical_wos_weeks:
            sev = _severity_for_wos(wos, ch.critical_wos_weeks, ch.target_wos_weeks)
            await _upsert(
                "stockout_imminent", r.store_id, r.variant_id, ch.id,
                message=(f"WOS crítico {wos:.1f} sem en {store_name} · "
                        f"{product_name or sku or 'SKU'} (mín {ch.critical_wos_weeks:.0f})."),
                severity=sev,
                wos=round(wos, 2), on_hand=on_hand, velocity=velocity,
                store_name=store_name, product_name=product_name, sku=sku,
            )
        # Regla overstock
        if velocity > 0 and wos > ch.overstock_wos_weeks and wos < WOS_INFINITY:
            await _upsert(
                "overstock", r.store_id, r.variant_id, ch.id,
                message=(f"Sobreinventario {wos:.1f} sem en {store_name} · "
                        f"{product_name or sku or 'SKU'} (máx {ch.overstock_wos_weeks:.0f})."),
                severity="medium",
                wos=round(wos, 2), on_hand=on_hand, velocity=velocity,
                store_name=store_name, product_name=product_name, sku=sku,
            )
        # Regla no_movement (con on_hand > 0 y ventana sin ventas)
        no_move_days = int(ch.no_movement_days or 21)
        no_move_threshold = now - timedelta(days=no_move_days)
        if on_hand > 0 and velocity == 0 and r.last_period_end is not None:
            last_sale_stmt = (
                select(func.max(models.SellOutReport.period_end))
                .where(
                    models.SellOutReport.store_id == r.store_id,
                    models.SellOutReport.variant_id == r.variant_id,
                    models.SellOutReport.units_sold > 0,
                )
            )
            last_sale = (await db.execute(last_sale_stmt)).scalar()
            if last_sale is not None and last_sale.tzinfo is None:
                last_sale = last_sale.replace(tzinfo=timezone.utc)
            if last_sale is None or last_sale < no_move_threshold:
                await _upsert(
                    "no_movement", r.store_id, r.variant_id, ch.id,
                    message=(f"Sin ventas > {no_move_days} días en {store_name} · "
                            f"{product_name or sku or 'SKU'} · {on_hand} en stock."),
                    severity="high",
                    wos=None, on_hand=on_hand, velocity=0.0,
                    store_name=store_name, product_name=product_name, sku=sku,
                )

    # Regla a nivel cadena: high_return_rate
    #   Suma devoluciones y ventas por cadena en la ventana. Si la tasa
    #   pasa del umbral configurado en la cadena, levanta alerta.
    returns_by_channel = (await db.execute(
        select(
            models.RetailStore.channel_id,
            func.coalesce(func.sum(models.SellOutReport.units_returned), 0).label("returned"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("sold"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(models.SellOutReport.period_start >= velocity_from,
                models.RetailStore.channel_id.in_(channel_ids))
        .group_by(models.RetailStore.channel_id)
    )).all()
    ret_map: Dict[int, Tuple[int, int]] = {
        int(cid): (int(ret or 0), int(sold or 0)) for cid, ret, sold in returns_by_channel
    }
    for ch in channels:
        ret_units, sold_units = ret_map.get(ch.id, (0, 0))
        if sold_units <= 0 or ret_units <= 0:
            continue
        rate = ret_units / sold_units * 100.0
        threshold = float(ch.return_rate_max_pct or 5.0)
        if rate < threshold:
            continue
        rep_store = (await db.execute(
            select(models.RetailStore.id, models.RetailStore.name).where(
                models.RetailStore.channel_id == ch.id,
                models.RetailStore.is_active.is_(True),
            ).limit(1)
        )).first()
        if rep_store is None:
            continue
        store_id_val, store_name_val = int(rep_store[0]), rep_store[1]
        sev = "urgent" if rate > threshold * 2 else "high"
        await _upsert(
            "high_return_rate", store_id_val, None, ch.id,
            message=(f"Tasa de devoluciones {rate:.1f}% en {ch.name} "
                     f"(máx {threshold:.1f}%). {ret_units} u devueltas / "
                     f"{sold_units} u vendidas en {VELOCITY_WINDOW_DAYS} días."),
            severity=sev,
            wos=None, on_hand=None, velocity=None,
            store_name=store_name_val, product_name=None, sku=None,
        )

    # Regla a nivel cadena: sell_through_low
    for ch in channels:
        so = sell_out_by_channel.get(ch.id, 0)
        si = sell_in_by_channel.get(ch.id, 0)
        if si <= 0:
            continue
        pct = so / si * 100.0
        if pct < ch.sell_through_min_pct:
            # store_id/variant_id no aplican; usamos alguna tienda representativa
            rep_store = (await db.execute(
                select(models.RetailStore.id, models.RetailStore.name).where(
                    models.RetailStore.channel_id == ch.id,
                    models.RetailStore.is_active.is_(True),
                ).limit(1)
            )).first()
            if rep_store is None:
                continue
            store_id_val, store_name_val = int(rep_store[0]), rep_store[1]
            await _upsert(
                "sell_through_low", store_id_val, None, ch.id,
                message=(f"Sell-through de {ch.name} en {pct:.1f}% "
                        f"(mín {ch.sell_through_min_pct:.0f}%). Sell-in {si} u "
                        f"vs sell-out {so} u en {VELOCITY_WINDOW_DAYS} días."),
                severity="medium",
                wos=None, on_hand=None, velocity=None,
                store_name=store_name_val, product_name=None, sku=None,
            )

    # Auto-resolve: alertas abiertas cuya condición YA NO SE CUMPLE
    for key, alert in open_by_key.items():
        if key in seen_keys:
            continue
        # No la volvimos a levantar → auto-resolve
        alert.status = "resolved"
        alert.resolved_at = now
        alert.resolution_notes = "Condición vuelta a zona sana"
        auto_resolved += 1

    await db.commit()

    total_open, urgent = await _summarize_alerts(db, channel_id)
    return schemas.EvaluateAlertsResponse(
        created=created, auto_resolved=auto_resolved,
        total_open=total_open, urgent_open=urgent,
    )


# ── Consulta / acciones sobre alertas ────────────────────────────────────

def _alerts_filter_conds(
    channel_id: Optional[int], status: Optional[str], severity: Optional[str],
    alert_type: Optional[str], q: Optional[str],
) -> list:
    conds = []
    if channel_id is not None:
        conds.append(models.RetailAlert.channel_id == channel_id)
    if status:
        conds.append(models.RetailAlert.status == status)
    if severity:
        conds.append(models.RetailAlert.severity == severity)
    if alert_type:
        conds.append(models.RetailAlert.alert_type == alert_type)
    if q:
        needle = f"%{q.strip()}%"
        conds.append(or_(
            models.RetailAlert.store_name.ilike(needle),
            models.RetailAlert.product_name.ilike(needle),
            models.RetailAlert.sku.ilike(needle),
            models.RetailAlert.message.ilike(needle),
        ))
    return conds


async def count_alerts(
    db: AsyncSession, channel_id: Optional[int] = None,
    status: Optional[str] = None, severity: Optional[str] = None,
    alert_type: Optional[str] = None, q: Optional[str] = None,
) -> int:
    stmt = select(func.count(models.RetailAlert.id))
    conds = _alerts_filter_conds(channel_id, status, severity, alert_type, q)
    if conds:
        stmt = stmt.where(and_(*conds))
    return int((await db.execute(stmt)).scalar() or 0)


async def list_alerts(
    db: AsyncSession, channel_id: Optional[int] = None,
    status: Optional[str] = None, severity: Optional[str] = None,
    alert_type: Optional[str] = None, q: Optional[str] = None,
    limit: int = 500, offset: int = 0,
) -> List[schemas.RetailAlertOut]:
    stmt = (
        select(
            models.RetailAlert,
            models.RetailChannel.name.label("channel_name"),
        )
        .join(models.RetailChannel, models.RetailAlert.channel_id == models.RetailChannel.id)
    )
    conds = _alerts_filter_conds(channel_id, status, severity, alert_type, q)
    if conds:
        stmt = stmt.where(and_(*conds))
    stmt = stmt.order_by(
        # Prioridad urgent primero, luego por fecha
        func.lower(models.RetailAlert.status) == "open",  # true (1) → open primero
        models.RetailAlert.severity == "urgent",
        models.RetailAlert.created_at.desc(),
    ).offset(offset).limit(limit)
    rows = (await db.execute(stmt)).all()
    return [
        schemas.RetailAlertOut(
            id=a.id, channel_id=a.channel_id, channel_name=cn,
            store_id=a.store_id, store_name=a.store_name,
            variant_id=a.variant_id, product_name=a.product_name, sku=a.sku,
            alert_type=a.alert_type, severity=a.severity, message=a.message,
            wos_snapshot=a.wos_snapshot,
            on_hand_snapshot=a.on_hand_snapshot,
            weekly_velocity_snapshot=a.weekly_velocity_snapshot,
            status=a.status,
            acknowledged_at=a.acknowledged_at,
            acknowledged_by_user_id=a.acknowledged_by_user_id,
            resolved_at=a.resolved_at,
            resolved_by_user_id=a.resolved_by_user_id,
            resolution_notes=a.resolution_notes,
            created_at=a.created_at,
        ) for a, cn in rows
    ]


async def alerts_summary(db: AsyncSession, channel_id: Optional[int] = None
                          ) -> schemas.AlertsSummary:
    def _count(cond) -> int:
        stmt = select(func.count(models.RetailAlert.id)).where(cond)
        if channel_id is not None:
            stmt = stmt.where(models.RetailAlert.channel_id == channel_id)
        # Ejecuta síncrono in-async via caller
        return stmt

    open_c = int((await db.execute(_count(models.RetailAlert.status == "open"))).scalar() or 0)
    ack_c = int((await db.execute(_count(models.RetailAlert.status == "acknowledged"))).scalar() or 0)

    def _sev(sev: str):
        return _count(and_(
            models.RetailAlert.status.in_(("open", "acknowledged")),
            models.RetailAlert.severity == sev,
        ))

    urgent = int((await db.execute(_sev("urgent"))).scalar() or 0)
    high = int((await db.execute(_sev("high"))).scalar() or 0)
    medium = int((await db.execute(_sev("medium"))).scalar() or 0)
    low = int((await db.execute(_sev("low"))).scalar() or 0)

    return schemas.AlertsSummary(
        open=open_c, urgent=urgent, high=high, medium=medium, low=low,
        acknowledged=ack_c,
    )


async def acknowledge_alert(db: AsyncSession, alert_id: int,
                              user_id: Optional[int], notes: Optional[str] = None
                              ) -> Optional[models.RetailAlert]:
    a = await db.get(models.RetailAlert, alert_id)
    if a is None:
        return None
    if a.status not in ("open", "acknowledged"):
        raise ValueError(f"La alerta ya está {a.status}")
    a.status = "acknowledged"
    a.acknowledged_at = datetime.now(timezone.utc)
    a.acknowledged_by_user_id = user_id
    if notes:
        a.resolution_notes = notes
    await db.commit()
    await db.refresh(a)
    return a


async def resolve_alert(db: AsyncSession, alert_id: int,
                          user_id: Optional[int], notes: Optional[str] = None
                          ) -> Optional[models.RetailAlert]:
    a = await db.get(models.RetailAlert, alert_id)
    if a is None:
        return None
    if a.status in ("resolved", "dismissed"):
        raise ValueError(f"La alerta ya está {a.status}")
    a.status = "resolved"
    a.resolved_at = datetime.now(timezone.utc)
    a.resolved_by_user_id = user_id
    if notes:
        a.resolution_notes = notes
    await db.commit()
    await db.refresh(a)
    return a


async def dismiss_alert(db: AsyncSession, alert_id: int,
                         user_id: Optional[int], notes: Optional[str] = None
                         ) -> Optional[models.RetailAlert]:
    a = await db.get(models.RetailAlert, alert_id)
    if a is None:
        return None
    if a.status in ("resolved", "dismissed"):
        raise ValueError(f"La alerta ya está {a.status}")
    a.status = "dismissed"
    a.resolved_at = datetime.now(timezone.utc)
    a.resolved_by_user_id = user_id
    if notes:
        a.resolution_notes = notes
    await db.commit()
    await db.refresh(a)
    return a


async def _alert_to_schema(db: AsyncSession, a: models.RetailAlert) -> schemas.RetailAlertOut:
    ch = await db.get(models.RetailChannel, a.channel_id)
    return schemas.RetailAlertOut(
        id=a.id, channel_id=a.channel_id,
        channel_name=ch.name if ch else None,
        store_id=a.store_id, store_name=a.store_name,
        variant_id=a.variant_id, product_name=a.product_name, sku=a.sku,
        alert_type=a.alert_type, severity=a.severity, message=a.message,
        wos_snapshot=a.wos_snapshot,
        on_hand_snapshot=a.on_hand_snapshot,
        weekly_velocity_snapshot=a.weekly_velocity_snapshot,
        status=a.status,
        acknowledged_at=a.acknowledged_at,
        acknowledged_by_user_id=a.acknowledged_by_user_id,
        resolved_at=a.resolved_at,
        resolved_by_user_id=a.resolved_by_user_id,
        resolution_notes=a.resolution_notes,
        created_at=a.created_at,
    )


# ── Analytics: heatmap tiendas × SKUs ────────────────────────────────────

async def heatmap(
    db: AsyncSession, channel_id: Optional[int] = None,
    metric: str = "wos", limit_variants: int = 40,
    store_search: Optional[str] = None,
    region: Optional[str] = None,
    state: Optional[str] = None,
    store_format: Optional[str] = None,
    store_offset: int = 0,
    store_limit: int = 100,
    sort_stores_by: str = "name",  # name | worst_wos | best_wos | most_sales
) -> schemas.HeatmapResponse:
    """Matriz tienda × SKU escalable para catálogos grandes.

    Diseño para soportar cadenas tipo Coppel (1100+ tiendas) y catálogos
    de 1000+ SKUs sin colapsar el frontend:
      - Paginación server-side de tiendas (default 100 por página).
      - Filtros por región, estado, formato de tienda y búsqueda por
        nombre/código externo.
      - Ordenamiento server-side por worst_wos (para atender primero
        las críticas), best_wos, most_sales o name.
      - Top-N variants (default 40) por ventas totales — el resto del
        catálogo queda oculto pero el UI puede pedir más con
        limit_variants más alto.
    El response incluye total_stores/total_variants para el paginador y
    para saber si hay que sugerir refinar el filtro.
    """
    now = datetime.now(timezone.utc)
    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)

    # 1) Base de tiendas + facetas
    base_stmt = (
        select(models.RetailStore, models.RetailChannel)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
        .where(models.RetailStore.is_active.is_(True))
    )
    if channel_id is not None:
        base_stmt = base_stmt.where(models.RetailStore.channel_id == channel_id)
    if region:
        base_stmt = base_stmt.where(models.RetailStore.region == region)
    if state:
        base_stmt = base_stmt.where(models.RetailStore.state == state)
    if store_format:
        base_stmt = base_stmt.where(models.RetailStore.store_format == store_format)
    if store_search:
        needle = f"%{store_search.strip()}%"
        base_stmt = base_stmt.where(
            or_(
                models.RetailStore.name.ilike(needle),
                models.RetailStore.external_code.ilike(needle),
                models.RetailStore.code.ilike(needle),
                models.RetailStore.city.ilike(needle),
            )
        )

    all_stores = (await db.execute(base_stmt)).all()
    total_stores = len(all_stores)

    if total_stores == 0:
        return schemas.HeatmapResponse(
            channel_id=channel_id, metric=metric,
            stores=[], variants=[], cells=[],
            total_stores=0, total_variants=0,
            store_offset=0, store_limit=store_limit,
        )

    # 2) Top-N variants por ventas en la ventana (restringido al mismo scope)
    all_store_ids = [s.id for s, _ in all_stores]

    top_stmt = (
        select(
            models.SellOutReport.variant_id,
            func.max(models.SellOutReport.sku).label("sku"),
            func.max(models.SellOutReport.product_name).label("product_name"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units"),
        )
        .where(
            models.SellOutReport.period_start >= velocity_from,
            models.SellOutReport.variant_id.isnot(None),
            models.SellOutReport.store_id.in_(all_store_ids),
        )
        .group_by(models.SellOutReport.variant_id)
        .order_by(func.sum(models.SellOutReport.units_sold).desc())
    )
    all_top = (await db.execute(top_stmt)).all()
    total_variants = len(all_top)
    top_slice = all_top[:limit_variants]
    variants = [
        schemas.HeatmapVariantRef(
            id=int(v.variant_id), sku=v.sku, product_name=v.product_name,
        ) for v in top_slice if v.variant_id
    ]
    variant_ids = [v.id for v in variants]

    if not variant_ids:
        # Sin ventas todavía en el scope; devuelve tiendas para que el UI
        # las muestre y ayude al usuario a entender el estado.
        page = all_stores[store_offset:store_offset + store_limit]
        return schemas.HeatmapResponse(
            channel_id=channel_id, metric=metric,
            stores=[schemas.HeatmapStoreRef(
                id=s.id, name=s.name, channel_name=ch.name,
            ) for s, ch in page],
            variants=[], cells=[],
            total_stores=total_stores, total_variants=0,
            store_offset=store_offset, store_limit=store_limit,
        )

    # 3) Ventas por (store, variant) en la ventana — solo top variants
    sales_map: Dict[Tuple[int, int], int] = {}
    sales_rows = (await db.execute(
        select(
            models.SellOutReport.store_id,
            models.SellOutReport.variant_id,
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units"),
        )
        .where(
            models.SellOutReport.period_start >= velocity_from,
            models.SellOutReport.variant_id.in_(variant_ids),
            models.SellOutReport.store_id.in_(all_store_ids),
        )
        .group_by(models.SellOutReport.store_id, models.SellOutReport.variant_id)
    )).all()
    for r in sales_rows:
        sales_map[(int(r.store_id), int(r.variant_id))] = int(r.units or 0)

    # 4) On-hand del último period_start por (store, variant)
    last_stmt = (
        select(
            models.SellOutReport.store_id,
            models.SellOutReport.variant_id,
            func.max(models.SellOutReport.period_start).label("last_period"),
        )
        .where(
            models.SellOutReport.variant_id.in_(variant_ids),
            models.SellOutReport.store_id.in_(all_store_ids),
        )
        .group_by(models.SellOutReport.store_id, models.SellOutReport.variant_id)
    )
    last_rows = (await db.execute(last_stmt)).all()
    on_hand_map: Dict[Tuple[int, int], int] = {}
    # Consolidar en una sola query para rendir con muchos combos
    if last_rows:
        # Construimos WHERE ((store_id=?, variant_id=?, period_start=?), ...)
        # con expresión SQL general: uso subquery IN para simplicidad.
        keys = [(int(r.store_id), int(r.variant_id), r.last_period) for r in last_rows if r.last_period]
        if keys:
            store_ids_k = list({k[0] for k in keys})
            variant_ids_k = list({k[1] for k in keys})
            oh_rows = (await db.execute(
                select(
                    models.SellOutReport.store_id,
                    models.SellOutReport.variant_id,
                    models.SellOutReport.period_start,
                    func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0).label("oh"),
                )
                .where(
                    models.SellOutReport.store_id.in_(store_ids_k),
                    models.SellOutReport.variant_id.in_(variant_ids_k),
                )
                .group_by(
                    models.SellOutReport.store_id,
                    models.SellOutReport.variant_id,
                    models.SellOutReport.period_start,
                )
            )).all()
            oh_by_key = {(int(r.store_id), int(r.variant_id), r.period_start): int(r.oh) for r in oh_rows}
            for (sid, vid, last_period) in keys:
                on_hand_map[(sid, vid)] = oh_by_key.get((sid, vid, last_period), 0)

    # 5) Métricas agregadas por tienda para ordenar
    def _store_score(store) -> Tuple[float, float, float]:
        """Regresa (worst_wos_finite, total_units, name_hash) del store en el
        conjunto de top variants. worst_wos = mínimo WOS de las celdas con
        ventas (rojo primero); ∞ si no hay ventas. total_units para
        most_sales."""
        min_wos = WOS_INFINITY
        tot = 0
        for vid in variant_ids:
            u = sales_map.get((store.id, vid), 0)
            oh = on_hand_map.get((store.id, vid), 0)
            v = (u / (VELOCITY_WINDOW_DAYS / 7.0)) if u > 0 else 0.0
            if v > 0:
                w = oh / v
                if w < min_wos:
                    min_wos = w
            tot += u
        return (min_wos, float(tot), 0.0)

    stores_with_score = [(s, ch, _store_score(s)) for (s, ch) in all_stores]

    if sort_stores_by == "worst_wos":
        stores_with_score.sort(key=lambda x: (x[2][0], -x[2][1]))
    elif sort_stores_by == "best_wos":
        stores_with_score.sort(key=lambda x: (-x[2][0], -x[2][1]))
    elif sort_stores_by == "most_sales":
        stores_with_score.sort(key=lambda x: (-x[2][1], x[2][0]))
    else:  # name
        stores_with_score.sort(key=lambda x: (x[0].name or "").lower())

    # 6) Paginar tiendas
    page_slice = stores_with_score[store_offset:store_offset + store_limit]

    # 7) Cells solo para la página
    cells: List[schemas.HeatmapCell] = []
    for store, ch, _sc in page_slice:
        for v in variants:
            units = sales_map.get((store.id, v.id), 0)
            on_hand = on_hand_map.get((store.id, v.id), 0)
            velocity = units / (VELOCITY_WINDOW_DAYS / 7.0) if units > 0 else 0.0
            wos = (on_hand / velocity) if velocity > 0 else WOS_INFINITY
            status = _wos_status(
                wos, ch.critical_wos_weeks, ch.target_wos_weeks,
                ch.overstock_wos_weeks, has_sales=velocity > 0,
            )
            if metric == "units_sold":
                value: Optional[float] = float(units)
            elif metric == "on_hand":
                value = float(on_hand)
            else:
                value = round(wos, 2) if wos < WOS_INFINITY else None
            cells.append(schemas.HeatmapCell(
                store_id=store.id, variant_id=v.id,
                value=value, on_hand=on_hand, units_sold=units,
                status=status,
            ))

    stores_out = [
        schemas.HeatmapStoreRef(id=s.id, name=s.name, channel_name=ch.name)
        for s, ch, _sc in page_slice
    ]
    return schemas.HeatmapResponse(
        channel_id=channel_id, metric=metric,
        stores=stores_out, variants=variants, cells=cells,
        total_stores=total_stores, total_variants=total_variants,
        store_offset=store_offset, store_limit=store_limit,
    )


async def heatmap_filters(
    db: AsyncSession, channel_id: Optional[int] = None,
) -> schemas.HeatmapFilters:
    """Devuelve valores únicos de region/state/store_format para las tiendas
    activas del scope, para poblar los selects del filtro."""
    stmt = select(
        models.RetailStore.region, models.RetailStore.state,
        models.RetailStore.store_format,
    ).where(models.RetailStore.is_active.is_(True))
    if channel_id is not None:
        stmt = stmt.where(models.RetailStore.channel_id == channel_id)
    rows = (await db.execute(stmt)).all()
    regions = sorted({r.region for r in rows if r.region})
    states = sorted({r.state for r in rows if r.state})
    formats = sorted({r.store_format for r in rows if r.store_format})
    return schemas.HeatmapFilters(regions=regions, states=states, formats=formats)


# ── Analytics: clasificación ABC ─────────────────────────────────────────

async def abc_classification(
    db: AsyncSession, channel_id: Optional[int] = None, days: int = 90,
) -> schemas.ABCResponse:
    """SKUs ordenados por revenue descendente. Suma acumulada:
    hasta 80% → clase A, hasta 95% → B, resto → C."""
    now = datetime.now(timezone.utc)
    since = now - timedelta(days=days)

    stmt = (
        select(
            models.SellOutReport.variant_id,
            func.max(models.SellOutReport.sku).label("sku"),
            func.max(models.SellOutReport.product_name).label("product_name"),
            func.count(func.distinct(models.SellOutReport.store_id)).label("stores"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units"),
            func.coalesce(func.sum(models.SellOutReport.revenue), 0.0).label("revenue"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(models.SellOutReport.period_start >= since)
        .group_by(models.SellOutReport.variant_id)
        .order_by(func.sum(models.SellOutReport.revenue).desc())
    )
    if channel_id is not None:
        stmt = stmt.where(models.RetailStore.channel_id == channel_id)
    rows = (await db.execute(stmt)).all()

    total_rev = float(sum(float(r.revenue or 0.0) for r in rows))
    result: List[schemas.ABCRow] = []
    if total_rev <= 0:
        return schemas.ABCResponse(
            channel_id=channel_id, total_revenue=0.0,
            class_a_count=0, class_b_count=0, class_c_count=0, rows=[],
        )

    cum = 0.0
    a_count = b_count = c_count = 0
    for i, r in enumerate(rows, start=1):
        rev = float(r.revenue or 0.0)
        pct = round(rev / total_rev * 100.0, 2)
        # Clasificación por acumulado ANTES de este SKU: si el previo era
        # <=80%, el SKU está dentro del top-80 → A. Regla clásica de Pareto:
        # el SKU cuya suma cruza la frontera aún cuenta en la clase donde
        # cayó su base.
        prev_pct = round(cum / total_rev * 100.0, 2)
        cum += rev
        cum_pct = round(cum / total_rev * 100.0, 2)
        if prev_pct < 80.0:
            cls = "A"; a_count += 1
        elif prev_pct < 95.0:
            cls = "B"; b_count += 1
        else:
            cls = "C"; c_count += 1
        result.append(schemas.ABCRow(
            rank=i,
            variant_id=int(r.variant_id) if r.variant_id else None,
            sku=r.sku, product_name=r.product_name,
            stores_count=int(r.stores or 0),
            total_units=int(r.units or 0),
            total_revenue=round(rev, 2),
            revenue_pct=pct,
            cumulative_pct=cum_pct,
            abc_class=cls,
        ))
    return schemas.ABCResponse(
        channel_id=channel_id, total_revenue=round(total_rev, 2),
        class_a_count=a_count, class_b_count=b_count, class_c_count=c_count,
        rows=result,
    )


# ── Analytics: tendencia (time-series) ───────────────────────────────────

_MONTHS_ES = ["ene", "feb", "mar", "abr", "may", "jun",
              "jul", "ago", "sep", "oct", "nov", "dic"]


def _period_label(dt: datetime, period_type: str) -> str:
    if period_type == "month":
        return f"{_MONTHS_ES[dt.month - 1]} {str(dt.year)[2:]}"
    return f"{dt.day:02d} {_MONTHS_ES[dt.month - 1]}"


async def trend(
    db: AsyncSession, channel_id: Optional[int] = None,
    variant_id: Optional[int] = None, store_id: Optional[int] = None,
    period_type: str = "week", weeks_back: int = 26,
) -> schemas.TrendResponse:
    """Serie temporal de sell-out agregada por periodo. Permite ver la
    evolución semana-a-semana (o mes-a-mes) de una cadena, SKU o tienda.

    Cada punto suma unidades, devoluciones e ingreso de todos los reportes
    cuyo period_start cae en ese periodo. On-hand toma el ÚLTIMO valor
    reportado dentro del periodo por combo (no se suma a lo largo del tiempo,
    se suma entre tiendas del mismo corte)."""
    now = datetime.now(timezone.utc)
    days_back = weeks_back * 7 if period_type == "week" else weeks_back * 31
    since = now - timedelta(days=days_back)

    conds = [models.SellOutReport.period_start >= since]
    if period_type in ("week", "month", "day"):
        conds.append(models.SellOutReport.period_type == period_type)
    if variant_id is not None:
        conds.append(models.SellOutReport.variant_id == variant_id)
    if store_id is not None:
        conds.append(models.SellOutReport.store_id == store_id)

    stmt = (
        select(
            models.SellOutReport.period_start,
            func.max(models.SellOutReport.period_end).label("period_end"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units"),
            func.coalesce(func.sum(models.SellOutReport.units_returned), 0).label("returned"),
            func.coalesce(func.sum(models.SellOutReport.revenue), 0.0).label("revenue"),
            func.coalesce(func.sum(models.SellOutReport.returns_amount), 0.0).label("returns_amount"),
            func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0).label("on_hand"),
            func.count(func.distinct(models.SellOutReport.store_id)).label("stores"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(and_(*conds))
        .group_by(models.SellOutReport.period_start)
        .order_by(models.SellOutReport.period_start.asc())
    )
    if channel_id is not None:
        stmt = stmt.where(models.RetailStore.channel_id == channel_id)

    rows = (await db.execute(stmt)).all()
    points: List[schemas.TrendPoint] = []
    total_units = 0
    total_revenue = 0.0
    for r in rows:
        u = int(r.units or 0)
        ru = int(r.returned or 0)
        rev = float(r.revenue or 0.0)
        ret_amt = float(r.returns_amount or 0.0)
        total_units += u
        total_revenue += rev
        points.append(schemas.TrendPoint(
            period_start=r.period_start, period_end=r.period_end,
            label=_period_label(r.period_start, period_type),
            units_sold=u, units_returned=ru, net_units=max(u - ru, 0),
            revenue=round(rev, 2), returns_amount=round(ret_amt, 2),
            net_revenue=round(max(rev - ret_amt, 0.0), 2),
            on_hand=int(r.on_hand or 0),
            stores_reporting=int(r.stores or 0),
        ))

    wow_units_pct: Optional[float] = None
    wow_revenue_pct: Optional[float] = None
    if len(points) >= 2:
        prev, last = points[-2], points[-1]
        if prev.units_sold > 0:
            wow_units_pct = round((last.units_sold - prev.units_sold) / prev.units_sold * 100.0, 1)
        if prev.revenue > 0:
            wow_revenue_pct = round((last.revenue - prev.revenue) / prev.revenue * 100.0, 1)

    return schemas.TrendResponse(
        channel_id=channel_id, variant_id=variant_id, store_id=store_id,
        period_type=period_type, points=points,
        total_units=total_units, total_revenue=round(total_revenue, 2),
        wow_units_pct=wow_units_pct, wow_revenue_pct=wow_revenue_pct,
    )


# ── Analytics: distribución numérica (voids) ─────────────────────────────

async def distribution(
    db: AsyncSession, channel_id: Optional[int] = None,
    days: int = 28, limit: int = 200,
) -> schemas.DistributionResponse:
    """Distribución numérica por SKU: en cuántas tiendas de la cadena se está
    vendiendo el producto vs el total de tiendas activas. Un 'void' es una
    tienda que no reporta venta del SKU (oportunidad de expansión).

    Es el KPI que usan Nielsen/las áreas comerciales para saber si un
    producto está bien distribuido o hay hueco de anaquel."""
    now = datetime.now(timezone.utc)
    since = now - timedelta(days=days)

    # Total tiendas activas del scope
    stores_stmt = select(func.count(models.RetailStore.id)).where(
        models.RetailStore.is_active.is_(True)
    )
    if channel_id is not None:
        stores_stmt = stores_stmt.where(models.RetailStore.channel_id == channel_id)
    total_stores = int((await db.execute(stores_stmt)).scalar() or 0)

    # Por variante: tiendas que venden (units>0), tiendas con stock, unidades
    stmt = (
        select(
            models.SellOutReport.variant_id,
            func.max(models.SellOutReport.sku).label("sku"),
            func.max(models.SellOutReport.product_name).label("product_name"),
            func.count(func.distinct(models.RetailStore.id)).label("stores_selling"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .where(
            models.SellOutReport.period_start >= since,
            models.SellOutReport.variant_id.isnot(None),
            models.SellOutReport.units_sold > 0,
            models.RetailStore.is_active.is_(True),
        )
        .group_by(models.SellOutReport.variant_id)
        .order_by(func.count(func.distinct(models.RetailStore.id)).desc())
        .limit(limit)
    )
    if channel_id is not None:
        stmt = stmt.where(models.RetailStore.channel_id == channel_id)
    rows = (await db.execute(stmt)).all()

    # Tiendas con stock (on_hand>0) por variante en el último corte
    variant_ids = [int(r.variant_id) for r in rows if r.variant_id]
    stocking_by_variant: Dict[int, int] = {}
    if variant_ids:
        stock_stmt = (
            select(
                models.SellOutReport.variant_id,
                func.count(func.distinct(models.RetailStore.id)).label("stores"),
            )
            .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
            .where(
                models.SellOutReport.period_start >= since,
                models.SellOutReport.variant_id.in_(variant_ids),
                models.SellOutReport.units_on_hand > 0,
                models.RetailStore.is_active.is_(True),
            )
            .group_by(models.SellOutReport.variant_id)
        )
        if channel_id is not None:
            stock_stmt = stock_stmt.where(models.RetailStore.channel_id == channel_id)
        for vid, cnt in (await db.execute(stock_stmt)).all():
            stocking_by_variant[int(vid)] = int(cnt or 0)

    out: List[schemas.DistributionRow] = []
    for r in rows:
        vid = int(r.variant_id) if r.variant_id else None
        selling = int(r.stores_selling or 0)
        pct = round(selling / total_stores * 100.0, 1) if total_stores > 0 else 0.0
        units = int(r.units or 0)
        if pct >= 80:
            status = "excellent"
        elif pct >= 50:
            status = "good"
        elif pct >= 25:
            status = "low"
        else:
            status = "critical"
        out.append(schemas.DistributionRow(
            variant_id=vid, sku=r.sku, product_name=r.product_name,
            stores_selling=selling,
            stores_stocking=stocking_by_variant.get(vid or -1, 0),
            total_stores=total_stores,
            distribution_pct=pct,
            void_stores=max(total_stores - selling, 0),
            total_units=units,
            avg_units_per_store=round(units / selling, 1) if selling > 0 else 0.0,
            status=status,
        ))
    # Ordena por más voids (mayor oportunidad) entre los que ya venden algo
    out.sort(key=lambda x: (-x.void_stores, -x.total_units))
    return schemas.DistributionResponse(
        channel_id=channel_id, total_stores=total_stores, rows=out,
    )


# ── Analytics: venta perdida por stockout ────────────────────────────────

async def lost_sales(
    db: AsyncSession, channel_id: Optional[int] = None, limit: int = 500,
) -> schemas.LostSalesResponse:
    """Estima la venta perdida por productos agotados. Para cada (tienda, SKU)
    cuyo último corte tiene on_hand=0 pero que traía velocidad de venta,
    proyecta cuántas unidades e ingreso se están perdiendo por semana × las
    semanas consecutivas sin stock.

    lost_units = velocidad_semanal × semanas_sin_stock
    lost_revenue = lost_units × precio (catálogo, o promedio histórico)."""
    now = datetime.now(timezone.utc)
    velocity_from = now - timedelta(days=VELOCITY_WINDOW_DAYS)

    # Combos (store, variant) con su último corte
    combos_stmt = (
        select(
            models.RetailStore.id.label("store_id"),
            models.RetailStore.name.label("store_name"),
            models.RetailChannel.id.label("channel_id"),
            models.RetailChannel.name.label("channel_name"),
            models.SellOutReport.variant_id,
            models.SellOutReport.product_name,
            models.SellOutReport.sku,
            func.max(models.SellOutReport.period_start).label("last_period"),
            func.coalesce(func.sum(models.SellOutReport.units_sold), 0).label("units_window"),
        )
        .join(models.RetailStore, models.SellOutReport.store_id == models.RetailStore.id)
        .join(models.RetailChannel, models.RetailStore.channel_id == models.RetailChannel.id)
        .where(
            models.SellOutReport.period_start >= velocity_from,
            models.SellOutReport.variant_id.isnot(None),
            models.RetailStore.is_active.is_(True),
        )
        .group_by(
            models.RetailStore.id, models.RetailStore.name,
            models.RetailChannel.id, models.RetailChannel.name,
            models.SellOutReport.variant_id,
            models.SellOutReport.product_name, models.SellOutReport.sku,
        )
    )
    if channel_id is not None:
        combos_stmt = combos_stmt.where(models.RetailStore.channel_id == channel_id)
    combos = (await db.execute(combos_stmt)).all()

    # Precios de catálogo por variante
    variant_ids = list({int(r.variant_id) for r in combos if r.variant_id})
    price_by_variant: Dict[int, float] = {}
    if variant_ids:
        prices = (await db.execute(
            select(inv_models.ProductVariant.id, inv_models.ProductVariant.price)
            .where(inv_models.ProductVariant.id.in_(variant_ids))
        )).all()
        price_by_variant = {int(vid): float(p or 0.0) for vid, p in prices}

    rows: List[schemas.LostSalesRow] = []
    total_lost_units = 0
    total_lost_revenue = 0.0

    for r in combos:
        vid = int(r.variant_id)
        # On-hand del último corte
        on_hand = int((await db.execute(
            select(func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0))
            .where(
                models.SellOutReport.store_id == r.store_id,
                models.SellOutReport.variant_id == vid,
                models.SellOutReport.period_start == r.last_period,
            )
        )).scalar() or 0)
        if on_hand > 0:
            continue  # hay stock → no hay venta perdida

        avg_weekly = float(r.units_window or 0) / (VELOCITY_WINDOW_DAYS / 7.0)
        if avg_weekly <= 0:
            continue  # sin velocidad histórica → no proyectamos pérdida

        # Semanas consecutivas sin stock: cuenta cortes recientes con on_hand=0
        recent = (await db.execute(
            select(
                models.SellOutReport.period_start,
                func.coalesce(func.sum(models.SellOutReport.units_on_hand), 0).label("oh"),
            )
            .where(
                models.SellOutReport.store_id == r.store_id,
                models.SellOutReport.variant_id == vid,
                models.SellOutReport.period_start >= velocity_from,
            )
            .group_by(models.SellOutReport.period_start)
            .order_by(models.SellOutReport.period_start.desc())
        )).all()
        weeks_out = 0
        for _ps, oh in recent:
            if int(oh or 0) == 0:
                weeks_out += 1
            else:
                break
        weeks_out = max(weeks_out, 1)

        # Precio: catálogo, o ingreso/unidad histórico
        price = price_by_variant.get(vid, 0.0)
        if price <= 0:
            hist = (await db.execute(
                select(
                    func.coalesce(func.sum(models.SellOutReport.revenue), 0.0),
                    func.coalesce(func.sum(models.SellOutReport.units_sold), 0),
                ).where(
                    models.SellOutReport.variant_id == vid,
                    models.SellOutReport.units_sold > 0,
                )
            )).one()
            hrev, hunits = float(hist[0] or 0.0), int(hist[1] or 0)
            price = round(hrev / hunits, 2) if hunits > 0 else 0.0

        lost_units = int(round(avg_weekly * weeks_out))
        if lost_units <= 0:
            continue
        lost_revenue = round(lost_units * price, 2)
        total_lost_units += lost_units
        total_lost_revenue += lost_revenue

        if weeks_out >= 3 or lost_revenue >= 10000:
            severity = "urgent"
        elif weeks_out >= 2:
            severity = "high"
        else:
            severity = "medium"

        rows.append(schemas.LostSalesRow(
            store_id=r.store_id, store_name=r.store_name,
            channel_id=r.channel_id, channel_name=r.channel_name,
            variant_id=vid, sku=r.sku, product_name=r.product_name,
            avg_weekly_units=round(avg_weekly, 2),
            weeks_out_of_stock=float(weeks_out),
            lost_units=lost_units, unit_price=round(price, 2),
            lost_revenue=lost_revenue, severity=severity,
        ))

    rows.sort(key=lambda x: -x.lost_revenue)
    rows = rows[:limit]
    return schemas.LostSalesResponse(
        channel_id=channel_id, generated_at=now,
        total_lost_units=total_lost_units,
        total_lost_revenue=round(total_lost_revenue, 2),
        affected_combos=len(rows), rows=rows,
    )


# ── Replenishment: crear traslado ────────────────────────────────────────

async def list_source_warehouses(db: AsyncSession) -> List[schemas.SourceWarehouseOption]:
    """Warehouses de origen para traslados (typical: type=own)."""
    res = await db.execute(
        select(inv_models.Warehouse)
        .where(inv_models.Warehouse.is_active.is_(True),
                inv_models.Warehouse.type != "consignment")
        .order_by(inv_models.Warehouse.name)
    )
    return [
        schemas.SourceWarehouseOption(
            id=w.id, name=w.name, location=w.location, type=w.type or "own",
        ) for w in res.scalars().all()
    ]


async def create_transfer(
    db: AsyncSession, req: schemas.TransferRequest, user_id: Optional[int] = None,
) -> schemas.TransferResponse:
    """Crea un par OUT+IN de StockMovements por cada item, del warehouse de
    origen al consignment_warehouse de la tienda. Cada item se procesa
    aisladamente: si falla uno, los demás siguen y se reporta status."""
    from app.modules.inventory import schemas as inv_schemas, service as inv_service

    src = await db.get(inv_models.Warehouse, req.source_warehouse_id)
    if src is None:
        raise ValueError("Almacén origen no encontrado")

    results: List[schemas.TransferItemResult] = []
    transferred_lines = 0
    warnings_ct = 0
    total_units = 0

    for item in req.items:
        store = await db.get(models.RetailStore, item.store_id)
        if store is None:
            results.append(schemas.TransferItemResult(
                store_id=item.store_id, variant_id=item.variant_id,
                units_requested=item.units, units_transferred=0,
                status="error", message="Tienda no encontrada",
            ))
            warnings_ct += 1; continue
        if store.consignment_warehouse_id is None:
            results.append(schemas.TransferItemResult(
                store_id=item.store_id, variant_id=item.variant_id,
                units_requested=item.units, units_transferred=0,
                status="no_consignment",
                message="La tienda no tiene almacén de consignación asignado",
            ))
            warnings_ct += 1; continue
        # Verifica stock origen
        stock = int((await db.execute(
            select(inv_models.StockLevel.quantity).where(
                inv_models.StockLevel.variant_id == item.variant_id,
                inv_models.StockLevel.warehouse_id == req.source_warehouse_id,
            )
        )).scalar() or 0)
        if stock < item.units:
            results.append(schemas.TransferItemResult(
                store_id=item.store_id, variant_id=item.variant_id,
                units_requested=item.units, units_transferred=0,
                status="insufficient_stock",
                message=f"Stock en origen {stock} < solicitado {item.units}",
            ))
            warnings_ct += 1; continue

        ref = f"retail_transfer:store:{store.id}"
        out = None
        try:
            out = await inv_service.adjust_stock(db, inv_schemas.StockMovementCreate(
                variant_id=item.variant_id,
                warehouse_id=req.source_warehouse_id,
                quantity=item.units, movement_type="out",
                reference=ref,
                notes=f"Traslado a consignación · {store.name}"
                       + (f" · {item.notes}" if item.notes else ""),
            ), user_id=user_id)
            in_ = await inv_service.adjust_stock(db, inv_schemas.StockMovementCreate(
                variant_id=item.variant_id,
                warehouse_id=store.consignment_warehouse_id,
                quantity=item.units, movement_type="in",
                unit_cost=out.unit_cost or 0.0,
                reference=ref,
                notes=f"Entrada por traslado desde {src.name}",
            ), user_id=user_id)
            results.append(schemas.TransferItemResult(
                store_id=item.store_id, variant_id=item.variant_id,
                units_requested=item.units, units_transferred=item.units,
                status="transferred",
                out_movement_id=out.id, in_movement_id=in_.id,
            ))
            transferred_lines += 1
            total_units += item.units
        except Exception as e:
            # Si el OUT ya se aplicó pero el IN falló, el stock salió
            # sin llegar al destino. Compensación: reingresar al origen
            # con un adjustment para evitar pérdida fantasma.
            rollback_note = None
            if out is not None:
                try:
                    await inv_service.adjust_stock(db, inv_schemas.StockMovementCreate(
                        variant_id=item.variant_id,
                        warehouse_id=req.source_warehouse_id,
                        quantity=item.units, movement_type="adjustment",
                        unit_cost=out.unit_cost or 0.0,
                        reference=f"{ref}:rollback",
                        notes="Reversión: OUT aplicado pero IN falló",
                    ), user_id=user_id)
                    rollback_note = " (stock devuelto al origen)"
                except Exception as re:
                    log.error(
                        "transfer rollback FAILED store=%s variant=%s: %s",
                        item.store_id, item.variant_id, re,
                    )
                    rollback_note = " (ROLLBACK MANUAL REQUERIDO)"
            log.warning(
                "transfer falló store=%s variant=%s: %s%s",
                item.store_id, item.variant_id, e, rollback_note or "",
            )
            results.append(schemas.TransferItemResult(
                store_id=item.store_id, variant_id=item.variant_id,
                units_requested=item.units, units_transferred=0,
                status="error", message=f"{e}{rollback_note or ''}",
            ))
            warnings_ct += 1

    return schemas.TransferResponse(
        source_warehouse_id=req.source_warehouse_id,
        source_warehouse_name=src.name,
        transferred_lines=transferred_lines,
        warnings=warnings_ct,
        total_units=total_units,
        results=results,
    )


# ── Perfiles de importación ──────────────────────────────────────────────

# Diccionario de heurísticas: para cada campo estándar, patrones (lowercase)
# que si aparecen dentro del nombre de columna del archivo, sugieren ese
# campo. Se elige el primer match. Cubre los portales típicos:
#   Walmart Retail Link, Costco POL, HEB, Chedraui, Amazon Vendor,
#   Soriana, Bodega Aurrera, Fresko, City Market, además de exports
#   genéricos en español e inglés.
_HEURISTICS: Dict[str, List[str]] = {
    "cadena_codigo": [
        "chain code", "banner code", "codigo cadena", "cadena_codigo",
    ],
    "cadena_nombre": [
        "chain", "banner", "cadena", "retailer",
    ],
    "tienda_codigo": [
        "store nbr", "store number", "store #", "store id", "store code",
        "warehouse nbr", "warehouse #", "warehouse id",
        "site id", "site nbr", "location id", "location code",
        "tienda_codigo", "codigo tienda", "num tienda", "num_tienda",
        "sucursal codigo", "sucursal_id", "no tienda", "clave tienda",
        "tienda id",
    ],
    "tienda_nombre": [
        "store name", "warehouse name", "location name",
        "tienda", "sucursal", "nombre tienda", "nombre_sucursal",
    ],
    "sku": [
        "item nbr", "item number", "item #", "item id", "item code",
        "product id", "product code", "product number",
        "sku", "upc", "gtin", "ean", "barcode", "codigo articulo",
        "codigo_producto", "codigo material", "material",
        "clave producto", "clave articulo",
    ],
    "producto_nombre": [
        "item desc", "item description", "product desc", "description",
        "product name", "producto", "descripcion", "nombre producto",
        "articulo",
    ],
    "periodo_inicio": [
        "week", "week ending", "week beginning", "period", "week start",
        "date", "fecha", "semana", "periodo", "fecha inicio",
        "fecha_inicio", "start date", "from date",
    ],
    "periodo_fin": [
        "week end", "period end", "fecha fin", "fecha_fin",
        "end date", "to date", "fin periodo",
    ],
    "periodo_tipo": [
        "period type", "tipo periodo", "granularity", "granularidad",
    ],
    "unidades_vendidas": [
        "pos sales units", "pos sales qty", "sold units", "units sold",
        "units", "qty", "quantity", "sales qty", "sales units",
        "unidades vendidas", "unidades_vendidas", "piezas vendidas",
        "venta unidades", "cantidad", "ventas",
    ],
    "unidades_stock": [
        "store on hand", "on hand", "on-hand", "onh", "inventory on hand",
        "current inventory", "inventory", "stock",
        "unidades stock", "unidades_stock", "inventario", "stock final",
        "existencia", "existencias",
    ],
    "unidades_devueltas": [
        "returns units", "return units", "returned units", "return qty",
        "returns qty", "sales returns units", "sales returns qty",
        "pos returns units", "pos returns qty",
        "unidades devueltas", "unidades_devueltas", "piezas devueltas",
        "devoluciones unidades", "devoluciones cantidad", "cantidad devuelta",
        "devoluciones",
    ],
    "importe_devoluciones": [
        "returns amount", "returns dollars", "return dollars",
        "sales returns amount", "returns revenue", "pos returns amount",
        "importe devoluciones", "importe_devoluciones",
        "monto devoluciones", "monto_devoluciones",
        "importe devuelto", "valor devoluciones",
    ],
    "ingreso": [
        "pos sales", "sales dollars", "sales amount", "revenue",
        "net sales", "gross sales", "total sales",
        "importe", "ingreso", "ventas totales", "monto", "total",
    ],
    "notas": ["notes", "notas", "observaciones", "comentarios", "comment"],
}


def _normalize_column_name(s: str) -> str:
    """Lowercase, sin espacios extras, sin acentos comunes."""
    s = (s or "").strip().lower()
    for a, b in [("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"),
                  ("ñ", "n")]:
        s = s.replace(a, b)
    # colapsa espacios
    s = " ".join(s.split())
    return s


def auto_detect_column_map(headers: List[str]) -> Dict[str, str]:
    """Recibe los encabezados del archivo, devuelve mapeo propuesto:
    standard_field → column_del_archivo (usando los nombres originales).

    Reglas para evitar falsos positivos con nombres solapados como
    "POS Sales" vs "POS Sales Units":
      1) Prioriza match EXACTO sobre parcial.
      2) Una columna sólo puede ser asignada a UN campo estándar.
      3) Cuando hay que romper empate por parcial, gana el patrón MÁS
         LARGO (más específico).
    """
    proposed: Dict[str, str] = {}
    used_columns: set = set()
    normalized_headers = [(h, _normalize_column_name(h)) for h in headers if h]

    # Pasada 1: match exacto para cada campo (usando el patrón más largo).
    for field, patterns in _HEURISTICS.items():
        best: Optional[Tuple[str, int]] = None  # (col_original, len_patrón)
        for pat in patterns:
            for orig, norm in normalized_headers:
                if orig in used_columns:
                    continue
                if norm == pat:
                    if best is None or len(pat) > best[1]:
                        best = (orig, len(pat))
        if best:
            proposed[field] = best[0]
            used_columns.add(best[0])

    # Pasada 2: match parcial (patrón contenido en el encabezado), con
    # patrón más largo primero, para "POS Sales Units" preferir el patrón
    # "pos sales units" a "pos sales".
    for field, patterns in _HEURISTICS.items():
        if field in proposed:
            continue
        sorted_pats = sorted(patterns, key=lambda p: -len(p))
        best_col: Optional[str] = None
        best_pat_len = 0
        for pat in sorted_pats:
            for orig, norm in normalized_headers:
                if orig in used_columns:
                    continue
                if pat in norm and len(pat) > best_pat_len:
                    best_col = orig
                    best_pat_len = len(pat)
                    break
            if best_col:
                break
        if best_col:
            proposed[field] = best_col
            used_columns.add(best_col)

    return proposed


async def list_profiles(db: AsyncSession, channel_id: Optional[int] = None
                          ) -> List[schemas.RetailImportProfileOut]:
    stmt = select(models.RetailImportProfile, models.RetailChannel.name).join(
        models.RetailChannel,
        models.RetailImportProfile.channel_id == models.RetailChannel.id,
    )
    if channel_id is not None:
        stmt = stmt.where(models.RetailImportProfile.channel_id == channel_id)
    stmt = stmt.order_by(
        models.RetailChannel.name.asc(),
        models.RetailImportProfile.is_default.desc(),
        models.RetailImportProfile.name.asc(),
    )
    rows = (await db.execute(stmt)).all()
    return [
        schemas.RetailImportProfileOut(
            id=p.id, channel_id=p.channel_id, channel_name=cn,
            name=p.name, notes=p.notes,
            is_active=p.is_active, is_default=p.is_default,
            file_format=p.file_format, sheet_name=p.sheet_name,
            header_row=p.header_row, encoding=p.encoding,
            delimiter=p.delimiter, date_format=p.date_format,
            decimal_separator=p.decimal_separator,
            thousands_separator=p.thousands_separator,
            units_multiplier=p.units_multiplier,
            revenue_multiplier=p.revenue_multiplier,
            default_period_type=p.default_period_type,
            column_map=dict(p.column_map or {}),
            ignore_row_pattern=p.ignore_row_pattern,
            default_channel_code=p.default_channel_code,
            created_at=p.created_at,
        )
        for p, cn in rows
    ]


async def get_profile(db: AsyncSession, profile_id: int
                        ) -> Optional[models.RetailImportProfile]:
    return await db.get(models.RetailImportProfile, profile_id)


async def create_profile(db: AsyncSession,
                          data: schemas.RetailImportProfileCreate
                          ) -> models.RetailImportProfile:
    payload = data.model_dump(exclude_unset=True)
    if payload.get("is_default"):
        # Al marcar como default, desmarcar los otros del mismo canal
        others = (await db.execute(
            select(models.RetailImportProfile).where(
                models.RetailImportProfile.channel_id == data.channel_id,
                models.RetailImportProfile.is_default.is_(True),
            )
        )).scalars().all()
        for o in others:
            o.is_default = False
    p = models.RetailImportProfile(**payload)
    db.add(p); await db.commit(); await db.refresh(p)
    return p


async def update_profile(db: AsyncSession, profile_id: int,
                          data: schemas.RetailImportProfileUpdate
                          ) -> Optional[models.RetailImportProfile]:
    p = await db.get(models.RetailImportProfile, profile_id)
    if p is None:
        return None
    payload = data.model_dump(exclude_unset=True)
    if payload.get("is_default"):
        others = (await db.execute(
            select(models.RetailImportProfile).where(
                models.RetailImportProfile.channel_id == p.channel_id,
                models.RetailImportProfile.id != profile_id,
                models.RetailImportProfile.is_default.is_(True),
            )
        )).scalars().all()
        for o in others:
            o.is_default = False
    for k, v in payload.items():
        setattr(p, k, v)
    await db.commit(); await db.refresh(p)
    return p


async def delete_profile(db: AsyncSession, profile_id: int) -> bool:
    p = await db.get(models.RetailImportProfile, profile_id)
    if p is None:
        return False
    await db.delete(p); await db.commit()
    return True


# ── Helpers de parsing con perfil ────────────────────────────────────────

def _read_file_rows(file_bytes: bytes, filename: str,
                     profile: Optional[models.RetailImportProfile] = None
                     ) -> Tuple[List[str], List[Dict[str, Any]], List[str]]:
    """Lee el archivo respetando el perfil (hoja, fila de encabezado,
    delimiter, encoding, decimal). Devuelve (headers, rows, sheet_names)."""
    name_lower = (filename or "").lower()
    is_xlsx = name_lower.endswith(".xlsx") or name_lower.endswith(".xlsm")
    if profile:
        is_xlsx = profile.file_format == "xlsx"

    sheet_names: List[str] = []
    header_row = int(getattr(profile, "header_row", 1) or 1) if profile else 1
    delimiter = getattr(profile, "delimiter", ",") if profile else ","
    encoding = getattr(profile, "encoding", "utf-8") if profile else "utf-8"

    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet_names = list(wb.sheetnames)
        target = None
        if profile and profile.sheet_name and profile.sheet_name in sheet_names:
            target = profile.sheet_name
        elif "SellOut" in sheet_names:
            target = "SellOut"
        else:
            target = wb.active.title
        ws = wb[target]
        rows_all = list(ws.iter_rows(values_only=True))
        if header_row - 1 >= len(rows_all):
            return [], [], sheet_names
        header = [str(c).strip() if c is not None else "" for c in rows_all[header_row - 1]]
        rows_out: List[Dict[str, Any]] = []
        for row in rows_all[header_row:]:
            if all(v is None or v == "" for v in row):
                continue
            data = {}
            for i, h in enumerate(header):
                if not h:
                    continue
                data[h] = row[i] if i < len(row) else None
            rows_out.append(data)
        return header, rows_out, sheet_names

    # CSV
    try:
        text = file_bytes.decode(encoding, errors="ignore")
    except LookupError:
        text = file_bytes.decode("utf-8", errors="ignore")
    text = text.lstrip("﻿")  # BOM
    lines = text.splitlines()
    if header_row - 1 >= len(lines):
        return [], [], []
    reader = csv.reader(io.StringIO(text), delimiter=(delimiter or ","))
    all_rows = list(reader)
    if not all_rows or header_row - 1 >= len(all_rows):
        return [], [], []
    header = [c.strip() for c in all_rows[header_row - 1]]
    rows_out = []
    for r in all_rows[header_row:]:
        if all((c is None or str(c).strip() == "") for c in r):
            continue
        data = {}
        for i, h in enumerate(header):
            if not h:
                continue
            data[h] = r[i] if i < len(r) else None
        rows_out.append(data)
    return header, rows_out, []


def _apply_multipliers(v: Any, mul: float) -> float:
    x = _parse_float(v)
    return round(x * mul, 4)


def _parse_date_flex(v: Any, fmt: str) -> Optional[datetime]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day, tzinfo=timezone.utc)
    s = str(v).strip()
    if fmt == "auto":
        return _parse_iso_date(s)
    tries = {
        "YYYY-MM-DD": "%Y-%m-%d",
        "DD/MM/YYYY": "%d/%m/%Y",
        "MM/DD/YYYY": "%m/%d/%Y",
        "YYYY/MM/DD": "%Y/%m/%d",
    }
    fmt_str = tries.get(fmt)
    if fmt_str:
        try:
            return datetime.strptime(s[:10], fmt_str).replace(tzinfo=timezone.utc)
        except ValueError:
            pass
    # Fallback auto
    return _parse_iso_date(s)


def _normalize_row_with_profile(
    row: Dict[str, Any], profile: models.RetailImportProfile,
) -> Tuple[Dict[str, Any], List[str]]:
    """Aplica el mapeo del perfil a una fila cruda. Devuelve
    (fila_normalizada, lista_de_errores)."""
    errors: List[str] = []
    normalized: Dict[str, Any] = {}
    cmap = profile.column_map or {}

    def _get(field: str) -> Any:
        col = cmap.get(field)
        return row.get(col) if col else None

    # Cadena
    chan_code = str(_get("cadena_codigo") or profile.default_channel_code or "").strip()
    normalized["cadena_codigo"] = chan_code
    normalized["cadena_nombre"] = str(_get("cadena_nombre") or "").strip()

    # Tienda
    normalized["tienda_codigo"] = str(_get("tienda_codigo") or "").strip()
    normalized["tienda_nombre"] = str(_get("tienda_nombre") or "").strip()

    # Producto
    normalized["sku"] = str(_get("sku") or "").strip()
    normalized["producto_nombre"] = str(_get("producto_nombre") or "").strip()

    # Periodo
    period_type_raw = str(_get("periodo_tipo") or "").strip().lower()
    if period_type_raw not in ("day", "week", "month"):
        period_type_raw = profile.default_period_type
    normalized["periodo_tipo"] = period_type_raw

    period_start = _parse_date_flex(_get("periodo_inicio"), profile.date_format)
    if period_start is None:
        errors.append("periodo_inicio inválido")
    normalized["periodo_inicio"] = period_start.isoformat() if period_start else None

    period_end = _parse_date_flex(_get("periodo_fin"), profile.date_format)
    if period_end is None and period_start is not None:
        period_end = _default_period_end(period_start, period_type_raw)
    normalized["periodo_fin"] = period_end.isoformat() if period_end else None

    # Cantidades
    normalized["unidades_vendidas"] = int(round(
        _apply_multipliers(_get("unidades_vendidas"), profile.units_multiplier)
    ))
    normalized["unidades_devueltas"] = int(round(
        _apply_multipliers(_get("unidades_devueltas"), profile.units_multiplier)
    ))
    normalized["unidades_stock"] = int(round(
        _apply_multipliers(_get("unidades_stock"), profile.units_multiplier)
    ))
    normalized["ingreso"] = round(
        _apply_multipliers(_get("ingreso"), profile.revenue_multiplier), 2,
    )
    normalized["importe_devoluciones"] = round(
        _apply_multipliers(_get("importe_devoluciones"), profile.revenue_multiplier), 2,
    )
    normalized["notas"] = str(_get("notas") or "").strip() or None

    # Validación mínima
    if not chan_code and not normalized["cadena_nombre"]:
        errors.append("sin cadena identificable")
    if not normalized["tienda_codigo"] and not normalized["tienda_nombre"]:
        errors.append("sin tienda identificable")
    if not normalized["sku"] and not normalized["producto_nombre"]:
        errors.append("sin producto identificable")
    return normalized, errors


# ── Endpoints principales del wizard ─────────────────────────────────────

async def detect_columns(
    db: AsyncSession, file_bytes: bytes, filename: str,
    profile_id: Optional[int] = None,
) -> schemas.DetectColumnsResponse:
    profile = await get_profile(db, profile_id) if profile_id else None
    headers, _rows, sheets = _read_file_rows(file_bytes, filename, profile)
    proposed = auto_detect_column_map(headers)
    active_sheet = None
    if profile and profile.sheet_name:
        active_sheet = profile.sheet_name
    elif sheets:
        active_sheet = sheets[0]
    return schemas.DetectColumnsResponse(
        detected_columns=headers,
        sheet_names=sheets,
        active_sheet=active_sheet,
        proposed_map=proposed,
    )


async def preview_with_profile(
    db: AsyncSession, profile_id: int, file_bytes: bytes, filename: str,
    limit: int = 10,
) -> schemas.PreviewResponse:
    profile = await get_profile(db, profile_id)
    if profile is None:
        raise ValueError("Perfil no encontrado")
    headers, rows, _ = _read_file_rows(file_bytes, filename, profile)
    warnings: List[str] = []
    if not headers:
        warnings.append("El archivo no tiene encabezados legibles con esta configuración.")

    required = ["tienda_codigo", "sku", "periodo_inicio", "unidades_vendidas"]
    cmap = profile.column_map or {}
    unmapped = []
    for field in required:
        col = cmap.get(field)
        if not col:
            # tienda_nombre / producto_nombre son fallbacks
            if field == "tienda_codigo" and cmap.get("tienda_nombre"):
                continue
            if field == "sku" and cmap.get("producto_nombre"):
                continue
            unmapped.append(field)

    preview_rows: List[schemas.PreviewRow] = []
    for i, r in enumerate(rows[:limit], start=int(profile.header_row) + 1):
        normalized, errs = _normalize_row_with_profile(r, profile)
        preview_rows.append(schemas.PreviewRow(
            row_number=i,
            raw={k: (str(v) if v is not None else "") for k, v in r.items()},
            normalized=normalized,
            errors=errs,
        ))

    return schemas.PreviewResponse(
        total_rows=len(rows),
        preview_rows=preview_rows,
        unmapped_required_fields=unmapped,
        warnings=warnings,
    )


async def import_with_profile(
    db: AsyncSession, profile_id: int, file_bytes: bytes, filename: str,
    user_id: Optional[int] = None,
) -> schemas.ImportSellOutResponse:
    """Aplica el perfil al archivo y llama a import_sellout con las filas
    normalizadas. Devuelve el mismo response que la ruta clásica."""
    profile = await get_profile(db, profile_id)
    if profile is None:
        raise ValueError("Perfil no encontrado")
    channel = await db.get(models.RetailChannel, profile.channel_id)

    headers, rows, _ = _read_file_rows(file_bytes, filename, profile)
    if not rows:
        return schemas.ImportSellOutResponse(
            total_rows=0, created=0, updated=0, skipped=0, errors=[],
        )

    # Convierte las filas al formato interno (nombres estándar) y arma un
    # CSV en memoria para reutilizar import_sellout con toda su lógica de
    # matcheo/upsert.
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    for r in rows:
        norm, _errs = _normalize_row_with_profile(r, profile)
        if not norm.get("cadena_codigo") and not norm.get("cadena_nombre"):
            norm["cadena_codigo"] = channel.code or ""
            norm["cadena_nombre"] = channel.name if channel else ""
        writer.writerow([
            norm.get("cadena_codigo") or "",
            norm.get("cadena_nombre") or "",
            norm.get("tienda_codigo") or "",
            norm.get("tienda_nombre") or "",
            norm.get("sku") or "",
            norm.get("producto_nombre") or "",
            norm.get("periodo_tipo") or "",
            norm.get("periodo_inicio") or "",
            norm.get("periodo_fin") or "",
            norm.get("unidades_vendidas") or 0,
            norm.get("unidades_devueltas") or 0,
            norm.get("unidades_stock") or 0,
            norm.get("ingreso") or 0,
            norm.get("importe_devoluciones") or 0,
            norm.get("notas") or "",
        ])
    csv_bytes = buf.getvalue().encode("utf-8-sig")
    fake_name = f"profile_{profile.id}.csv"
    return await import_sellout(db, csv_bytes, fake_name, user_id=user_id)
