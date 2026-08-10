"""Conciliación bancaria — importar extracto y matching automático.

Flujo:
  1. Usuario sube el CSV/XLSX del banco.
  2. Parser detecta columnas por sinónimos (fecha/concepto/cargo/abono/monto).
  3. Cada fila crea (o localiza) un `BankTransaction`.
  4. Se busca match automático contra `Transaction` (income/expense) del
     sistema por fecha (±3 días) + monto exacto.
  5. Se marca `reconciled=True` en las matcheadas.
  6. Se reporta resumen: importadas / matcheadas / sin match / duplicadas.

Formato mínimo del CSV/XLSX:
  Fecha, Concepto, Cargo, Abono
    ó
  Fecha, Descripción, Monto (positivo=abono, negativo=cargo)
"""
from __future__ import annotations
import io, csv
from datetime import datetime, timedelta, timezone
from typing import Optional
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy import and_

from app.modules.finance import models as fin_models


COLUMN_ALIASES = {
    "date":        ("fecha", "date", "fecha operacion", "fecha operación", "f.oper", "fecha_valor"),
    "description": ("concepto", "descripcion", "descripción", "detalle", "referencia", "movimiento", "description"),
    "debit":       ("cargo", "cargos", "debe", "debit", "salida", "retiro"),
    "credit":      ("abono", "abonos", "haber", "credit", "entrada", "deposito", "depósito"),
    "amount":      ("monto", "importe", "amount"),
    "ref":         ("folio", "ref", "referencia banco", "id operacion", "id operación"),
}


def _norm(s: str) -> str:
    return (s or "").strip().lower()


def _find_col(headers: list[str], keys: tuple[str, ...]) -> Optional[int]:
    for i, h in enumerate(headers):
        if _norm(h) in keys:
            return i
    return None


def _to_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", "").replace("$", "").strip() or 0.0)
    except Exception:
        return 0.0


def _to_date(v) -> Optional[datetime]:
    if not v:
        return None
    if isinstance(v, datetime):
        return v if v.tzinfo else v.replace(tzinfo=timezone.utc)
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s, fmt).replace(tzinfo=timezone.utc)
        except ValueError:
            continue
    return None


def _parse_csv(file_bytes: bytes) -> list[dict]:
    """Retorna lista de dicts {date, description, amount, ref}. amount>0=abono, <0=cargo."""
    text = file_bytes.decode("utf-8", errors="replace")
    # Detectar delimitador
    sample = text[:1024]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.reader(io.StringIO(text), delimiter=delim)
    rows = list(reader)
    if not rows:
        return []
    headers = rows[0]
    col_date = _find_col(headers, COLUMN_ALIASES["date"])
    col_desc = _find_col(headers, COLUMN_ALIASES["description"])
    col_debit = _find_col(headers, COLUMN_ALIASES["debit"])
    col_credit = _find_col(headers, COLUMN_ALIASES["credit"])
    col_amount = _find_col(headers, COLUMN_ALIASES["amount"])
    col_ref = _find_col(headers, COLUMN_ALIASES["ref"])
    if col_date is None:
        raise ValueError("El CSV no tiene columna de fecha reconocible (Fecha, Date, etc.)")
    if col_debit is None and col_credit is None and col_amount is None:
        raise ValueError("El CSV no tiene columnas de importe (Cargo/Abono o Monto)")

    out: list[dict] = []
    for r in rows[1:]:
        if not r or all((c or "").strip() == "" for c in r):
            continue
        get = lambda i: (r[i] if (i is not None and i < len(r)) else None)
        date = _to_date(get(col_date))
        if not date:
            continue
        amt = 0.0
        if col_debit is not None and _to_float(get(col_debit)) > 0:
            amt = -_to_float(get(col_debit))
        elif col_credit is not None and _to_float(get(col_credit)) > 0:
            amt = _to_float(get(col_credit))
        elif col_amount is not None:
            amt = _to_float(get(col_amount))
        if amt == 0.0:
            continue
        out.append({
            "date": date,
            "description": (get(col_desc) or "").strip(),
            "amount": amt,
            "ref": (get(col_ref) or "").strip() or None,
        })
    return out


def _parse_xlsx(file_bytes: bytes) -> list[dict]:
    """Igual que _parse_csv pero para xlsx."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "") for c in rows[0]]
    # Mismo mapeo
    col_date = _find_col(headers, COLUMN_ALIASES["date"])
    col_desc = _find_col(headers, COLUMN_ALIASES["description"])
    col_debit = _find_col(headers, COLUMN_ALIASES["debit"])
    col_credit = _find_col(headers, COLUMN_ALIASES["credit"])
    col_amount = _find_col(headers, COLUMN_ALIASES["amount"])
    col_ref = _find_col(headers, COLUMN_ALIASES["ref"])
    if col_date is None:
        raise ValueError("El XLSX no tiene columna de fecha reconocible")
    out: list[dict] = []
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        get = lambda i: (r[i] if (i is not None and i < len(r)) else None)
        date = _to_date(get(col_date))
        if not date:
            continue
        amt = 0.0
        if col_debit is not None and _to_float(get(col_debit)) > 0:
            amt = -_to_float(get(col_debit))
        elif col_credit is not None and _to_float(get(col_credit)) > 0:
            amt = _to_float(get(col_credit))
        elif col_amount is not None:
            amt = _to_float(get(col_amount))
        if amt == 0.0:
            continue
        out.append({
            "date": date,
            "description": str(get(col_desc) or "").strip(),
            "amount": amt,
            "ref": str(get(col_ref) or "").strip() or None,
        })
    return out


async def import_statement(
    db: AsyncSession, bank_account_id: int, file_bytes: bytes,
    filename: str, match_window_days: int = 3,
) -> dict:
    """Importa extracto bancario y hace matching automático.

    Retorna {imported, matched, unmatched, duplicated, details[...]}.
    """
    account = await db.get(fin_models.BankAccount, bank_account_id)
    if not account:
        raise ValueError("Cuenta bancaria no encontrada")

    fname_lower = filename.lower()
    if fname_lower.endswith(".xlsx") or fname_lower.endswith(".xls"):
        rows = _parse_xlsx(file_bytes)
    else:
        rows = _parse_csv(file_bytes)

    if not rows:
        return {"imported": 0, "matched": 0, "unmatched": 0, "duplicated": 0,
                "details": [], "error": "No se detectaron movimientos en el archivo"}

    # Pre-cargar transacciones del sistema en el rango del extracto para matching
    min_date = min(r["date"] for r in rows) - timedelta(days=match_window_days)
    max_date = max(r["date"] for r in rows) + timedelta(days=match_window_days)
    res_tx = await db.execute(
        select(fin_models.Transaction).where(
            fin_models.Transaction.created_at >= min_date,
            fin_models.Transaction.created_at <= max_date,
        )
    )
    system_txs = list(res_tx.scalars().all())

    matched_ids: set[int] = set()  # ya usados
    imported = 0; matched = 0; unmatched = 0; duplicated = 0
    details: list[dict] = []

    for row in rows:
        ext_ref = row["ref"] or f"{row['date'].strftime('%Y%m%d')}_{int(row['amount']*100)}_{(row['description'] or '')[:20]}"
        # ¿Ya importado antes?
        dup = (await db.execute(
            select(fin_models.BankTransaction).where(
                fin_models.BankTransaction.bank_account_id == bank_account_id,
                fin_models.BankTransaction.external_ref == ext_ref,
            )
        )).scalars().first()
        if dup:
            duplicated += 1
            details.append({"status": "duplicated", "date": row["date"].isoformat(),
                            "amount": row["amount"], "desc": row["description"]})
            continue

        tx_type = "deposit" if row["amount"] > 0 else "withdrawal"
        bt = fin_models.BankTransaction(
            bank_account_id=bank_account_id,
            type=tx_type, amount=abs(row["amount"]),
            description=row["description"] or None,
            reference=row["ref"], reconciled=False,
            bank_date=row["date"], source=f"import:{account.bank or 'bank'}",
            external_ref=ext_ref,
        )
        # Match: mismo signo (income vs deposit; expense vs withdrawal) + monto exacto + ±window
        target_type = "income" if row["amount"] > 0 else "expense"
        target_amount = round(abs(row["amount"]), 2)
        best = None
        for t in system_txs:
            if t.id in matched_ids:
                continue
            if t.type != target_type:
                continue
            if round(t.amount or 0, 2) != target_amount:
                continue
            # Ventana de fechas
            t_date = t.created_at
            if t_date.tzinfo is None:
                t_date = t_date.replace(tzinfo=timezone.utc)
            if abs((t_date - row["date"]).days) <= match_window_days:
                best = t
                break
        if best:
            bt.reconciled = True
            bt.matched_transaction_id = best.id
            matched_ids.add(best.id)
            matched += 1
            details.append({"status": "matched", "date": row["date"].isoformat(),
                            "amount": row["amount"], "desc": row["description"],
                            "matched_tx_id": best.id})
        else:
            unmatched += 1
            details.append({"status": "unmatched", "date": row["date"].isoformat(),
                            "amount": row["amount"], "desc": row["description"]})
        db.add(bt)
        imported += 1

    await db.commit()
    return {
        "imported": imported, "matched": matched,
        "unmatched": unmatched, "duplicated": duplicated,
        "match_rate": round(matched / imported * 100, 1) if imported > 0 else 0.0,
        "details": details[:100],
    }


# ═══════════════════════════════════════════════════════════════════════════
# Conciliación dedicada — workspace por sesión mensual
# ═══════════════════════════════════════════════════════════════════════════

from calendar import monthrange
from datetime import date
from io import BytesIO
from sqlalchemy import func, or_


def _period_bounds(year: int, month: int):
    first = datetime(year, month, 1, tzinfo=timezone.utc)
    last_day = monthrange(year, month)[1]
    end_excl = datetime(year, month, last_day, 23, 59, 59, tzinfo=timezone.utc) + timedelta(seconds=1)
    return first, end_excl


async def _computed_system_balance(
    db: AsyncSession, bank_account_id: int, until: datetime,
) -> float:
    """Saldo del sistema para la cuenta bancaria al corte (incluye todas las
    transactions ligadas a esta cuenta con fecha <= until).

    Nota: usa Transaction.bank_account_id + type (income/expense) para no
    depender del `balance` cacheado en BankAccount.
    """
    from sqlalchemy import case
    q = (
        select(func.coalesce(func.sum(
            case((fin_models.Transaction.type == "income", fin_models.Transaction.amount),
                 else_=-fin_models.Transaction.amount)
        ), 0.0))
        .where(fin_models.Transaction.bank_account_id == bank_account_id)
        .where(fin_models.Transaction.created_at <= until)
    )
    return float((await db.execute(q)).scalar() or 0.0)


async def get_or_create_reconciliation(
    db: AsyncSession, bank_account_id: int, year: int, month: int,
    user_id: Optional[int] = None,
) -> fin_models.BankReconciliation:
    q = select(fin_models.BankReconciliation).where(
        fin_models.BankReconciliation.bank_account_id == bank_account_id,
        fin_models.BankReconciliation.period_year == year,
        fin_models.BankReconciliation.period_month == month,
    )
    rec = (await db.execute(q)).scalars().first()
    if rec:
        return rec
    _, end_excl = _period_bounds(year, month)
    system_bal = await _computed_system_balance(db, bank_account_id, end_excl)
    rec = fin_models.BankReconciliation(
        bank_account_id=bank_account_id,
        period_year=year, period_month=month,
        opening_balance=0.0,
        closing_balance_bank=0.0,
        closing_balance_system=round(system_bal, 2),
        difference=0.0,
        status="draft",
        created_by_id=user_id,
    )
    db.add(rec)
    await db.commit()
    await db.refresh(rec)
    return rec


async def get_workspace(
    db: AsyncSession, bank_account_id: int, year: int, month: int,
    user_id: Optional[int] = None,
) -> dict:
    """Devuelve el workspace completo:
      - Sesión (BankReconciliation)
      - Movimientos del banco del periodo (matched + unmatched)
      - Transactions internas del periodo ligadas a la cuenta
      - Totales por lado
    """
    rec = await get_or_create_reconciliation(db, bank_account_id, year, month, user_id)
    start, end_excl = _period_bounds(year, month)

    # 1) Movimientos del banco del periodo
    q_bank = (
        select(fin_models.BankTransaction)
        .where(fin_models.BankTransaction.bank_account_id == bank_account_id)
        .where(
            or_(
                # con bank_date en el periodo
                (fin_models.BankTransaction.bank_date >= start)
                & (fin_models.BankTransaction.bank_date < end_excl),
                # sin bank_date, usa created_at
                fin_models.BankTransaction.bank_date.is_(None)
                & (fin_models.BankTransaction.created_at >= start)
                & (fin_models.BankTransaction.created_at < end_excl),
            )
        )
        .order_by(fin_models.BankTransaction.bank_date.desc().nulls_last(),
                    fin_models.BankTransaction.created_at.desc())
    )
    bank_rows = (await db.execute(q_bank)).scalars().all()

    # 2) Transactions del sistema del periodo (para la misma cuenta)
    q_sys = (
        select(fin_models.Transaction)
        .where(fin_models.Transaction.bank_account_id == bank_account_id)
        .where(fin_models.Transaction.created_at >= start)
        .where(fin_models.Transaction.created_at < end_excl)
        .order_by(fin_models.Transaction.created_at.desc())
    )
    sys_rows = (await db.execute(q_sys)).scalars().all()

    matched_ids = {r.matched_transaction_id for r in bank_rows if r.matched_transaction_id}

    def _bank_dict(b):
        return {
            "id": b.id, "type": b.type, "amount": float(b.amount or 0.0),
            "description": b.description, "reference": b.reference,
            "reconciled": bool(b.reconciled),
            "bank_date": (b.bank_date.isoformat() if b.bank_date else None),
            "created_at": (b.created_at.isoformat() if b.created_at else None),
            "matched_transaction_id": b.matched_transaction_id,
            "source": b.source,
        }

    def _sys_dict(t):
        return {
            "id": t.id, "type": t.type, "amount": float(t.amount or 0.0),
            "category": t.category, "description": t.description,
            "reference": t.reference,
            "matched": t.id in matched_ids,
            "created_at": (t.created_at.isoformat() if t.created_at else None),
        }

    bank_serialized = [_bank_dict(b) for b in bank_rows]
    sys_serialized = [_sys_dict(t) for t in sys_rows]

    # Totales
    bank_deposits = sum(b["amount"] for b in bank_serialized if b["amount"] > 0)
    bank_withdrawals = sum(-b["amount"] for b in bank_serialized if b["amount"] < 0)
    sys_income = sum(t["amount"] for t in sys_serialized if t["type"] == "income")
    sys_expense = sum(t["amount"] for t in sys_serialized if t["type"] == "expense")

    matched_ct = len(matched_ids)
    unmatched_bank = sum(1 for b in bank_serialized if not b["matched_transaction_id"])
    unmatched_sys = sum(1 for t in sys_serialized if not t["matched"])

    # Recalcula system balance actual (por si hubo movimientos nuevos)
    current_system_bal = await _computed_system_balance(db, bank_account_id, end_excl)

    return {
        "reconciliation": {
            "id": rec.id, "bank_account_id": rec.bank_account_id,
            "period_year": rec.period_year, "period_month": rec.period_month,
            "opening_balance": rec.opening_balance,
            "closing_balance_bank": rec.closing_balance_bank,
            "closing_balance_system": round(current_system_bal, 2),
            "difference": round(rec.closing_balance_bank - current_system_bal, 2),
            "status": rec.status, "notes": rec.notes,
            "completed_at": (rec.completed_at.isoformat() if rec.completed_at else None),
        },
        "bank_transactions": bank_serialized,
        "system_transactions": sys_serialized,
        "totals": {
            "bank_deposits": round(bank_deposits, 2),
            "bank_withdrawals": round(bank_withdrawals, 2),
            "sys_income": round(sys_income, 2),
            "sys_expense": round(sys_expense, 2),
            "matched_count": matched_ct,
            "unmatched_bank_count": unmatched_bank,
            "unmatched_sys_count": unmatched_sys,
        },
    }


async def manual_match(
    db: AsyncSession, bank_tx_id: int, system_tx_id: int,
) -> dict:
    """Match manual entre un movimiento del banco y una transacción del sistema."""
    b = (await db.execute(
        select(fin_models.BankTransaction).where(fin_models.BankTransaction.id == bank_tx_id)
    )).scalars().first()
    if not b:
        return {"ok": False, "error": "Movimiento bancario no encontrado"}
    t = (await db.execute(
        select(fin_models.Transaction).where(fin_models.Transaction.id == system_tx_id)
    )).scalars().first()
    if not t:
        return {"ok": False, "error": "Transacción del sistema no encontrada"}
    # Validación suave: mismo signo
    b_amt = float(b.amount or 0.0)
    t_amt = float(t.amount or 0.0)
    b_income = b_amt > 0
    t_income = t.type == "income"
    if b_income != t_income:
        return {"ok": False, "error": "El movimiento bancario y la transacción tienen signos opuestos"}
    b.matched_transaction_id = t.id
    b.reconciled = True
    await db.commit()
    return {"ok": True, "matched": True, "bank_tx_id": b.id, "system_tx_id": t.id}


async def manual_unmatch(db: AsyncSession, bank_tx_id: int) -> dict:
    b = (await db.execute(
        select(fin_models.BankTransaction).where(fin_models.BankTransaction.id == bank_tx_id)
    )).scalars().first()
    if not b:
        return {"ok": False, "error": "Movimiento bancario no encontrado"}
    b.matched_transaction_id = None
    b.reconciled = False
    await db.commit()
    return {"ok": True, "unmatched": True, "bank_tx_id": b.id}


async def update_reconciliation(
    db: AsyncSession, rec_id: int, opening_balance: Optional[float] = None,
    closing_balance_bank: Optional[float] = None, notes: Optional[str] = None,
) -> Optional[fin_models.BankReconciliation]:
    rec = (await db.execute(
        select(fin_models.BankReconciliation).where(fin_models.BankReconciliation.id == rec_id)
    )).scalars().first()
    if not rec:
        return None
    if opening_balance is not None:
        rec.opening_balance = float(opening_balance)
    if closing_balance_bank is not None:
        rec.closing_balance_bank = float(closing_balance_bank)
    if notes is not None:
        rec.notes = notes
    # Recalcular diferencia
    _, end_excl = _period_bounds(rec.period_year, rec.period_month)
    sys_bal = await _computed_system_balance(db, rec.bank_account_id, end_excl)
    rec.closing_balance_system = round(sys_bal, 2)
    rec.difference = round(rec.closing_balance_bank - sys_bal, 2)
    await db.commit()
    await db.refresh(rec)
    return rec


async def complete_reconciliation(
    db: AsyncSession, rec_id: int, user_id: Optional[int] = None,
) -> Optional[fin_models.BankReconciliation]:
    rec = (await db.execute(
        select(fin_models.BankReconciliation).where(fin_models.BankReconciliation.id == rec_id)
    )).scalars().first()
    if not rec:
        return None
    _, end_excl = _period_bounds(rec.period_year, rec.period_month)
    sys_bal = await _computed_system_balance(db, rec.bank_account_id, end_excl)
    rec.closing_balance_system = round(sys_bal, 2)
    rec.difference = round(rec.closing_balance_bank - sys_bal, 2)
    rec.status = "completed"
    rec.completed_at = datetime.now(timezone.utc)
    rec.completed_by_id = user_id
    await db.commit()
    await db.refresh(rec)
    return rec


async def reopen_reconciliation(
    db: AsyncSession, rec_id: int, user_id: Optional[int] = None,
) -> Optional[fin_models.BankReconciliation]:
    rec = (await db.execute(
        select(fin_models.BankReconciliation).where(fin_models.BankReconciliation.id == rec_id)
    )).scalars().first()
    if not rec:
        return None
    rec.status = "reopened"
    rec.reopened_at = datetime.now(timezone.utc)
    rec.reopened_by_id = user_id
    await db.commit()
    await db.refresh(rec)
    return rec


def build_reconciliation_pdf(company: dict, workspace: dict, bank_account: dict) -> bytes:
    """PDF oficial de la conciliación bancaria."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
    )

    MONTHS = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
              "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]

    def _mxn(n):
        return "$" + f"{(n or 0):,.2f}"

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=LETTER,
        leftMargin=18*mm, rightMargin=18*mm, topMargin=14*mm, bottomMargin=14*mm)

    styles = getSampleStyleSheet()
    brand = colors.HexColor(company.get("brand_color") or "#33B2F5")
    story = []

    rec = workspace["reconciliation"]
    totals = workspace["totals"]

    # Header
    empresa_lines = []
    if company.get("commercial_name"):
        empresa_lines.append(f"<b>{company['commercial_name']}</b>")
    if company.get("legal_name") and company.get("legal_name") != company.get("commercial_name"):
        empresa_lines.append(company["legal_name"])
    if company.get("rfc"):
        empresa_lines.append(f"RFC: {company['rfc']}")

    style_small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8.5, leading=11)
    style_title = ParagraphStyle("title", parent=styles["Heading1"], fontSize=15,
                                  alignment=TA_RIGHT, textColor=brand, leading=18)
    style_sub = ParagraphStyle("sub", parent=styles["Normal"], fontSize=9.5,
                                alignment=TA_RIGHT, textColor=colors.HexColor("#666666"))

    period_lbl = f"{MONTHS[rec['period_month']]} de {rec['period_year']}"
    title_para = Paragraph("CONCILIACIÓN BANCARIA", style_title)
    sub_para = Paragraph(period_lbl, style_sub)

    header_tbl = Table(
        [[Paragraph("<br/>".join(empresa_lines) or "&nbsp;", style_small),
          [title_para, sub_para]]],
        colWidths=[110*mm, 70*mm],
    )
    header_tbl.setStyle(TableStyle([("VALIGN", (0,0), (-1,-1), "TOP")]))
    story.append(header_tbl)
    story.append(Spacer(1, 4*mm))

    bar = Table([[""]], colWidths=[180*mm], rowHeights=[1.4*mm])
    bar.setStyle(TableStyle([("BACKGROUND", (0,0), (-1,-1), brand)]))
    story.append(bar)
    story.append(Spacer(1, 5*mm))

    # Datos bancarios + saldos
    balance_data = [
        ["Cuenta bancaria", bank_account.get("name", "")],
        ["Banco", bank_account.get("bank", "")],
        ["No. de cuenta", bank_account.get("account_number", "")],
        ["", ""],
        ["Saldo inicial (banco)", _mxn(rec["opening_balance"])],
        ["Saldo final (banco)", _mxn(rec["closing_balance_bank"])],
        ["Saldo final (sistema)", _mxn(rec["closing_balance_system"])],
        ["Diferencia", _mxn(rec["difference"])],
    ]
    tbl = Table(balance_data, colWidths=[60*mm, 60*mm])
    tbl.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 10),
        ("ALIGN", (1,0), (1,-1), "RIGHT"),
        ("FONTNAME", (0,4), (-1,-1), "Helvetica-Bold"),
        ("BACKGROUND", (0,7), (-1,7),
         colors.HexColor("#E8F7E8") if abs(rec["difference"]) < 1 else colors.HexColor("#FBEAEA")),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 6*mm))

    # Resumen
    story.append(Paragraph("<b>Resumen del periodo</b>",
                            ParagraphStyle("h2", parent=styles["Heading3"], textColor=brand)))
    resumen = [
        ["Concepto", "Banco", "Sistema"],
        ["Depósitos / Ingresos", _mxn(totals["bank_deposits"]), _mxn(totals["sys_income"])],
        ["Retiros / Egresos", _mxn(totals["bank_withdrawals"]), _mxn(totals["sys_expense"])],
        ["Movimientos conciliados", str(totals["matched_count"]), ""],
        ["Movimientos sin conciliar (banco)", str(totals["unmatched_bank_count"]), ""],
        ["Movimientos sin conciliar (sistema)", "", str(totals["unmatched_sys_count"])],
    ]
    r_tbl = Table(resumen, colWidths=[75*mm, 45*mm, 45*mm])
    r_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("BACKGROUND", (0,0), (-1,0), brand),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("ALIGN", (1,0), (-1,-1), "RIGHT"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#E0E0E0")),
    ]))
    story.append(r_tbl)
    story.append(Spacer(1, 8*mm))

    if rec.get("notes"):
        story.append(Paragraph(f"<b>Notas:</b> {rec['notes']}",
                                ParagraphStyle("n", parent=styles["Normal"], fontSize=9)))
        story.append(Spacer(1, 5*mm))

    # Firmas
    sign_line = "_" * 32
    sign_tbl = Table([
        [sign_line, sign_line, sign_line],
        ["Conciliado por", "Revisado por", "Autorizado por"],
    ], colWidths=[58*mm, 58*mm, 58*mm])
    sign_tbl.setStyle(TableStyle([
        ("FONTSIZE", (0,0), (-1,-1), 9),
        ("ALIGN", (0,0), (-1,-1), "CENTER"),
        ("TEXTCOLOR", (0,0), (-1,0), colors.HexColor("#999999")),
        ("FONTNAME", (0,1), (-1,1), "Helvetica-Bold"),
    ]))
    story.append(sign_tbl)
    story.append(Spacer(1, 6*mm))

    footer = (f"Generado el {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}"
              f" · Folio interno: CB-{rec['id']:06d}"
              f" · Estado: {rec['status'].upper()}")
    story.append(Paragraph(f"<font color='#999999' size='7.5'>{footer}</font>",
                            ParagraphStyle("f", parent=styles["Normal"], alignment=TA_CENTER)))

    doc.build(story)
    return buf.getvalue()


def build_reconciliation_xlsx(company: dict, workspace: dict, bank_account: dict) -> bytes:
    """XLSX detallado de la conciliación: 3 hojas (Resumen, Banco, Sistema)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rec = workspace["reconciliation"]
    totals = workspace["totals"]
    brand = (company.get("brand_color") or "#33B2F5").lstrip("#")
    brand_fill = PatternFill("solid", fgColor=brand)
    white_bold = Font(bold=True, color="FFFFFF")
    header_font = Font(bold=True, size=13, color=brand)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    ws["A1"] = f"Conciliación Bancaria — {rec['period_year']}-{rec['period_month']:02d}"
    ws["A1"].font = header_font
    ws["A2"] = f"Cuenta: {bank_account.get('name','')} · {bank_account.get('bank','')}"
    ws["A3"] = f"No. cuenta: {bank_account.get('account_number','')}"
    ws["A4"] = f"Empresa: {company.get('commercial_name','')}"

    rows = [
        ("Saldo inicial (banco)", rec["opening_balance"]),
        ("Saldo final (banco)", rec["closing_balance_bank"]),
        ("Saldo final (sistema)", rec["closing_balance_system"]),
        ("Diferencia", rec["difference"]),
        ("", ""),
        ("Depósitos (banco)", totals["bank_deposits"]),
        ("Retiros (banco)", totals["bank_withdrawals"]),
        ("Ingresos (sistema)", totals["sys_income"]),
        ("Egresos (sistema)", totals["sys_expense"]),
        ("", ""),
        ("Movimientos conciliados", totals["matched_count"]),
        ("Sin conciliar (banco)", totals["unmatched_bank_count"]),
        ("Sin conciliar (sistema)", totals["unmatched_sys_count"]),
    ]
    for i, (lbl, val) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=lbl).font = Font(bold=True) if val != "" else Font()
        c = ws.cell(row=i, column=2, value=val)
        if isinstance(val, (int, float)):
            c.number_format = '"$"#,##0.00' if abs(val) > 0 else "#,##0"
            c.alignment = Alignment(horizontal="right")
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22

    # Hoja Banco
    ws_b = wb.create_sheet("Movimientos banco")
    headers_b = ["Fecha", "Tipo", "Descripción", "Referencia", "Monto", "Conciliado", "Match ID"]
    for i, h in enumerate(headers_b, start=1):
        c = ws_b.cell(row=1, column=i, value=h)
        c.fill = brand_fill; c.font = white_bold
        c.alignment = Alignment(horizontal="center")
    for i, b in enumerate(workspace["bank_transactions"], start=2):
        ws_b.cell(row=i, column=1, value=(b["bank_date"] or b["created_at"] or "")[:10])
        ws_b.cell(row=i, column=2, value=b["type"])
        ws_b.cell(row=i, column=3, value=b["description"] or "")
        ws_b.cell(row=i, column=4, value=b["reference"] or "")
        c = ws_b.cell(row=i, column=5, value=b["amount"])
        c.number_format = '"$"#,##0.00'; c.alignment = Alignment(horizontal="right")
        ws_b.cell(row=i, column=6, value="Sí" if b["reconciled"] else "No")
        ws_b.cell(row=i, column=7, value=b["matched_transaction_id"] or "")
    for i, w in enumerate([12, 12, 40, 18, 14, 12, 10], start=1):
        ws_b.column_dimensions[chr(ord("A") + i - 1)].width = w

    # Hoja Sistema
    ws_s = wb.create_sheet("Transacciones sistema")
    headers_s = ["Fecha", "Tipo", "Categoría", "Descripción", "Referencia", "Monto", "Conciliada"]
    for i, h in enumerate(headers_s, start=1):
        c = ws_s.cell(row=1, column=i, value=h)
        c.fill = brand_fill; c.font = white_bold
        c.alignment = Alignment(horizontal="center")
    for i, t in enumerate(workspace["system_transactions"], start=2):
        ws_s.cell(row=i, column=1, value=(t["created_at"] or "")[:10])
        ws_s.cell(row=i, column=2, value=t["type"])
        ws_s.cell(row=i, column=3, value=t["category"] or "")
        ws_s.cell(row=i, column=4, value=t["description"] or "")
        ws_s.cell(row=i, column=5, value=t["reference"] or "")
        c = ws_s.cell(row=i, column=6, value=t["amount"])
        c.number_format = '"$"#,##0.00'; c.alignment = Alignment(horizontal="right")
        ws_s.cell(row=i, column=7, value="Sí" if t["matched"] else "No")
    for i, w in enumerate([12, 10, 18, 40, 18, 14, 12], start=1):
        ws_s.column_dimensions[chr(ord("A") + i - 1)].width = w

    buf = BytesIO(); wb.save(buf); return buf.getvalue()
