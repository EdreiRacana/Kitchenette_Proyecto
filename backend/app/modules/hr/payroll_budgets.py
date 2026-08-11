"""Presupuesto anual de nómina — planeación de costo laboral por empleado.

Módulo estándar de recursos humanos: cada empleado tiene un presupuesto
anual con 12 columnas mensuales (costo total proyectado, i.e. bruto +
prestaciones + carga patronal). El sistema compara mes a mes contra la
suma de PayrollDetail reales (regular + aguinaldo + prima_vacacional +
finiquito) de todos los períodos que se solapan con el mes.

Referencia real por mes = SUM(total_gross + imss_employer +
infonavit_employer + state_payroll_tax) de los PayrollDetail cuyo período
tiene start_date en el mes (aproximación estándar en nómina mexicana:
se prorratea al mes del inicio del período; catorcenal y semanal caen
al mes que arrancan).

Reglas del modelo:
- UniqueConstraint (employee_id, period_year) — un presupuesto por
  empleado/año.
- No hay dimensión de sucursal aquí (el empleado ya tiene department).
"""
from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import List, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.hr import models


MONTH_COLS = [f"m{i}" for i in range(1, 13)]
MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
         "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]


# ─── Helpers ────────────────────────────────────────────────────────────────


def _budget_to_dict(b: models.PayrollBudget, emp: Optional[models.Employee]) -> dict:
    item = {
        "id": b.id,
        "period_year": b.period_year,
        "employee_id": b.employee_id,
        "employee_number": emp.employee_number if emp else None,
        "employee_name": f"{emp.name} {emp.last_name}" if emp else None,
        "department": emp.department if emp else None,
        "position": emp.position if emp else None,
        "notes": b.notes,
        "annual_total": round(sum(getattr(b, c) or 0.0 for c in MONTH_COLS), 2),
    }
    for c in MONTH_COLS:
        item[c] = float(getattr(b, c) or 0.0)
    return item


def _month_from_iso(iso_date: str) -> Optional[int]:
    """Extrae el mes 1..12 de una fecha ISO YYYY-MM-DD. None si es inválida."""
    if not iso_date:
        return None
    try:
        return int(iso_date[5:7])
    except (ValueError, IndexError):
        return None


# ─── CRUD ───────────────────────────────────────────────────────────────────


async def list_budgets(db: AsyncSession, year: int) -> List[dict]:
    """Devuelve todos los presupuestos del año + info del empleado."""
    q = (
        select(models.PayrollBudget, models.Employee)
        .join(models.Employee, models.PayrollBudget.employee_id == models.Employee.id)
        .where(models.PayrollBudget.period_year == year)
        .order_by(models.Employee.department, models.Employee.name)
    )
    rows = (await db.execute(q)).all()
    return [_budget_to_dict(b, e) for b, e in rows]


async def upsert_budget(
    db: AsyncSession, *, year: int, employee_id: int, months: dict,
    notes: Optional[str] = None, user_id: Optional[int] = None,
) -> dict:
    """Crea o actualiza el presupuesto de un empleado/año."""
    q_emp = select(models.Employee).where(models.Employee.id == employee_id)
    emp = (await db.execute(q_emp)).scalars().first()
    if emp is None:
        raise ValueError(f"Empleado {employee_id} no existe")

    q = select(models.PayrollBudget).where(
        models.PayrollBudget.period_year == year,
        models.PayrollBudget.employee_id == employee_id,
    )
    b = (await db.execute(q)).scalars().first()
    if b is None:
        b = models.PayrollBudget(
            period_year=year, employee_id=employee_id,
            created_by_id=user_id,
        )
        db.add(b)
    for c in MONTH_COLS:
        if c in months and months[c] is not None:
            setattr(b, c, float(months[c] or 0.0))
    if notes is not None:
        b.notes = notes
    b.updated_by_id = user_id
    await db.commit()
    await db.refresh(b)
    return _budget_to_dict(b, emp)


async def delete_budget(db: AsyncSession, budget_id: int) -> bool:
    q = select(models.PayrollBudget).where(models.PayrollBudget.id == budget_id)
    b = (await db.execute(q)).scalars().first()
    if not b:
        return False
    await db.delete(b)
    await db.commit()
    return True


async def copy_from_year(
    db: AsyncSession, source_year: int, target_year: int,
    factor: float = 1.0, user_id: Optional[int] = None,
) -> int:
    """Copia los presupuestos del año origen al destino multiplicando por
    `factor` (útil para proyectar aumentos salariales: factor=1.05 = +5%).
    No sobreescribe: si ya existe presupuesto en target_year, se respeta."""
    q_src = select(models.PayrollBudget).where(
        models.PayrollBudget.period_year == source_year,
    )
    sources = (await db.execute(q_src)).scalars().all()
    if not sources:
        return 0

    q_existing = select(models.PayrollBudget.employee_id).where(
        models.PayrollBudget.period_year == target_year,
    )
    existing_ids = {r for r, in (await db.execute(q_existing)).all()}

    created = 0
    for src in sources:
        if src.employee_id in existing_ids:
            continue
        new_b = models.PayrollBudget(
            period_year=target_year,
            employee_id=src.employee_id,
            notes=(f"Copiado de {source_year}"
                   + (f" x {factor:.2%}" if factor != 1.0 else "")),
            created_by_id=user_id,
        )
        for c in MONTH_COLS:
            setattr(new_b, c, round(float(getattr(src, c) or 0.0) * factor, 2))
        db.add(new_b)
        created += 1
    await db.commit()
    return created


async def seed_from_active_employees(
    db: AsyncSession, year: int, user_id: Optional[int] = None,
) -> int:
    """Inicializa presupuesto en 0 para todos los empleados activos que
    aún no tengan presupuesto en ese año. Útil al iniciar la planeación."""
    q_emp = select(models.Employee.id).where(models.Employee.is_active == True)
    all_ids = {r for r, in (await db.execute(q_emp)).all()}

    q_existing = select(models.PayrollBudget.employee_id).where(
        models.PayrollBudget.period_year == year,
    )
    existing = {r for r, in (await db.execute(q_existing)).all()}

    to_create = all_ids - existing
    if not to_create:
        return 0

    for emp_id in to_create:
        db.add(models.PayrollBudget(
            period_year=year, employee_id=emp_id, created_by_id=user_id,
        ))
    await db.commit()
    return len(to_create)


# ─── Variance (Presupuesto vs Real) ─────────────────────────────────────────


async def compute_variance(db: AsyncSession, year: int) -> dict:
    """Comparativo mes a mes entre presupuesto y costo real de nómina.

    Real = SUM(total_gross + imss_employer + infonavit_employer +
                state_payroll_tax) por empleado × mes (usando el mes del
    start_date del período de nómina).
    """
    budgets = await list_budgets(db, year)
    if not budgets:
        return {"period_year": year, "rows": [], "totals": _empty_totals()}

    real_by_emp = await _real_per_employee_per_month(db, year)

    rows = []
    for b in budgets:
        emp_id = b["employee_id"]
        real_monthly = real_by_emp.get(emp_id, [0.0] * 12)

        budget_arr = [float(b[c] or 0.0) for c in MONTH_COLS]
        real_arr = [float(real_monthly[i] or 0.0) for i in range(12)]
        variance = [round(real_arr[i] - budget_arr[i], 2) for i in range(12)]
        variance_pct = [
            round(variance[i] / abs(budget_arr[i]) * 100, 1)
            if abs(budget_arr[i]) > 0.01 else None
            for i in range(12)
        ]

        rows.append({
            **b,
            "budget_monthly": budget_arr,
            "real_monthly": real_arr,
            "variance_monthly": variance,
            "variance_pct_monthly": variance_pct,
            "budget_ytd": round(sum(budget_arr), 2),
            "real_ytd": round(sum(real_arr), 2),
            "variance_ytd": round(sum(variance), 2),
            "variance_pct_ytd": (
                round(sum(variance) / abs(sum(budget_arr)) * 100, 1)
                if abs(sum(budget_arr)) > 0.01 else None
            ),
        })

    totals = _compute_totals(rows)
    return {"period_year": year, "rows": rows, "totals": totals}


def _empty_totals() -> dict:
    return {
        "budget_monthly": [0.0] * 12, "real_monthly": [0.0] * 12,
        "variance_monthly": [0.0] * 12,
        "budget_ytd": 0.0, "real_ytd": 0.0, "variance_ytd": 0.0,
    }


def _compute_totals(rows: List[dict]) -> dict:
    if not rows:
        return _empty_totals()
    bm = [round(sum(r["budget_monthly"][i] for r in rows), 2) for i in range(12)]
    rm = [round(sum(r["real_monthly"][i] for r in rows), 2) for i in range(12)]
    vm = [round(rm[i] - bm[i], 2) for i in range(12)]
    return {
        "budget_monthly": bm, "real_monthly": rm, "variance_monthly": vm,
        "budget_ytd": round(sum(bm), 2),
        "real_ytd": round(sum(rm), 2),
        "variance_ytd": round(sum(vm), 2),
    }


async def _real_per_employee_per_month(
    db: AsyncSession, year: int,
) -> dict:
    """{employee_id: [real_m1, ..., real_m12]} sumando costo total de nómina
    (total_gross + carga patronal) por mes del start_date del período."""
    q = (
        select(
            models.PayrollDetail.employee_id,
            models.PayrollPeriod.start_date,
            models.PayrollDetail.total_gross,
            models.PayrollDetail.imss_employer,
            models.PayrollDetail.infonavit_employer,
            models.PayrollDetail.state_payroll_tax,
        )
        .join(models.PayrollPeriod,
              models.PayrollDetail.period_id == models.PayrollPeriod.id)
    )
    out: dict[int, list[float]] = {}
    year_prefix = f"{year:04d}-"
    for emp_id, start_date, gross, imss_p, inf_p, isn in (await db.execute(q)).all():
        if not start_date or not start_date.startswith(year_prefix):
            continue
        m = _month_from_iso(start_date)
        if m is None:
            continue
        arr = out.setdefault(int(emp_id), [0.0] * 12)
        arr[m - 1] += (
            float(gross or 0.0)
            + float(imss_p or 0.0)
            + float(inf_p or 0.0)
            + float(isn or 0.0)
        )
    return out


# ─── XLSX export ────────────────────────────────────────────────────────────


def build_variance_xlsx(variance: dict) -> bytes:
    """XLSX profesional del comparativo presupuesto vs real de nómina."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    year = variance["period_year"]
    ws.title = f"Nomina {year}"

    brand = "33B2F5"
    header_fill = PatternFill("solid", fgColor=brand)
    section_fill = PatternFill("solid", fgColor="F0F5FA")
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    ws["A1"] = f"Presupuesto de Nomina vs Real — {year}"
    ws["A1"].font = Font(bold=True, size=14, color=brand)

    row0 = 3
    for col_i, title in enumerate(("Empleado", "Depto", "Puesto"), start=1):
        c = ws.cell(row=row0, column=col_i, value=title)
        c.fill = header_fill; c.font = header_font

    col = 4
    for m in MESES:
        for sub in ("Presup.", "Real", "Var."):
            c = ws.cell(row=row0, column=col, value=f"{m} {sub}")
            c.fill = header_fill; c.font = header_font
            c.alignment = center
            col += 1
    for sub in ("Presup. YTD", "Real YTD", "Var. YTD", "Var. %"):
        c = ws.cell(row=row0, column=col, value=sub)
        c.fill = header_fill; c.font = header_font
        c.alignment = center
        col += 1

    r_idx = row0 + 1
    for r in variance["rows"]:
        emp_label = f"{r.get('employee_number') or ''} · {r.get('employee_name') or ''}".strip(" ·")
        ws.cell(row=r_idx, column=1, value=emp_label)
        ws.cell(row=r_idx, column=2, value=r.get("department") or "")
        ws.cell(row=r_idx, column=3, value=r.get("position") or "")
        col = 4
        for i in range(12):
            for val in (r["budget_monthly"][i], r["real_monthly"][i], r["variance_monthly"][i]):
                c = ws.cell(row=r_idx, column=col, value=val)
                c.number_format = "#,##0.00"; c.alignment = right
                col += 1
        ws.cell(row=r_idx, column=col, value=r["budget_ytd"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col + 1, value=r["real_ytd"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col + 2, value=r["variance_ytd"]).number_format = "#,##0.00"
        vpct = r["variance_pct_ytd"]
        ws.cell(row=r_idx, column=col + 3,
                value=(f"{vpct:.1f}%" if vpct is not None else "—"))
        r_idx += 1

    totals = variance["totals"]
    r_idx += 1
    ws.cell(row=r_idx, column=1, value="TOTALES").font = Font(bold=True, size=12)
    for col_i in (1, 2, 3):
        ws.cell(row=r_idx, column=col_i).fill = section_fill
    col = 4
    for i in range(12):
        for val in (totals["budget_monthly"][i], totals["real_monthly"][i],
                    totals["variance_monthly"][i]):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.number_format = "#,##0.00"; c.alignment = right
            c.font = bold; c.fill = section_fill
            col += 1
    for val in (totals["budget_ytd"], totals["real_ytd"], totals["variance_ytd"]):
        c = ws.cell(row=r_idx, column=col, value=val)
        c.number_format = "#,##0.00"; c.font = bold; c.fill = section_fill
        col += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    for i in range(4, col + 1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = 13

    ws.freeze_panes = "D4"

    buf = BytesIO(); wb.save(buf); return buf.getvalue()
