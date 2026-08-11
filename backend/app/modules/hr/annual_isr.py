"""Ajuste Anual del ISR sobre salarios (LISR arts. 97 y 116).

Obligación del patrón: al cierre del ejercicio comparar el ISR retenido
mes a mes contra el ISR causado sobre los ingresos gravables anuales.
Si retuvo de más → devuelve al empleado (saldo a favor).
Si retuvo de menos → retiene la diferencia en las siguientes nóminas
del año siguiente (saldo a cargo).

**Exclusiones art. 97-A / 97-B**: el patrón NO realiza el ajuste cuando:

  I.   El empleado dejó de laborar antes del 1° de diciembre.
  II.  El empleado obtuvo ingresos anuales por salarios que excedan
       $400,000 (fracción III del art. 98).
  III. El empleado comunicó por escrito que hará su propia declaración
       anual (art. 97-B).
  IV.  El empleado inició la relación laboral después del 1° de enero
       (no hay ejercicio completo).

**Tarifa anual**: se construye multiplicando la mensual del Anexo 8 RMF
por 12 (li, ls y fi × 12; pct permanece igual). Fuente: el ISR anual
es acumulativo, por eso las bases crecen 12×.

Este módulo consume `_ISR_TABLE_MONTHLY` de `service.py` para no
duplicar la fuente de verdad.
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import List, Optional

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.hr import models
from app.modules.hr import service as _svc


# Umbral art. 98 frac. III LISR — ingresos por salarios sobre los cuales el
# empleado debe presentar declaración anual propia (2026: $400,000).
INGRESO_LIMITE_ANUAL = 400_000.0

# Tarifa anual construida desde la mensual (li, ls y fi × 12).
_ISR_TABLE_ANNUAL = [
    {"li": r["li"] * 12, "ls": r["ls"] * 12, "fi": r["fi"] * 12, "pct": r["pct"]}
    for r in _svc._ISR_TABLE_MONTHLY
]


# ─── Cálculo ────────────────────────────────────────────────────────────


def calc_isr_annual(gravable_anual: float) -> float:
    """Aplica la tarifa anual al ingreso gravable acumulado del ejercicio."""
    if (gravable_anual or 0.0) <= 0:
        return 0.0
    row = next(
        (r for r in _ISR_TABLE_ANNUAL if r["li"] <= gravable_anual <= r["ls"]),
        _ISR_TABLE_ANNUAL[-1],
    )
    isr = (gravable_anual - row["li"]) * row["pct"] + row["fi"]
    return round(max(isr, 0.0), 2)


def _full_name(emp: models.Employee) -> str:
    return f"{emp.name} {emp.last_name}".strip()


async def _accum_by_employee(db: AsyncSession, year: int) -> dict:
    """Acumula por empleado los conceptos del año (nómina calculada,
    aprobada o dispersada) desde PayrollDetail.

    Se distingue **total** (bruto) de **gravable estimado** (bruto menos
    los conceptos que suelen entrar exentos: vales de despensa hasta 7%
    UMA×30, aguinaldo hasta 30 UMA, prima vacacional hasta 15 UMA, fondo
    de ahorro topado a 13% UMA×30). Los cortes exactos se calculan sobre
    el valor UMA vigente del año.
    """
    q = (
        select(
            models.PayrollDetail.employee_id,
            models.PayrollDetail.salary_earned,
            models.PayrollDetail.overtime_double,
            models.PayrollDetail.overtime_triple,
            models.PayrollDetail.bonus,
            models.PayrollDetail.vacation_premium,
            models.PayrollDetail.food_vouchers,
            models.PayrollDetail.savings_fund,
            models.PayrollDetail.aguinaldo,
            models.PayrollDetail.subsidy_applied,
            models.PayrollDetail.isr,
            models.PayrollDetail.total_gross,
            models.PayrollDetail.days_worked,
        )
        .join(models.PayrollPeriod,
              models.PayrollDetail.period_id == models.PayrollPeriod.id)
        .where(
            models.PayrollPeriod.start_date >= f"{year:04d}-01-01",
            models.PayrollPeriod.start_date <= f"{year:04d}-12-31",
            models.PayrollPeriod.status.in_(["calculated", "approved", "dispersed"]),
        )
    )
    rows = (await db.execute(q)).all()

    # Topes de exención anuales
    uma_daily = _svc.UMA_2026
    uma_monthly = _svc.UMA_2026_MONTHLY
    tope_aguinaldo = 30 * uma_daily
    tope_prima_vac = 15 * uma_daily
    tope_vales_pct = 0.07 * uma_monthly * 12   # anual
    tope_fondo_ahorro = 0.13 * uma_monthly * 12  # anual

    per_emp: dict[int, dict] = {}
    for r in rows:
        emp_id = int(r[0])
        acc = per_emp.setdefault(emp_id, {
            "employee_id": emp_id,
            "salary": 0.0, "double": 0.0, "triple": 0.0, "bonus": 0.0,
            "prima_vac": 0.0, "vales": 0.0, "fondo": 0.0, "aguinaldo": 0.0,
            "sae_pagado": 0.0, "isr_retenido": 0.0, "total_gross": 0.0,
            "days": 0.0, "periods": 0,
        })
        acc["salary"] += float(r[1] or 0)
        acc["double"] += float(r[2] or 0)
        acc["triple"] += float(r[3] or 0)
        acc["bonus"] += float(r[4] or 0)
        acc["prima_vac"] += float(r[5] or 0)
        acc["vales"] += float(r[6] or 0)
        acc["fondo"] += float(r[7] or 0)
        acc["aguinaldo"] += float(r[8] or 0)
        acc["sae_pagado"] += float(r[9] or 0)
        acc["isr_retenido"] += float(r[10] or 0)
        acc["total_gross"] += float(r[11] or 0)
        acc["days"] += float(r[12] or 0)
        acc["periods"] += 1

    # Calcular gravable estimado por empleado (bruto - exentos)
    for a in per_emp.values():
        exento = 0.0
        # Aguinaldo exento hasta 30 UMA
        exento += min(a["aguinaldo"], tope_aguinaldo)
        # Prima vacacional exenta hasta 15 UMA
        exento += min(a["prima_vac"], tope_prima_vac)
        # Vales despensa exentos hasta 7% UMA×30×12
        exento += min(a["vales"], tope_vales_pct)
        # Fondo de ahorro exento hasta 13% UMA×30×12
        exento += min(a["fondo"], tope_fondo_ahorro)
        # Horas extra: las dobles son exentas al 100% hasta 5 UMA semanales,
        # pero simplificamos y las contamos gravables aquí (aproximación
        # conservadora — el patrón puede editar el manual override si
        # necesita el detalle preciso).
        gravable = max(a["total_gross"] - exento, 0.0)
        a["gravable"] = round(gravable, 2)
        a["exento_estimado"] = round(exento, 2)
    return per_emp


def _exclusion_reason(
    emp: models.Employee, total_ingresos: float, year: int,
) -> Optional[str]:
    """Devuelve la razón de exclusión (art. 97-A / 97-B) o None si aplica ajuste."""
    if emp.declares_own_annual:
        return "declara_propia"       # art. 97-B
    if total_ingresos > INGRESO_LIMITE_ANUAL:
        return "ingresos_excede_400k"  # art. 98-III
    # Alta después del 1° enero del año
    if emp.hire_date and emp.hire_date > f"{year:04d}-01-01":
        return "alta_mid_year"        # art. 97-A frac. IV (no ejercicio completo)
    # Baja antes del 1° diciembre del año — inactivo o contract_end
    is_baja = (emp.status or "") == "baja" or (emp.is_active is False)
    if is_baja:
        return "baja_pre_diciembre"   # art. 97-A frac. I
    if emp.contract_end and emp.contract_end < f"{year:04d}-12-01":
        return "baja_pre_diciembre"
    return None


async def calculate_adjustment(db: AsyncSession, year: int) -> dict:
    """Preview del ajuste anual (sin persistir)."""
    accum = await _accum_by_employee(db, year)
    if not accum:
        return {
            "period_year": year, "rows": [], "totals": _empty_totals(),
            "warnings": ["No hay nómina calculada / aprobada / dispersada en el año."],
        }

    # Trae Employee para exclusiones + nombres
    ids = list(accum.keys())
    q_emp = select(models.Employee).where(models.Employee.id.in_(ids))
    emps = {e.id: e for e in (await db.execute(q_emp)).scalars().all()}

    rows = []
    for emp_id, a in accum.items():
        emp = emps.get(emp_id)
        if not emp:
            continue

        reason = _exclusion_reason(emp, a["total_gross"], year)
        excluded = reason is not None

        isr_causado = calc_isr_annual(a["gravable"]) if not excluded else 0.0
        diferencia = round(a["isr_retenido"] - isr_causado, 2) if not excluded else 0.0

        rows.append({
            "employee_id": emp_id,
            "employee_number": emp.employee_number,
            "employee_name": _full_name(emp),
            "department": emp.department,
            "position": emp.position,
            "hire_date": emp.hire_date,
            "declares_own_annual": bool(emp.declares_own_annual),
            "total_ingresos": round(a["total_gross"], 2),
            "total_gravable": a["gravable"],
            "total_exento": a["exento_estimado"],
            "total_isr_retenido": round(a["isr_retenido"], 2),
            "total_sae_pagado": round(a["sae_pagado"], 2),
            "days_worked_year": round(a["days"], 2),
            "periods_count": a["periods"],
            "isr_causado_anual": isr_causado,
            "diferencia": diferencia,   # >0 saldo a favor, <0 saldo a cargo
            "excluded": excluded,
            "excluded_reason": reason,
        })

    rows.sort(key=lambda r: (r["department"] or "", r["employee_name"]))

    totals = _compute_totals(rows)
    return {
        "period_year": year, "rows": rows, "totals": totals,
        "warnings": [],
    }


def _empty_totals() -> dict:
    return {
        "total_employees": 0, "total_excluded": 0,
        "total_ingresos": 0.0, "total_isr_retenido": 0.0,
        "total_isr_causado": 0.0,
        "total_saldo_a_favor": 0.0, "total_saldo_a_cargo": 0.0,
    }


def _compute_totals(rows: List[dict]) -> dict:
    if not rows:
        return _empty_totals()
    included = [r for r in rows if not r["excluded"]]
    a_favor = sum(r["diferencia"] for r in included if r["diferencia"] > 0)
    a_cargo = sum(-r["diferencia"] for r in included if r["diferencia"] < 0)
    return {
        "total_employees": len(rows),
        "total_excluded": len(rows) - len(included),
        "total_ingresos": round(sum(r["total_ingresos"] for r in rows), 2),
        "total_isr_retenido": round(sum(r["total_isr_retenido"] for r in rows), 2),
        "total_isr_causado": round(sum(r["isr_causado_anual"] for r in rows), 2),
        "total_saldo_a_favor": round(a_favor, 2),
        "total_saldo_a_cargo": round(a_cargo, 2),
    }


# ─── Persistencia ──────────────────────────────────────────────────────


async def save_adjustment(
    db: AsyncSession, year: int, user_id: Optional[int] = None,
) -> dict:
    """Corre + persiste. Sobreescribe el draft del año si existe."""
    result = await calculate_adjustment(db, year)

    # Si ya hay un draft del año, lo reemplaza. Un applied se preserva.
    q = select(models.AnnualISRAdjustment).where(
        models.AnnualISRAdjustment.period_year == year,
    )
    existing = (await db.execute(q)).scalars().first()
    if existing:
        if existing.status == "applied":
            raise ValueError(
                f"El ajuste anual {year} ya está aplicado — no se puede recalcular."
            )
        await db.delete(existing)
        await db.flush()

    totals = result["totals"]
    adj = models.AnnualISRAdjustment(
        period_year=year, status="draft",
        total_employees=totals["total_employees"],
        total_excluded=totals["total_excluded"],
        total_saldo_a_favor=totals["total_saldo_a_favor"],
        total_saldo_a_cargo=totals["total_saldo_a_cargo"],
        calculated_by_id=user_id,
    )
    db.add(adj); await db.flush()

    for r in result["rows"]:
        db.add(models.AnnualISRDetail(
            adjustment_id=adj.id, employee_id=r["employee_id"],
            total_ingresos=r["total_ingresos"],
            total_gravable=r["total_gravable"],
            total_isr_retenido=r["total_isr_retenido"],
            total_sae_pagado=r["total_sae_pagado"],
            days_worked_year=r["days_worked_year"],
            periods_count=r["periods_count"],
            isr_causado_anual=r["isr_causado_anual"],
            diferencia=r["diferencia"],
            excluded=r["excluded"],
            excluded_reason=r["excluded_reason"],
        ))
    await db.commit()
    await db.refresh(adj)
    return await get_adjustment(db, adj.id)


async def get_adjustment(db: AsyncSession, adj_id: int) -> Optional[dict]:
    q = select(models.AnnualISRAdjustment).where(models.AnnualISRAdjustment.id == adj_id)
    adj = (await db.execute(q)).scalars().first()
    if not adj:
        return None
    q_d = (
        select(models.AnnualISRDetail, models.Employee)
        .join(models.Employee, models.AnnualISRDetail.employee_id == models.Employee.id)
        .where(models.AnnualISRDetail.adjustment_id == adj_id)
        .order_by(models.Employee.department, models.Employee.name)
    )
    rows = []
    for d, emp in (await db.execute(q_d)).all():
        rows.append({
            "id": d.id, "employee_id": emp.id,
            "employee_number": emp.employee_number,
            "employee_name": _full_name(emp),
            "department": emp.department, "position": emp.position,
            "hire_date": emp.hire_date,
            "declares_own_annual": bool(emp.declares_own_annual),
            "total_ingresos": d.total_ingresos,
            "total_gravable": d.total_gravable,
            "total_isr_retenido": d.total_isr_retenido,
            "total_sae_pagado": d.total_sae_pagado,
            "days_worked_year": d.days_worked_year,
            "periods_count": d.periods_count,
            "isr_causado_anual": d.isr_causado_anual,
            "diferencia": d.diferencia,
            "excluded": d.excluded, "excluded_reason": d.excluded_reason,
            "applied_to_period_id": d.applied_to_period_id,
        })
    totals = _compute_totals(rows)
    return {
        "id": adj.id, "period_year": adj.period_year, "status": adj.status,
        "notes": adj.notes,
        "total_employees": adj.total_employees, "total_excluded": adj.total_excluded,
        "total_saldo_a_favor": adj.total_saldo_a_favor,
        "total_saldo_a_cargo": adj.total_saldo_a_cargo,
        "approved_at": adj.approved_at.isoformat() if adj.approved_at else None,
        "created_at": adj.created_at.isoformat() if adj.created_at else None,
        "rows": rows, "totals": totals,
    }


async def list_adjustments(db: AsyncSession) -> List[dict]:
    q = select(models.AnnualISRAdjustment).order_by(models.AnnualISRAdjustment.period_year.desc())
    adjs = (await db.execute(q)).scalars().all()
    return [{
        "id": a.id, "period_year": a.period_year, "status": a.status,
        "total_employees": a.total_employees, "total_excluded": a.total_excluded,
        "total_saldo_a_favor": a.total_saldo_a_favor,
        "total_saldo_a_cargo": a.total_saldo_a_cargo,
        "created_at": a.created_at.isoformat() if a.created_at else None,
    } for a in adjs]


async def approve_adjustment(
    db: AsyncSession, adj_id: int, user_id: Optional[int] = None,
) -> dict:
    q = select(models.AnnualISRAdjustment).where(models.AnnualISRAdjustment.id == adj_id)
    adj = (await db.execute(q)).scalars().first()
    if not adj:
        raise ValueError("Ajuste no encontrado")
    if adj.status != "draft":
        raise ValueError(f"El ajuste está en estado '{adj.status}'. Ya no se puede aprobar.")
    adj.status = "approved"
    adj.approved_by_id = user_id
    adj.approved_at = datetime.utcnow()
    await db.commit()
    return await get_adjustment(db, adj.id)


async def delete_adjustment(db: AsyncSession, adj_id: int) -> bool:
    q = select(models.AnnualISRAdjustment).where(models.AnnualISRAdjustment.id == adj_id)
    adj = (await db.execute(q)).scalars().first()
    if not adj:
        return False
    if adj.status == "applied":
        raise ValueError("No se puede eliminar un ajuste ya aplicado a nómina.")
    await db.delete(adj)
    await db.commit()
    return True


# ─── XLSX export ────────────────────────────────────────────────────────


EXCLUSION_LABELS = {
    "declara_propia": "Empleado presenta declaración propia (art. 97-B)",
    "ingresos_excede_400k": "Ingresos anuales > $400,000 (art. 98-III)",
    "alta_mid_year": "Ingresó después del 1° enero (art. 97-A IV)",
    "baja_pre_diciembre": "Dejó de laborar antes del 1° diciembre (art. 97-A I)",
}


def build_adjustment_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook(); ws = wb.active
    year = data["period_year"]
    ws.title = f"Ajuste ISR {year}"

    brand = "33B2F5"
    header_fill = PatternFill("solid", fgColor=brand)
    excl_fill = PatternFill("solid", fgColor="FEF3C7")
    fav_fill = PatternFill("solid", fgColor="DCFCE7")
    car_fill = PatternFill("solid", fgColor="FEE2E2")
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    ws["A1"] = f"Ajuste Anual del ISR — Ejercicio {year}"
    ws["A1"].font = Font(bold=True, size=14, color=brand)
    ws["A2"] = "LISR arts. 97 y 116 · Comparativo ISR retenido vs. ISR causado anual"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    row0 = 4
    headers = [
        "Empleado", "Depto", "Ingresos anuales", "Gravable estimado",
        "ISR retenido", "ISR causado", "Diferencia", "Estado",
        "Observaciones",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row0, column=i, value=h)
        c.fill = header_fill; c.font = header_font

    r_idx = row0 + 1
    for r in data["rows"]:
        emp_label = f"{r['employee_number']} · {r['employee_name']}"
        if r["excluded"]:
            estado = "NO APLICA"; fill = excl_fill
            obs = EXCLUSION_LABELS.get(r["excluded_reason"] or "", r["excluded_reason"] or "")
        elif r["diferencia"] > 0.5:
            estado = "SALDO A FAVOR"; fill = fav_fill
            obs = f"Devolver ${r['diferencia']:,.2f} al empleado"
        elif r["diferencia"] < -0.5:
            estado = "SALDO A CARGO"; fill = car_fill
            obs = f"Retener ${-r['diferencia']:,.2f} adicionales en próximas nóminas"
        else:
            estado = "SIN DIFERENCIA"; fill = None
            obs = ""
        vals = [
            emp_label, r["department"] or "",
            r["total_ingresos"], r["total_gravable"],
            r["total_isr_retenido"], r["isr_causado_anual"], r["diferencia"],
            estado, obs,
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r_idx, column=i, value=v)
            if i >= 3 and i <= 7:
                c.number_format = "#,##0.00"; c.alignment = right
            if fill:
                c.fill = fill
        r_idx += 1

    # Totales
    r_idx += 1
    t = data["totals"]
    ws.cell(row=r_idx, column=1, value="TOTALES").font = Font(bold=True, size=12)
    ws.cell(row=r_idx, column=3, value=t["total_ingresos"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=5, value=t["total_isr_retenido"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=6, value=t["total_isr_causado"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=7, value=t["total_saldo_a_favor"] - t["total_saldo_a_cargo"]).number_format = "#,##0.00"
    for c in (1, 3, 5, 6, 7):
        ws.cell(row=r_idx, column=c).font = bold
        ws.cell(row=r_idx, column=c).alignment = right

    # Resumen abajo
    r_idx += 2
    ws.cell(row=r_idx, column=1, value=f"Total saldo a favor: ${t['total_saldo_a_favor']:,.2f}").font = Font(bold=True, color="15803D")
    ws.cell(row=r_idx + 1, column=1, value=f"Total saldo a cargo: ${t['total_saldo_a_cargo']:,.2f}").font = Font(bold=True, color="B91C1C")

    widths = {"A": 34, "B": 16, "C": 16, "D": 16, "E": 14, "F": 14, "G": 14, "H": 16, "I": 42}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = BytesIO(); wb.save(buf); return buf.getvalue()


# ─── PDF: constancia individual (art. 99 LISR) ─────────────────────────


async def build_constancia_pdf(
    db: AsyncSession, company: dict, year: int, employee_id: int,
) -> Optional[bytes]:
    """Constancia individual de retenciones (art. 99 LISR).

    Se entrega al trabajador para que sustente su declaración anual.
    Contiene: total ingresos, gravables, exentos, ISR retenido, subsidio
    al empleo pagado, saldo del ajuste anual si aplica.
    """
    q = select(models.Employee).where(models.Employee.id == employee_id)
    emp = (await db.execute(q)).scalars().first()
    if not emp:
        return None
    accum = await _accum_by_employee(db, year)
    a = accum.get(employee_id)
    if not a:
        return None

    isr_causado = calc_isr_annual(a["gravable"])
    diferencia = round(a["isr_retenido"] - isr_causado, 2)
    reason = _exclusion_reason(emp, a["total_gross"], year)
    excluded = reason is not None

    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.units import mm

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=LETTER,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=18 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    body = styles["BodyText"]; body.fontSize = 9
    small = ParagraphStyle("small", parent=body, fontSize=7.5, textColor=colors.grey)
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=13,
                        textColor=colors.HexColor("#0F172A"), spaceAfter=3)

    story = []

    company_name = (company or {}).get("legal_name") or "Empresa"
    company_rfc = (company or {}).get("rfc") or ""

    story.append(Paragraph(f"<b>{company_name}</b> — RFC {company_rfc}", body))
    story.append(Paragraph(
        f"Constancia individual de retenciones · Ejercicio {year}", h1,
    ))
    story.append(Paragraph(
        "Fundamento: Ley del ISR arts. 97 y 99 · Se entrega al trabajador para "
        "sustentar su declaración anual y/o el ajuste realizado por el patrón.",
        small,
    ))
    story.append(Spacer(1, 8))

    # Datos empleado
    emp_data = [
        ["Nombre", _full_name(emp)],
        ["RFC / CURP", f"{emp.rfc or ''}  ·  {emp.curp or ''}"],
        ["No. empleado", emp.employee_number],
        ["Fecha de ingreso", emp.hire_date],
        ["Puesto / Depto", f"{emp.position} — {emp.department}"],
    ]
    tbl_emp = Table(emp_data, colWidths=[45 * mm, 130 * mm])
    tbl_emp.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#E2E8F0")),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl_emp); story.append(Spacer(1, 8))

    # Ingresos y retenciones
    ingresos_rows = [
        ["Concepto", "Monto anual"],
        ["Total ingresos por salarios (bruto)", f"${a['total_gross']:,.2f}"],
        ["Ingresos gravables (estimado)", f"${a['gravable']:,.2f}"],
        ["Ingresos exentos (estimado)", f"${a['exento_estimado']:,.2f}"],
        ["ISR retenido durante el año", f"${a['isr_retenido']:,.2f}"],
        ["Subsidio al empleo entregado", f"${a['sae_pagado']:,.2f}"],
        ["ISR causado anual (tarifa art. 152)", f"${isr_causado:,.2f}"],
    ]
    tbl_ing = Table(ingresos_rows, colWidths=[110 * mm, 65 * mm])
    tbl_ing.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 1), (1, -1), "RIGHT"),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#E2E8F0")),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl_ing); story.append(Spacer(1, 8))

    # Resultado del ajuste
    if excluded:
        resultado = ("NO APLICA AJUSTE — "
                     f"{EXCLUSION_LABELS.get(reason, reason)}")
        color = "#B45309"
    elif diferencia > 0.5:
        resultado = f"SALDO A FAVOR DEL EMPLEADO: ${diferencia:,.2f}"
        color = "#15803D"
    elif diferencia < -0.5:
        resultado = f"SALDO A CARGO (a retener): ${-diferencia:,.2f}"
        color = "#B91C1C"
    else:
        resultado = "SIN DIFERENCIA — ISR causado ≈ ISR retenido"
        color = "#334155"

    story.append(Paragraph(
        f"<font color='{color}'><b>Resultado del ajuste anual:</b><br/>{resultado}</font>",
        ParagraphStyle("res", parent=body, fontSize=11, spaceAfter=8),
    ))

    # Firmas
    firms = Table([
        ["_____________________________", "", "_____________________________"],
        ["Patrón / Representante legal", "", "Trabajador"],
        ["", "", ""],
        [company_name, "", _full_name(emp)],
        [f"RFC {company_rfc}", "", f"RFC {emp.rfc or ''}"],
    ], colWidths=[70 * mm, 20 * mm, 70 * mm])
    firms.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(Spacer(1, 24)); story.append(firms)

    doc.build(story)
    return buf.getvalue()
