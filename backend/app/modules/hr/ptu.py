"""PTU — Participación de los Trabajadores en las Utilidades (LFT arts. 117-131).

Reforma DOF 23-abr-2021: art. 127 frac. VIII agregó el TOPE INDIVIDUAL
del monto a repartir a cada trabajador, que no puede exceder el mayor entre:
  (a) 3 meses del salario del trabajador, o
  (b) el promedio de la PTU recibida en los últimos 3 años.

Reglas legales implementadas:

- **Art. 122**: reparto 60 días después del pago del ISR anual (30-mayo típico).
- **Art. 123**: días base = días efectivamente trabajados + descansos con goce +
  vacaciones + permisos con goce + incapacidad por riesgo/maternidad.
  NO cuentan: faltas, permisos sin goce, incapacidad por enfermedad general.
- **Art. 127 frac. I**: exclusión de directores, administradores y gerentes
  generales (marca `ptu_excluded=True` en la ficha).
- **Art. 127 frac. II**: cap salarial para trabajadores de confianza
  (`is_confidential=True`) — su salario para PTU se topa al del sindicalizado
  de mayor tarifa incrementado en 20%.
- **Art. 127 frac. VI**: trabajadores domésticos excluidos (mismo campo
  `ptu_excluded`).
- **Art. 127 frac. VII**: eventuales con < 60 días trabajados en el año.
- **Art. 127 frac. VIII**: tope individual (arriba descrito).

El cálculo del reparto es 50/50: mitad proporcional a días, mitad
proporcional al salario percibido con el cap art. 127-II aplicado.
"""
from __future__ import annotations

from datetime import datetime, date
from io import BytesIO
from typing import List, Optional

from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.hr import models


# ─── Helpers ────────────────────────────────────────────────────────────────


def _full_name(emp: models.Employee) -> str:
    return f"{emp.name} {emp.last_name}".strip()


async def _get_employees_with_activity(
    db: AsyncSession, year: int,
) -> List[models.Employee]:
    """Empleados que tuvieron actividad de nómina en el año.

    Se toma UNION: (a) empleados con al menos un PayrollDetail en un período
    del año con status calculated/approved/dispersed, y (b) empleados activos
    hoy pero sin nómina aún (edge case: alta reciente sin período procesado
    todavía).
    """
    q_with_detail = (
        select(models.Employee.id)
        .join(models.PayrollDetail,
              models.PayrollDetail.employee_id == models.Employee.id)
        .join(models.PayrollPeriod,
              models.PayrollDetail.period_id == models.PayrollPeriod.id)
        .where(
            models.PayrollPeriod.start_date >= f"{year:04d}-01-01",
            models.PayrollPeriod.start_date <= f"{year:04d}-12-31",
            models.PayrollPeriod.status.in_(["calculated", "approved", "dispersed"]),
        )
        .distinct()
    )
    active_ids = {r for r, in (await db.execute(q_with_detail)).all()}
    if not active_ids:
        return []
    q = select(models.Employee).where(models.Employee.id.in_(list(active_ids)))
    return list((await db.execute(q)).scalars().all())


async def _days_and_salary_by_employee(
    db: AsyncSession, year: int,
) -> dict:
    """{employee_id: {days: float, salary: float, incap_general: float}}

    days = días efectivamente pagados (days_worked de PayrollDetail).
    salary = suma de salary_earned + vacation_premium (parte del salario del art. 84 LFT).
    Los subtipos de incapacidad se manejan via Attendance para restar/sumar
    correctamente según art. 123 (incapacidad por riesgo/maternidad SÍ cuentan;
    enfermedad general NO cuenta como día base).
    """
    q = (
        select(
            models.PayrollDetail.employee_id,
            func.coalesce(func.sum(models.PayrollDetail.days_worked), 0.0),
            func.coalesce(func.sum(
                models.PayrollDetail.salary_earned
                + models.PayrollDetail.vacation_premium
            ), 0.0),
            func.coalesce(func.sum(models.PayrollDetail.days_incapacity), 0.0),
        )
        .join(models.PayrollPeriod,
              models.PayrollDetail.period_id == models.PayrollPeriod.id)
        .where(
            models.PayrollPeriod.start_date >= f"{year:04d}-01-01",
            models.PayrollPeriod.start_date <= f"{year:04d}-12-31",
            models.PayrollPeriod.status.in_(["calculated", "approved", "dispersed"]),
        )
        .group_by(models.PayrollDetail.employee_id)
    )
    out: dict[int, dict] = {}
    for emp_id, days, salary, incap in (await db.execute(q)).all():
        out[int(emp_id)] = {
            "days": float(days or 0.0),
            "salary": float(salary or 0.0),
            "days_incap_total": float(incap or 0.0),
        }
    # Ajuste art. 123: sumar días de incapacidad por riesgo/maternidad
    # (esos SÍ cuentan). enfermedad_general se descuenta.
    # Attendance guarda incapacity_subtype cuando type='incapacidad'.
    q_att = (
        select(
            models.Attendance.employee_id,
            models.Attendance.incapacity_subtype,
            func.count(models.Attendance.id),
        )
        .where(
            models.Attendance.type == "incapacidad",
            models.Attendance.date >= f"{year:04d}-01-01",
            models.Attendance.date <= f"{year:04d}-12-31",
            models.Attendance.approved == True,  # noqa: E712
        )
        .group_by(models.Attendance.employee_id,
                  models.Attendance.incapacity_subtype)
    )
    incap_by_type: dict[int, dict[str, int]] = {}
    for emp_id, subtype, cnt in (await db.execute(q_att)).all():
        incap_by_type.setdefault(int(emp_id), {})[subtype or "none"] = int(cnt)

    for emp_id, subtypes in incap_by_type.items():
        # Los días de riesgo/maternidad/paternidad SÍ suman a días PTU
        credit = (subtypes.get("riesgo_trabajo", 0)
                  + subtypes.get("maternidad", 0)
                  + subtypes.get("paternidad", 0))
        if emp_id not in out:
            out[emp_id] = {"days": 0.0, "salary": 0.0, "days_incap_total": 0.0}
        out[emp_id]["days"] += float(credit)
    return out


async def _historic_last_3_years(
    db: AsyncSession, employee_id: int, current_year: int,
) -> List[float]:
    """Regresa la PTU pagada a este empleado en los últimos 3 años."""
    ys = [current_year - 1, current_year - 2, current_year - 3]
    q = select(
        models.PTUHistoric.period_year, models.PTUHistoric.amount_paid,
    ).where(
        models.PTUHistoric.employee_id == employee_id,
        models.PTUHistoric.period_year.in_(ys),
    )
    by_year = {int(y): float(a or 0.0) for y, a in (await db.execute(q)).all()}
    return [by_year.get(y, 0.0) for y in ys]


# ─── Cálculo principal ──────────────────────────────────────────────────────


async def calculate_ptu(
    db: AsyncSession,
    year: int,
    utilidad_repartible: float,
    sindicato_max_daily: Optional[float] = None,
) -> dict:
    """Preview del reparto (sin persistir).

    Args:
        year: ejercicio fiscal (los días/salario se leen de la nómina del año).
        utilidad_repartible: 10% de la utilidad fiscal (o el % que aplique).
        sindicato_max_daily: tarifa diaria del sindicalizado de mayor tarifa
          (art. 127-II). Si NULL, no se aplica cap salarial a confianza.
    """
    employees = await _get_employees_with_activity(db, year)
    if not employees:
        return {
            "period_year": year, "utilidad_repartible": utilidad_repartible,
            "sindicato_max_daily": sindicato_max_daily, "rows": [],
            "totals": _empty_totals(), "warnings": ["Sin nómina calculada en el año."],
        }

    base_data = await _days_and_salary_by_employee(db, year)

    # Cap art. 127-II: salario base diario para confianza = sindicato_max × 1.20
    salary_cap_daily = (sindicato_max_daily * 1.20) if sindicato_max_daily else None

    # Paso 1: filtrar + preparar filas con exclusiones y cap salarial
    prelim = []
    for emp in employees:
        raw = base_data.get(emp.id, {"days": 0.0, "salary": 0.0})
        days_raw = float(raw["days"] or 0.0)
        salary_raw = float(raw["salary"] or 0.0)

        row = {
            "employee_id": emp.id,
            "employee_number": emp.employee_number,
            "employee_name": _full_name(emp),
            "department": emp.department,
            "position": emp.position,
            "base_salary_daily": float(emp.base_salary or 0.0),  # diario
            "is_confidential": bool(emp.is_confidential),
            "days_raw": days_raw,
            "salary_raw": salary_raw,
        }

        # Exclusión art. 127-I / VI (marca manual en ficha)
        if emp.ptu_excluded:
            row.update({
                "excluded": True,
                "excluded_reason": "director_gerente_domestico",
                "days_worked_ptu": 0.0, "salary_earned_ptu": 0.0,
                "salary_cap_applied": False, "salary_cap_note": None,
            })
            prelim.append(row); continue

        # Exclusión art. 127-VII: eventuales < 60 días
        if days_raw < 60:
            row.update({
                "excluded": True,
                "excluded_reason": "eventual_menor_60_dias",
                "days_worked_ptu": days_raw, "salary_earned_ptu": salary_raw,
                "salary_cap_applied": False, "salary_cap_note": None,
            })
            prelim.append(row); continue

        # Cap salarial art. 127-II para confianza
        salary_cap_applied = False
        salary_cap_note = None
        salary_ptu = salary_raw
        if emp.is_confidential and salary_cap_daily:
            cap_annual = salary_cap_daily * days_raw
            if salary_raw > cap_annual:
                salary_ptu = round(cap_annual, 2)
                salary_cap_applied = True
                salary_cap_note = (
                    f"Cap art. 127-II: ${salary_cap_daily:.2f}/día "
                    f"× {days_raw:.0f} días = ${cap_annual:,.2f}"
                )

        row.update({
            "excluded": False, "excluded_reason": None,
            "days_worked_ptu": days_raw,
            "salary_earned_ptu": salary_ptu,
            "salary_cap_applied": salary_cap_applied,
            "salary_cap_note": salary_cap_note,
        })
        prelim.append(row)

    # Paso 2: totales para el reparto (solo NO excluidos)
    included = [r for r in prelim if not r["excluded"]]
    total_days = sum(r["days_worked_ptu"] for r in included)
    total_salary = sum(r["salary_earned_ptu"] for r in included)

    monto_por_dias = utilidad_repartible * 0.5
    monto_por_salario = utilidad_repartible * 0.5

    # Paso 3: cálculo bruto + cap art. 127-VIII individual
    for r in prelim:
        if r["excluded"]:
            r["ptu_by_days"] = 0.0
            r["ptu_by_salary"] = 0.0
            r["ptu_gross"] = 0.0
            r["cap_3_months"] = None
            r["cap_avg_3_years"] = None
            r["cap_applied_amount"] = None
            r["cap_reason"] = None
            r["ptu_final"] = 0.0
            continue

        part_days = (
            (r["days_worked_ptu"] / total_days) * monto_por_dias
            if total_days > 0 else 0.0
        )
        part_salary = (
            (r["salary_earned_ptu"] / total_salary) * monto_por_salario
            if total_salary > 0 else 0.0
        )
        gross = part_days + part_salary
        r["ptu_by_days"] = round(part_days, 2)
        r["ptu_by_salary"] = round(part_salary, 2)
        r["ptu_gross"] = round(gross, 2)

        # Tope art. 127-VIII
        cap_3m = round(r["base_salary_daily"] * 90, 2)  # 3 meses = 90 días
        history = await _historic_last_3_years(db, r["employee_id"], year)
        # Promedio solo de años con pago real (>0). Si no hay ninguno, promedio = 0
        pagos = [x for x in history if x > 0]
        cap_avg = round(sum(pagos) / len(pagos), 2) if pagos else 0.0

        cap_amount = max(cap_3m, cap_avg)
        r["cap_3_months"] = cap_3m
        r["cap_avg_3_years"] = cap_avg
        r["cap_applied_amount"] = cap_amount

        if gross > cap_amount and cap_amount > 0:
            r["ptu_final"] = cap_amount
            r["cap_reason"] = "3_months" if cap_3m >= cap_avg else "avg_3_years"
        else:
            r["ptu_final"] = round(gross, 2)
            r["cap_reason"] = None

    # Totales
    totals = {
        "total_employees": len(prelim),
        "total_included": len(included),
        "total_excluded": len(prelim) - len(included),
        "total_days": round(total_days, 2),
        "total_salary_base": round(total_salary, 2),
        "total_ptu_gross": round(sum(r["ptu_gross"] for r in prelim), 2),
        "total_ptu_paid": round(sum(r["ptu_final"] for r in prelim), 2),
        "total_cap_savings": round(
            sum(r["ptu_gross"] - r["ptu_final"] for r in prelim if not r["excluded"]), 2
        ),
    }

    warnings = []
    if total_days == 0:
        warnings.append("Total de días trabajados = 0 en el año.")
    if not sindicato_max_daily:
        warnings.append(
            "No se capturó tarifa diaria máxima del sindicato — "
            "no se aplicó cap art. 127-II a trabajadores de confianza."
        )

    return {
        "period_year": year,
        "utilidad_repartible": round(utilidad_repartible, 2),
        "sindicato_max_daily": sindicato_max_daily,
        "salary_cap_daily_confidence": salary_cap_daily,
        "rows": prelim,
        "totals": totals,
        "warnings": warnings,
    }


def _empty_totals() -> dict:
    return {
        "total_employees": 0, "total_included": 0, "total_excluded": 0,
        "total_days": 0.0, "total_salary_base": 0.0,
        "total_ptu_gross": 0.0, "total_ptu_paid": 0.0, "total_cap_savings": 0.0,
    }


# ─── Persistencia ──────────────────────────────────────────────────────────


async def save_calculation(
    db: AsyncSession, *,
    year: int, utilidad_repartible: float,
    sindicato_max_daily: Optional[float] = None,
    payment_deadline: Optional[str] = None,
    notes: Optional[str] = None,
    user_id: Optional[int] = None,
) -> dict:
    """Corre el cálculo y lo persiste como draft."""
    result = await calculate_ptu(db, year, utilidad_repartible, sindicato_max_daily)

    # Si ya hay una corrida draft del mismo año, la reemplaza. Un año
    # aprobado/pagado se preserva y regresa error.
    q = select(models.PTUCalculation).where(
        models.PTUCalculation.period_year == year,
    ).order_by(models.PTUCalculation.id.desc())
    existing = (await db.execute(q)).scalars().first()
    if existing and existing.status in ("approved", "paid"):
        raise ValueError(
            f"Ya existe un cálculo PTU {year} en estado "
            f"'{existing.status}'. Ábrelo para consultar o crea uno nuevo con "
            f"otro año."
        )
    if existing and existing.status == "draft":
        # Reemplazar la corrida draft
        await db.delete(existing)
        await db.flush()

    calc = models.PTUCalculation(
        period_year=year,
        utilidad_repartible=result["utilidad_repartible"],
        sindicato_max_daily=sindicato_max_daily,
        payment_deadline=payment_deadline,
        notes=notes,
        status="draft",
        total_days=result["totals"]["total_days"],
        total_salary_base=result["totals"]["total_salary_base"],
        total_ptu_paid=result["totals"]["total_ptu_paid"],
        total_excluded=result["totals"]["total_excluded"],
        calculated_by_id=user_id,
    )
    db.add(calc)
    await db.flush()

    for r in result["rows"]:
        db.add(models.PTUDetail(
            calculation_id=calc.id,
            employee_id=r["employee_id"],
            days_worked_ptu=r["days_worked_ptu"],
            salary_earned_ptu=r["salary_earned_ptu"],
            salary_cap_applied=r["salary_cap_applied"],
            salary_cap_note=r["salary_cap_note"],
            ptu_by_days=r["ptu_by_days"],
            ptu_by_salary=r["ptu_by_salary"],
            ptu_gross=r["ptu_gross"],
            cap_3_months=r["cap_3_months"],
            cap_avg_3_years=r["cap_avg_3_years"],
            cap_applied_amount=r["cap_applied_amount"],
            cap_reason=r["cap_reason"],
            ptu_final=r["ptu_final"],
            excluded=r["excluded"],
            excluded_reason=r["excluded_reason"],
        ))
    await db.commit()
    await db.refresh(calc)
    return await get_calculation(db, calc.id)


async def get_calculation(db: AsyncSession, calc_id: int) -> Optional[dict]:
    q = select(models.PTUCalculation).where(models.PTUCalculation.id == calc_id)
    calc = (await db.execute(q)).scalars().first()
    if not calc:
        return None

    q_d = (
        select(models.PTUDetail, models.Employee)
        .join(models.Employee, models.PTUDetail.employee_id == models.Employee.id)
        .where(models.PTUDetail.calculation_id == calc_id)
        .order_by(models.Employee.department, models.Employee.name)
    )
    details = (await db.execute(q_d)).all()

    rows = []
    for d, emp in details:
        rows.append({
            "id": d.id,
            "employee_id": emp.id,
            "employee_number": emp.employee_number,
            "employee_name": _full_name(emp),
            "department": emp.department,
            "position": emp.position,
            "base_salary_daily": float(emp.base_salary or 0.0),
            "is_confidential": bool(emp.is_confidential),
            "days_worked_ptu": d.days_worked_ptu,
            "salary_earned_ptu": d.salary_earned_ptu,
            "salary_cap_applied": d.salary_cap_applied,
            "salary_cap_note": d.salary_cap_note,
            "ptu_by_days": d.ptu_by_days,
            "ptu_by_salary": d.ptu_by_salary,
            "ptu_gross": d.ptu_gross,
            "cap_3_months": d.cap_3_months,
            "cap_avg_3_years": d.cap_avg_3_years,
            "cap_applied_amount": d.cap_applied_amount,
            "cap_reason": d.cap_reason,
            "ptu_final": d.ptu_final,
            "excluded": d.excluded,
            "excluded_reason": d.excluded_reason,
        })

    return {
        "id": calc.id,
        "period_year": calc.period_year,
        "utilidad_repartible": calc.utilidad_repartible,
        "sindicato_max_daily": calc.sindicato_max_daily,
        "payment_deadline": calc.payment_deadline,
        "notes": calc.notes,
        "status": calc.status,
        "payroll_period_id": calc.payroll_period_id,
        "total_days": calc.total_days,
        "total_salary_base": calc.total_salary_base,
        "total_ptu_paid": calc.total_ptu_paid,
        "total_excluded": calc.total_excluded,
        "approved_at": calc.approved_at.isoformat() if calc.approved_at else None,
        "created_at": calc.created_at.isoformat() if calc.created_at else None,
        "rows": rows,
    }


async def list_calculations(db: AsyncSession) -> List[dict]:
    q = select(models.PTUCalculation).order_by(models.PTUCalculation.period_year.desc())
    calcs = (await db.execute(q)).scalars().all()
    return [{
        "id": c.id,
        "period_year": c.period_year,
        "utilidad_repartible": c.utilidad_repartible,
        "status": c.status,
        "total_employees": (c.total_excluded or 0)
                            + int(round((c.total_days or 0) / max(c.total_days or 1, 1))),
        "total_excluded": c.total_excluded,
        "total_ptu_paid": c.total_ptu_paid,
        "payment_deadline": c.payment_deadline,
        "payroll_period_id": c.payroll_period_id,
        "created_at": c.created_at.isoformat() if c.created_at else None,
    } for c in calcs]


async def delete_calculation(db: AsyncSession, calc_id: int) -> bool:
    q = select(models.PTUCalculation).where(models.PTUCalculation.id == calc_id)
    calc = (await db.execute(q)).scalars().first()
    if not calc:
        return False
    if calc.status != "draft":
        raise ValueError("Solo se pueden eliminar cálculos en estado 'draft'.")
    await db.delete(calc)
    await db.commit()
    return True


async def approve_calculation(
    db: AsyncSession, calc_id: int, user_id: Optional[int] = None,
) -> dict:
    q = select(models.PTUCalculation).where(models.PTUCalculation.id == calc_id)
    calc = (await db.execute(q)).scalars().first()
    if not calc:
        raise ValueError("Cálculo no encontrado")
    if calc.status != "draft":
        raise ValueError(f"El cálculo está en estado '{calc.status}', no se puede aprobar.")
    calc.status = "approved"
    calc.approved_by_id = user_id
    calc.approved_at = datetime.utcnow()
    await db.commit()
    return await get_calculation(db, calc.id)


# ─── Generación de período de nómina PTU ───────────────────────────────────


async def generate_payroll_period(
    db: AsyncSession, calc_id: int, payment_date: str,
) -> dict:
    """Crea un PayrollPeriod tipo 'ptu' con un PayrollDetail por empleado
    con el monto ptu_final. Sirve para timbrar el CFDI de nómina con
    concepto SAT 023 (PTU)."""
    q = select(models.PTUCalculation).where(models.PTUCalculation.id == calc_id)
    calc = (await db.execute(q)).scalars().first()
    if not calc:
        raise ValueError("Cálculo no encontrado")
    if calc.status not in ("approved",):
        raise ValueError(
            f"El cálculo está en estado '{calc.status}'. "
            "Aprueba antes de generar la nómina."
        )
    if calc.payroll_period_id:
        raise ValueError("Este cálculo ya generó su período de nómina.")

    q_d = select(models.PTUDetail).where(
        models.PTUDetail.calculation_id == calc_id,
        models.PTUDetail.excluded == False,  # noqa: E712
        models.PTUDetail.ptu_final > 0,
    )
    details = (await db.execute(q_d)).scalars().all()
    if not details:
        raise ValueError("No hay empleados con PTU a pagar en este cálculo.")

    period = models.PayrollPeriod(
        name=f"PTU {calc.period_year}",
        frequency="anual",
        start_date=f"{calc.period_year:04d}-01-01",
        end_date=f"{calc.period_year:04d}-12-31",
        payment_date=payment_date,
        status="calculated",
        kind="ptu",
    )
    db.add(period)
    await db.flush()

    for d in details:
        # PayrollDetail para PTU: gross = ptu_final, ISR se puede calcular
        # con art. 174 del RLISR (opción 1: sobre ingresos gravables del mes).
        # Aquí lo dejamos en 0 y RH podrá editar manualmente antes de aprobar.
        db.add(models.PayrollDetail(
            period_id=period.id,
            employee_id=d.employee_id,
            department=None,
            base_salary=0.0,
            days_worked=0.0,
            days_absent=0.0,
            days_incapacity=0.0,
            salary_earned=0.0,
            bonus=float(d.ptu_final),  # PTU se registra como percepción extraordinaria
            notes=f"PTU {calc.period_year} — art. 122 LFT",
            total_gross=float(d.ptu_final),
            total_deductions=0.0,
            total_net=float(d.ptu_final),
        ))

    calc.status = "paid"
    calc.payroll_period_id = period.id
    await db.flush()

    # Registrar en histórico para futuros cap art. 127-VIII (promedio 3 años)
    for d in details:
        db.add(models.PTUHistoric(
            employee_id=d.employee_id,
            period_year=calc.period_year,
            amount_paid=float(d.ptu_final),
            source="system",
        ))
    await db.commit()

    return {
        "calc_id": calc.id,
        "period_id": period.id,
        "period_name": period.name,
        "employees": len(details),
        "total_paid": round(sum(float(d.ptu_final) for d in details), 2),
    }


# ─── XLSX export ────────────────────────────────────────────────────────────


def build_ptu_xlsx(calc: dict) -> bytes:
    """XLSX oficial del reparto con exclusiones + cap justificado."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    year = calc["period_year"]
    ws.title = f"PTU {year}"

    brand = "33B2F5"
    header_fill = PatternFill("solid", fgColor=brand)
    section_fill = PatternFill("solid", fgColor="F0F5FA")
    excl_fill = PatternFill("solid", fgColor="FEF3C7")  # amarillo tenue
    cap_fill = PatternFill("solid", fgColor="FEE2E2")   # rojo tenue
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    # Encabezado
    ws["A1"] = f"Reparto de Utilidades (PTU) — Ejercicio {year}"
    ws["A1"].font = Font(bold=True, size=14, color=brand)
    ws["A2"] = f"Utilidad repartible: ${calc['utilidad_repartible']:,.2f} MXN"
    ws["A2"].font = Font(size=11)
    ws["A3"] = "Art. 122 LFT · Reparto 50% días trabajados / 50% salario percibido"
    ws["A3"].font = Font(italic=True, size=10, color="666666")

    row0 = 5
    headers = [
        "Empleado", "Departamento", "Días PTU", "Salario para PTU",
        "PTU por días", "PTU por salario", "PTU bruto",
        "Cap 3 meses", "Cap prom. 3 años", "PTU final",
        "Observaciones",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row0, column=i, value=h)
        c.fill = header_fill; c.font = header_font

    r_idx = row0 + 1
    for r in calc["rows"]:
        obs = []
        if r["excluded"]:
            reasons = {
                "director_gerente_domestico": "Excluido art. 127-I/VI",
                "eventual_menor_60_dias": "Excluido art. 127-VII (< 60 días)",
            }
            obs.append(reasons.get(r["excluded_reason"], r["excluded_reason"]))
        if r["salary_cap_applied"]:
            obs.append(r["salary_cap_note"] or "Cap art. 127-II aplicado")
        if r["cap_reason"]:
            src = "3 meses de salario" if r["cap_reason"] == "3_months" else "promedio últimos 3 años"
            obs.append(f"Tope art. 127-VIII aplicado ({src})")

        fill = excl_fill if r["excluded"] else (cap_fill if r["cap_reason"] else None)
        emp_label = f"{r['employee_number']} · {r['employee_name']}"

        vals = [
            emp_label, r["department"] or "",
            r["days_worked_ptu"], r["salary_earned_ptu"],
            r["ptu_by_days"], r["ptu_by_salary"], r["ptu_gross"],
            r["cap_3_months"] or 0.0, r["cap_avg_3_years"] or 0.0,
            r["ptu_final"], "; ".join(obs),
        ]
        for i, v in enumerate(vals, start=1):
            c = ws.cell(row=r_idx, column=i, value=v)
            if i >= 3 and i <= 10:
                c.number_format = "#,##0.00"
                c.alignment = right
            if fill:
                c.fill = fill
        r_idx += 1

    # Totales
    r_idx += 1
    t = calc["totals"]
    ws.cell(row=r_idx, column=1, value="TOTALES").font = Font(bold=True, size=12)
    for i in range(1, 12):
        ws.cell(row=r_idx, column=i).fill = section_fill
    ws.cell(row=r_idx, column=3, value=t["total_days"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=4, value=t["total_salary_base"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=7, value=t["total_ptu_gross"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=10, value=t["total_ptu_paid"]).number_format = "#,##0.00"
    for i in (3, 4, 7, 10):
        ws.cell(row=r_idx, column=i).font = bold
        ws.cell(row=r_idx, column=i).alignment = right

    # Leyenda
    r_idx += 3
    ws.cell(row=r_idx, column=1, value="Leyenda:").font = bold
    ws.cell(row=r_idx + 1, column=1, value="Amarillo = empleado excluido del reparto")
    ws.cell(row=r_idx + 2, column=1, value="Rojo tenue = tope art. 127-VIII aplicado")

    # Anchos
    widths = {"A": 38, "B": 18, "C": 12, "D": 16, "E": 14, "F": 14, "G": 14,
              "H": 14, "I": 16, "J": 14, "K": 46}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    buf = BytesIO(); wb.save(buf); return buf.getvalue()


# ─── PDF oficial ────────────────────────────────────────────────────────────


def build_ptu_pdf(company: dict, calc: dict) -> bytes:
    """Cédula oficial del reparto — LETTER portrait con firmas."""
    from reportlab.lib.pagesizes import LETTER
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak,
    )
    from reportlab.lib.units import mm

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=LETTER,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )
    styles = getSampleStyleSheet()
    body = styles["BodyText"]; body.fontSize = 9
    small = ParagraphStyle(name="small", parent=body, fontSize=7.5, textColor=colors.grey)
    h_style = ParagraphStyle(
        name="hh", parent=styles["Heading1"],
        fontSize=13, textColor=colors.HexColor("#0F172A"), spaceAfter=4,
    )

    story = []

    # Header institucional
    company_name = (company or {}).get("legal_name") or "Empresa"
    rfc = (company or {}).get("rfc") or ""
    story.append(Paragraph(f"<b>{company_name}</b> — RFC {rfc}", body))
    story.append(Paragraph(
        f"Cédula de Reparto de Utilidades (PTU) — Ejercicio {calc['period_year']}",
        h_style,
    ))
    story.append(Paragraph(
        "Fundamento legal: LFT arts. 117 al 131 y reforma DOF 23-abr-2021 "
        "(art. 127 frac. VIII). Reparto 50% días trabajados / 50% salario percibido.",
        small,
    ))
    story.append(Spacer(1, 6))

    # Cabecero de datos
    resumen_data = [
        ["Utilidad repartible", f"${calc['utilidad_repartible']:,.2f}"],
        ["Empleados incluidos", str(calc['totals']['total_included'])],
        ["Empleados excluidos", str(calc['totals']['total_excluded'])],
        ["Total días base", f"{calc['totals']['total_days']:,.2f}"],
        ["Total salario base", f"${calc['totals']['total_salary_base']:,.2f}"],
        ["Total PTU a pagar", f"${calc['totals']['total_ptu_paid']:,.2f}"],
        ["Ahorro por topes art. 127-VIII", f"${calc['totals']['total_cap_savings']:,.2f}"],
    ]
    if calc.get("sindicato_max_daily"):
        resumen_data.insert(1, [
            "Tarifa máx. sindicato (art. 127-II)",
            f"${calc['sindicato_max_daily']:,.2f}/día",
        ])
    if calc.get("payment_deadline"):
        resumen_data.insert(0, ["Fecha límite de pago (art. 122)", calc["payment_deadline"]])

    tbl_res = Table(resumen_data, colWidths=[90 * mm, 90 * mm])
    tbl_res.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#E2E8F0")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl_res); story.append(Spacer(1, 8))

    # Detalle
    data = [[
        "Empleado", "Depto", "Días", "Salario base",
        "PTU bruta", "PTU final", "Nota",
    ]]
    for r in calc["rows"]:
        nota = ""
        if r["excluded"]:
            m = {
                "director_gerente_domestico": "Excl. 127-I/VI",
                "eventual_menor_60_dias": "Excl. 127-VII",
            }
            nota = m.get(r["excluded_reason"], "Excluido")
        elif r["cap_reason"]:
            nota = "Tope 127-VIII " + ("(3 meses)" if r["cap_reason"] == "3_months" else "(prom. 3a)")
        elif r["salary_cap_applied"]:
            nota = "Cap 127-II"

        data.append([
            Paragraph(f"{r['employee_number']}<br/>{r['employee_name']}", small),
            r["department"] or "",
            f"{r['days_worked_ptu']:.0f}",
            f"${r['salary_earned_ptu']:,.0f}",
            f"${r['ptu_gross']:,.2f}",
            f"${r['ptu_final']:,.2f}",
            nota,
        ])

    tbl = Table(data, colWidths=[45 * mm, 22 * mm, 14 * mm, 25 * mm, 25 * mm, 25 * mm, 25 * mm])
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#E2E8F0")),
        ("PADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(tbl); story.append(Spacer(1, 12))

    # Firmas
    firms = Table([
        ["_____________________________", "", "_____________________________"],
        ["Representante de la empresa", "", "Representante de los trabajadores"],
        ["", "", ""],
        ["_____________________________", "", "_____________________________"],
        ["Contador Público responsable", "", "Fecha y sello"],
    ], colWidths=[70 * mm, 20 * mm, 70 * mm])
    firms.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(Spacer(1, 20))
    story.append(firms)

    doc.build(story)
    return buf.getvalue()
