"""Cédulas de determinación IMSS / SAR / INFONAVIT.

Documentos de apoyo para el pago de cuotas al IMSS y aportaciones al
INFONAVIT/SAR. NO reemplazan al SUA (Sistema Único de Autodeterminación),
sino que sirven de cédula PRE para validar/capturar los montos.

**Ramos legales del IMSS (Art. 25, 106, 147, 168 LSS)** — mensual:

  EGM (Enfermedad y Maternidad):
    · Especie — cuota fija patrón 20.40% del SMDF, excedente 1.10%
      patrón / 0.40% obrero sobre (SBC − 3 SMDF)
    · Dinero — 0.70% patrón / 0.25% obrero
  IV (Invalidez y Vida):        1.75% patrón / 0.625% obrero
  GPS (Guarderías y Prest Soc.): 1.00% patrón / — obrero
  RT (Riesgo de Trabajo):        variable por prima (default 0.54355%)

**Ramos legales bimestrales**:

  Retiro (SAR — Art. 168 fracción I LSS):       2.00% patrón
  CV (Cesantía y Vejez — Art. 168 II LSS):      3.150% patrón / 1.125% obrero
  INFONAVIT aportación (Art. 29 LINFONAVIT):    5.00% patrón
  INFONAVIT amortización de crédito:            según discount_type del empleado

Todos los SBC se topan a 25 UMAs diarias.

**Pagos**: mensual y bimestral se pagan a más tardar el día 17 del mes
siguiente al periodo (o siguiente hábil si cae en fin de semana o festivo).
"""
from __future__ import annotations

from calendar import monthrange
from datetime import date, timedelta
from io import BytesIO
from typing import List, Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.hr import models
from app.modules.hr import service as _svc


UMA = _svc.UMA_2026
TOPE_SBC = 25 * UMA

# Tasas ramos mensuales (patrón, obrero)
RATE = {
    "em_especie_cuota_fija":  (0.2040, 0.0000),  # sobre SMDF (UMA)
    "em_especie_excedente":   (0.0110, 0.0040),  # sobre (SBC − 3 UMA)
    "em_dinero":              (0.0070, 0.0025),
    "iv":                     (0.0175, 0.00625),
    "gps":                    (0.0100, 0.0000),
    "rt_default":             (0.0054355, 0.0000),  # el patrón real la lee del Anexo 5.9
}
# Tasas bimestrales
RATE_BIMESTRAL = {
    "retiro":                 (0.0200, 0.0000),
    "cv":                     (0.03150, 0.01125),
    "infonavit_aportacion":   (0.0500, 0.0000),
}

RAMO_LABEL = {
    "em_especie_cuota_fija":  "E y M · Especie (cuota fija)",
    "em_especie_excedente":   "E y M · Especie (excedente)",
    "em_dinero":              "E y M · Dinero",
    "iv":                     "Invalidez y Vida",
    "gps":                    "Guarderías y Prest. Sociales",
    "rt_default":             "Riesgo de Trabajo",
    "retiro":                 "Retiro (SAR)",
    "cv":                     "Cesantía y Vejez",
    "infonavit_aportacion":   "INFONAVIT (aportación 5%)",
}


# ─── Helpers ────────────────────────────────────────────────────────────


def _full_name(emp: models.Employee) -> str:
    return f"{emp.name} {emp.last_name}".strip()


def _bimester_bounds(year: int, bimester: int) -> tuple[str, str]:
    """1 → ene-feb, 2 → mar-abr, ..., 6 → nov-dic. Devuelve (start, end) ISO."""
    start_month = (bimester - 1) * 2 + 1
    end_month = start_month + 1
    end_day = monthrange(year, end_month)[1]
    return (f"{year:04d}-{start_month:02d}-01",
            f"{year:04d}-{end_month:02d}-{end_day:02d}")


def _month_bounds(year: int, month: int) -> tuple[str, str]:
    end_day = monthrange(year, month)[1]
    return (f"{year:04d}-{month:02d}-01", f"{year:04d}-{month:02d}-{end_day:02d}")


def _payment_deadline(period_end_iso: str) -> str:
    """Día 17 del mes siguiente al periodo (o siguiente hábil si cae en fin de semana)."""
    y, m, d = [int(x) for x in period_end_iso.split("-")]
    next_month = m + 1
    next_year = y
    if next_month > 12:
        next_month = 1; next_year = y + 1
    dl = date(next_year, next_month, 17)
    # Si cae sábado (5) o domingo (6), mover al lunes siguiente.
    while dl.weekday() > 4:
        dl += timedelta(days=1)
    return dl.isoformat()


async def _aggregate_by_employee(
    db: AsyncSession, start_iso: str, end_iso: str,
) -> List[dict]:
    """SBC × días cotizados por empleado en el rango, tomando PayrollDetail
    de periodos con start_date dentro del rango y status calc/appr/disp."""
    q = (
        select(
            models.PayrollDetail.employee_id,
            models.PayrollDetail.days_worked,
            models.PayrollDetail.days_incapacity,
            models.PayrollDetail.infonavit,
        )
        .join(models.PayrollPeriod,
              models.PayrollDetail.period_id == models.PayrollPeriod.id)
        .where(
            models.PayrollPeriod.start_date >= start_iso,
            models.PayrollPeriod.start_date <= end_iso,
            models.PayrollPeriod.status.in_(["calculated", "approved", "dispersed"]),
            models.PayrollPeriod.kind == "regular",  # excluye aguinaldo/finiquito/PTU
        )
    )
    per_emp: dict[int, dict] = {}
    for emp_id, days, incap, inf_desc in (await db.execute(q)).all():
        acc = per_emp.setdefault(int(emp_id), {
            "employee_id": int(emp_id),
            "days_cotizados": 0.0,
            "days_incapacity": 0.0,
            "infonavit_descuento": 0.0,
        })
        acc["days_cotizados"] += float(days or 0)
        acc["days_incapacity"] += float(incap or 0)
        acc["infonavit_descuento"] += float(inf_desc or 0)

    if not per_emp:
        return []

    q_emp = select(models.Employee).where(models.Employee.id.in_(list(per_emp.keys())))
    emps = {e.id: e for e in (await db.execute(q_emp)).scalars().all()}
    out = []
    for emp_id, d in per_emp.items():
        emp = emps.get(emp_id)
        if not emp:
            continue
        # Días para IMSS = días cotizados; los días de incapacidad por
        # riesgo/maternidad los subsidia el IMSS pero el patrón NO paga
        # cuotas por ellos (art. 31 LSS). Aquí usamos days_worked de la
        # PayrollDetail que ya considera esa exclusión.
        out.append({
            **d,
            "employee_number": emp.employee_number,
            "employee_name": _full_name(emp),
            "curp": emp.curp,
            "nss": emp.nss,
            "rfc": emp.rfc,
            "sbc": float(emp.sbc or 0.0),
            "sbc_topado": min(float(emp.sbc or 0.0), TOPE_SBC),
            "infonavit_credit": emp.infonavit_credit,
        })
    return sorted(out, key=lambda r: r["employee_number"])


# ─── Cálculo por empleado × ramo ────────────────────────────────────────


def _calc_monthly_per_employee(r: dict, prima_rt: Optional[float] = None) -> dict:
    """Regresa {ramo: {patron, obrero, base}} y totales."""
    sbc = r["sbc_topado"]
    dias = r["days_cotizados"]
    ramos = {}

    # E y M especie cuota fija — base UMA (SMDF)
    p, o = RATE["em_especie_cuota_fija"]
    ramos["em_especie_cuota_fija"] = {
        "base": UMA, "dias": dias,
        "patron": round(p * UMA * dias, 2),
        "obrero": 0.0,
    }
    # E y M especie excedente — base (SBC − 3 UMA)
    exc = max(sbc - 3 * UMA, 0)
    p, o = RATE["em_especie_excedente"]
    ramos["em_especie_excedente"] = {
        "base": exc, "dias": dias,
        "patron": round(p * exc * dias, 2),
        "obrero": round(o * exc * dias, 2),
    }
    # Ramos sobre SBC
    for k in ("em_dinero", "iv", "gps"):
        p, o = RATE[k]
        ramos[k] = {
            "base": sbc, "dias": dias,
            "patron": round(p * sbc * dias, 2),
            "obrero": round(o * sbc * dias, 2),
        }
    # Riesgo de trabajo (prima variable — default 0.54355%)
    rt_rate = prima_rt if (prima_rt and prima_rt > 0) else RATE["rt_default"][0]
    ramos["rt_default"] = {
        "base": sbc, "dias": dias, "prima": rt_rate,
        "patron": round(rt_rate * sbc * dias, 2),
        "obrero": 0.0,
    }

    total_patron = round(sum(v["patron"] for v in ramos.values()), 2)
    total_obrero = round(sum(v["obrero"] for v in ramos.values()), 2)
    return {
        "ramos": ramos,
        "total_patron": total_patron,
        "total_obrero": total_obrero,
        "total_pagar_imss": round(total_patron + total_obrero, 2),
    }


def _calc_bimonthly_per_employee(r: dict) -> dict:
    sbc = r["sbc_topado"]
    dias = r["days_cotizados"]
    ramos = {}
    for k in ("retiro", "cv", "infonavit_aportacion"):
        p, o = RATE_BIMESTRAL[k]
        ramos[k] = {
            "base": sbc, "dias": dias,
            "patron": round(p * sbc * dias, 2),
            "obrero": round(o * sbc * dias, 2),
        }
    total_patron = round(sum(v["patron"] for v in ramos.values()), 2)
    total_obrero = round(sum(v["obrero"] for v in ramos.values()), 2)
    inf_desc = round(float(r.get("infonavit_descuento") or 0), 2)
    return {
        "ramos": ramos,
        "infonavit_descuento": inf_desc,   # amortización trabajador
        "total_patron": total_patron,
        "total_obrero": total_obrero,
        "total_pagar": round(total_patron + total_obrero + inf_desc, 2),
    }


# ─── API pública ────────────────────────────────────────────────────────


async def cedula_mensual(
    db: AsyncSession, year: int, month: int, prima_rt: Optional[float] = None,
) -> dict:
    start, end = _month_bounds(year, month)
    emps = await _aggregate_by_employee(db, start, end)
    rows = []
    for r in emps:
        calc = _calc_monthly_per_employee(r, prima_rt=prima_rt)
        rows.append({**r, **calc})
    totals = _totals(rows, list(RATE.keys()))
    return {
        "period_year": year, "period_month": month,
        "start_date": start, "end_date": end,
        "prima_rt": prima_rt or RATE["rt_default"][0],
        "payment_deadline": _payment_deadline(end),
        "rows": rows, "totals": totals,
        "ramos_labels": {k: RAMO_LABEL[k] for k in RATE},
    }


async def cedula_bimestral(
    db: AsyncSession, year: int, bimester: int,
) -> dict:
    start, end = _bimester_bounds(year, bimester)
    emps = await _aggregate_by_employee(db, start, end)
    rows = []
    for r in emps:
        calc = _calc_bimonthly_per_employee(r)
        rows.append({**r, **calc})
    totals = _totals(rows, list(RATE_BIMESTRAL.keys()))
    totals["total_infonavit_descuento"] = round(
        sum(r["infonavit_descuento"] for r in rows), 2
    )
    totals["total_pagar_infonavit"] = round(
        sum(r["ramos"]["infonavit_aportacion"]["patron"] for r in rows)
        + totals["total_infonavit_descuento"], 2
    )
    return {
        "period_year": year, "bimester": bimester,
        "start_date": start, "end_date": end,
        "payment_deadline": _payment_deadline(end),
        "rows": rows, "totals": totals,
        "ramos_labels": {k: RAMO_LABEL[k] for k in RATE_BIMESTRAL},
    }


def _totals(rows: List[dict], ramo_keys: List[str]) -> dict:
    if not rows:
        return {
            "total_employees": 0, "total_days": 0.0,
            "total_patron": 0.0, "total_obrero": 0.0, "total_pagar_imss": 0.0,
            "por_ramo": {k: {"patron": 0.0, "obrero": 0.0} for k in ramo_keys},
        }
    por_ramo: dict = {k: {"patron": 0.0, "obrero": 0.0} for k in ramo_keys}
    for r in rows:
        for k in ramo_keys:
            v = r["ramos"].get(k, {})
            por_ramo[k]["patron"] += float(v.get("patron", 0))
            por_ramo[k]["obrero"] += float(v.get("obrero", 0))
    return {
        "total_employees": len(rows),
        "total_days": round(sum(r["days_cotizados"] for r in rows), 2),
        "total_patron": round(sum(r["total_patron"] for r in rows), 2),
        "total_obrero": round(sum(r["total_obrero"] for r in rows), 2),
        "total_pagar_imss": round(
            sum(r.get("total_pagar_imss") or r.get("total_pagar") or 0 for r in rows), 2
        ),
        "por_ramo": {k: {"patron": round(v["patron"], 2), "obrero": round(v["obrero"], 2)}
                      for k, v in por_ramo.items()},
    }


# ─── XLSX export ────────────────────────────────────────────────────────


def build_xlsx(cedula: dict, kind: str) -> bytes:
    """kind = 'mensual' | 'bimestral'."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook(); ws = wb.active
    if kind == "mensual":
        title = f"Cedula IMSS mensual {cedula['period_year']}-{cedula['period_month']:02d}"
        ramos_ord = ["em_especie_cuota_fija", "em_especie_excedente", "em_dinero",
                     "iv", "gps", "rt_default"]
    else:
        title = f"Cedula bimestral {cedula['period_year']} B{cedula['bimester']}"
        ramos_ord = ["retiro", "cv", "infonavit_aportacion"]

    ws.title = title[:31]

    brand = "33B2F5"
    header_fill = PatternFill("solid", fgColor=brand)
    section_fill = PatternFill("solid", fgColor="F0F5FA")
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    right = Alignment(horizontal="right")

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14, color=brand)
    ws["A2"] = (f"Periodo: {cedula['start_date']} a {cedula['end_date']} · "
                f"Pago límite (día 17 mes sig.): {cedula['payment_deadline']}")
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    row0 = 4
    # Headers: Empleado | NSS | Días | SBC | (ramos con 2 subcols cada uno) | Total pat | Total obr | Total
    ws.cell(row=row0, column=1, value="Empleado").fill = header_fill
    ws.cell(row=row0, column=1).font = header_font
    ws.cell(row=row0, column=2, value="NSS").fill = header_fill
    ws.cell(row=row0, column=2).font = header_font
    ws.cell(row=row0, column=3, value="Días").fill = header_fill
    ws.cell(row=row0, column=3).font = header_font
    ws.cell(row=row0, column=4, value="SBC topado").fill = header_fill
    ws.cell(row=row0, column=4).font = header_font

    col = 5
    ramo_col_start = {}
    for r_key in ramos_ord:
        ramo_col_start[r_key] = col
        label = RAMO_LABEL[r_key]
        for sub in ("Patrón", "Obrero"):
            c = ws.cell(row=row0, column=col, value=f"{label} · {sub}")
            c.fill = header_fill; c.font = header_font
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            col += 1
    for sub in ("Total patrón", "Total obrero", "Total"):
        c = ws.cell(row=row0, column=col, value=sub)
        c.fill = header_fill; c.font = header_font
        col += 1
    if kind == "bimestral":
        c = ws.cell(row=row0, column=col, value="INFONAVIT descuento")
        c.fill = header_fill; c.font = header_font
        col += 1
        c = ws.cell(row=row0, column=col, value="Total a pagar")
        c.fill = header_fill; c.font = header_font
        col += 1

    r_idx = row0 + 1
    for r in cedula["rows"]:
        emp_label = f"{r['employee_number']} · {r['employee_name']}"
        ws.cell(row=r_idx, column=1, value=emp_label)
        ws.cell(row=r_idx, column=2, value=r.get("nss") or "")
        ws.cell(row=r_idx, column=3, value=r["days_cotizados"]).alignment = right
        ws.cell(row=r_idx, column=4, value=r["sbc_topado"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=4).alignment = right
        for r_key in ramos_ord:
            v = r["ramos"].get(r_key, {})
            c = ramo_col_start[r_key]
            ws.cell(row=r_idx, column=c, value=v.get("patron", 0)).number_format = "#,##0.00"
            ws.cell(row=r_idx, column=c).alignment = right
            ws.cell(row=r_idx, column=c + 1, value=v.get("obrero", 0)).number_format = "#,##0.00"
            ws.cell(row=r_idx, column=c + 1).alignment = right
        col_t = ramo_col_start[ramos_ord[-1]] + 2
        ws.cell(row=r_idx, column=col_t, value=r["total_patron"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col_t).alignment = right; ws.cell(row=r_idx, column=col_t).font = bold
        ws.cell(row=r_idx, column=col_t + 1, value=r["total_obrero"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col_t + 1).alignment = right; ws.cell(row=r_idx, column=col_t + 1).font = bold
        total_r = r.get("total_pagar_imss") or r.get("total_pagar") or 0
        ws.cell(row=r_idx, column=col_t + 2, value=total_r).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col_t + 2).alignment = right; ws.cell(row=r_idx, column=col_t + 2).font = bold
        if kind == "bimestral":
            ws.cell(row=r_idx, column=col_t + 3,
                    value=r["infonavit_descuento"]).number_format = "#,##0.00"
            ws.cell(row=r_idx, column=col_t + 3).alignment = right
            grand = round(r["total_patron"] + r["total_obrero"] + r["infonavit_descuento"], 2)
            ws.cell(row=r_idx, column=col_t + 4, value=grand).number_format = "#,##0.00"
            ws.cell(row=r_idx, column=col_t + 4).alignment = right
            ws.cell(row=r_idx, column=col_t + 4).font = bold
        r_idx += 1

    # Totales
    r_idx += 1
    t = cedula["totals"]
    for c in range(1, col):
        ws.cell(row=r_idx, column=c).fill = section_fill
    ws.cell(row=r_idx, column=1, value="TOTALES").font = Font(bold=True, size=12)
    ws.cell(row=r_idx, column=3, value=t["total_days"]).number_format = "#,##0.00"
    for r_key in ramos_ord:
        v = t["por_ramo"][r_key]
        c = ramo_col_start[r_key]
        ws.cell(row=r_idx, column=c, value=v["patron"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=c).font = bold
        ws.cell(row=r_idx, column=c + 1, value=v["obrero"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=c + 1).font = bold
    col_t = ramo_col_start[ramos_ord[-1]] + 2
    ws.cell(row=r_idx, column=col_t, value=t["total_patron"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=col_t).font = bold
    ws.cell(row=r_idx, column=col_t + 1, value=t["total_obrero"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=col_t + 1).font = bold
    ws.cell(row=r_idx, column=col_t + 2, value=t["total_pagar_imss"]).number_format = "#,##0.00"
    ws.cell(row=r_idx, column=col_t + 2).font = bold
    if kind == "bimestral":
        ws.cell(row=r_idx, column=col_t + 3,
                value=t.get("total_infonavit_descuento", 0)).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col_t + 3).font = bold
        ws.cell(row=r_idx, column=col_t + 4,
                value=t.get("total_pagar_infonavit", 0)).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col_t + 4).font = bold

    # Anchos
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    for i in range(5, col):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = 13
    ws.row_dimensions[row0].height = 42
    ws.freeze_panes = "E5"

    buf = BytesIO(); wb.save(buf); return buf.getvalue()


# ─── PDF cédula oficial ─────────────────────────────────────────────────


def build_pdf(company: dict, cedula: dict, kind: str) -> bytes:
    """Cédula oficial en LETTER landscape."""
    from reportlab.lib.pagesizes import LETTER, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.units import mm

    if kind == "mensual":
        title = (f"Cédula IMSS mensual · {cedula['period_year']}-"
                 f"{cedula['period_month']:02d}")
        ramos_ord = ["em_especie_cuota_fija", "em_especie_excedente", "em_dinero",
                     "iv", "gps", "rt_default"]
    else:
        title = f"Cédula bimestral · Ejercicio {cedula['period_year']} · Bimestre {cedula['bimester']}"
        ramos_ord = ["retiro", "cv", "infonavit_aportacion"]

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(LETTER),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
    )
    styles = getSampleStyleSheet()
    body = styles["BodyText"]; body.fontSize = 8
    small = ParagraphStyle("small", parent=body, fontSize=7, textColor=colors.grey)
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=12,
                        textColor=colors.HexColor("#0F172A"), spaceAfter=2)

    story = []
    company_name = (company or {}).get("legal_name") or "Empresa"
    company_rfc = (company or {}).get("rfc") or ""

    story.append(Paragraph(f"<b>{company_name}</b> — RFC {company_rfc}", body))
    story.append(Paragraph(title, h1))
    story.append(Paragraph(
        f"Periodo: {cedula['start_date']} a {cedula['end_date']} · "
        f"Pago límite: {cedula['payment_deadline']}",
        small,
    ))
    story.append(Spacer(1, 4))

    # Encabezado tabla
    header = ["Empleado", "NSS", "Días", "SBC"]
    for k in ramos_ord:
        header.append(RAMO_LABEL[k])
    header += ["Total P.", "Total O.", "Total"]
    if kind == "bimestral":
        header += ["Amort. INFONAVIT", "A pagar"]

    data = [header]
    for r in cedula["rows"]:
        row = [
            Paragraph(f"<b>{r['employee_number']}</b><br/>{r['employee_name']}", small),
            r.get("nss") or "",
            f"{r['days_cotizados']:.0f}",
            f"${r['sbc_topado']:,.2f}",
        ]
        for k in ramos_ord:
            v = r["ramos"].get(k, {})
            row.append(f"${v.get('patron', 0):,.0f} / ${v.get('obrero', 0):,.0f}")
        row.append(f"${r['total_patron']:,.2f}")
        row.append(f"${r['total_obrero']:,.2f}")
        row.append(f"${(r.get('total_pagar_imss') or r.get('total_pagar')):,.2f}")
        if kind == "bimestral":
            row.append(f"${r['infonavit_descuento']:,.2f}")
            grand = r['total_patron'] + r['total_obrero'] + r['infonavit_descuento']
            row.append(f"${grand:,.2f}")
        data.append(row)

    # Totales
    t = cedula["totals"]
    trow = ["TOTALES", "", f"{t['total_days']:.0f}", ""]
    for k in ramos_ord:
        v = t["por_ramo"][k]
        trow.append(f"${v['patron']:,.0f} / ${v['obrero']:,.0f}")
    trow.append(f"${t['total_patron']:,.2f}")
    trow.append(f"${t['total_obrero']:,.2f}")
    trow.append(f"${t['total_pagar_imss']:,.2f}")
    if kind == "bimestral":
        trow.append(f"${t.get('total_infonavit_descuento', 0):,.2f}")
        trow.append(f"${t.get('total_pagar_infonavit', 0):,.2f}")
    data.append(trow)

    n_ramos = len(ramos_ord)
    # Ancho de columnas
    fixed = [35 * mm, 16 * mm, 10 * mm, 16 * mm]
    ramo_w = 18 * mm
    tail = [16 * mm, 16 * mm, 20 * mm]
    if kind == "bimestral":
        tail += [20 * mm, 20 * mm]
    col_widths = fixed + [ramo_w] * n_ramos + tail

    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F172A")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("BOX", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("INNERGRID", (0, 0), (-1, -1), 0.2, colors.HexColor("#E2E8F0")),
        ("PADDING", (0, 0), (-1, -1), 2),
        # Fila de totales
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#F1F5F9")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
    ]))
    story.append(tbl); story.append(Spacer(1, 12))

    # Nota fiscal + firmas
    story.append(Paragraph(
        "Cédula de determinación de cuotas obrero-patronales — apoyo de cálculo. "
        "El pago oficial debe realizarse desde el Sistema Único de Autodeterminación "
        "(SUA) del IMSS con las líneas de captura emitidas al valorar la propuesta.",
        small,
    ))
    story.append(Spacer(1, 16))
    firms = Table([
        ["_____________________________", "", "_____________________________"],
        ["Contador Público responsable", "", "Representante legal del patrón"],
    ], colWidths=[80 * mm, 30 * mm, 80 * mm])
    firms.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
    ]))
    story.append(firms)

    doc.build(story)
    return buf.getvalue()
