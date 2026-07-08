"""Carga a granel de bonos/vales/ahorro/préstamos/notas en una nómina calculada.

Flujo típico:
  1. Operador descarga la plantilla del período (XLSX o CSV). Ya viene
     pre-llenada con los empleados del período y sus valores actuales.
  2. La llena en Excel (mucho más rápido que fila por fila en la UI).
  3. La sube y el sistema aplica los cambios mediante el mismo servicio de
     edición manual (update_payroll_detail), que recalcula ISR, ISN y neto.

Diseño:
  - El match es por `no_empleado` primero y por `rfc` como fallback. Si
    ninguno coincide, la fila se reporta como error (no se aplica).
  - Valores vacíos NO borran el valor previo — para poner en cero hay que
    escribir explícitamente `0`. Así el operador puede editar solo algunas
    columnas sin tocar las demás.
  - Se acumula un reporte de "aplicados / omitidos / errores por fila" para
    que la UI lo muestre después de la subida.
"""

from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from typing import Iterable, List, Optional


BULK_HEADERS = [
    "no_empleado",     # match primario (EMP-001)
    "rfc",             # match secundario (opcional)
    "nombre",          # solo referencia — no se usa en el match
    "bono",            # percepción gravable
    "vales",           # vales de despensa (gravable)
    "ahorro",          # fondo de ahorro (exento LISR)
    "prestamo",        # deducción
    "notas",           # justificación (queda en el recibo PDF)
]


@dataclass
class BulkImportError:
    row: int
    reason: str


@dataclass
class BulkImportSummary:
    applied: int = 0
    skipped: int = 0
    errors: List[BulkImportError] = None

    def __post_init__(self):
        if self.errors is None:
            self.errors = []


def _to_float(v) -> Optional[float]:
    """Convierte un valor de celda a float, devolviendo None si viene vacío
    (para distinguir 'no lo tocaste' de 'lo pusiste en cero')."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s == "" or s == "-" or s == "—":
        return None
    # tolera separadores mexicanos: $1,234.56
    s = s.replace("$", "").replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        raise ValueError(f"'{v}' no es un número válido")


def parse_bulk_file(content: bytes, filename: str) -> tuple[List[dict], List[BulkImportError]]:
    """Parsea XLSX/CSV y devuelve (filas parseadas, errores de parseo)."""
    name = (filename or "").lower()
    is_xlsx = name.endswith(".xlsx") or name.endswith(".xlsm")

    rows_raw: List[dict] = []
    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        ws = wb["Detalle"] if "Detalle" in wb.sheetnames else wb.active
        header = None
        for row in ws.iter_rows(values_only=True):
            if header is None:
                header = [(str(c).strip() if c is not None else "") for c in row]
                continue
            rows_raw.append(dict(zip(header, row)))
    else:
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        for r in csv.DictReader(io.StringIO(text)):
            rows_raw.append(dict(r))

    parsed: List[dict] = []
    errors: List[BulkImportError] = []
    for i, raw in enumerate(rows_raw, start=2):  # excel row 2 = primera de datos
        no_emp = str(raw.get("no_empleado") or "").strip()
        rfc = str(raw.get("rfc") or "").strip().upper()
        if not no_emp and not rfc:
            # Fila vacía — solo la ignoramos, no cuenta como error
            continue
        try:
            parsed.append({
                "row": i,
                "no_empleado": no_emp,
                "rfc": rfc,
                "bono": _to_float(raw.get("bono")),
                "vales": _to_float(raw.get("vales")),
                "ahorro": _to_float(raw.get("ahorro")),
                "prestamo": _to_float(raw.get("prestamo")),
                "notas": str(raw.get("notas") or "").strip() or None,
            })
        except ValueError as e:
            errors.append(BulkImportError(row=i, reason=str(e)))
    return parsed, errors


def build_template_xlsx(period_name: str, rows: Iterable[dict]) -> bytes:
    """Genera plantilla XLSX profesional pre-llenada con los empleados
    actuales del período. `rows` son diccionarios con
    no_empleado / rfc / nombre / bono / vales / ahorro / prestamo / notas."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Detalle"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    center = Alignment(horizontal="center")
    grey = PatternFill("solid", fgColor="F1F5F9")
    italic = Font(italic=True, color="64748B")
    money_fmt = "$#,##0.00"

    # Headers
    ws.append(BULK_HEADERS)
    for col_idx in range(1, len(BULK_HEADERS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    widths = [14, 16, 32, 14, 14, 14, 14, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Filas pre-llenadas (col nombre es solo referencia; el sistema la ignora)
    for r in rows:
        ws.append([
            r.get("no_empleado", ""),
            r.get("rfc", ""),
            r.get("nombre", ""),
            r.get("bono", 0) or 0,
            r.get("vales", 0) or 0,
            r.get("ahorro", 0) or 0,
            r.get("prestamo", 0) or 0,
            r.get("notas", "") or "",
        ])

    # La columna "nombre" es solo referencia — pintarla en gris/itálica
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=3).fill = grey
        ws.cell(row=row_idx, column=3).font = italic
        # Formato dinero en las 4 columnas numéricas
        for col in (4, 5, 6, 7):
            ws.cell(row=row_idx, column=col).number_format = money_fmt

    ws.freeze_panes = "A2"

    # Hoja de instrucciones al inicio
    ws_i = wb.create_sheet("Instrucciones", 0)
    lines = [
        f"Carga a granel — {period_name}",
        "",
        "Cómo llenar:",
        "  1) En la hoja 'Detalle' están todos los empleados del período pre-llenados.",
        "  2) Ajusta las columnas 'bono', 'vales', 'ahorro', 'prestamo' y/o 'notas'.",
        "  3) Puedes cambiar solo las columnas que necesites — dejar una en blanco NO borra el valor previo.",
        "     Para poner algo en cero, escribe explícitamente 0.",
        "  4) La columna 'nombre' es solo referencia (el sistema la ignora).",
        "  5) El match del empleado es por 'no_empleado' (primario) y 'rfc' (fallback).",
        "  6) Vuelve a la app → botón 'Cargar bonos/vales' → sube este archivo.",
        "",
        "El sistema recalcula ISR, ISN y neto automáticamente por cada empleado modificado.",
        "Solo funciona cuando la nómina está en estado 'Calculada'; una vez aprobada no acepta cambios.",
    ]
    for row_ix, line in enumerate(lines, start=1):
        c = ws_i.cell(row=row_ix, column=1, value=line)
        if row_ix == 1:
            c.font = Font(bold=True, size=14, color="1E3A8A")
    ws_i.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_csv(rows: Iterable[dict]) -> bytes:
    """CSV con BOM UTF-8 para compatibilidad con Excel."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(BULK_HEADERS)
    for r in rows:
        w.writerow([
            r.get("no_empleado", ""),
            r.get("rfc", ""),
            r.get("nombre", ""),
            r.get("bono", 0) or 0,
            r.get("vales", 0) or 0,
            r.get("ahorro", 0) or 0,
            r.get("prestamo", 0) or 0,
            r.get("notas", "") or "",
        ])
    return buf.getvalue().encode("utf-8-sig")
