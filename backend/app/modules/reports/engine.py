"""Motor de ejecución de reportes. Recibe una configuración validada y
construye una query SQLAlchemy segura sobre el dataset elegido.

Config schema:
{
  "dataset": "accounting_movements",
  "columns": ["account_code", "account_name", "debit", "credit"],   # cols a mostrar
  "filters": [
    {"field": "account_type", "op": "eq", "value": "gasto"},
    {"field": "date", "op": "between", "value": ["2026-01-01", "2026-12-31"]}
  ],
  "group_by": ["account_code", "account_name"],                     # opcional
  "aggregations": {"debit": "sum", "credit": "sum", "line_id": "count"},
  "sort": [{"field": "debit", "direction": "desc"}],                # aliases
  "limit": 5000                                                     # cap defensivo
}
"""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import and_, asc, desc, func, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.reports import datasets as ds_mod


ALLOWED_OPS = {"eq", "ne", "gt", "gte", "lt", "lte", "in", "not_in",
               "like", "ilike", "between", "is_null", "not_null"}

DEFAULT_LIMIT = 5000
MAX_LIMIT = 20000


def _agg_func(fn: str, col):
    fn = fn.lower()
    if fn == "sum":
        return func.coalesce(func.sum(col), 0.0)
    if fn == "count":
        return func.count(col)
    if fn == "avg":
        return func.coalesce(func.avg(col), 0.0)
    if fn == "min":
        return func.min(col)
    if fn == "max":
        return func.max(col)
    raise ValueError(f"Agregación no soportada: {fn}")


def _apply_op(col, op: str, value):
    if op == "eq":
        return col == value
    if op == "ne":
        return col != value
    if op == "gt":
        return col > value
    if op == "gte":
        return col >= value
    if op == "lt":
        return col < value
    if op == "lte":
        return col <= value
    if op == "in":
        return col.in_(value if isinstance(value, list) else [value])
    if op == "not_in":
        return ~col.in_(value if isinstance(value, list) else [value])
    if op == "like":
        return col.like(f"%{value}%")
    if op == "ilike":
        return col.ilike(f"%{value}%")
    if op == "between":
        if not isinstance(value, list) or len(value) != 2:
            raise ValueError("between requiere [min, max]")
        return col.between(value[0], value[1])
    if op == "is_null":
        return col.is_(None)
    if op == "not_null":
        return col.is_not(None)
    raise ValueError(f"Operador no soportado: {op}")


def _validate_config(config: dict) -> Tuple[ds_mod.Dataset, dict]:
    ds_key = config.get("dataset")
    ds = ds_mod.get(ds_key)
    if ds is None:
        raise ValueError(f"Dataset desconocido: {ds_key}")

    columns = config.get("columns") or []
    for c in columns:
        if c not in ds.fields:
            raise ValueError(f"Columna no permitida en dataset '{ds_key}': {c}")

    group_by = config.get("group_by") or []
    for g in group_by:
        if g not in ds.groupable:
            raise ValueError(f"Grouping no permitido: {g}")

    aggregations = config.get("aggregations") or {}
    for alias, agg in aggregations.items():
        f = ds.field(alias)
        if f is None:
            raise ValueError(f"Campo de agregación desconocido: {alias}")
        if agg not in f.allowed_aggs:
            raise ValueError(f"Agregación '{agg}' no permitida en '{alias}'")

    filters = config.get("filters") or []
    for flt in filters:
        if not isinstance(flt, dict):
            raise ValueError("Filtro inválido")
        if flt.get("field") not in ds.fields:
            raise ValueError(f"Campo de filtro desconocido: {flt.get('field')}")
        if flt.get("op") not in ALLOWED_OPS:
            raise ValueError(f"Operador de filtro no permitido: {flt.get('op')}")

    sort = config.get("sort") or []
    for s in sort:
        if not isinstance(s, dict):
            raise ValueError("Sort inválido")
        field = s.get("field")
        if field not in ds.fields and field not in aggregations:
            raise ValueError(f"Sort por campo no permitido: {field}")

    limit = min(int(config.get("limit") or DEFAULT_LIMIT), MAX_LIMIT)

    validated = {
        "dataset": ds_key,
        "columns": columns,
        "filters": filters,
        "group_by": group_by,
        "aggregations": aggregations,
        "sort": sort,
        "limit": limit,
    }
    return ds, validated


def _build_select(ds: ds_mod.Dataset, config: dict):
    """Construye la sentencia SQLAlchemy respetando group_by / aggregations."""
    columns = config["columns"] or []
    group_by = config["group_by"] or []
    aggregations = config["aggregations"] or {}
    filters = config["filters"] or []
    sort = config["sort"] or []
    limit = config["limit"]

    # Bloque SELECT
    select_cols: List[Any] = []
    labels_used: set = set()

    if group_by:
        for g in group_by:
            f = ds.fields[g]
            select_cols.append(f.sqla_col.label(g))
            labels_used.add(g)
        for alias, agg in aggregations.items():
            f = ds.fields[alias]
            select_cols.append(_agg_func(agg, f.sqla_col).label(f"{alias}_{agg}"))
            labels_used.add(f"{alias}_{agg}")
    else:
        # Sin group_by: si hay agregaciones, es un grand total sin filas por línea
        if aggregations and not columns:
            for alias, agg in aggregations.items():
                f = ds.fields[alias]
                select_cols.append(_agg_func(agg, f.sqla_col).label(f"{alias}_{agg}"))
                labels_used.add(f"{alias}_{agg}")
        else:
            for c in columns:
                f = ds.fields[c]
                select_cols.append(f.sqla_col.label(c))
                labels_used.add(c)

    if not select_cols:
        # fallback: primera columna del dataset
        first = next(iter(ds.fields.values()))
        select_cols.append(first.sqla_col.label(first.alias))
        labels_used.add(first.alias)

    # Base source builder ya trae joins + where posted-only para movimientos
    base = ds.source_builder()

    # Sustituimos las columnas seleccionadas
    stmt = base.with_only_columns(*select_cols)

    # Filtros
    for flt in filters:
        f = ds.fields[flt["field"]]
        stmt = stmt.where(_apply_op(f.sqla_col, flt["op"], flt.get("value")))

    # Grouping
    if group_by:
        for g in group_by:
            stmt = stmt.group_by(ds.fields[g].sqla_col)

    # Sort
    for s in sort:
        field = s["field"]
        direction = (s.get("direction") or "asc").lower()
        # Puede ser un alias de campo directo o alias_agg de agregación
        if field in ds.fields and (not aggregations or field in (group_by or [])):
            col = ds.fields[field].sqla_col
        elif field in aggregations:
            f = ds.fields[field]
            col = _agg_func(aggregations[field], f.sqla_col)
        elif "_" in field and field.rsplit("_", 1)[0] in ds.fields:
            # alias tipo "debit_sum" pasado directamente
            alias, agg = field.rsplit("_", 1)
            f = ds.fields[alias]
            col = _agg_func(agg, f.sqla_col)
        else:
            # sort en columna literal seleccionada
            col = ds.fields[field].sqla_col
        stmt = stmt.order_by(desc(col) if direction == "desc" else asc(col))

    stmt = stmt.limit(limit)
    return stmt, list(labels_used)


async def run_report(db: AsyncSession, config: dict) -> dict:
    """Ejecuta un reporte y devuelve rows + metadatos."""
    ds, validated = _validate_config(config)
    stmt, labels = _build_select(ds, validated)
    result = (await db.execute(stmt)).all()

    # Serializar
    columns_meta = []
    for label in labels:
        if label in ds.fields:
            f = ds.fields[label]
            columns_meta.append({"alias": f.alias, "label": f.label_es, "type": f.type})
        elif "_" in label:
            alias, agg = label.rsplit("_", 1)
            f = ds.fields.get(alias)
            if f:
                columns_meta.append({
                    "alias": label,
                    "label": f"{f.label_es} ({agg})",
                    "type": "money" if f.type == "money" else "number",
                })

    rows = []
    for r in result:
        row_dict = {}
        for i, label in enumerate(labels):
            val = r[i]
            if isinstance(val, (date, datetime)):
                val = val.isoformat()
            elif hasattr(val, "__float__") and not isinstance(val, (int, bool)):
                try:
                    val = float(val)
                except Exception:
                    pass
            row_dict[label] = val
        rows.append(row_dict)

    return {
        "dataset": ds.key,
        "columns": columns_meta,
        "rows": rows,
        "row_count": len(rows),
        "truncated": len(rows) >= validated["limit"],
    }


def build_xlsx(result: dict, report_name: str = "Reporte") -> bytes:
    """Exporta el resultado a XLSX profesional."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = report_name[:31] or "Reporte"

    brand = "33B2F5"
    header_fill = PatternFill("solid", fgColor=brand)
    white_bold = Font(bold=True, color="FFFFFF")

    ws["A1"] = report_name
    ws["A1"].font = Font(bold=True, size=14, color=brand)

    columns = result["columns"]
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=3, column=i, value=col["label"])
        c.fill = header_fill; c.font = white_bold
        c.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(result["rows"], start=4):
        for c_idx, col in enumerate(columns, start=1):
            v = row.get(col["alias"])
            c = ws.cell(row=r_idx, column=c_idx, value=v)
            if col["type"] in ("money",):
                c.number_format = '"$"#,##0.00'
                c.alignment = Alignment(horizontal="right")
            elif col["type"] == "number":
                c.number_format = "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            elif col["type"] == "int":
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")

    for i, col in enumerate(columns, start=1):
        letter = chr(ord("A") + i - 1) if i <= 26 else "A" + chr(ord("A") + i - 27)
        ws.column_dimensions[letter].width = max(len(col["label"]) + 2, 14)

    buf = BytesIO(); wb.save(buf); return buf.getvalue()
