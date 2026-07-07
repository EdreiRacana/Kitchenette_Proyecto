"""Business logic para el módulo Forecast.

Diseño:
- CRUD atómico para planes y líneas.
- baseline(): lee historial real (orders + order_items) del cliente/producto,
  agrupa las unidades por mes calendario para respetar estacionalidad, aplica
  el % de crecimiento del plan y crea una línea por combinación
  cliente/producto/vendedor detectada.
- rollup(): concentra las líneas del plan por cada dimensión.
- attainment(): compara la meta mensual del plan contra las ventas reales
  del año (calendario) del backend de sales.
"""

from __future__ import annotations

import csv
import io
from datetime import date
from typing import Dict, List, Optional, Tuple

from sqlalchemy import and_, delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.modules.customers import models as cust_models
from app.modules.inventory import models as inv_models
from app.modules.sales import models as sales_models
from app.modules.auth import models as auth_models

from . import models, schemas


MONTH_COLS = ("m1", "m2", "m3", "m4", "m5", "m6", "m7", "m8", "m9", "m10", "m11", "m12")


def _line_totals(line: models.ForecastLine) -> Tuple[int, float]:
    units = sum(int(getattr(line, c) or 0) for c in MONTH_COLS)
    amount = round(units * float(line.unit_price or 0.0), 2)
    return units, amount


def _line_to_schema(line: models.ForecastLine) -> schemas.ForecastLineInDB:
    units, amount = _line_totals(line)
    data = {c: int(getattr(line, c) or 0) for c in MONTH_COLS}
    return schemas.ForecastLineInDB(
        id=line.id,
        plan_id=line.plan_id,
        customer_id=line.customer_id,
        variant_id=line.variant_id,
        salesperson_id=line.salesperson_id,
        product_name=line.product_name,
        sku=line.sku,
        customer_name=line.customer_name,
        salesperson_name=line.salesperson_name,
        unit_price=float(line.unit_price or 0.0),
        total_units=units,
        total_amount=amount,
        **data,
    )


# ── Planes ───────────────────────────────────────────────────────────────────

async def list_plans(db: AsyncSession) -> List[models.ForecastPlan]:
    stmt = select(models.ForecastPlan).order_by(
        models.ForecastPlan.year.desc(), models.ForecastPlan.id.desc()
    )
    res = await db.execute(stmt)
    return list(res.scalars().all())


async def get_plan(db: AsyncSession, plan_id: int) -> Optional[models.ForecastPlan]:
    return await db.get(models.ForecastPlan, plan_id)


async def create_plan(db: AsyncSession, data: schemas.ForecastPlanCreate) -> models.ForecastPlan:
    plan = models.ForecastPlan(**data.model_dump(exclude_unset=True))
    db.add(plan)
    await db.commit()
    await db.refresh(plan)
    return plan


async def update_plan(
    db: AsyncSession, plan_id: int, data: schemas.ForecastPlanUpdate
) -> Optional[models.ForecastPlan]:
    plan = await db.get(models.ForecastPlan, plan_id)
    if plan is None:
        return None
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(plan, k, v)
    await db.commit()
    await db.refresh(plan)
    return plan


async def delete_plan(db: AsyncSession, plan_id: int) -> bool:
    plan = await db.get(models.ForecastPlan, plan_id)
    if plan is None:
        return False
    await db.delete(plan)
    await db.commit()
    return True


# ── Líneas ───────────────────────────────────────────────────────────────────

async def list_lines(db: AsyncSession, plan_id: int) -> List[schemas.ForecastLineInDB]:
    stmt = (
        select(models.ForecastLine)
        .where(models.ForecastLine.plan_id == plan_id)
        .order_by(models.ForecastLine.id.asc())
    )
    res = await db.execute(stmt)
    return [_line_to_schema(r) for r in res.scalars().all()]


async def _fill_snapshots(db: AsyncSession, line: models.ForecastLine) -> None:
    """Llena los nombres a partir de las FKs cuando el cliente no los mandó."""
    if line.customer_id and not line.customer_name:
        cust = await db.get(cust_models.Customer, line.customer_id)
        if cust:
            line.customer_name = cust.name
    if line.variant_id and (not line.product_name or not line.sku):
        variant = await db.get(inv_models.ProductVariant, line.variant_id)
        if variant:
            if not line.sku:
                line.sku = variant.sku
            if not line.product_name and variant.product_id:
                prod = await db.get(inv_models.Product, variant.product_id)
                if prod:
                    line.product_name = prod.name
            if not line.unit_price and variant.price:
                line.unit_price = float(variant.price)
    if line.salesperson_id and not line.salesperson_name:
        user = await db.get(auth_models.User, line.salesperson_id)
        if user:
            line.salesperson_name = user.full_name or user.email


async def create_line(
    db: AsyncSession, plan_id: int, data: schemas.ForecastLineCreate
) -> schemas.ForecastLineInDB:
    line = models.ForecastLine(plan_id=plan_id, **data.model_dump(exclude_unset=True))
    await _fill_snapshots(db, line)
    db.add(line)
    await db.commit()
    await db.refresh(line)
    return _line_to_schema(line)


async def update_line(
    db: AsyncSession, line_id: int, data: schemas.ForecastLineUpdate
) -> Optional[schemas.ForecastLineInDB]:
    line = await db.get(models.ForecastLine, line_id)
    if line is None:
        return None
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(line, k, v)
    await _fill_snapshots(db, line)
    await db.commit()
    await db.refresh(line)
    return _line_to_schema(line)


async def delete_line(db: AsyncSession, line_id: int) -> bool:
    line = await db.get(models.ForecastLine, line_id)
    if line is None:
        return False
    await db.delete(line)
    await db.commit()
    return True


# ── Baseline ────────────────────────────────────────────────────────────────

async def build_baseline(
    db: AsyncSession, req: schemas.BaselineRequest
) -> schemas.BaselineResponse:
    """Genera líneas del plan a partir del historial real del año anterior."""
    plan = await db.get(models.ForecastPlan, req.plan_id)
    if plan is None:
        raise ValueError("Plan no encontrado")

    year_source = req.year_source if req.year_source is not None else (plan.year - 1)
    growth_pct = req.growth_pct if req.growth_pct is not None else float(plan.growth_pct or 0.0)
    factor = 1.0 + (growth_pct / 100.0)

    lines_deleted = 0
    if req.replace:
        # Recuento antes de borrar (para el reporte de la respuesta)
        cnt = await db.execute(
            select(func.count(models.ForecastLine.id)).where(
                models.ForecastLine.plan_id == plan.id
            )
        )
        lines_deleted = int(cnt.scalar() or 0)
        await db.execute(
            delete(models.ForecastLine).where(models.ForecastLine.plan_id == plan.id)
        )

    O = sales_models.Order
    OI = sales_models.OrderItem

    conds = [
        O.kind == "order",
        O.status != "cancelled",
        func.extract("year", O.created_at) == year_source,
    ]
    if req.customer_id is not None:
        conds.append(O.customer_id == req.customer_id)
    if req.salesperson_id is not None:
        conds.append(O.user_id == req.salesperson_id)

    month_expr = func.extract("month", O.created_at).label("month")

    stmt = (
        select(
            O.customer_id.label("customer_id"),
            OI.variant_id.label("variant_id"),
            O.user_id.label("salesperson_id"),
            func.coalesce(OI.product_name, "—").label("product_name"),
            func.coalesce(OI.sku, "").label("sku"),
            month_expr,
            func.coalesce(func.sum(OI.quantity), 0).label("units"),
            func.coalesce(func.avg(OI.unit_price), 0.0).label("unit_price"),
        )
        .join(O, OI.order_id == O.id)
        .where(and_(*conds))
        .group_by(
            O.customer_id,
            OI.variant_id,
            O.user_id,
            OI.product_name,
            OI.sku,
            month_expr,
        )
    )

    rows = (await db.execute(stmt)).all()

    # Agrupación por (customer, variant, salesperson, product_name, sku) para
    # construir una línea con las 12 columnas de mes.
    grouped: Dict[tuple, Dict] = {}
    for r in rows:
        key = (
            r.customer_id,
            r.variant_id,
            r.salesperson_id,
            r.product_name,
            r.sku,
        )
        g = grouped.setdefault(
            key,
            {
                "customer_id": r.customer_id,
                "variant_id": r.variant_id,
                "salesperson_id": r.salesperson_id,
                "product_name": r.product_name or "—",
                "sku": r.sku or "",
                "unit_price": float(r.unit_price or 0.0),
                "months": [0] * 12,
            },
        )
        m = int(r.month or 0)
        if 1 <= m <= 12:
            g["months"][m - 1] += int(r.units or 0)

    # Snapshots de nombres — cache local para no hacer N queries.
    cust_ids = {g["customer_id"] for g in grouped.values() if g["customer_id"]}
    user_ids = {g["salesperson_id"] for g in grouped.values() if g["salesperson_id"]}

    cust_name_by_id: Dict[int, str] = {}
    if cust_ids:
        cres = await db.execute(
            select(cust_models.Customer.id, cust_models.Customer.name).where(
                cust_models.Customer.id.in_(cust_ids)
            )
        )
        cust_name_by_id = {row.id: row.name for row in cres}

    user_name_by_id: Dict[int, str] = {}
    if user_ids:
        ures = await db.execute(
            select(
                auth_models.User.id,
                auth_models.User.full_name,
                auth_models.User.email,
            ).where(auth_models.User.id.in_(user_ids))
        )
        user_name_by_id = {row.id: (row.full_name or row.email) for row in ures}

    created_lines: List[models.ForecastLine] = []
    for g in grouped.values():
        months_grown = [max(0, int(round(u * factor))) for u in g["months"]]
        line = models.ForecastLine(
            plan_id=plan.id,
            customer_id=g["customer_id"],
            variant_id=g["variant_id"],
            salesperson_id=g["salesperson_id"],
            product_name=g["product_name"],
            sku=g["sku"] or None,
            customer_name=cust_name_by_id.get(g["customer_id"]) if g["customer_id"] else None,
            salesperson_name=user_name_by_id.get(g["salesperson_id"]) if g["salesperson_id"] else None,
            unit_price=g["unit_price"],
        )
        for i, col in enumerate(MONTH_COLS):
            setattr(line, col, months_grown[i])
        db.add(line)
        created_lines.append(line)

    await db.commit()
    for line in created_lines:
        await db.refresh(line)

    return schemas.BaselineResponse(
        plan_id=plan.id,
        year_source=year_source,
        growth_pct=growth_pct,
        lines_created=len(created_lines),
        lines_deleted=lines_deleted,
        lines=[_line_to_schema(l) for l in created_lines],
    )


# ── Rollup ───────────────────────────────────────────────────────────────────

async def rollup(db: AsyncSession, plan_id: int) -> schemas.RollupResponse:
    stmt = (
        select(models.ForecastLine)
        .where(models.ForecastLine.plan_id == plan_id)
        .order_by(models.ForecastLine.id.asc())
    )
    lines = list((await db.execute(stmt)).scalars().all())

    by_customer: Dict[str, Dict] = {}
    by_product: Dict[str, Dict] = {}
    by_salesperson: Dict[str, Dict] = {}
    monthly_amount = [0.0] * 12
    monthly_units = [0] * 12
    total_units = 0
    total_amount = 0.0

    for l in lines:
        units, amount = _line_totals(l)
        total_units += units
        total_amount += amount
        for i, col in enumerate(MONTH_COLS):
            u = int(getattr(l, col) or 0)
            monthly_units[i] += u
            monthly_amount[i] += u * float(l.unit_price or 0.0)

        cust_key = f"cust:{l.customer_id}" if l.customer_id else f"custname:{l.customer_name or 'Sin cliente'}"
        cust_label = l.customer_name or "Sin cliente"
        by_customer.setdefault(cust_key, {"label": cust_label, "units": 0, "amount": 0.0})
        by_customer[cust_key]["units"] += units
        by_customer[cust_key]["amount"] += amount

        prod_key = f"var:{l.variant_id}" if l.variant_id else f"text:{l.product_name or '—'}"
        prod_label = l.product_name or "—"
        if l.sku:
            prod_label = f"{prod_label} · {l.sku}"
        by_product.setdefault(prod_key, {"label": prod_label, "units": 0, "amount": 0.0})
        by_product[prod_key]["units"] += units
        by_product[prod_key]["amount"] += amount

        sp_key = f"user:{l.salesperson_id}" if l.salesperson_id else f"spname:{l.salesperson_name or 'Sin vendedor'}"
        sp_label = l.salesperson_name or "Sin vendedor"
        by_salesperson.setdefault(sp_key, {"label": sp_label, "units": 0, "amount": 0.0})
        by_salesperson[sp_key]["units"] += units
        by_salesperson[sp_key]["amount"] += amount

    def _sorted(d: Dict[str, Dict]) -> List[schemas.RollupRow]:
        rows = [
            schemas.RollupRow(
                key=k, label=v["label"], units=int(v["units"]), amount=round(v["amount"], 2)
            )
            for k, v in d.items()
        ]
        rows.sort(key=lambda r: r.amount, reverse=True)
        return rows

    monthly_amount_r = [round(x, 2) for x in monthly_amount]

    return schemas.RollupResponse(
        plan_id=plan_id,
        by_customer=_sorted(by_customer),
        by_product=_sorted(by_product),
        by_salesperson=_sorted(by_salesperson),
        monthly_amount=monthly_amount_r,
        monthly_units=monthly_units,
        total_units=total_units,
        total_amount=round(total_amount, 2),
    )


# ── Attainment (meta vs ventas reales) ──────────────────────────────────────

async def attainment(db: AsyncSession, plan_id: int) -> schemas.AttainmentResponse:
    plan = await db.get(models.ForecastPlan, plan_id)
    if plan is None:
        raise ValueError("Plan no encontrado")

    roll = await rollup(db, plan_id)

    O = sales_models.Order
    month_expr = func.extract("month", O.created_at).label("month")
    stmt = (
        select(
            month_expr,
            func.coalesce(func.sum(O.total_amount), 0.0).label("total"),
        )
        .where(
            O.kind == "order",
            O.status != "cancelled",
            func.extract("year", O.created_at) == plan.year,
        )
        .group_by(month_expr)
    )
    real_by_month = {int(r.month or 0): float(r.total or 0.0) for r in (await db.execute(stmt)).all()}

    months: List[schemas.AttainmentMonth] = []
    goal_year = 0.0
    real_year = 0.0
    for i in range(12):
        m = i + 1
        goal = float(roll.monthly_amount[i])
        real = float(real_by_month.get(m, 0.0))
        pct = round((real / goal * 100.0), 2) if goal > 0 else 0.0
        months.append(
            schemas.AttainmentMonth(
                month=m,
                goal_amount=round(goal, 2),
                real_amount=round(real, 2),
                attainment_pct=pct,
            )
        )
        goal_year += goal
        real_year += real

    att_year = round((real_year / goal_year * 100.0), 2) if goal_year > 0 else 0.0

    return schemas.AttainmentResponse(
        plan_id=plan_id,
        year=plan.year,
        months=months,
        goal_year=round(goal_year, 2),
        real_year=round(real_year, 2),
        attainment_year_pct=att_year,
    )


# ── Goal for range (usado por el tablero) ────────────────────────────────────

def _months_in_range(start: date, end: date) -> List[Tuple[int, int]]:
    """Regresa (year, month) por cada mes entre start y end inclusive."""
    out: List[Tuple[int, int]] = []
    y, m = start.year, start.month
    while (y, m) <= (end.year, end.month):
        out.append((y, m))
        m += 1
        if m == 13:
            m = 1
            y += 1
    return out


async def goal_for_range(
    db: AsyncSession, start: date, end: date
) -> schemas.GoalForRangeResponse:
    """Meta agregada del rango tomando el plan activo (no 'closed') más reciente
    cuyo año caiga en el rango. Regresa 0 si no hay plan aplicable — la UI
    decide si cae a otro origen (p. ej. presupuestos de Finanzas)."""
    months = _months_in_range(start, end)
    if not months:
        return schemas.GoalForRangeResponse(goal_amount=0.0, months_covered=[])

    years_in_range = sorted({y for (y, _m) in months}, reverse=True)

    plan_stmt = (
        select(models.ForecastPlan)
        .where(
            models.ForecastPlan.status != "closed",
            models.ForecastPlan.year.in_(years_in_range),
        )
        .order_by(
            models.ForecastPlan.year.desc(),
            models.ForecastPlan.updated_at.desc().nulls_last(),
            models.ForecastPlan.id.desc(),
        )
    )
    plan = (await db.execute(plan_stmt)).scalars().first()
    if plan is None:
        return schemas.GoalForRangeResponse(
            goal_amount=0.0,
            months_covered=[f"{y:04d}-{m:02d}" for (y, m) in months],
        )

    # Meses del rango que aplican a ESTE plan (mismo año)
    plan_months = [m for (y, m) in months if y == plan.year]
    if not plan_months:
        return schemas.GoalForRangeResponse(
            goal_amount=0.0, plan_id=plan.id, plan_name=plan.name, plan_year=plan.year,
            months_covered=[f"{y:04d}-{m:02d}" for (y, m) in months],
        )

    cols = [getattr(models.ForecastLine, f"m{m}") for m in plan_months]
    month_sum = sum(cols[1:], cols[0])  # suma columnar
    total_expr = func.coalesce(
        func.sum(month_sum * models.ForecastLine.unit_price),
        0.0,
    )
    total = float(
        (await db.execute(
            select(total_expr).where(models.ForecastLine.plan_id == plan.id)
        )).scalar() or 0.0
    )

    return schemas.GoalForRangeResponse(
        goal_amount=round(total, 2),
        plan_id=plan.id,
        plan_name=plan.name,
        plan_year=plan.year,
        months_covered=[f"{y:04d}-{m:02d}" for (y, m) in months],
    )


# ── Bulk template + export + import ─────────────────────────────────────────

TEMPLATE_HEADERS = [
    "cliente_rfc", "cliente_nombre",
    "sku", "producto_texto_libre",
    "vendedor_email",
    "precio_unitario",
    "m1", "m2", "m3", "m4", "m5", "m6",
    "m7", "m8", "m9", "m10", "m11", "m12",
]


async def _catalog_hints(db: AsyncSession) -> Dict[str, list]:
    """Regresa listas breves de catálogos para incluir como hojas de referencia."""
    cust_rows = (await db.execute(
        select(
            cust_models.Customer.name,
            cust_models.Customer.rfc,
            cust_models.Customer.client_number,
        ).where(cust_models.Customer.is_active.is_(True)).limit(2000)
    )).all()

    var_rows = (await db.execute(
        select(
            inv_models.ProductVariant.sku,
            inv_models.ProductVariant.price,
            inv_models.Product.name,
        )
        .join(inv_models.Product, inv_models.ProductVariant.product_id == inv_models.Product.id)
        .where(inv_models.ProductVariant.is_active.is_(True))
        .limit(2000)
    )).all()

    user_rows = (await db.execute(
        select(auth_models.User.email, auth_models.User.full_name)
        .where(auth_models.User.is_active.is_(True))
        .limit(500)
    )).all()

    return {
        "clientes": [{"nombre": r.name, "rfc": r.rfc or "", "codigo": r.client_number or ""} for r in cust_rows],
        "productos": [{"sku": r.sku, "producto": r.name, "precio_sugerido": float(r.price or 0.0)} for r in var_rows],
        "vendedores": [{"email": r.email, "nombre": r.full_name or ""} for r in user_rows],
    }


async def build_template_xlsx(db: AsyncSession, year: int) -> bytes:
    """Plantilla profesional multi-hoja: Forecast + Clientes + Productos + Vendedores."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    hints = await _catalog_hints(db)

    wb = Workbook()
    ws = wb.active
    ws.title = "Forecast"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    center = Alignment(horizontal="center")

    ws.append(TEMPLATE_HEADERS)
    for col_idx, _h in enumerate(TEMPLATE_HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    widths = [16, 30, 16, 30, 26, 14, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Fila-guía de ejemplo (comentada visualmente en gris claro)
    example = [
        "CRO180921AB2", "Constructora Robles",
        "CRSHR-BLK", "",
        "vendedor@empresa.com",
        1290,
        5, 5, 7, 10, 10, 12, 15, 18, 20, 25, 30, 28,
    ]
    ws.append(example)
    grey = PatternFill("solid", fgColor="F1F5F9")
    italic = Font(italic=True, color="64748B")
    for col_idx in range(1, len(example) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill = grey
        cell.font = italic

    ws.freeze_panes = "A2"

    # Hoja Clientes
    ws_c = wb.create_sheet("Clientes")
    ws_c.append(["Nombre", "RFC", "No. Cliente"])
    for row in hints["clientes"]:
        ws_c.append([row["nombre"], row["rfc"], row["codigo"]])
    ws_c.column_dimensions["A"].width = 34
    ws_c.column_dimensions["B"].width = 16
    ws_c.column_dimensions["C"].width = 14
    for c in ws_c[1]:
        c.font = header_font; c.fill = header_fill

    # Hoja Productos
    ws_p = wb.create_sheet("Productos")
    ws_p.append(["SKU", "Producto", "Precio sugerido"])
    for row in hints["productos"]:
        ws_p.append([row["sku"], row["producto"], row["precio_sugerido"]])
    ws_p.column_dimensions["A"].width = 18
    ws_p.column_dimensions["B"].width = 34
    ws_p.column_dimensions["C"].width = 16
    for c in ws_p[1]:
        c.font = header_font; c.fill = header_fill

    # Hoja Vendedores
    ws_v = wb.create_sheet("Vendedores")
    ws_v.append(["Email", "Nombre"])
    for row in hints["vendedores"]:
        ws_v.append([row["email"], row["nombre"]])
    ws_v.column_dimensions["A"].width = 30
    ws_v.column_dimensions["B"].width = 30
    for c in ws_v[1]:
        c.font = header_font; c.fill = header_fill

    # Hoja Instrucciones
    ws_i = wb.create_sheet("Instrucciones", 0)  # al inicio
    lines = [
        f"Plantilla de Forecast — año {year}",
        "",
        "Cómo llenarla:",
        "  1) Ve a la hoja 'Forecast' y agrega una fila por combinación cliente + producto/servicio + vendedor.",
        "  2) Puedes matchear por cliente_rfc (recomendado) o por cliente_nombre exacto.",
        "  3) El producto se matchea por SKU. Si no tienes SKU en catálogo, deja SKU vacío y usa 'producto_texto_libre'.",
        "  4) El vendedor se matchea por su email. Si no lo pones, la línea queda 'Sin vendedor'.",
        "  5) Llena m1..m12 con las unidades esperadas por mes calendario.",
        "  6) precio_unitario en pesos (MXN). Si el SKU está en catálogo se usa ese si dejas 0.",
        "",
        "Si un dato no matchea con el catálogo, la línea NO falla: se guarda como snapshot de texto libre.",
    ]
    for row_ix, line in enumerate(lines, start=1):
        c = ws_i.cell(row=row_ix, column=1, value=line)
        if row_ix == 1:
            c.font = Font(bold=True, size=14, color="1E3A8A")
    ws_i.column_dimensions["A"].width = 100

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_template_csv() -> bytes:
    """Plantilla mínima en CSV (sin catálogo de referencia)."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow([
        "CRO180921AB2", "Constructora Robles",
        "CRSHR-BLK", "",
        "vendedor@empresa.com",
        1290,
        5, 5, 7, 10, 10, 12, 15, 18, 20, 25, 30, 28,
    ])
    return buf.getvalue().encode("utf-8-sig")


async def export_plan_xlsx(db: AsyncSession, plan_id: int) -> bytes:
    """Descarga el plan como XLSX (mismas columnas que la plantilla + totales)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    plan = await db.get(models.ForecastPlan, plan_id)
    if plan is None:
        raise ValueError("Plan no encontrado")

    lines_stmt = (
        select(models.ForecastLine)
        .where(models.ForecastLine.plan_id == plan_id)
        .order_by(models.ForecastLine.id.asc())
    )
    lines = list((await db.execute(lines_stmt)).scalars().all())

    wb = Workbook()
    ws = wb.active
    ws.title = f"Forecast {plan.year}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1E3A8A")

    headers = TEMPLATE_HEADERS + ["total_unidades", "total_importe"]
    ws.append(headers)
    for i, _h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=i)
        cell.font = header_font
        cell.fill = header_fill

    # Cache nombres para el export
    cust_ids = {l.customer_id for l in lines if l.customer_id}
    var_ids = {l.variant_id for l in lines if l.variant_id}
    user_ids = {l.salesperson_id for l in lines if l.salesperson_id}

    rfc_by_cust: Dict[int, str] = {}
    if cust_ids:
        res = await db.execute(
            select(cust_models.Customer.id, cust_models.Customer.rfc).where(
                cust_models.Customer.id.in_(cust_ids)
            )
        )
        rfc_by_cust = {r.id: (r.rfc or "") for r in res}

    email_by_user: Dict[int, str] = {}
    if user_ids:
        res = await db.execute(
            select(auth_models.User.id, auth_models.User.email).where(
                auth_models.User.id.in_(user_ids)
            )
        )
        email_by_user = {r.id: r.email for r in res}

    for l in lines:
        units, amount = _line_totals(l)
        rfc = rfc_by_cust.get(l.customer_id, "") if l.customer_id else ""
        email = email_by_user.get(l.salesperson_id, "") if l.salesperson_id else ""
        ws.append([
            rfc, l.customer_name or "",
            l.sku or "", l.product_name or "" if not l.variant_id else "",
            email,
            float(l.unit_price or 0.0),
            int(l.m1), int(l.m2), int(l.m3), int(l.m4), int(l.m5), int(l.m6),
            int(l.m7), int(l.m8), int(l.m9), int(l.m10), int(l.m11), int(l.m12),
            units, amount,
        ])

    widths = [16, 30, 16, 30, 26, 14] + [8] * 12 + [12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_int(v) -> int:
    if v is None or v == "":
        return 0
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0


def _parse_float(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


async def _resolve_hints(db: AsyncSession) -> Tuple[Dict[str, int], Dict[str, Tuple[int, str, float]], Dict[str, Tuple[int, str]]]:
    """Regresa mapas para matchear rápido en la importación:
      - rfc → customer_id  (también incluye nombres a lower como fallback)
      - sku → (variant_id, product_name, price)
      - email → (user_id, full_name)
    """
    cust_rows = (await db.execute(
        select(cust_models.Customer.id, cust_models.Customer.name, cust_models.Customer.rfc)
    )).all()
    cust_by_key: Dict[str, int] = {}
    for r in cust_rows:
        if r.rfc:
            cust_by_key[r.rfc.strip().upper()] = r.id
        if r.name:
            cust_by_key[f"NAME::{r.name.strip().lower()}"] = r.id

    var_rows = (await db.execute(
        select(inv_models.ProductVariant.id, inv_models.ProductVariant.sku,
               inv_models.ProductVariant.price, inv_models.Product.name)
        .join(inv_models.Product, inv_models.ProductVariant.product_id == inv_models.Product.id)
    )).all()
    var_by_sku: Dict[str, Tuple[int, str, float]] = {
        r.sku.strip().upper(): (r.id, r.name, float(r.price or 0.0))
        for r in var_rows if r.sku
    }

    user_rows = (await db.execute(
        select(auth_models.User.id, auth_models.User.email, auth_models.User.full_name)
    )).all()
    user_by_email: Dict[str, Tuple[int, str]] = {
        r.email.strip().lower(): (r.id, r.full_name or "")
        for r in user_rows if r.email
    }

    return cust_by_key, var_by_sku, user_by_email


async def import_lines(
    db: AsyncSession,
    plan_id: int,
    file_bytes: bytes,
    filename: str,
) -> schemas.ImportResponse:
    """Ingesta bulk. Detecta xlsx/csv por extensión y crea líneas."""
    plan = await db.get(models.ForecastPlan, plan_id)
    if plan is None:
        raise ValueError("Plan no encontrado")

    name_lower = (filename or "").lower()
    is_xlsx = name_lower.endswith(".xlsx") or name_lower.endswith(".xlsm")

    rows: List[Dict[str, object]] = []
    if is_xlsx:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        ws = wb["Forecast"] if "Forecast" in wb.sheetnames else wb.active
        header_row = None
        for row in ws.iter_rows(values_only=True):
            if header_row is None:
                header_row = [str(c).strip() if c is not None else "" for c in row]
                continue
            data = dict(zip(header_row, row))
            rows.append(data)
    else:
        try:
            text = file_bytes.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = file_bytes.decode("latin-1")
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            rows.append(dict(r))

    cust_by_key, var_by_sku, user_by_email = await _resolve_hints(db)

    created: List[models.ForecastLine] = []
    errors: List[schemas.ImportRowError] = []
    skipped = 0

    for row_idx, row in enumerate(rows, start=2):  # start=2 → después del header
        rfc = str(row.get("cliente_rfc") or "").strip().upper()
        cust_name = str(row.get("cliente_nombre") or "").strip()
        sku = str(row.get("sku") or "").strip().upper()
        free_text = str(row.get("producto_texto_libre") or "").strip()
        email = str(row.get("vendedor_email") or "").strip().lower()

        # ¿Fila vacía?
        months_raw = {f"m{i}": _parse_int(row.get(f"m{i}")) for i in range(1, 13)}
        total_units = sum(months_raw.values())
        if not (rfc or cust_name or sku or free_text) and total_units == 0:
            skipped += 1
            continue

        customer_id = cust_by_key.get(rfc) if rfc else None
        if customer_id is None and cust_name:
            customer_id = cust_by_key.get(f"NAME::{cust_name.lower()}")

        variant_id: Optional[int] = None
        product_name: Optional[str] = free_text or None
        sku_snapshot: Optional[str] = None
        default_price: float = 0.0
        if sku and sku in var_by_sku:
            vid, pname, price = var_by_sku[sku]
            variant_id = vid
            product_name = pname
            sku_snapshot = sku
            default_price = price
        elif sku:
            # SKU escrito pero no existe → snapshot con el SKU tal cual
            sku_snapshot = sku
            if not product_name:
                product_name = f"SKU {sku}"

        if not product_name:
            errors.append(schemas.ImportRowError(
                row=row_idx, reason="Sin producto: pon un SKU válido o 'producto_texto_libre'."
            ))
            continue

        salesperson_id: Optional[int] = None
        salesperson_name: Optional[str] = None
        if email and email in user_by_email:
            uid, fname = user_by_email[email]
            salesperson_id = uid
            salesperson_name = fname or email
        elif email:
            salesperson_name = email  # snapshot

        unit_price = _parse_float(row.get("precio_unitario")) or default_price

        line = models.ForecastLine(
            plan_id=plan_id,
            customer_id=customer_id,
            variant_id=variant_id,
            salesperson_id=salesperson_id,
            product_name=product_name,
            sku=sku_snapshot,
            customer_name=cust_name or None,
            salesperson_name=salesperson_name,
            unit_price=unit_price,
        )
        for i in range(1, 13):
            setattr(line, f"m{i}", months_raw[f"m{i}"])

        # Rellenar snapshot de nombre desde catálogo si vino solo el id
        await _fill_snapshots(db, line)
        db.add(line)
        created.append(line)

    await db.commit()
    for l in created:
        await db.refresh(l)

    return schemas.ImportResponse(
        plan_id=plan_id,
        lines_created=len(created),
        lines_skipped=skipped,
        errors=errors,
        lines=[_line_to_schema(l) for l in created],
    )
