"""Business logic for the Sales / CRM module.

Cross-module side effects are intentional and centralized here:
  - Inventory: stock is decremented (+ StockMovement OUT) when an order is
    committed, and restored on cancel.
  - Finance: a Transaction (income) is recorded for every Payment, so the
    accounts-receivable picture and the P&L stay in sync.

Inventory/finance models are imported lazily inside functions to avoid
circular imports (same pattern the original module used).
"""

from __future__ import annotations

import csv
import io
from calendar import monthrange
from datetime import datetime, timedelta, timezone
from typing import List, Optional, Tuple

from sqlalchemy import and_, case, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload

from app.modules.sales import models, schemas


# ── Small helpers ─────────────────────────────────────────────────────────────

def _r(x: float) -> float:
    return round(float(x or 0.0), 2)


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _compute_line(item) -> Tuple[float, float]:
    """Return (subtotal, total) for a line, net of discount, plus line tax."""
    gross = (item.unit_price or 0.0) * (item.quantity or 0)
    subtotal = max(gross - (item.discount_amount or 0.0), 0.0)
    total = subtotal * (1 + (item.tax_rate or 0.0) / 100.0)
    return _r(subtotal), _r(total)


def _compute_totals(order: models.Order, items: List[models.OrderItem]) -> None:
    """Recompute money breakdown on `order` from its items + header config."""
    subtotal = sum((it.subtotal or 0.0) for it in items)

    if order.discount_type == "percent":
        discount_amount = subtotal * (order.discount_value or 0.0) / 100.0
    else:
        discount_amount = order.discount_value or 0.0
    discount_amount = min(discount_amount, subtotal)

    taxable = max(subtotal - discount_amount, 0.0)
    tax_amount = taxable * (order.tax_rate or 0.0) / 100.0
    total = taxable + tax_amount + (order.shipping_amount or 0.0)

    order.subtotal = _r(subtotal)
    order.discount_amount = _r(discount_amount)
    order.tax_amount = _r(tax_amount)
    order.total_amount = _r(total)


async def _generate_folio(db: AsyncSession, kind: str) -> str:
    prefix = "COT" if kind == "quote" else "ORD"
    result = await db.execute(
        select(func.count(models.Order.id)).where(models.Order.kind == kind)
    )
    n = (result.scalar() or 0) + 1
    return f"{prefix}-{n:06d}"


def _log_event(db: AsyncSession, order_id: int, event_type: str, *,
               from_status: str = None, to_status: str = None,
               message: str = None, user_id: int = None) -> None:
    db.add(models.OrderEvent(
        order_id=order_id, event_type=event_type,
        from_status=from_status, to_status=to_status,
        message=message, user_id=user_id,
    ))


# ── Inventory integration ─────────────────────────────────────────────────────

async def _resolve_warehouse_id(db: AsyncSession, warehouse_id: Optional[int]) -> Optional[int]:
    from app.modules.inventory import models as inv
    if warehouse_id:
        return warehouse_id
    res = await db.execute(
        select(inv.Warehouse.id).where(inv.Warehouse.is_active == True).order_by(inv.Warehouse.id).limit(1)  # noqa: E712
    )
    return res.scalar()


async def _move_stock(db: AsyncSession, *, variant_id: int, warehouse_id: Optional[int],
                      qty: int, direction: str, order_id: int, user_id: Optional[int]) -> None:
    """direction='out' decrements, 'in' increments. No-ops if stock isn't set up."""
    if not variant_id or not warehouse_id or not qty:
        return
    from app.modules.inventory import models as inv

    res = await db.execute(
        select(inv.StockLevel).where(
            inv.StockLevel.variant_id == variant_id,
            inv.StockLevel.warehouse_id == warehouse_id,
        )
    )
    level = res.scalars().first()
    delta = -qty if direction == "out" else qty
    if level is None:
        # Create a level lazily so movements are always traceable.
        level = inv.StockLevel(variant_id=variant_id, warehouse_id=warehouse_id, quantity=0)
        db.add(level)
    level.quantity = (level.quantity or 0) + delta

    db.add(inv.StockMovement(
        variant_id=variant_id, warehouse_id=warehouse_id,
        quantity=delta, movement_type="out" if direction == "out" else "in",
        reference=f"order:{order_id}", user_id=user_id,
        notes=f"Auto {'venta' if direction == 'out' else 'cancelación/devolución'} pedido #{order_id}",
    ))


async def _apply_stock_for_items(db: AsyncSession, order: models.Order,
                                 items: List[models.OrderItem], direction: str,
                                 user_id: Optional[int]) -> None:
    if order.kind != "order":
        return  # quotes never touch inventory
    wh = await _resolve_warehouse_id(db, order.warehouse_id)
    for it in items:
        if it.variant_id:
            await _move_stock(
                db, variant_id=it.variant_id, warehouse_id=wh,
                qty=it.quantity or 0, direction=direction,
                order_id=order.id, user_id=user_id,
            )


# ── Finance integration ───────────────────────────────────────────────────────

async def _record_finance_income(db: AsyncSession, order: models.Order, amount: float) -> None:
    if amount <= 0:
        return
    from app.modules.finance import models as fin
    db.add(fin.Transaction(
        type="income", amount=_r(amount), category="sales",
        description=f"Pago pedido {order.folio or '#' + str(order.id)}",
        reference=f"order:{order.id}",
    ))


async def _reverse_finance_income(db: AsyncSession, order: models.Order) -> None:
    """Counter-entry so cancelling a paid order doesn't inflate revenue."""
    from app.modules.finance import models as fin
    if (order.paid_amount or 0) <= 0:
        return
    db.add(fin.Transaction(
        type="expense", amount=_r(order.paid_amount), category="sales_reversal",
        description=f"Reverso por cancelación pedido {order.folio or '#' + str(order.id)}",
        reference=f"order:{order.id}:reversal",
    ))


# ── Item materialization (snapshots from catalog) ──────────────────────────────

async def _build_items(db: AsyncSession, items_in: List[schemas.OrderItemCreate]) -> List[models.OrderItem]:
    from app.modules.inventory import models as inv

    out: List[models.OrderItem] = []
    for raw in items_in:
        name, sku = raw.product_name, raw.sku
        if raw.variant_id and (not name or not sku):
            res = await db.execute(
                select(inv.ProductVariant)
                .where(inv.ProductVariant.id == raw.variant_id)
                .options(selectinload(inv.ProductVariant.product))
            )
            variant = res.scalars().first()
            if variant:
                sku = sku or variant.sku
                name = name or (variant.product.name if variant.product else variant.sku)
        it = models.OrderItem(
            variant_id=raw.variant_id,
            product_name=name or "Producto",
            sku=sku,
            quantity=raw.quantity,
            unit_price=_r(raw.unit_price),
            discount_amount=_r(raw.discount_amount),
            tax_rate=raw.tax_rate or 0.0,
        )
        it.subtotal, it.total = _compute_line(it)
        out.append(it)
    return out


# ── Queries ───────────────────────────────────────────────────────────────────

_LOAD = (
    selectinload(models.Order.items),
    selectinload(models.Order.payments),
    selectinload(models.Order.customer),
    selectinload(models.Order.seller),
)


async def get_order(db: AsyncSession, order_id: int) -> Optional[models.Order]:
    res = await db.execute(
        select(models.Order).where(models.Order.id == order_id).options(*_LOAD)
    )
    return res.scalars().first()


async def get_order_detail(db: AsyncSession, order_id: int) -> Optional[models.Order]:
    res = await db.execute(
        select(models.Order).where(models.Order.id == order_id)
        .options(*_LOAD, selectinload(models.Order.events))
    )
    return res.scalars().first()


async def get_orders(
    db: AsyncSession, *,
    skip: int = 0, limit: int = 50,
    kind: Optional[str] = None,
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    seller_id: Optional[int] = None,
    payment_method: Optional[str] = None,
    channel: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
    sort_by: str = "created_at",
    sort_dir: str = "desc",
    branch_warehouse_ids: Optional[List[int]] = None,
) -> Tuple[List[models.Order], int]:
    from app.modules.customers import models as cust

    base = select(models.Order)
    count_q = select(func.count(models.Order.id))

    conds = []
    # Aislamiento por sucursal: el pedido se ancla en su almacén (location).
    # Pedidos sin almacén se consideran globales/compartidos.
    if branch_warehouse_ids is not None:
        conds.append(or_(
            models.Order.warehouse_id.in_(branch_warehouse_ids),
            models.Order.warehouse_id.is_(None),
        ))
    if kind:
        conds.append(models.Order.kind == kind)
    if status:
        conds.append(models.Order.status == status)
    if customer_id:
        conds.append(models.Order.customer_id == customer_id)
    if seller_id:
        conds.append(models.Order.user_id == seller_id)
    if payment_method:
        conds.append(models.Order.payment_method == payment_method)
    if channel:
        conds.append(models.Order.channel == channel)
    if date_from:
        conds.append(models.Order.created_at >= date_from)
    if date_to:
        conds.append(models.Order.created_at <= date_to)

    needs_join = bool(q)
    if needs_join:
        base = base.outerjoin(cust.Customer, models.Order.customer_id == cust.Customer.id)
        count_q = count_q.outerjoin(cust.Customer, models.Order.customer_id == cust.Customer.id)
        like = f"%{q}%"
        conds.append(or_(
            models.Order.folio.ilike(like),
            models.Order.notes.ilike(like),
            models.Order.status.ilike(like),
            cust.Customer.name.ilike(like),
        ))

    for c in conds:
        base = base.where(c)
        count_q = count_q.where(c)

    sortable = {
        "created_at": models.Order.created_at,
        "total_amount": models.Order.total_amount,
        "folio": models.Order.folio,
        "status": models.Order.status,
        "due_date": models.Order.due_date,
    }
    col = sortable.get(sort_by, models.Order.created_at)
    base = base.order_by(col.asc() if sort_dir == "asc" else col.desc())

    total = (await db.execute(count_q)).scalar() or 0
    res = await db.execute(base.offset(skip).limit(limit).options(*_LOAD))
    return res.scalars().unique().all(), total


# ── Mutations ─────────────────────────────────────────────────────────────────

async def create_order(db: AsyncSession, order_in: schemas.OrderCreate,
                        user_id: Optional[int] = None) -> models.Order:
    kind = order_in.kind or "order"
    status = order_in.status or ("draft" if kind == "quote" else "pending")

    order = models.Order(
        folio=await _generate_folio(db, kind),
        kind=kind,
        customer_id=order_in.customer_id,
        warehouse_id=order_in.warehouse_id,
        user_id=user_id,
        status=status,
        payment_method=order_in.payment_method,
        channel=order_in.channel,
        currency=order_in.currency or "MXN",
        discount_type=order_in.discount_type or "amount",
        discount_value=order_in.discount_value or 0.0,
        tax_rate=order_in.tax_rate or 0.0,
        shipping_amount=order_in.shipping_amount or 0.0,
        due_date=order_in.due_date,
        valid_until=order_in.valid_until,
        notes=order_in.notes,
        bill_rfc=order_in.bill_rfc, bill_name=order_in.bill_name,
        bill_use=order_in.bill_use, bill_regime=order_in.bill_regime,
        bill_zip=order_in.bill_zip,
        cfdi_status="none",
        paid_amount=0.0,
    )
    db.add(order)
    await db.flush()

    items = await _build_items(db, order_in.items)
    for it in items:
        it.order_id = order.id
        db.add(it)
    _compute_totals(order, items)

    _log_event(db, order.id, "created", to_status=status,
               message=f"{'Cotización' if kind == 'quote' else 'Pedido'} creado", user_id=user_id)

    # Commit stock for real orders (not drafts / quotes)
    if kind == "order" and status not in ("draft", "cancelled"):
        await _apply_stock_for_items(db, order, items, "out", user_id)

    # If created already paid, register the full payment (finance + ledger)
    if kind == "order" and status == "paid":
        await _settle_payment(db, order, order.total_amount,
                              method=order.payment_method, user_id=user_id,
                              reference="auto", note="Pago al crear")

    await db.commit()
    return await get_order(db, order.id)


async def update_order(db: AsyncSession, order_id: int,
                       data: schemas.OrderUpdate, user_id: Optional[int] = None) -> Optional[models.Order]:
    order = await get_order(db, order_id)
    if not order:
        return None
    if order.status in ("cancelled", "paid") and data.items is not None:
        raise ValueError("No se pueden editar las partidas de un pedido pagado o cancelado")

    # Header fields
    for f in ("customer_id", "warehouse_id", "payment_method", "channel",
              "discount_type", "discount_value", "tax_rate", "shipping_amount",
              "due_date", "valid_until", "notes",
              "bill_rfc", "bill_name", "bill_use", "bill_regime", "bill_zip"):
        val = getattr(data, f)
        if val is not None:
            setattr(order, f, val)

    # Item replacement (with stock re-sync)
    if data.items is not None:
        if order.kind == "order" and order.status not in ("draft",):
            await _apply_stock_for_items(db, order, order.items, "in", user_id)  # return old
        for old in list(order.items):
            await db.delete(old)
        await db.flush()
        new_items = await _build_items(db, data.items)
        for it in new_items:
            it.order_id = order.id
            db.add(it)
        await db.flush()
        order = await get_order(db, order_id)
        if order.kind == "order" and order.status not in ("draft", "cancelled"):
            await _apply_stock_for_items(db, order, order.items, "out", user_id)  # take new

    if data.status is not None:
        order.status = data.status

    _compute_totals(order, order.items)
    _log_event(db, order.id, "edited", message="Pedido actualizado", user_id=user_id)
    await db.commit()
    return await get_order(db, order_id)


async def change_status(db: AsyncSession, order_id: int, new_status: str,
                        message: Optional[str] = None, user_id: Optional[int] = None) -> Optional[models.Order]:
    order = await get_order(db, order_id)
    if not order:
        return None
    old = order.status
    if old == new_status:
        return order

    # Cancellation: restock + reverse any recorded revenue
    if new_status == "cancelled":
        if order.kind == "order" and old not in ("draft", "cancelled"):
            await _apply_stock_for_items(db, order, order.items, "in", user_id)
        await _reverse_finance_income(db, order)

    # Re-activating a draft order commits stock
    if old == "draft" and new_status in ("pending", "partial", "paid") and order.kind == "order":
        await _apply_stock_for_items(db, order, order.items, "out", user_id)

    order.status = new_status
    if new_status == "converted":
        order.status = "converted"
    _log_event(db, order.id, "status_change", from_status=old, to_status=new_status,
               message=message, user_id=user_id)
    await db.commit()
    return await get_order(db, order_id)


async def _settle_payment(db: AsyncSession, order: models.Order, amount: float, *,
                          method: Optional[str], user_id: Optional[int],
                          reference: Optional[str] = None, note: Optional[str] = None) -> None:
    """Internal: append a payment, bump paid_amount, auto-advance status, hit finance."""
    db.add(models.Payment(
        order_id=order.id, amount=_r(amount), method=method,
        reference=reference, note=note, user_id=user_id,
    ))
    order.paid_amount = _r((order.paid_amount or 0.0) + amount)
    await _record_finance_income(db, order, amount)

    prev = order.status
    if order.paid_amount + 0.001 >= order.total_amount and order.total_amount > 0:
        order.status = "paid"
    elif order.paid_amount > 0:
        order.status = "partial"
    _log_event(db, order.id, "payment", from_status=prev, to_status=order.status,
               message=f"Pago de ${_r(amount):,.2f}", user_id=user_id)


async def register_payment(db: AsyncSession, order_id: int, pay: schemas.PaymentCreate,
                           user_id: Optional[int] = None) -> Optional[models.Order]:
    order = await get_order(db, order_id)
    if not order:
        return None
    if order.kind != "order":
        raise ValueError("Solo los pedidos admiten pagos (no las cotizaciones)")
    if order.status == "cancelled":
        raise ValueError("No se puede pagar un pedido cancelado")
    if pay.amount > order.balance + 0.001:
        raise ValueError(f"El pago (${pay.amount:,.2f}) excede el saldo (${order.balance:,.2f})")
    await _settle_payment(db, order, pay.amount, method=pay.method or order.payment_method,
                          user_id=user_id, reference=pay.reference, note=pay.note)
    await db.commit()
    return await get_order(db, order_id)


async def convert_quote_to_order(db: AsyncSession, quote_id: int,
                                 user_id: Optional[int] = None) -> Optional[models.Order]:
    quote = await get_order(db, quote_id)
    if not quote:
        return None
    if quote.kind != "quote":
        raise ValueError("El documento no es una cotización")

    order = models.Order(
        folio=await _generate_folio(db, "order"),
        kind="order", customer_id=quote.customer_id, warehouse_id=quote.warehouse_id,
        user_id=user_id or quote.user_id, status="pending",
        payment_method=quote.payment_method, channel=quote.channel, currency=quote.currency,
        discount_type=quote.discount_type, discount_value=quote.discount_value,
        tax_rate=quote.tax_rate, shipping_amount=quote.shipping_amount,
        notes=quote.notes, bill_rfc=quote.bill_rfc, bill_name=quote.bill_name,
        bill_use=quote.bill_use, bill_regime=quote.bill_regime, bill_zip=quote.bill_zip,
        cfdi_status="none", paid_amount=0.0,
    )
    db.add(order)
    await db.flush()

    items = []
    for src in quote.items:
        it = models.OrderItem(
            order_id=order.id, variant_id=src.variant_id, product_name=src.product_name,
            sku=src.sku, quantity=src.quantity, unit_price=src.unit_price,
            discount_amount=src.discount_amount, tax_rate=src.tax_rate,
            subtotal=src.subtotal, total=src.total,
        )
        db.add(it)
        items.append(it)
    _compute_totals(order, items)
    await _apply_stock_for_items(db, order, items, "out", user_id)

    quote.status = "converted"
    _log_event(db, quote.id, "status_change", from_status="quote", to_status="converted",
               message=f"Convertida a pedido {order.folio}", user_id=user_id)
    _log_event(db, order.id, "created", to_status="pending",
               message=f"Generado desde cotización {quote.folio}", user_id=user_id)
    await db.commit()
    return await get_order(db, order.id)


async def cancel_order(db: AsyncSession, order_id: int, user_id: Optional[int] = None):
    return await change_status(db, order_id, "cancelled", message="Cancelado", user_id=user_id)


# ── Analytics ─────────────────────────────────────────────────────────────────

async def get_stats(db: AsyncSession, start: Optional[datetime] = None, end: Optional[datetime] = None,
                    branch_warehouse_ids: Optional[List[int]] = None) -> schemas.SalesStats:
    """All KPIs in a SINGLE aggregated query (no rows pulled into Python).

    Uses portable conditional aggregation (CASE inside COUNT/SUM) so it runs the
    same on SQLite (dev) and Postgres (prod).
    """
    O = models.Order
    active = and_(O.kind == "order", O.status != "cancelled")
    pending = and_(active, O.status.in_(("pending", "partial")))

    date_filters = []
    if start is not None:
        date_filters.append(O.created_at >= start)
    if end is not None:
        date_filters.append(O.created_at < end)
    if branch_warehouse_ids is not None:
        date_filters.append(or_(O.warehouse_id.in_(branch_warehouse_ids), O.warehouse_id.is_(None)))

    stmt = select(
        func.count(case((active, 1))).label("orders_count"),
        func.coalesce(func.sum(case((active, O.paid_amount), else_=0.0)), 0.0).label("total_sold"),
        func.count(case((and_(active, O.status == "paid"), 1))).label("paid_count"),
        func.count(case((pending, 1))).label("pending_orders"),
        func.coalesce(func.sum(case((pending, O.total_amount - O.paid_amount), else_=0.0)), 0.0).label("pending_amount"),
        func.coalesce(func.sum(case((active, O.total_amount), else_=0.0)), 0.0).label("active_total"),
        func.count(case((O.kind == "quote", 1))).label("quotes_count"),
    ).where(*date_filters)
    r = (await db.execute(stmt)).one()
    oc = r.orders_count or 0

    return schemas.SalesStats(
        total_sold=_r(r.total_sold), orders_count=oc,
        pending_orders=r.pending_orders or 0, pending_amount=_r(r.pending_amount),
        paid_rate=round(r.paid_count / oc * 100, 1) if oc else 0.0,
        avg_ticket=round(r.active_total / oc, 2) if oc else 0.0,
        quotes_count=r.quotes_count or 0,
    )


def _branch_cond(O, branch_warehouse_ids):
    """Condición de aislamiento por sucursal (pedido anclado en su almacén;
    pedidos sin almacén son globales)."""
    return or_(O.warehouse_id.in_(branch_warehouse_ids), O.warehouse_id.is_(None))


async def sales_trend(db: AsyncSession, granularity: str = "day", days: int = 30, end: Optional[datetime] = None,
                       customer_id: Optional[int] = None, branch_warehouse_ids: Optional[List[int]] = None) -> List[schemas.TrendPoint]:
    """Aggregate revenue by day in SQL, bounded to the relevant date window so we
    never scan the whole history; then roll days up to week/month in Python.

    `end` lets the caller anchor the window somewhere other than "now" (e.g. to
    build a comparable previous-period series for BI charts). `customer_id`
    restricts both the sales and returns series to a single customer.

    "Returns" has no dedicated ledger in this app today; cancelled orders are
    the closest existing concept (see Order.status), so they're used as the
    returns series.
    """
    O = models.Order
    mult = {"day": 1, "week": 7, "month": 31}.get(granularity, 1)
    end_bound = end or _now()
    cutoff = end_bound - timedelta(days=days * mult + 1)

    sales_conds = [O.kind == "order", O.status != "cancelled", O.created_at >= cutoff, O.created_at < end_bound]
    returns_conds = [O.kind == "order", O.status == "cancelled", O.created_at >= cutoff, O.created_at < end_bound]
    if customer_id is not None:
        sales_conds.append(O.customer_id == customer_id)
        returns_conds.append(O.customer_id == customer_id)
    if branch_warehouse_ids is not None:
        sales_conds.append(_branch_cond(O, branch_warehouse_ids))
        returns_conds.append(_branch_cond(O, branch_warehouse_ids))

    stmt = (
        select(
            func.date(O.created_at).label("d"),
            func.coalesce(func.sum(O.total_amount), 0.0).label("total"),
            func.count(O.id).label("count"),
        )
        .where(*sales_conds)
        .group_by(func.date(O.created_at))
        .order_by(func.date(O.created_at))
    )
    daily = [(str(row.d)[:10], float(row.total or 0.0), int(row.count or 0))
             for row in (await db.execute(stmt)).all()]

    returns_stmt = (
        select(
            func.date(O.created_at).label("d"),
            func.coalesce(func.sum(O.total_amount), 0.0).label("returns_total"),
        )
        .where(*returns_conds)
        .group_by(func.date(O.created_at))
    )
    returns_daily = {str(row.d)[:10]: float(row.returns_total or 0.0)
                     for row in (await db.execute(returns_stmt)).all()}

    goal_by_month = await _sales_goal_by_month(db, cutoff, end_bound)

    def bucket_key(dstr: str) -> str:
        dt = datetime.strptime(dstr, "%Y-%m-%d")
        if granularity == "month":
            return dt.strftime("%Y-%m")
        if granularity == "week":
            return f"{dt.isocalendar().year}-W{dt.isocalendar().week:02d}"
        return dstr

    buckets: dict[str, list] = {}
    for dstr, total, count in daily:
        e = buckets.setdefault(bucket_key(dstr), [0.0, 0, 0.0, 0.0])
        e[0] += total
        e[1] += count
    for dstr, rtotal in returns_daily.items():
        e = buckets.setdefault(bucket_key(dstr), [0.0, 0, 0.0, 0.0])
        e[2] += rtotal
    for dstr in {d for d, *_ in daily} | set(returns_daily.keys()):
        dt = datetime.strptime(dstr, "%Y-%m-%d")
        month_key = dt.strftime("%Y-%m")
        monthly_goal = goal_by_month.get(month_key)
        if monthly_goal is not None:
            days_in_month = monthrange(dt.year, dt.month)[1]
            per_day_goal = monthly_goal / days_in_month
            e = buckets.setdefault(bucket_key(dstr), [0.0, 0, 0.0, 0.0])
            e[3] += per_day_goal

    points = [
        schemas.TrendPoint(period=k, total=_r(v[0]), count=v[1], returns_total=_r(v[2]),
                            goal=_r(v[3]) if v[3] else None)
        for k, v in sorted(buckets.items())
    ]
    return points[-days:]


async def _sales_goal_by_month(db: AsyncSession, start: datetime, end: datetime) -> dict[str, float]:
    """Pull the income budget tagged as a sales goal (category containing
    "venta"/"sales") for each "YYYY-MM" period overlapping the window."""
    from app.modules.finance import models as fin

    periods = set()
    cur = start.replace(day=1)
    while cur <= end:
        periods.add(cur.strftime("%Y-%m"))
        cur = (cur.replace(day=28) + timedelta(days=4)).replace(day=1)

    if not periods:
        return {}
    stmt = select(fin.Budget).where(
        fin.Budget.type == "income",
        fin.Budget.period.in_(periods),
        or_(fin.Budget.category.ilike("%venta%"), fin.Budget.category.ilike("%sales%")),
    )
    rows = (await db.execute(stmt)).scalars().all()
    out: dict[str, float] = {}
    for b in rows:
        out[b.period] = out.get(b.period, 0.0) + b.amount
    return out


async def get_average_returns(db: AsyncSession, customer_id: Optional[int] = None,
                              branch_warehouse_ids: Optional[List[int]] = None) -> schemas.AverageReturns:
    """Average size of cancelled ("returned") orders, optionally scoped to one
    customer. Cancellations are the closest existing concept to a return."""
    O = models.Order
    conds = [O.kind == "order", O.status == "cancelled"]
    if customer_id is not None:
        conds.append(O.customer_id == customer_id)
    if branch_warehouse_ids is not None:
        conds.append(_branch_cond(O, branch_warehouse_ids))
    stmt = select(
        func.coalesce(func.avg(O.total_amount), 0.0).label("avg_amount"),
        func.count(O.id).label("count"),
    ).where(*conds)
    row = (await db.execute(stmt)).one()
    return schemas.AverageReturns(customer_id=customer_id, average_amount=_r(row.avg_amount or 0.0), count=row.count or 0)


async def get_customer_forecast(db: AsyncSession, customer_id: int, months: int = 6) -> schemas.CustomerForecast:
    """Simple per-customer sales forecast: average of the last `months` of
    actual sales (cancelled orders excluded), used to project next month, and
    weighed against the customer's historical share of the company-wide
    monthly sales goal (Finance `Budget`, type=income, category~venta/sales).

    This is a moving-average forecast, not a statistical model — there isn't
    enough order volume in this app to justify anything more sophisticated,
    and a transparent average is easier to audit than a black-box estimate.
    """
    from app.modules.customers import models as cust

    O = models.Order
    now = _now()
    start = (now.replace(day=1) - timedelta(days=1)).replace(day=1)
    for _ in range(months - 1):
        start = (start - timedelta(days=1)).replace(day=1)

    cust_row = (await db.execute(select(cust.Customer).where(cust.Customer.id == customer_id))).scalars().first()
    customer_name = cust_row.name if cust_row else "Cliente"

    # Group by day in SQL (portable across SQLite/Postgres) and roll up to
    # months in Python, same approach as sales_trend() above.
    month_stmt = (
        select(
            func.date(O.created_at).label("d"),
            func.coalesce(func.sum(O.total_amount), 0.0).label("total"),
        )
        .where(O.kind == "order", O.status != "cancelled", O.customer_id == customer_id, O.created_at >= start)
        .group_by(func.date(O.created_at))
        .order_by(func.date(O.created_at))
    )
    rows = (await db.execute(month_stmt)).all()
    monthly: dict[str, float] = {}
    for r in rows:
        month_key = str(r.d)[:7]
        monthly[month_key] = monthly.get(month_key, 0.0) + float(r.total or 0.0)
    history_months = sorted(monthly.keys())
    history_totals = [_r(monthly[m]) for m in history_months]

    avg_monthly = _r(sum(history_totals) / len(history_totals)) if history_totals else 0.0
    if len(history_totals) >= 3:
        forecast_next_month = _r(sum(history_totals[-3:]) / 3)
    elif history_totals:
        forecast_next_month = avg_monthly
    else:
        forecast_next_month = 0.0

    trend_pct = None
    if len(history_totals) >= 2:
        prev = history_totals[:-1]
        prev_avg = sum(prev) / len(prev)
        if prev_avg:
            trend_pct = _r((history_totals[-1] - prev_avg) / prev_avg * 100)

    total_stmt = (
        select(func.coalesce(func.sum(O.total_amount), 0.0))
        .where(O.kind == "order", O.status != "cancelled", O.created_at >= start)
    )
    total_all_customers = float((await db.execute(total_stmt)).scalar() or 0.0)
    goal_share_pct = _r(sum(history_totals) / total_all_customers) if total_all_customers else None

    next_month_dt = (now.replace(day=1) + timedelta(days=32)).replace(day=1)
    goal_month = next_month_dt.strftime("%Y-%m")
    goal_by_month = await _sales_goal_by_month(db, next_month_dt, next_month_dt)
    goal_amount = goal_by_month.get(goal_month)

    goal_allocated = _r(goal_amount * goal_share_pct) if (goal_amount is not None and goal_share_pct is not None) else None
    variance_vs_goal = _r(forecast_next_month - goal_allocated) if goal_allocated is not None else None

    return schemas.CustomerForecast(
        customer_id=customer_id, customer_name=customer_name,
        history_months=history_months, history_totals=history_totals,
        avg_monthly=avg_monthly, forecast_next_month=forecast_next_month, trend_pct=trend_pct,
        goal_month=goal_month, goal_amount=_r(goal_amount) if goal_amount is not None else None,
        goal_share_pct=goal_share_pct, goal_allocated=goal_allocated, variance_vs_goal=variance_vs_goal,
    )


async def top_customers(db: AsyncSession, limit: int = 5, start: Optional[datetime] = None, end: Optional[datetime] = None,
                        branch_warehouse_ids: Optional[List[int]] = None) -> List[schemas.TopCustomer]:
    from app.modules.customers import models as cust
    O = models.Order
    conds = [O.kind == "order", O.status != "cancelled"]
    if start is not None:
        conds.append(O.created_at >= start)
    if end is not None:
        conds.append(O.created_at < end)
    if branch_warehouse_ids is not None:
        conds.append(_branch_cond(O, branch_warehouse_ids))
    stmt = (
        select(
            O.customer_id,
            func.coalesce(cust.Customer.name, "Sin cliente").label("name"),
            func.coalesce(func.sum(O.total_amount), 0.0).label("total"),
            func.count(O.id).label("orders"),
        )
        .outerjoin(cust.Customer, O.customer_id == cust.Customer.id)
        .where(*conds)
        .group_by(O.customer_id, cust.Customer.name)
        .order_by(func.sum(O.total_amount).desc())
        .limit(limit)
    )
    return [
        schemas.TopCustomer(customer_id=r.customer_id, name=r.name,
                            total=_r(r.total), orders=r.orders)
        for r in (await db.execute(stmt)).all()
    ]


async def top_products(db: AsyncSession, limit: int = 5, start: Optional[datetime] = None, end: Optional[datetime] = None,
                       branch_warehouse_ids: Optional[List[int]] = None) -> List[schemas.TopProduct]:
    O = models.Order
    OI = models.OrderItem
    conds = [O.kind == "order", O.status != "cancelled"]
    if start is not None:
        conds.append(O.created_at >= start)
    if end is not None:
        conds.append(O.created_at < end)
    if branch_warehouse_ids is not None:
        conds.append(_branch_cond(O, branch_warehouse_ids))
    stmt = (
        select(
            OI.variant_id,
            func.coalesce(OI.product_name, "—").label("name"),
            func.coalesce(func.sum(OI.quantity), 0).label("qty"),
            func.coalesce(func.sum(OI.total), 0.0).label("total"),
        )
        .join(O, OI.order_id == O.id)
        .where(*conds)
        .group_by(OI.variant_id, OI.product_name)
        .order_by(func.sum(OI.total).desc())
        .limit(limit)
    )
    return [
        schemas.TopProduct(variant_id=r.variant_id, name=r.name,
                           quantity=r.qty, total=_r(r.total))
        for r in (await db.execute(stmt)).all()
    ]


async def sales_by_seller(db: AsyncSession, start: Optional[datetime] = None, end: Optional[datetime] = None,
                          branch_warehouse_ids: Optional[List[int]] = None) -> List[schemas.SalesBySeller]:
    from app.modules.auth import models as auth_models
    O = models.Order
    conds = [O.kind == "order", O.status != "cancelled"]
    if start is not None:
        conds.append(O.created_at >= start)
    if end is not None:
        conds.append(O.created_at < end)
    if branch_warehouse_ids is not None:
        conds.append(_branch_cond(O, branch_warehouse_ids))
    stmt = (
        select(
            O.user_id,
            func.coalesce(auth_models.User.full_name, auth_models.User.email, "Sin vendedor").label("name"),
            func.coalesce(func.sum(O.total_amount), 0.0).label("total"),
            func.count(O.id).label("orders"),
        )
        .outerjoin(auth_models.User, O.user_id == auth_models.User.id)
        .where(*conds)
        .group_by(O.user_id, auth_models.User.full_name, auth_models.User.email)
        .order_by(func.sum(O.total_amount).desc())
    )
    return [
        schemas.SalesBySeller(user_id=r.user_id, name=r.name, total=_r(r.total), orders=r.orders)
        for r in (await db.execute(stmt)).all()
    ]


async def sales_by_channel(db: AsyncSession, start: Optional[datetime] = None, end: Optional[datetime] = None,
                           branch_warehouse_ids: Optional[List[int]] = None) -> List[schemas.SalesByChannel]:
    O = models.Order
    conds = [O.kind == "order", O.status != "cancelled"]
    if start is not None:
        conds.append(O.created_at >= start)
    if end is not None:
        conds.append(O.created_at < end)
    if branch_warehouse_ids is not None:
        conds.append(_branch_cond(O, branch_warehouse_ids))
    stmt = (
        select(
            func.coalesce(O.channel, "Sin canal").label("channel"),
            func.coalesce(func.sum(O.total_amount), 0.0).label("total"),
            func.count(O.id).label("orders"),
        )
        .where(*conds)
        .group_by(O.channel)
        .order_by(func.sum(O.total_amount).desc())
    )
    return [
        schemas.SalesByChannel(channel=r.channel, total=_r(r.total), orders=r.orders)
        for r in (await db.execute(stmt)).all()
    ]


async def customer_360(db: AsyncSession, customer_id: int) -> Optional[schemas.Customer360]:
    from app.modules.customers import models as cust
    cres = await db.execute(select(cust.Customer).where(cust.Customer.id == customer_id))
    customer = cres.scalars().first()
    if not customer:
        return None

    res = await db.execute(
        select(models.Order).where(
            models.Order.customer_id == customer_id, models.Order.kind == "order"
        ).options(*_LOAD).order_by(models.Order.created_at.desc())
    )
    orders = res.scalars().unique().all()
    active = [o for o in orders if o.status != "cancelled"]

    total_spent = sum(o.paid_amount or 0 for o in active)
    open_balance = sum((o.total_amount or 0) - (o.paid_amount or 0)
                       for o in active if o.status in ("pending", "partial"))
    avg_ticket = round(sum(o.total_amount or 0 for o in active) / len(active), 2) if active else 0.0

    return schemas.Customer360(
        customer=schemas.CustomerLite.model_validate(customer),
        total_spent=_r(total_spent), orders_count=len(active),
        open_balance=_r(open_balance), avg_ticket=avg_ticket,
        last_order_at=orders[0].created_at if orders else None,
        recent_orders=[schemas.OrderInDB.model_validate(_attach_balance(o)) for o in orders[:5]],
    )


# ── Export & invoice ──────────────────────────────────────────────────────────

EXPORT_HEADERS = ["Folio", "Tipo", "Cliente", "Fecha", "Estado", "Metodo",
                  "Subtotal", "Descuento", "Impuesto", "Envio", "Total", "Pagado", "Saldo"]


def _export_rows(orders) -> List[list]:
    return [
        [
            o.folio, o.kind, o.customer.name if o.customer else "",
            o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else "",
            o.status, o.payment_method or "",
            o.subtotal, o.discount_amount, o.tax_amount, o.shipping_amount,
            o.total_amount, o.paid_amount, _r((o.total_amount or 0) - (o.paid_amount or 0)),
        ]
        for o in orders
    ]


async def _export_orders(db: AsyncSession, **filters):
    filters.pop("skip", None)
    filters.pop("limit", None)
    orders, _ = await get_orders(db, skip=0, limit=100000, **filters)
    return orders


async def export_csv(db: AsyncSession, **filters) -> str:
    orders = await _export_orders(db, **filters)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(EXPORT_HEADERS)
    w.writerows(_export_rows(orders))
    # BOM al inicio para que Excel detecte UTF-8 y muestre acentos correctamente.
    return "﻿" + buf.getvalue()


async def export_xlsx(db: AsyncSession, **filters) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    orders = await _export_orders(db, **filters)
    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"
    ws.append(EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E2E5C", end_color="1E2E5C", fill_type="solid")
    money_cols = {7, 8, 9, 10, 11, 12, 13}  # 1-indexed: Subtotal..Saldo
    for row in _export_rows(orders):
        ws.append(row)
        for col_idx in money_cols:
            ws.cell(row=ws.max_row, column=col_idx).number_format = "#,##0.00"
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in _export_rows(orders)] + [8])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_invoice_payload(order: models.Order) -> dict:
    """Assemble a CFDI 4.0-style payload. Stamping (timbrado) must be done by a
    PAC (Facturama, Finkok, etc.) with the taxpayer's CSD — wire credentials in
    a dedicated billing service. This returns the document ready to stamp."""
    return {
        "Receptor": {
            "Rfc": order.bill_rfc, "Nombre": order.bill_name,
            "UsoCFDI": order.bill_use or "G03",
            "RegimenFiscalReceptor": order.bill_regime,
            "DomicilioFiscalReceptor": order.bill_zip,
        },
        "Moneda": order.currency or "MXN",
        "MetodoPago": "PUE" if order.status == "paid" else "PPD",
        "FormaPago": {
            "cash": "01", "check": "02", "transfer": "03", "card": "04", "credit": "99",
        }.get(order.payment_method or "", "99"),
        "Conceptos": [
            {
                "Descripcion": it.product_name, "ClaveProdServ": "50000000",
                "NoIdentificacion": it.sku, "Cantidad": it.quantity,
                "ValorUnitario": it.unit_price, "Importe": _r((it.unit_price or 0) * (it.quantity or 0)),
                "Descuento": it.discount_amount,
                "Impuestos": {"Traslados": [{"Impuesto": "002", "TasaOCuota": (it.tax_rate or 0) / 100.0}]},
            }
            for it in order.items
        ],
        "Subtotal": order.subtotal, "Descuento": order.discount_amount,
        "Total": order.total_amount, "_meta": {"folio": order.folio, "order_id": order.id},
    }


def _attach_balance(order: models.Order) -> models.Order:
    # `balance` is a @property; OrderInDB reads it via from_attributes. No-op kept
    # for readability of intent at call sites.
    return order
