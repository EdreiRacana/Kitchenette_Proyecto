"""REST API for the Sales / CRM module."""

from __future__ import annotations

from datetime import datetime
from typing import Annotated, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
import io

from sqlalchemy.ext.asyncio import AsyncSession

from app.api import deps
from app.modules.auth.models import User
from app.modules.sales import schemas, service

router = APIRouter()

DB = Annotated[AsyncSession, Depends(deps.get_db)]
CurrentUser = Annotated[User, Depends(deps.get_current_active_user)]


# ── Analytics (declared before /{order_id} so paths don't collide) ────────────

@router.get("/stats", response_model=schemas.SalesStats)
async def stats(db: DB, current_user: CurrentUser, start: Optional[datetime] = None, end: Optional[datetime] = None,
                status: Optional[str] = None, payment_method: Optional[str] = None, q: Optional[str] = None,
                relationship_type: Optional[str] = None, client_type: Optional[str] = None,
                channel: Optional[str] = None):
    from app.modules.inventory.branch_scope import visible_warehouse_ids
    ids = await visible_warehouse_ids(db, current_user)
    return await service.get_stats(db, start=start, end=end, branch_warehouse_ids=ids,
                                   status=status, payment_method=payment_method, q=q,
                                   relationship_type=relationship_type, client_type=client_type,
                                   channel=channel)


@router.get("/pipeline-stats", response_model=schemas.PipelineStats)
async def pipeline_stats(db: DB, current_user: CurrentUser,
                         start: Optional[datetime] = None, end: Optional[datetime] = None,
                         relationship_type: Optional[str] = None,
                         client_type: Optional[str] = None,
                         channel: Optional[str] = None):
    """Universo COMPLETO por (kind, status). El Pipeline lo usa para no depender
    de la página 1 de la Lista."""
    from app.modules.inventory.branch_scope import visible_warehouse_ids
    ids = await visible_warehouse_ids(db, current_user)
    return await service.get_pipeline_stats(
        db, branch_warehouse_ids=ids,
        relationship_type=relationship_type, client_type=client_type,
        channel=channel, date_from=start, date_to=end,
    )


async def _branch_ids(db, user):
    from app.modules.inventory.branch_scope import visible_warehouse_ids
    return await visible_warehouse_ids(db, user)


@router.get("/analytics/trend", response_model=List[schemas.TrendPoint])
async def trend(db: DB, current_user: CurrentUser,
                granularity: str = Query("day", pattern="^(day|week|month)$"),
                days: int = Query(30, ge=1, le=365),
                end: Optional[datetime] = None,
                customer_id: Optional[int] = None,
                relationship_type: Optional[str] = None,
                client_type: Optional[str] = None,
                channel: Optional[str] = None):
    ids = await _branch_ids(db, current_user)
    return await service.sales_trend(db, granularity=granularity, days=days, end=end,
                                     customer_id=customer_id, branch_warehouse_ids=ids,
                                     relationship_type=relationship_type,
                                     client_type=client_type, channel=channel)


@router.get("/sellers", response_model=List[schemas.SellerLite])
async def sellers(db: DB, _: CurrentUser):
    return await service.list_sellers(db)


# ── Agentes de venta / comisionistas ──────────────────────────────────────────
@router.get("/agents", response_model=List[schemas.SalesAgentInDB])
async def list_agents(db: DB, _: CurrentUser, include_inactive: bool = False):
    return await service.list_agents(db, include_inactive=include_inactive)


@router.post("/agents", response_model=schemas.SalesAgentInDB, status_code=201)
async def create_agent(data: schemas.SalesAgentCreate, db: DB, _: CurrentUser):
    return await service.create_agent(db, data)


@router.patch("/agents/{agent_id}", response_model=schemas.SalesAgentInDB)
async def update_agent(agent_id: int, data: schemas.SalesAgentUpdate, db: DB, _: CurrentUser):
    agent = await service.update_agent(db, agent_id, data)
    if not agent:
        raise HTTPException(status_code=404, detail="Agente no encontrado")
    return agent


@router.delete("/agents/{agent_id}", status_code=204)
async def delete_agent(agent_id: int, db: DB, _: CurrentUser):
    ok = await service.delete_agent(db, agent_id)
    if not ok:
        raise HTTPException(status_code=404, detail="Agente no encontrado")


@router.get("/agents/commissions", response_model=schemas.AgentCommissionReport)
async def agent_commissions(db: DB, current_user: CurrentUser,
                            start: Optional[datetime] = None, end: Optional[datetime] = None):
    ids = await _branch_ids(db, current_user)
    return await service.agent_commissions(db, start=start, end=end, branch_warehouse_ids=ids)


@router.get("/analytics/returns-avg", response_model=schemas.AverageReturns)
async def returns_avg(db: DB, current_user: CurrentUser, customer_id: Optional[int] = None):
    ids = await _branch_ids(db, current_user)
    return await service.get_average_returns(db, customer_id=customer_id, branch_warehouse_ids=ids)


@router.get("/analytics/forecast/{customer_id}", response_model=schemas.CustomerForecast)
async def customer_forecast(customer_id: int, db: DB, _: CurrentUser, months: int = Query(6, ge=2, le=24)):
    return await service.get_customer_forecast(db, customer_id, months=months)


@router.get("/analytics/top-customers", response_model=List[schemas.TopCustomer])
async def top_customers(db: DB, current_user: CurrentUser, limit: int = Query(5, ge=1, le=50),
                         start: Optional[datetime] = None, end: Optional[datetime] = None):
    ids = await _branch_ids(db, current_user)
    return await service.top_customers(db, limit=limit, start=start, end=end, branch_warehouse_ids=ids)


@router.get("/analytics/top-products", response_model=List[schemas.TopProduct])
async def top_products(db: DB, current_user: CurrentUser, limit: int = Query(5, ge=1, le=50),
                        start: Optional[datetime] = None, end: Optional[datetime] = None):
    ids = await _branch_ids(db, current_user)
    return await service.top_products(db, limit=limit, start=start, end=end, branch_warehouse_ids=ids)


@router.get("/analytics/by-seller", response_model=List[schemas.SalesBySeller])
async def by_seller(db: DB, current_user: CurrentUser, start: Optional[datetime] = None, end: Optional[datetime] = None):
    ids = await _branch_ids(db, current_user)
    return await service.sales_by_seller(db, start=start, end=end, branch_warehouse_ids=ids)


@router.get("/analytics/by-channel", response_model=List[schemas.SalesByChannel])
async def by_channel(db: DB, current_user: CurrentUser, start: Optional[datetime] = None, end: Optional[datetime] = None):
    ids = await _branch_ids(db, current_user)
    return await service.sales_by_channel(db, start=start, end=end, branch_warehouse_ids=ids)


@router.get("/analytics/heatmap", response_model=List[schemas.HeatmapCell])
async def sales_heatmap(db: DB, current_user: CurrentUser, start: Optional[datetime] = None, end: Optional[datetime] = None):
    """Actividad de ventas por día-de-semana × hora. Devuelve solo las
    celdas con al menos 1 pedido; el frontend completa el grid 7×24."""
    ids = await _branch_ids(db, current_user)
    return await service.sales_heatmap(db, start=start, end=end, branch_warehouse_ids=ids)


@router.get("/customers/{customer_id}/360", response_model=schemas.Customer360)
async def customer_360(customer_id: int, db: DB, _: CurrentUser):
    data = await service.customer_360(db, customer_id)
    if not data:
        raise HTTPException(404, "Cliente no encontrado")
    return data


@router.get("/customers/{customer_id}/pnl", response_model=schemas.CustomerPnLReport)
async def customer_pnl(
    customer_id: int, db: DB, _: CurrentUser,
    start: datetime, end: datetime,
):
    if end <= start:
        raise HTTPException(400, "El fin del periodo debe ser posterior al inicio")
    data = await service.customer_pnl_report(db, customer_id, start, end)
    if not data:
        raise HTTPException(404, "Cliente no encontrado")
    return data


@router.get("/export")
async def export_orders(
    db: DB, _: CurrentUser,
    formato: str = Query("csv", pattern="^(csv|xlsx)$"),
    kind: Optional[str] = None, status: Optional[str] = None,
    customer_id: Optional[int] = None, seller_id: Optional[int] = None,
    payment_method: Optional[str] = None, channel: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[datetime] = None, date_to: Optional[datetime] = None,
):
    filtros = dict(
        kind=kind, status=status, customer_id=customer_id, seller_id=seller_id,
        payment_method=payment_method, channel=channel, q=q,
        date_from=date_from, date_to=date_to,
    )
    if formato == "xlsx":
        contenido = await service.export_xlsx(db, **filtros)
        return StreamingResponse(
            io.BytesIO(contenido),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=ventas.xlsx"},
        )
    csv_text = await service.export_csv(db, **filtros)
    return StreamingResponse(
        io.StringIO(csv_text), media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=ventas.csv"},
    )


# ── CRUD / listing ────────────────────────────────────────────────────────────

@router.post("/", response_model=schemas.OrderDetail, status_code=201)
async def create_order(order_in: schemas.OrderCreate, db: DB, user: CurrentUser):
    order = await service.create_order(db, order_in, user_id=user.id)
    return await service.get_order_detail(db, order.id)


@router.get("/", response_model=schemas.PaginatedOrders)
async def read_orders(
    db: DB, current_user: CurrentUser,
    skip: int = Query(0, ge=0), limit: int = Query(50, ge=1, le=500),
    kind: Optional[str] = None, status: Optional[str] = None,
    customer_id: Optional[int] = None, seller_id: Optional[int] = None,
    payment_method: Optional[str] = None, channel: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[datetime] = None, date_to: Optional[datetime] = None,
    sort_by: str = Query("created_at"), sort_dir: str = Query("desc", pattern="^(asc|desc)$"),
    relationship_type: Optional[str] = None, client_type: Optional[str] = None,
):
    from app.modules.inventory.branch_scope import visible_warehouse_ids
    ids = await visible_warehouse_ids(db, current_user)
    items, total = await service.get_orders(
        db, skip=skip, limit=limit, kind=kind, status=status, customer_id=customer_id,
        seller_id=seller_id, payment_method=payment_method, channel=channel, q=q,
        date_from=date_from, date_to=date_to, sort_by=sort_by, sort_dir=sort_dir,
        branch_warehouse_ids=ids,
        relationship_type=relationship_type, client_type=client_type,
    )
    return schemas.PaginatedOrders(items=items, total=total, skip=skip, limit=limit)


@router.get("/{order_id}", response_model=schemas.OrderDetail)
async def read_order(order_id: int, db: DB, _: CurrentUser):
    order = await service.get_order_detail(db, order_id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    return order


@router.put("/{order_id}", response_model=schemas.OrderDetail)
async def update_order(order_id: int, data: schemas.OrderUpdate, db: DB, user: CurrentUser):
    try:
        order = await service.update_order(db, order_id, data, user_id=user.id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    return await service.get_order_detail(db, order_id)


@router.patch("/{order_id}/fiscal-data", response_model=schemas.OrderDetail)
async def update_fiscal_data(order_id: int, payload: schemas.FiscalDataPatch,
                              db: DB, user: CurrentUser):
    """Actualiza los datos fiscales del pedido (RFC, razon social, regimen,
    uso CFDI, CP) y opcionalmente los guarda tambien en la ficha del cliente
    para que no haga falta volver a capturarlos en futuras ventas."""
    from sqlalchemy import select as _sel
    from app.modules.customers import models as customer_models
    res = await db.execute(_sel(models.Order).where(models.Order.id == order_id)
                             .execution_options(skip_tenant_filter=True))
    order = res.scalars().first()
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    if order.cfdi_uuid:
        raise HTTPException(400, "No se puede editar: la factura ya esta timbrada.")
    # Aplicar al pedido
    if payload.rfc is not None:
        order.bill_rfc = (payload.rfc or "").upper().strip()
    if payload.name is not None:
        order.bill_name = payload.name
    if payload.regime is not None:
        order.bill_regime = payload.regime
    if payload.use is not None:
        order.bill_use = payload.use
    if payload.zip is not None:
        order.bill_zip = payload.zip
    # Propagar al cliente si se pidio
    if payload.save_to_customer and order.customer_id:
        r_cust = await db.execute(_sel(customer_models.Customer)
                                     .where(customer_models.Customer.id == order.customer_id)
                                     .execution_options(skip_tenant_filter=True))
        cust = r_cust.scalars().first()
        if cust:
            if payload.rfc: cust.rfc = payload.rfc.upper().strip()
            if payload.regime: cust.regimen_fiscal = payload.regime
            if payload.use: cust.uso_cfdi = payload.use
            if payload.zip: cust.codigo_postal = payload.zip
            # Nombre fiscal solo se guarda en cust si no rompe el existente
    await db.commit()
    return await service.get_order_detail(db, order_id)


@router.patch("/{order_id}/status", response_model=schemas.OrderDetail)
async def update_status(order_id: int, payload: schemas.StatusUpdate, db: DB, user: CurrentUser):
    order = await service.change_status(db, order_id, payload.status, payload.message, user_id=user.id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    return await service.get_order_detail(db, order_id)


@router.post("/{order_id}/payments", response_model=schemas.OrderDetail)
async def add_payment(order_id: int, pay: schemas.PaymentCreate, db: DB, user: CurrentUser):
    try:
        order = await service.register_payment(db, order_id, pay, user_id=user.id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    return await service.get_order_detail(db, order_id)


@router.post("/{order_id}/convert", response_model=schemas.OrderDetail)
async def convert_quote(order_id: int, db: DB, user: CurrentUser):
    try:
        order = await service.convert_quote_to_order(db, order_id, user_id=user.id)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not order:
        raise HTTPException(404, "Cotización no encontrada")
    return await service.get_order_detail(db, order.id)


@router.post("/{order_id}/cancel", response_model=schemas.OrderDetail)
async def cancel(order_id: int, db: DB, user: CurrentUser):
    order = await service.cancel_order(db, order_id, user_id=user.id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    return await service.get_order_detail(db, order_id)


@router.get("/{order_id}/invoice")
async def invoice_payload(order_id: int, db: DB, _: CurrentUser):
    order = await service.get_order_detail(db, order_id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    if not order.bill_rfc:
        raise HTTPException(400, "El pedido no tiene datos de facturación (RFC)")
    return service.build_invoice_payload(order)


@router.get("/{order_id}/batches")
async def get_order_batches(order_id: int, db: DB, _: CurrentUser):
    """Devuelve los lotes específicos que se despacharon en esta orden, con
    código y caducidad — para el panel 'Lotes despachados' en el OrderDrawer
    y para recall retroactivo. Solo aplica a productos perecederos."""
    from app.modules.inventory import batch_service
    order = await service.get_order_detail(db, order_id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    batches = await batch_service.get_batches_for_order(db, order_id)
    # Enriquecer con nombre del producto por comodidad del frontend
    variant_names: dict = {}
    for it in order.items:
        variant_names[it.variant_id] = it.product_name
    out = []
    for vid, rows in batches.items():
        for r in rows:
            out.append({
                "variant_id": vid,
                "product_name": variant_names.get(vid, "?"),
                **r,
            })
    return {"order_id": order_id, "batches": out}


@router.get("/{order_id}/ticket/text")
async def get_ticket_text(order_id: int, db: DB, _: CurrentUser):
    """Devuelve el ticket como texto plano listo para WhatsApp (con emojis,
    *bold* y saldo). Pensado para pre-llenar wa.me?text=... desde el POS."""
    from app.modules.sales import ticket as ticket_mod
    from app.modules.core_config.service import get_company_profile
    from app.modules.inventory import batch_service

    order = await service.get_order_detail(db, order_id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")
    company = await get_company_profile(db)
    # Caducidades por variant (imprime "Cad: dd/mm/yyyy · Lote X" bajo cada
    # perecedero — requisito legal en varios países para alimentos y farma)
    batches_by_variant = await batch_service.get_batches_for_order(db, order_id)
    text = ticket_mod.render_ticket_text(order, company, batches_by_variant)
    return {
        "text": text,
        "phone": order.customer.phone if order.customer else None,
        "customer_name": order.customer.name if order.customer else None,
    }


@router.post("/{order_id}/ticket/email", response_model=schemas.TicketEmailResult)
async def send_ticket_email(order_id: int, payload: schemas.TicketSendRequest, db: DB, _: CurrentUser):
    """Envía el ticket de la orden por correo (usa Resend/plataforma o SMTP).
    Adjunta el ticket térmico en PDF y renderiza el cuerpo HTML con logo
    embebido en base64 — así se ve el ticket completo en cualquier cliente
    de correo aunque las imágenes remotas estén bloqueadas."""
    from app.modules.sales import ticket as ticket_mod
    from app.modules.core_config.service import get_company_profile
    from app.core.email import send_email

    order = await service.get_order_detail(db, order_id)
    if not order:
        raise HTTPException(404, "Pedido no encontrado")

    recipient = (payload.to or "").strip() or (order.customer.email if order.customer else "")
    if not recipient:
        return schemas.TicketEmailResult(sent=False, reason="El cliente no tiene correo registrado")

    company = await get_company_profile(db)
    html = ticket_mod.render_ticket_html(order, company)
    biz = (company.legal_name if company and company.legal_name else "Sthenova")
    folio_txt = order.folio or f"#{order.id}"

    # Clasificar cliente para elegir el PDF adjunto correcto:
    #   • B2B (distribuidor, mayorista, retail, marketplace) → remisión
    #     formal tamaño carta con datos del comprador y del emisor. Se
    #     entrega junto con la mercancía; la factura CFDI se emite después
    #     via Facturama u otro PAC.
    #   • POS / mostrador → ticket térmico 80mm.
    tone = ticket_mod._classify_customer_tone(order)
    is_b2b = tone == "b2b"
    subject = (
        f"Remisión {folio_txt} — {biz}" if is_b2b
        else f"Ticket {folio_txt} — {biz}"
    )

    attachments = []
    # 1) B2B → intenta la remisión formal primero
    if is_b2b:
        try:
            from app.modules.sales import universal_service
            pdf_bytes = await universal_service.generate_document_pdf(db, order_id, "remission")
            if pdf_bytes:
                attachments = [(f"remision_{order.folio or order.id}.pdf", pdf_bytes, "pdf")]
        except Exception as e:
            print(f"[b2b-email] no se pudo generar remisión, intento ticket como respaldo: {e}")

    # 2) POS o fallback si la remisión falló → ticket térmico 80mm
    if not attachments:
        try:
            from app.modules.pos import service as pos_service, pdf_ticket
            from app.modules.inventory import batch_service as _bs
            data = await pos_service.prepare_ticket_data(db, order_id)
            if data:
                batches_by_variant = await _bs.get_batches_for_order(db, order_id)
                pdf_bytes = pdf_ticket.build_thermal_ticket(
                    company=data["company"], order=data["order"],
                    items=data["items"], payments=data["payments"],
                    session=data["session"], width_mm=80,
                    batches_by_variant=batches_by_variant,
                )
                fname = (f"remision_{order.folio or order.id}.pdf" if is_b2b
                         else f"ticket_{order.folio or order.id}.pdf")
                attachments = [(fname, pdf_bytes, "pdf")]
        except Exception as e:
            # El HTML del correo ya trae el resumen, así que si el PDF
            # falla no rompemos el envío — solo se pierde el adjunto.
            print(f"[ticket-email] no se pudo generar PDF adjunto: {e}")

    # 3) CFDI 4.0 — si la venta ya esta timbrada, adjuntar PDF+XML.
    # Practica estandar en MX: el cliente y su contador necesitan ambos.
    if getattr(order, "cfdi_uuid", None):
        base = f"CFDI_{order.cfdi_serie or 'F'}-{order.cfdi_folio or order.id}_{(order.cfdi_uuid or '')[:8]}"
        if getattr(order, "cfdi_pdf", None):
            attachments = [*attachments, (f"{base}.pdf", bytes(order.cfdi_pdf), "pdf")]
        if getattr(order, "cfdi_xml", None):
            attachments = [*attachments, (f"{base}.xml", bytes(order.cfdi_xml), "xml")]

    # Logo inline para que se vea en el body del correo (Gmail/Outlook
    # bloquean data:base64 en <img>, pero CID inline se renderiza sin problema).
    logo_att = ticket_mod.logo_inline_attachment(company)
    if logo_att:
        attachments = [*attachments, logo_att]

    ok = await send_email(db, to=recipient, subject=subject, body_html=html,
                          attachments=attachments or None)
    if not ok:
        return schemas.TicketEmailResult(sent=False, to=recipient,
                                         reason="No se pudo enviar el correo (revisa la configuración de correo)")
    return schemas.TicketEmailResult(sent=True, to=recipient)


# ── Universal ERP: PDFs, importadores marketplace, P&L cliente ─────────────
from fastapi import UploadFile, File, Form
from fastapi.responses import Response
from app.modules.sales import universal_service


@router.get("/{order_id}/document/{kind}.pdf")
async def download_document_pdf(order_id: int, kind: str, db: DB, _: CurrentUser):
    """Descarga PDF del documento asociado a la orden.
    kind ∈ {quote, remission, proforma}. Se genera con logo, colores y datos
    de la empresa (CompanyProfile)."""
    if kind not in ("quote", "remission", "proforma"):
        raise HTTPException(400, "kind debe ser quote | remission | proforma")
    pdf_bytes = await universal_service.generate_document_pdf(db, order_id, kind)
    if not pdf_bytes:
        raise HTTPException(404, "Pedido no encontrado")
    filenames = {"quote": "cotizacion", "remission": "remision", "proforma": "pre_factura"}
    fname = f"{filenames[kind]}_{order_id}.pdf"
    return Response(content=pdf_bytes, media_type="application/pdf",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


@router.get("/marketplace/parsers")
async def list_parsers(_: CurrentUser):
    """Lista de plataformas soportadas por el importador de reportes."""
    return {"parsers": list(universal_service.PARSERS.keys()) + ["custom"]}


@router.post("/marketplace/import")
async def import_marketplace(
    db: DB, current_user: CurrentUser,
    customer_id: int = Form(...),
    platform: str = Form(...),
    file: UploadFile = File(...),
    mapping_json: Optional[str] = Form(None),
):
    """Sube un XLSX de reporte marketplace (Liverpool, Amazon, etc.) y crea
    órdenes + devoluciones automáticamente. Idempotente por external_order_id.
    """
    import json as _json
    contents = await file.read()
    mapping = _json.loads(mapping_json) if mapping_json else None
    try:
        result = await universal_service.import_marketplace_report(
            db, customer_id=customer_id, platform=platform,
            file_bytes=contents, filename=file.filename or "reporte.xlsx",
            mapping=mapping, user_id=current_user.id,
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    return result


@router.get("/customers/{customer_id}/pnl-universal")
async def customer_pnl_universal(customer_id: int, db: DB, _: CurrentUser,
                                  start: Optional[datetime] = Query(default=None),
                                  end: Optional[datetime] = Query(default=None)):
    """Estado de resultados por cliente con desglose completo (Universal ERP):
    Venta bruta − comisiones − logísticos − CEDIS − portal − descuentos −
    devoluciones − retenciones ISR/IVA − COGS = Margen bruto. Usa la config
    comercial del cliente (relationship_type, commission_base_pct, etc.)."""
    return await universal_service.compute_customer_pnl(db, customer_id, start=start, end=end)


@router.get("/customers/{customer_id}/settlement")
async def marketplace_settlement(customer_id: int, db: DB, _: CurrentUser,
                                 start: Optional[datetime] = Query(default=None),
                                 end: Optional[datetime] = Query(default=None),
                                 deposited_amount: Optional[float] = Query(default=None)):
    """Reconciliación de liquidación marketplace. Compara lo depositado por
    la plataforma (Liverpool, Amazon, ML) contra lo esperado según órdenes
    − devoluciones. Detecta variance para reclamaciones."""
    from app.modules.sales.marketplace_settlement import compute_settlement
    return await compute_settlement(db, customer_id, start=start, end=end,
                                     deposited_amount=deposited_amount)


@router.post("/returns/{return_id}/receive")
async def receive_return_endpoint(return_id: int, payload: dict, db: DB, current_user: CurrentUser):
    """Recibe físicamente la devolución en almacén y marca condition
    (sellable/damaged) por cada partida."""
    warehouse_id = payload.get("warehouse_id")
    items_condition = payload.get("items_condition", {})
    if not warehouse_id:
        raise HTTPException(400, "warehouse_id requerido")
    # Convertir keys a int
    items_condition = {int(k): v for k, v in items_condition.items()}
    result = await universal_service.receive_return(
        db, return_id=return_id, warehouse_id=warehouse_id,
        items_condition=items_condition, notes=payload.get("notes"),
        user_id=current_user.id,
    )
    if not result:
        raise HTTPException(404, "Devolución no encontrada")
    return result


# ── #7 · Export XLSX de ventas con filtros ────────────────────────────

@router.get("/export.xlsx")
async def export_orders_xlsx(
    db: DB, current_user: CurrentUser,
    kind: Optional[str] = None,
    status: Optional[str] = None,
    customer_id: Optional[int] = None,
    seller_id: Optional[int] = None,
    payment_method: Optional[str] = None,
    channel: Optional[str] = None,
    q: Optional[str] = None,
    date_from: Optional[datetime] = None,
    date_to: Optional[datetime] = None,
):
    """XLSX detallado (1 fila por partida) de ventas con los mismos
    filtros que la tabla del CRM: cliente, producto, fecha, canal,
    forma de pago, vendedor, estado, tipo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    orders, _ = await service.get_orders(
        db, skip=0, limit=10000,
        kind=kind, status=status, customer_id=customer_id,
        seller_id=seller_id, payment_method=payment_method,
        channel=channel, q=q, date_from=date_from, date_to=date_to,
    )

    wb = Workbook(); ws = wb.active
    ws.title = "Ventas"
    header = [
        "Folio", "Fecha", "Estado", "Tipo", "Cliente", "RFC", "Vendedor",
        "Canal", "Forma pago", "SKU", "Producto", "Cantidad",
        "Precio unit.", "Descuento", "Subtotal partida",
        "Total orden", "Pagado", "Saldo", "Notas",
    ]
    brand = "33B2F5"
    for i, h in enumerate(header, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = PatternFill("solid", fgColor=brand)
        c.font = Font(bold=True, color="FFFFFF")

    row = 2
    for o in orders:
        cliente = getattr(o.customer, "name", None) if getattr(o, "customer", None) else None
        rfc = getattr(o.customer, "rfc", None) if getattr(o, "customer", None) else None
        vendedor = getattr(o.user, "email", None) if getattr(o, "user", None) else None
        fecha = o.created_at.strftime("%Y-%m-%d %H:%M") if o.created_at else ""
        items = list(getattr(o, "items", []) or [])
        if not items:
            items = [None]
        for it in items:
            vals = [
                o.folio or f"ORD-{o.id}", fecha, o.status or "", o.kind or "",
                cliente or "—", rfc or "", vendedor or "",
                o.channel or "", o.payment_method or "",
                (it.sku if it else "") or "",
                (it.product_name if it else "") or "",
                float(it.quantity or 0) if it else 0,
                float(it.unit_price or 0) if it else 0,
                float(it.discount_amount or 0) if it else 0,
                float(it.total or 0) if it else 0,
                float(o.total or 0), float(o.paid or 0),
                float((o.total or 0) - (o.paid or 0)),
                (o.notes or "") if it is items[0] else "",
            ]
            for i, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=i, value=v)
                if i >= 12 and isinstance(v, (int, float)):
                    cell.number_format = "#,##0.00"
                    cell.alignment = Alignment(horizontal="right")
            row += 1

    widths = [14, 17, 12, 12, 32, 15, 22, 14, 14, 16, 34, 10, 12, 12, 14, 14, 12, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    fname = f"ventas_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
