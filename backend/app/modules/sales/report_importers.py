"""Importadores de reportes de ventas de marketplaces y cadenas.

Diseño:
  - Cada plataforma tiene su propio parser (ParserLiverpool, ParserAmazon, ...)
    que hereda de MarketplaceParser y traduce las columnas del reporte a un
    esquema normalizado.
  - El core del importador es agnóstico a la plataforma: toma filas
    normalizadas, crea/actualiza órdenes y devoluciones automáticamente.
  - Idempotente: usar external_order_id como clave. Si ya existe, se actualiza.

Esquema normalizado:
  external_order_id: str          - identificador único del pedido en la plataforma
  created_at: datetime            - fecha de creación
  sku: str                        - SKU del producto (nuestro o de la plataforma)
  product_name: str
  brand: str | None
  category: str | None            - categoría de producto (para % comisión por categoría)
  quantity: int
  unit_price: float               - PVP en la plataforma
  subtotal: float
  commission_amount: float        - comisión que se queda la plataforma (si viene desglosada)
  net_to_seller: float            - lo que recibe el seller neto
  channel: str                    - "marketplace" | "chain_physical"
  fulfillment: str | None         - "seller" | "platform"
  return_partial: bool            - hay devolución parcial
  return_total: bool              - hay devolución total
  returned_qty: int               - piezas devueltas
  delivery_status: str | None     - estado logístico
"""
from __future__ import annotations
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional, List, Dict, Any
import io
import json


@dataclass
class NormalizedRow:
    external_order_id: str
    created_at: Optional[datetime] = None
    sku: str = ""
    product_name: str = ""
    brand: Optional[str] = None
    category: Optional[str] = None
    quantity: int = 1
    unit_price: float = 0.0
    subtotal: float = 0.0
    commission_amount: float = 0.0
    net_to_seller: float = 0.0
    channel: str = "marketplace"
    fulfillment: Optional[str] = None
    return_partial: bool = False
    return_total: bool = False
    returned_qty: int = 0
    delivery_status: Optional[str] = None
    raw_row: Dict[str, Any] = field(default_factory=dict)


@dataclass
class ImportResult:
    rows_read: int = 0
    orders_created: int = 0
    orders_updated: int = 0
    returns_created: int = 0
    errors: List[Dict[str, Any]] = field(default_factory=list)


class MarketplaceParser:
    """Interfaz base. Cada parser implementa parse(file_bytes) → List[NormalizedRow]."""
    platform: str = "unknown"

    def parse(self, file_bytes: bytes) -> List[NormalizedRow]:
        raise NotImplementedError


class LiverpoolParser(MarketplaceParser):
    """Reporte oficial de Liverpool Marketplace (OrderReport.xlsx).

    Columnas del reporte:
      - "Id del pedido" → external_order_id
      - "Fecha de creación" → created_at
      - "SKU de seller" ó "No de SKU Liverpool" → sku
      - "Nombre del producto"
      - "Marca"
      - "Tipo de artículo" → category (Soft Line = ropa, Hard Line = electrónica, etc.)
      - "Cantidad"
      - "Precio por unidad"
      - "Subtotal"
      - "Total de la orden en pesos"
      - "Total de la orden a pagar al seller" → net_to_seller
      - "Reembolso parcial" / "Reembolso total" → flags
      - "Piezas devueltas por indicador"
      - "Fulfillment (Si/No)"
      - "Estado" → delivery_status
    """
    platform = "liverpool"

    def parse(self, file_bytes: bytes) -> List[NormalizedRow]:
        import openpyxl
        # NO usar read_only: en algunos reportes (Liverpool) el atributo dimension
        # del XML dice max_row=1 y openpyxl termina la iteración antes de tiempo.
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows: List[NormalizedRow] = []
        header: List[str] = []
        for i, r in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                header = [str(c or "").strip() for c in r]
                continue
            if not r or all(c is None for c in r):
                continue
            row_dict = {header[j]: r[j] for j in range(min(len(header), len(r)))}
            try:
                rows.append(self._normalize(row_dict))
            except Exception as e:
                rows.append(NormalizedRow(
                    external_order_id="ERROR",
                    raw_row={"error": str(e), "row": row_dict, "row_index": i},
                ))
        return rows

    @staticmethod
    def _to_float(v) -> float:
        if v is None or v == "":
            return 0.0
        try:
            return float(str(v).replace(",", "").replace("$", "").strip() or 0.0)
        except Exception:
            return 0.0

    @staticmethod
    def _to_int(v) -> int:
        try:
            return int(float(str(v).replace(",", "").strip() or 0))
        except Exception:
            return 0

    @staticmethod
    def _to_bool(v) -> bool:
        s = str(v or "").strip().lower()
        return s in ("si", "sí", "yes", "true", "1")

    @staticmethod
    def _parse_date(v) -> Optional[datetime]:
        if not v:
            return None
        if isinstance(v, datetime):
            return v
        # Liverpool usa formato "01/01/2026 - 12:21 AM"
        s = str(v).strip()
        for fmt in ("%d/%m/%Y - %I:%M %p", "%d/%m/%Y %I:%M %p", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    def _normalize(self, row: Dict[str, Any]) -> NormalizedRow:
        subtotal = self._to_float(row.get("Subtotal"))
        net_to_seller = self._to_float(row.get("Total de la orden a pagar al seller"))
        # Si Liverpool no puso comisión desglosada, la calculamos como diferencia
        commission_amount = self._to_float(row.get("Valor de la comisión (monto)"))
        if commission_amount == 0.0 and subtotal > 0 and net_to_seller > 0 and net_to_seller < subtotal:
            commission_amount = round(subtotal - net_to_seller, 2)

        returned_qty = self._to_int(row.get("Piezas devueltas por indicador"))
        return NormalizedRow(
            external_order_id=str(row.get("Id del pedido") or "").strip(),
            created_at=self._parse_date(row.get("Fecha de creación")),
            sku=str(row.get("SKU de seller") or row.get("No de SKU Liverpool") or row.get("No de SKU") or "").strip(),
            product_name=str(row.get("Nombre del producto") or "").strip(),
            brand=str(row.get("Marca") or "").strip() or None,
            category=str(row.get("Tipo de artículo") or "").strip() or None,
            quantity=self._to_int(row.get("Cantidad")) or 1,
            unit_price=self._to_float(row.get("Precio por unidad")),
            subtotal=subtotal,
            commission_amount=commission_amount,
            net_to_seller=net_to_seller if net_to_seller > 0 else subtotal,
            channel="marketplace",
            fulfillment="platform" if self._to_bool(row.get("Fulfillment (Si/No)")) else "seller",
            return_partial=self._to_bool(row.get("Reembolso parcial")),
            return_total=self._to_bool(row.get("Reembolso total")),
            returned_qty=returned_qty,
            delivery_status=str(row.get("Estado") or "").strip() or None,
            raw_row=row,
        )


class GenericMarketplaceParser(MarketplaceParser):
    """Parser configurable por mapeo. Se usa cuando el usuario define su propia
    plantilla de columnas. Recibe `mapping: Dict[campo_normalizado, columna_reporte]`.
    """
    platform = "custom"

    def __init__(self, mapping: Dict[str, str]):
        self.mapping = mapping

    def parse(self, file_bytes: bytes) -> List[NormalizedRow]:
        import openpyxl
        # NO usar read_only: en algunos reportes (Liverpool) el atributo dimension
        # del XML dice max_row=1 y openpyxl termina la iteración antes de tiempo.
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows: List[NormalizedRow] = []
        header: List[str] = []
        for i, r in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                header = [str(c or "").strip() for c in r]
                continue
            if not r or all(c is None for c in r):
                continue
            row_dict = {header[j]: r[j] for j in range(min(len(header), len(r)))}
            get = lambda field: row_dict.get(self.mapping.get(field, ""))
            rows.append(NormalizedRow(
                external_order_id=str(get("external_order_id") or "").strip(),
                sku=str(get("sku") or "").strip(),
                product_name=str(get("product_name") or "").strip(),
                quantity=LiverpoolParser._to_int(get("quantity")) or 1,
                unit_price=LiverpoolParser._to_float(get("unit_price")),
                subtotal=LiverpoolParser._to_float(get("subtotal")),
                commission_amount=LiverpoolParser._to_float(get("commission_amount")),
                net_to_seller=LiverpoolParser._to_float(get("net_to_seller")),
                raw_row=row_dict,
            ))
        return rows


class ChainSellThroughParser(MarketplaceParser):
    """Sell-through de cadenas físicas (Sears, Chedraui, Costco, Walmart, Sanborns…).

    En consignación/distribución en cadena, el fabricante embarca stock y sólo
    reconoce ingreso cuando la cadena vende al público. Este parser toma el
    reporte semanal de venta a piso y crea una "orden" sintética por línea
    (SKU × tienda × fecha) para poder liquidar y conciliar.

    Columnas esperadas (flexibles, se acepta cualquier variante en español):
      - "SKU" / "Modelo" / "Código"           → sku
      - "Producto" / "Descripción"            → product_name
      - "Tienda" / "Sucursal" / "Store"       → store (va a delivery_status)
      - "Fecha" / "Semana" / "Periodo"        → created_at
      - "Unidades" / "Cantidad" / "Piezas"    → quantity
      - "Precio" / "Precio unitario" / "PVP"  → unit_price
      - "Total" / "Importe" / "Subtotal"      → subtotal
      - "Comisión" / "Fee" / "% cadena"       → commission_amount (opcional)
      - "Neto a pagar" / "Neto seller"        → net_to_seller (opcional)
      - "Devueltas" / "Piezas devueltas"      → returned_qty (opcional)
    """
    platform = "chain_sellthrough"

    # Mapeo de sinónimos → campo canónico (todo en minúsculas para comparar)
    COLUMN_ALIASES = {
        "sku": ("sku", "modelo", "codigo", "código", "clave", "cod", "no de sku"),
        "product_name": ("producto", "descripcion", "descripción", "nombre", "nombre del producto", "articulo", "artículo"),
        "store": ("tienda", "sucursal", "store", "punto de venta", "pdv"),
        "date": ("fecha", "semana", "periodo", "período", "week", "fecha venta"),
        "quantity": ("unidades", "cantidad", "piezas", "qty", "units", "vendido"),
        "unit_price": ("precio", "precio unitario", "pvp", "precio publico", "precio público"),
        "subtotal": ("total", "importe", "subtotal", "venta", "venta total"),
        "commission": ("comision", "comisión", "fee", "cadena fee", "% cadena", "descuento cadena"),
        "net_to_seller": ("neto a pagar", "neto seller", "neto al seller", "pago al proveedor", "pago proveedor"),
        "returned_qty": ("devueltas", "piezas devueltas", "returned", "devoluciones"),
    }

    def parse(self, file_bytes: bytes) -> List[NormalizedRow]:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows: List[NormalizedRow] = []
        header: List[str] = []
        col_map: Dict[str, Optional[int]] = {}
        for i, r in enumerate(ws.iter_rows(values_only=True), 1):
            if i == 1:
                header = [str(c or "").strip() for c in r]
                col_map = self._map_columns(header)
                continue
            if not r or all(c is None for c in r):
                continue
            try:
                rows.append(self._normalize(r, col_map))
            except Exception as e:
                rows.append(NormalizedRow(
                    external_order_id="ERROR",
                    raw_row={"error": str(e), "row": list(r), "row_index": i},
                ))
        return rows

    def _map_columns(self, header: List[str]) -> Dict[str, Optional[int]]:
        lc = [h.lower().strip() for h in header]
        out: Dict[str, Optional[int]] = {}
        for canonical, aliases in self.COLUMN_ALIASES.items():
            found = None
            for alias in aliases:
                if alias in lc:
                    found = lc.index(alias)
                    break
            out[canonical] = found
        return out

    def _normalize(self, row: tuple, col_map: Dict[str, Optional[int]]) -> NormalizedRow:
        def col(name: str):
            idx = col_map.get(name)
            if idx is None or idx >= len(row):
                return None
            return row[idx]

        sku = str(col("sku") or "").strip()
        store = str(col("store") or "").strip() or "N/D"
        date_raw = col("date")
        created_at = LiverpoolParser._parse_date(date_raw)
        # ID sintético: idempotente por (SKU, tienda, fecha)
        date_key = created_at.strftime("%Y%m%d") if created_at else str(date_raw or "sin_fecha").strip().replace(" ", "")
        store_key = store.replace(" ", "_")[:20]
        external_id = f"SELLTHRU-{sku}-{store_key}-{date_key}"

        quantity = LiverpoolParser._to_int(col("quantity")) or 1
        unit_price = LiverpoolParser._to_float(col("unit_price"))
        subtotal = LiverpoolParser._to_float(col("subtotal"))
        if subtotal == 0.0 and unit_price > 0 and quantity > 0:
            subtotal = round(unit_price * quantity, 2)
        commission_amount = LiverpoolParser._to_float(col("commission"))
        net_to_seller = LiverpoolParser._to_float(col("net_to_seller"))
        if net_to_seller == 0.0 and subtotal > 0:
            net_to_seller = round(subtotal - commission_amount, 2) if commission_amount > 0 else subtotal
        if commission_amount == 0.0 and subtotal > 0 and net_to_seller > 0 and net_to_seller < subtotal:
            commission_amount = round(subtotal - net_to_seller, 2)

        returned_qty = LiverpoolParser._to_int(col("returned_qty"))

        raw = {header_key: row[idx] for header_key, idx in [
            ("sku", col_map.get("sku") or -1),
            ("store", col_map.get("store") or -1),
        ] if idx is not None and idx >= 0 and idx < len(row)}
        raw["store"] = store
        raw["date_raw"] = date_raw

        return NormalizedRow(
            external_order_id=external_id,
            created_at=created_at,
            sku=sku,
            product_name=str(col("product_name") or "").strip() or f"SKU {sku}",
            quantity=quantity,
            unit_price=unit_price,
            subtotal=subtotal,
            commission_amount=commission_amount,
            net_to_seller=net_to_seller if net_to_seller > 0 else subtotal,
            channel="chain_sellthrough",
            fulfillment="platform",
            return_partial=returned_qty > 0,
            returned_qty=returned_qty,
            delivery_status=f"Tienda: {store}",
            raw_row=raw,
        )


PARSERS = {
    "liverpool": LiverpoolParser,
    "chain_sellthrough": ChainSellThroughParser,
    "sears": ChainSellThroughParser,
    "chedraui": ChainSellThroughParser,
    "costco": ChainSellThroughParser,
    "walmart": ChainSellThroughParser,
    "sanborns": ChainSellThroughParser,
    # extender: "amazon": AmazonParser, "mercadolibre": MLParser, "shopify": ShopifyParser
}


def get_parser(platform: str, mapping: Optional[Dict[str, str]] = None) -> MarketplaceParser:
    if platform in PARSERS:
        return PARSERS[platform]()
    if mapping:
        return GenericMarketplaceParser(mapping)
    raise ValueError(f"Plataforma '{platform}' no soportada y no se dio mapping personalizado.")
