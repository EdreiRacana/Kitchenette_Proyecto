from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.future import select
from sqlalchemy.orm import selectinload
from typing import List, Optional
from datetime import datetime, timezone
from io import BytesIO
import logging
import os
import uuid
import pandas as pd

log = logging.getLogger(__name__)
from app.modules.inventory.models import (
    Product, ProductVariant, Warehouse, StockLevel, StockMovement, StockMovementType,
    Supplier, SupplierDocument, StockLot, PurchaseOrder, PurchaseOrderItem, PurchaseOrderStatus,
    Recipe, RecipeItem, ProductionOrder, ProductionOrderStatus,
)
from app.modules.inventory import schemas

# Umbral por default para "stock bajo" cuando una variante no tiene
# reorder_point ni safety_stock configurados. Sirve para que el KPI del
# Tablero alerte útilmente en la etapa de setup del cliente, antes de
# haber terminado de configurar cada SKU. Valor sensible para catálogos
# típicos de distribución/venta minorista.
DEFAULT_LOW_STOCK_THRESHOLD = 10

# --- Product Services ---
async def create_product(db: AsyncSession, product_in: schemas.ProductCreate) -> Product:
    db_product = Product(**product_in.model_dump())
    db.add(db_product)
    await db.commit()
    await db.refresh(db_product)
    return db_product

async def get_products(db: AsyncSession, skip: int = 0, limit: int = 100, item_type: Optional[str] = None) -> List[Product]:
    query = select(Product).offset(skip).limit(limit).options(
        selectinload(Product.variants).selectinload(ProductVariant.stock_levels).selectinload(StockLevel.warehouse),
        selectinload(Product.media),
    )
    if item_type:
        query = query.where(Product.item_type == item_type)
    result = await db.execute(query)
    return result.scalars().all()


EXPORT_HEADERS = ["SKU", "Codigo de barras", "Producto", "Tipo", "Categoria",
                  "Talla", "Color", "Material", "Almacen", "Stock disponible",
                  "Stock reservado", "Precio", "Costo"]


async def _export_rows(db: AsyncSession, warehouse_id: Optional[int] = None) -> List[list]:
    query = select(Product).options(
        selectinload(Product.variants).selectinload(ProductVariant.stock_levels).selectinload(StockLevel.warehouse),
    )
    result = await db.execute(query)
    products = result.scalars().unique().all()

    rows: List[list] = []
    for p in products:
        for v in p.variants:
            levels = v.stock_levels
            if warehouse_id:
                levels = [sl for sl in levels if sl.warehouse_id == warehouse_id]
            if not levels:
                if warehouse_id:
                    continue
                rows.append([
                    v.sku, v.barcode or "", p.name, p.item_type, p.category or "",
                    v.size or "", v.color or "", v.material or "", "—", 0, 0,
                    v.price, v.cost_price or 0,
                ])
                continue
            for sl in levels:
                rows.append([
                    v.sku, v.barcode or "", p.name, p.item_type, p.category or "",
                    v.size or "", v.color or "", v.material or "",
                    sl.warehouse.name if sl.warehouse else "—",
                    sl.quantity, sl.reserved_quantity,
                    v.price, v.cost_price or 0,
                ])
    return rows


async def export_inventory_csv(db: AsyncSession, warehouse_id: Optional[int] = None) -> str:
    import csv, io
    rows = await _export_rows(db, warehouse_id)
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(EXPORT_HEADERS)
    w.writerows(rows)
    return "﻿" + buf.getvalue()


async def export_inventory_xlsx(db: AsyncSession, warehouse_id: Optional[int] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    rows = await _export_rows(db, warehouse_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    ws.append(EXPORT_HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E2E5C", end_color="1E2E5C", fill_type="solid")
    money_cols = {12, 13}
    for row in rows:
        ws.append(row)
        for col_idx in money_cols:
            ws.cell(row=ws.max_row, column=col_idx).number_format = "#,##0.00"
    for col_idx, header in enumerate(EXPORT_HEADERS, start=1):
        max_len = max([len(str(header))] + [len(str(r[col_idx - 1])) for r in rows] + [8])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def save_compressed_image(content: bytes, filename: str, url_prefix: str) -> str:
    """Comprime/convierte la imagen a WebP y la sube al almacenamiento configurado."""
    from PIL import Image
    from app.core.storage import upload_bytes
    img = Image.open(BytesIO(content))
    img = img.convert("RGB")
    img.thumbnail((1200, 1200))
    buf = BytesIO()
    img.save(buf, "WEBP", quality=80)
    return await upload_bytes(buf.getvalue(), "imagen.webp", folder=url_prefix)

async def get_product(db: AsyncSession, product_id: int) -> Optional[Product]:
    result = await db.execute(
        select(Product).where(Product.id == product_id).options(
            selectinload(Product.variants).selectinload(ProductVariant.stock_levels).selectinload(StockLevel.warehouse),
            selectinload(Product.media),
        )
    )
    return result.scalars().first()

async def update_product(db: AsyncSession, product_id: int, product_in: schemas.ProductUpdate) -> Optional[Product]:
    product = await db.get(Product, product_id)
    if not product:
        return None
    for k, v in product_in.model_dump(exclude_unset=True).items():
        setattr(product, k, v)
    await db.commit()
    await db.refresh(product)
    return product

# --- Variant Services ---
async def create_variant(db: AsyncSession, variant_in: schemas.VariantCreate) -> ProductVariant:
    db_variant = ProductVariant(**variant_in.model_dump())
    db.add(db_variant)
    await db.commit()
    await db.refresh(db_variant)
    return db_variant

async def update_variant(db: AsyncSession, variant_id: int, variant_in: schemas.VariantUpdate) -> Optional[ProductVariant]:
    variant = await db.get(ProductVariant, variant_id)
    if not variant:
        return None
    for k, v in variant_in.model_dump(exclude_unset=True).items():
        setattr(variant, k, v)
    await db.commit()
    await db.refresh(variant)
    return variant

# --- Supplier Services ---
async def _load_supplier(db: AsyncSession, supplier_id: int) -> Optional[Supplier]:
    """Recarga el proveedor con `documents` eager-loaded — indispensable porque
    SupplierInDB serializa esa relación y en async SQLAlchemy tocar un lazy
    attribute revienta con 500 y hace pensar que la creación falló (aunque el
    commit ya haya ocurrido). Este helper garantiza response válido."""
    res = await db.execute(
        select(Supplier).where(Supplier.id == supplier_id)
        .options(selectinload(Supplier.documents))
    )
    return res.scalars().first()


async def create_supplier(db: AsyncSession, supplier_in: schemas.SupplierCreate) -> Supplier:
    db_supplier = Supplier(**supplier_in.model_dump())
    db.add(db_supplier)
    await db.commit()
    await db.refresh(db_supplier)
    reloaded = await _load_supplier(db, db_supplier.id)
    return reloaded or db_supplier

async def get_suppliers(db: AsyncSession) -> List[Supplier]:
    result = await db.execute(
        select(Supplier).order_by(Supplier.name).options(selectinload(Supplier.documents))
    )
    return result.scalars().all()

async def update_supplier(db: AsyncSession, supplier_id: int, supplier_in: schemas.SupplierUpdate) -> Optional[Supplier]:
    supplier = await db.get(Supplier, supplier_id)
    if not supplier:
        return None
    for k, v in supplier_in.model_dump(exclude_unset=True).items():
        setattr(supplier, k, v)
    await db.commit()
    reloaded = await _load_supplier(db, supplier_id)
    return reloaded or supplier

async def delete_supplier(db: AsyncSession, supplier_id: int) -> bool:
    supplier = await db.get(Supplier, supplier_id)
    if not supplier:
        return False
    existing_po = await db.execute(
        select(PurchaseOrder).where(PurchaseOrder.supplier_id == supplier_id).limit(1)
    )
    if existing_po.scalars().first():
        raise ValueError("No se puede eliminar: el proveedor tiene órdenes de compra asociadas. Desactívalo en su lugar.")
    await db.delete(supplier)
    await db.commit()
    return True

async def add_supplier_document(db: AsyncSession, supplier_id: int, doc_type: str, file_url: str, file_name: Optional[str]) -> Optional[SupplierDocument]:
    supplier = await db.get(Supplier, supplier_id)
    if not supplier:
        return None
    doc = SupplierDocument(supplier_id=supplier_id, doc_type=doc_type, file_url=file_url, file_name=file_name)
    db.add(doc)
    await db.commit()
    await db.refresh(doc)
    return doc

async def delete_supplier_document(db: AsyncSession, supplier_id: int, document_id: int) -> bool:
    doc = await db.get(SupplierDocument, document_id)
    if not doc or doc.supplier_id != supplier_id:
        return False
    await db.delete(doc)
    await db.commit()
    return True

# --- Warehouse Services ---
async def create_warehouse(db: AsyncSession, warehouse_in: schemas.WarehouseCreate) -> Warehouse:
    db_warehouse = Warehouse(**warehouse_in.model_dump())
    db.add(db_warehouse)
    await db.commit()
    await db.refresh(db_warehouse)
    return db_warehouse

async def get_warehouses(db: AsyncSession, warehouse_ids: Optional[List[int]] = None) -> List[Warehouse]:
    query = select(Warehouse)
    if warehouse_ids is not None:
        query = query.where(Warehouse.id.in_(warehouse_ids))
    result = await db.execute(query)
    return result.scalars().all()

async def update_warehouse(db: AsyncSession, warehouse_id: int, warehouse_in: schemas.WarehouseUpdate) -> Optional[Warehouse]:
    warehouse = await db.get(Warehouse, warehouse_id)
    if not warehouse:
        return None
    for k, v in warehouse_in.model_dump(exclude_unset=True).items():
        setattr(warehouse, k, v)
    await db.commit()
    await db.refresh(warehouse)
    return warehouse


# --- Costeo FIFO -------------------------------------------------------------
async def _create_lot(db: AsyncSession, variant_id: int, warehouse_id: int, quantity: int, unit_cost: float, reference: Optional[str]) -> StockLot:
    lot = StockLot(
        variant_id=variant_id, warehouse_id=warehouse_id,
        quantity_received=quantity, quantity_remaining=quantity,
        unit_cost=unit_cost, reference=reference,
    )
    db.add(lot)
    return lot

async def _consume_fifo(db: AsyncSession, variant_id: int, warehouse_id: int, quantity: int) -> float:
    """Consume `quantity` unidades de los lotes más antiguos disponibles en ese
    almacén y devuelve el costo promedio ponderado de lo consumido (para
    registrar en el movimiento de salida). Si no hay lotes suficientes, el
    remanente se valúa al último costo conocido del lote más reciente (o 0)."""
    result = await db.execute(
        select(StockLot)
        .where(StockLot.variant_id == variant_id, StockLot.warehouse_id == warehouse_id, StockLot.quantity_remaining > 0)
        .order_by(StockLot.received_at.asc())
    )
    lots = result.scalars().all()
    remaining = quantity
    total_cost = 0.0
    total_consumed = 0
    last_cost = 0.0
    for lot in lots:
        if remaining <= 0:
            break
        take = min(remaining, lot.quantity_remaining)
        lot.quantity_remaining -= take
        total_cost += take * lot.unit_cost
        total_consumed += take
        last_cost = lot.unit_cost
        remaining -= take
    if remaining > 0:
        # No había lotes suficientes (datos históricos sin lote, ajuste, etc.)
        total_cost += remaining * last_cost
        total_consumed += remaining
    return (total_cost / total_consumed) if total_consumed else 0.0

async def _recalc_weighted_avg_cost(db: AsyncSession, variant_id: int) -> None:
    result = await db.execute(
        select(StockLot).where(StockLot.variant_id == variant_id, StockLot.quantity_remaining > 0)
    )
    lots = result.scalars().all()
    total_qty = sum(l.quantity_remaining for l in lots)
    if total_qty <= 0:
        return
    total_cost = sum(l.quantity_remaining * l.unit_cost for l in lots)
    variant = await db.get(ProductVariant, variant_id)
    if variant:
        variant.cost_price = round(total_cost / total_qty, 4)


# --- Stock Services ---
async def adjust_stock(db: AsyncSession, movement_in: schemas.StockMovementCreate, user_id: Optional[int] = None) -> StockMovement:
    result = await db.execute(
        select(StockLevel).where(
            StockLevel.variant_id == movement_in.variant_id,
            StockLevel.warehouse_id == movement_in.warehouse_id
        )
    )
    stock_level = result.scalars().first()
    if not stock_level:
        stock_level = StockLevel(variant_id=movement_in.variant_id, warehouse_id=movement_in.warehouse_id, quantity=0)
        db.add(stock_level)

    unit_cost = movement_in.unit_cost

    if movement_in.movement_type == StockMovementType.IN:
        cost = unit_cost if unit_cost is not None else 0.0
        await _create_lot(db, movement_in.variant_id, movement_in.warehouse_id, movement_in.quantity, cost, movement_in.reference)
        stock_level.quantity += movement_in.quantity
        await _recalc_weighted_avg_cost(db, movement_in.variant_id)
    elif movement_in.movement_type == StockMovementType.OUT:
        unit_cost = await _consume_fifo(db, movement_in.variant_id, movement_in.warehouse_id, movement_in.quantity)
        stock_level.quantity -= movement_in.quantity
        await _recalc_weighted_avg_cost(db, movement_in.variant_id)
    elif movement_in.movement_type == StockMovementType.ADJUSTMENT:
        # Delta de corrección. Si es positivo se trata como una entrada a costo
        # actual (no genera lote nuevo); si es negativo, se descuenta del costo
        # promedio vigente sin tocar lotes (merma/ajuste de conteo).
        stock_level.quantity += movement_in.quantity
        if unit_cost is None:
            variant = await db.get(ProductVariant, movement_in.variant_id)
            unit_cost = variant.cost_price if variant else 0.0

    db_movement = StockMovement(
        variant_id=movement_in.variant_id,
        warehouse_id=movement_in.warehouse_id,
        quantity=movement_in.quantity if movement_in.movement_type != StockMovementType.OUT else -abs(movement_in.quantity),
        movement_type=movement_in.movement_type,
        unit_cost=unit_cost,
        reference=movement_in.reference,
        notes=movement_in.notes,
        user_id=user_id
    )
    db.add(db_movement)

    await db.commit()
    await db.refresh(db_movement)
    return db_movement

async def get_stock_levels(db: AsyncSession, variant_id: Optional[int] = None) -> List[StockLevel]:
    query = select(StockLevel).options(selectinload(StockLevel.warehouse), selectinload(StockLevel.variant))
    if variant_id:
        query = query.where(StockLevel.variant_id == variant_id)
    result = await db.execute(query)
    return result.scalars().all()

async def get_movements(db: AsyncSession, skip: int = 0, limit: int = 100, warehouse_ids: Optional[List[int]] = None) -> List[dict]:
    query = (
        select(StockMovement)
        .options(selectinload(StockMovement.variant).selectinload(ProductVariant.product), selectinload(StockMovement.warehouse))
        .order_by(StockMovement.created_at.desc())
    )
    if warehouse_ids is not None:
        query = query.where(StockMovement.warehouse_id.in_(warehouse_ids))
    result = await db.execute(query.offset(skip).limit(limit))
    movements = result.scalars().all()
    out = []
    for m in movements:
        out.append({
            "id": m.id, "variant_id": m.variant_id, "warehouse_id": m.warehouse_id,
            "quantity": m.quantity, "movement_type": m.movement_type, "unit_cost": m.unit_cost,
            "reference": m.reference, "notes": m.notes, "created_at": m.created_at,
            "product_name": m.variant.product.name if m.variant and m.variant.product else None,
            "sku": m.variant.sku if m.variant else None,
            "warehouse_name": m.warehouse.name if m.warehouse else None,
        })
    return out


# --- Reorder alerts -----------------------------------------------------------
async def get_reorder_alerts(db: AsyncSession, warehouse_ids: Optional[List[int]] = None) -> List[schemas.ReorderAlert]:
    query = (
        select(StockLevel)
        .options(
            selectinload(StockLevel.variant).selectinload(ProductVariant.product),
            selectinload(StockLevel.variant).selectinload(ProductVariant.preferred_supplier),
            selectinload(StockLevel.warehouse),
        )
    )
    if warehouse_ids is not None:
        query = query.where(StockLevel.warehouse_id.in_(warehouse_ids))
    result = await db.execute(query)
    levels = result.scalars().all()
    alerts: List[schemas.ReorderAlert] = []
    for lvl in levels:
        v = lvl.variant
        if not v:
            continue
        available = lvl.quantity - lvl.reserved_quantity
        # Umbral efectivo — misma regla que get_inventory_stats para que el
        # KPI del Tablero y esta lista siempre coincidan (si el Tablero dice
        # "3 stock bajo", esta lista tiene exactamente esos 3 SKUs).
        effective_reorder = v.reorder_point
        if effective_reorder is None:
            effective_reorder = v.safety_stock
        if effective_reorder is None:
            effective_reorder = DEFAULT_LOW_STOCK_THRESHOLD
        # Solo alertar si hay stock disponible y está bajo o al ras del umbral;
        # los agotados (available <= 0) aparecen en su propio KPI y no en esta
        # lista de "reorden inminente".
        if available <= 0 or available > effective_reorder:
            continue
        safety = v.safety_stock or 0
        level = "red" if available <= safety else "yellow"
        alerts.append(schemas.ReorderAlert(
            variant_id=v.id, sku=v.sku, product_name=v.product.name if v.product else "",
            warehouse_id=lvl.warehouse_id, warehouse_name=lvl.warehouse.name if lvl.warehouse else "",
            available=available, reserved=lvl.reserved_quantity,
            reorder_point=effective_reorder, safety_stock=safety, level=level,
            preferred_supplier_id=v.preferred_supplier_id,
            preferred_supplier_name=v.preferred_supplier.name if v.preferred_supplier else None,
            lead_time_days=v.lead_time_days,
        ))
    alerts.sort(key=lambda a: (a.level != "red", a.available))
    return alerts


async def get_inventory_stats(db: AsyncSession, warehouse_ids: Optional[List[int]] = None) -> schemas.InventoryStats:
    # Se parte de TODAS las variantes activas (no solo las que ya tienen un
    # registro de StockLevel) para que una variante nunca surtida también
    # cuente como agotada, igual que el cálculo a nivel producto del frontend.
    result = await db.execute(
        select(ProductVariant).where(ProductVariant.is_active == True).options(
            selectinload(ProductVariant.product),
            selectinload(ProductVariant.stock_levels),
        )
    )
    all_variants = result.scalars().all()

    by_category: dict[str, float] = {}
    total_value = 0.0
    total_units = 0
    out_of_stock = 0
    low_stock = 0

    for v in all_variants:
        levels = [lvl for lvl in (v.stock_levels or []) if warehouse_ids is None or lvl.warehouse_id in warehouse_ids]
        available = sum((lvl.quantity - lvl.reserved_quantity) for lvl in levels)
        unit_cost = v.cost_price if v.cost_price is not None else v.price
        value = available * unit_cost
        category = (v.product.category if v.product and v.product.category else "Sin categoría")

        by_category[category] = by_category.get(category, 0.0) + value
        total_value += value
        total_units += available

        if available <= 0:
            out_of_stock += 1
        else:
            # Umbral de stock bajo — regla profesional en 3 pasos:
            #  1. Si la variante tiene punto de reorden configurado, ese es
            #     el umbral autoritario (lo definió el usuario para este SKU).
            #  2. Si no, usar safety_stock (el "colchón mínimo seguro" es
            #     también un umbral de alerta razonable).
            #  3. Si tampoco, usar DEFAULT_LOW_STOCK_THRESHOLD como último
            #     recurso, para que el sistema alerte útilmente en la etapa
            #     de setup del cliente antes de que hayan configurado cada
            #     SKU. Si un cliente serio no quiere alertas por default,
            #     configura reorder_point=0 en el SKU correspondiente y
            #     nunca aparece como bajo (mientras haya al menos 1 unidad).
            threshold = v.reorder_point
            if threshold is None:
                threshold = v.safety_stock
            if threshold is None:
                threshold = DEFAULT_LOW_STOCK_THRESHOLD
            if available <= threshold:
                low_stock += 1

    cats = [
        schemas.CategoryValue(
            category=cat, value=round(val, 2),
            pct=round((val / total_value) * 100, 1) if total_value else 0.0,
        )
        for cat, val in by_category.items()
    ]
    cats.sort(key=lambda c: c.value, reverse=True)

    return schemas.InventoryStats(
        total_value=round(total_value, 2), total_units=total_units,
        out_of_stock=out_of_stock, low_stock=low_stock, by_category=cats,
    )


# --- Purchase Orders -----------------------------------------------------------
async def _next_folio(db: AsyncSession, model, prefix: str) -> str:
    result = await db.execute(select(model.id).order_by(model.id.desc()).limit(1))
    last_id = result.scalar()
    return f"{prefix}-{(last_id or 0) + 1:05d}"

_ALLOCATION_METHODS = ("by_value", "by_quantity")


def _validate_allocation(method: str) -> str:
    if method not in _ALLOCATION_METHODS:
        raise ValueError(f"landed_cost_allocation debe ser uno de: {', '.join(_ALLOCATION_METHODS)}")
    return method


def _extras_total(extras: Optional[List[dict]]) -> float:
    if not extras:
        return 0.0
    return round(sum(float((c or {}).get("amount", 0.0) or 0.0) for c in extras), 2)


def _compute_landed_unit_costs(
    items: List[dict], extras_total: float, method: str,
) -> List[float]:
    """Prorratea `extras_total` entre las partidas y devuelve el costo unitario
    integrado (landed) por partida, en el mismo orden que `items`.

    Cada partida debe tener {'quantity', 'unit_cost'}. Si no hay extras o
    todas las partidas suman cero (denominador cero), regresa el costo unitario
    original — sin prorrateo, sin dividir entre cero.

    - "by_value":    peso = quantity × unit_cost / total_value
    - "by_quantity": peso = quantity / total_quantity

    El extra prorrateado se divide entre `quantity` para volver a costo unit.
    Redondeo a 4 decimales (mismo que el kardex FIFO).

    **Cierre exacto (nivel contable):** después de prorratear, la suma de
    (qty × extra_prorrateado) puede desviarse por centavos del extras_total
    original por errores de coma flotante (típico cuando el peso es 1/3, 1/6,
    etc.). Para que el kardex jamás pierda o gane un centavo, la última
    partida elegible absorbe el residuo. Así el total prorrateado === total
    de extras, dígito por dígito."""
    if extras_total <= 0 or not items:
        return [round(float(it.get("unit_cost") or 0.0), 4) for it in items]

    method = _validate_allocation(method)
    if method == "by_value":
        total_value = sum(float(it.get("quantity") or 0) * float(it.get("unit_cost") or 0.0) for it in items)
    else:  # by_quantity
        total_value = sum(float(it.get("quantity") or 0) for it in items)

    if total_value <= 0:
        return [round(float(it.get("unit_cost") or 0.0), 4) for it in items]

    # Paso 1: reparto en dinero por partida a 2 decimales (moneda real).
    # Trabajamos siempre con dinero, no con costo unitario, para tener control
    # exacto sobre el cierre a nivel centavo.
    shares_money = []
    for it in items:
        qty = float(it.get("quantity") or 0)
        unit_cost = float(it.get("unit_cost") or 0.0)
        if qty <= 0:
            shares_money.append(0.0)
            continue
        line_weight = (qty * unit_cost) if method == "by_value" else qty
        shares_money.append(round(extras_total * (line_weight / total_value), 2))

    # Paso 2: cierre por residuo — la última partida elegible absorbe la
    # diferencia por redondeo para que sum(shares) === extras_total.
    diff = round(extras_total - sum(shares_money), 2)
    if abs(diff) >= 0.01:
        for idx in range(len(items) - 1, -1, -1):
            if float(items[idx].get("quantity") or 0) > 0:
                shares_money[idx] = round(shares_money[idx] + diff, 2)
                break

    # Paso 3: costo unitario integrado. Sin redondeo intermedio a 4 decimales
    # aquí — con qty grandes eso reintroduce el residuo. El StockLot.unit_cost
    # es Float (doble precisión), así que aguantamos la precisión completa.
    # La invariante contable que importa:
    #   sum_i(qty_i * (landed_i - unit_cost_i)) === extras_total
    # se mantiene exacta a nivel centavo porque partimos de shares_money (2 dec)
    # y unit_cost integrado = unit_cost + share_money / qty (division exacta
    # en float64 mientras qty sea entero, que es el caso — quantity es Integer).
    landed = []
    for it, share in zip(items, shares_money):
        qty = float(it.get("quantity") or 0)
        unit_cost = float(it.get("unit_cost") or 0.0)
        if qty <= 0:
            landed.append(unit_cost)
            continue
        landed.append(unit_cost + (share / qty))
    return landed


async def create_purchase_order(db: AsyncSession, po_in: schemas.PurchaseOrderCreate, user_id: Optional[int] = None) -> PurchaseOrder:
    folio = await _next_folio(db, PurchaseOrder, "OC")
    extras = [c.model_dump() for c in (po_in.extra_costs or [])]
    method = _validate_allocation(po_in.landed_cost_allocation or "by_value")
    # total_amount = SOLO mercancía. Es lo que el negocio le debe al proveedor
    # de la OC (la CxP se salda al pagarle su factura). Los extras (flete,
    # aduana, seguros) son deudas con terceros distintos — cada uno con su
    # propia factura y su propia CxP — así que NO entran en el total_amount
    # de esta OC. Sí se usan para prorratear el landed cost del inventario,
    # y el UI los muestra por separado como 'Extras' + 'Total desembolso'.
    total_amount = round(sum(item.quantity * item.unit_cost for item in po_in.items), 2)
    po = PurchaseOrder(
        folio=folio, supplier_id=po_in.supplier_id, warehouse_id=po_in.warehouse_id,
        notes=po_in.notes, status=PurchaseOrderStatus.ORDERED.value, user_id=user_id,
        total_amount=total_amount, paid_amount=0.0, due_date=po_in.due_date,
        extra_costs=extras, landed_cost_allocation=method,
        currency=(po_in.currency or "MXN"),
        fx_rate=float(po_in.fx_rate or 1.0),
    )
    db.add(po)
    await db.flush()
    for item in po_in.items:
        db.add(PurchaseOrderItem(purchase_order_id=po.id, variant_id=item.variant_id, quantity=item.quantity, unit_cost=item.unit_cost))
    await db.commit()
    result = await db.execute(select(PurchaseOrder).where(PurchaseOrder.id == po.id).options(selectinload(PurchaseOrder.items)))
    return result.scalars().first()

async def update_purchase_order(db: AsyncSession, po_id: int, po_in: schemas.PurchaseOrderUpdate) -> Optional[PurchaseOrder]:
    """Edit a PO's supplier/warehouse/notes/due_date/items. Only allowed while
    the order hasn't been received yet — receiving is what mutates stock, so
    any edit afterwards would desync the stock ledger from the PO."""
    result = await db.execute(select(PurchaseOrder).where(PurchaseOrder.id == po_id).options(selectinload(PurchaseOrder.items)))
    po = result.scalars().first()
    if not po:
        return None
    if po.status not in (PurchaseOrderStatus.DRAFT.value, PurchaseOrderStatus.ORDERED.value):
        raise ValueError("Solo se pueden editar órdenes en borrador o pendientes de recibir")

    if po_in.supplier_id is not None:
        po.supplier_id = po_in.supplier_id
    if po_in.warehouse_id is not None:
        po.warehouse_id = po_in.warehouse_id
    if po_in.notes is not None:
        po.notes = po_in.notes
    if po_in.due_date is not None:
        po.due_date = po_in.due_date
    if po_in.extra_costs is not None:
        po.extra_costs = [c.model_dump() for c in po_in.extra_costs]
    if po_in.landed_cost_allocation is not None:
        po.landed_cost_allocation = _validate_allocation(po_in.landed_cost_allocation)
    if po_in.items is not None:
        for item in list(po.items):
            await db.delete(item)
        await db.flush()
        for item in po_in.items:
            db.add(PurchaseOrderItem(purchase_order_id=po.id, variant_id=item.variant_id, quantity=item.quantity, unit_cost=item.unit_cost))
        await db.flush()
        await db.refresh(po, attribute_names=["items"])
    # Recalcular total_amount: SOLO mercancía (deuda con este proveedor).
    # Los extras se muestran por separado y no afectan la CxP de esta OC.
    po.total_amount = round(sum((it.quantity or 0) * (it.unit_cost or 0.0) for it in po.items), 2)

    await db.commit()
    result = await db.execute(select(PurchaseOrder).where(PurchaseOrder.id == po.id).options(selectinload(PurchaseOrder.items)))
    return result.scalars().first()

async def get_purchase_orders(db: AsyncSession, warehouse_ids: Optional[List[int]] = None) -> List[PurchaseOrder]:
    query = select(PurchaseOrder).options(selectinload(PurchaseOrder.items)).order_by(PurchaseOrder.created_at.desc())
    if warehouse_ids is not None:
        query = query.where(PurchaseOrder.warehouse_id.in_(warehouse_ids))
    result = await db.execute(query)
    return result.scalars().all()

async def receive_purchase_order(db: AsyncSession, po_id: int, user_id: Optional[int] = None) -> Optional[PurchaseOrder]:
    """Recibe la orden de compra: prorratea los costos extra (landed cost)
    entre las partidas y crea los StockLots FIFO con el costo INTEGRADO
    (unit_cost de factura + fracción de flete/aduana/etc.). Guarda el costo
    integrado en `landed_unit_cost` de cada partida como snapshot histórico
    para trazabilidad contable — el `unit_cost` original (factura) se preserva
    intacto."""
    result = await db.execute(select(PurchaseOrder).where(PurchaseOrder.id == po_id).options(selectinload(PurchaseOrder.items)))
    po = result.scalars().first()
    if not po or po.status == PurchaseOrderStatus.RECEIVED.value:
        return po

    # Prorrateo de extras — los mismos que se ven en el frontend de la OC.
    extras_total = _extras_total(po.extra_costs)
    items_data = [
        {"quantity": it.quantity, "unit_cost": it.unit_cost} for it in po.items
    ]
    landed_costs = _compute_landed_unit_costs(
        items_data, extras_total, po.landed_cost_allocation or "by_value",
    )

    for item, landed_unit_cost in zip(po.items, landed_costs):
        # Snapshot del costo integrado en la partida (auditable).
        item.landed_unit_cost = landed_unit_cost
        # El FIFO / kardex entran con el costo integrado — desde este punto
        # cada venta que consuma este lote reporta un margen honesto.
        note = f"Recepción de orden de compra {po.folio}"
        if extras_total > 0 and abs(landed_unit_cost - (item.unit_cost or 0.0)) > 0.005:
            note += f" · landed cost prorrateado (factura {item.unit_cost:.2f} → integrado {landed_unit_cost:.2f})"
        movement_in = schemas.StockMovementCreate(
            variant_id=item.variant_id, warehouse_id=po.warehouse_id, quantity=item.quantity,
            movement_type=StockMovementType.IN.value, unit_cost=landed_unit_cost,
            reference=po.folio, notes=note,
        )
        await adjust_stock(db, movement_in, user_id=user_id)
    po.status = PurchaseOrderStatus.RECEIVED.value
    po.received_at = datetime.now(timezone.utc)

    # ── Hook contable: póliza automática de compra ────────────────────────
    # (Cargo Inventarios + IVA acreditable / Abono Proveedores + retenciones)
    # Defensivo: si contabilidad falla, la recepción NO se cae.
    try:
        goods_total = sum(
            (it.quantity or 0) * (it.landed_unit_cost or it.unit_cost or 0.0) for it in po.items
        )
        # IVA típico 16% sobre goods (sin extras — los extras rara vez traen IVA propio;
        # si el proveedor factura IVA sobre flete, ya viene en el subtotal integrado).
        # Nota: la implementación de la Fase 4B mejorada permitirá desglosar IVA por
        # partida cuando la OC tenga un campo tax_amount explícito.
        from app.modules.accounting import service as acc
        await acc.record_purchase_receipt(
            db, po_id=po.id, goods_total=goods_total, tax_total=0.0,
            concept=f"Compra {po.folio} recibida", user_id=user_id,
        )
    except Exception as e:
        log.warning("hook contable compra falló", extra={"po_id": po.id, "error": str(e)}, exc_info=True)

    await db.commit()
    await db.refresh(po)
    return po

async def cancel_purchase_order(db: AsyncSession, po_id: int) -> Optional[PurchaseOrder]:
    result = await db.execute(select(PurchaseOrder).where(PurchaseOrder.id == po_id).options(selectinload(PurchaseOrder.items)))
    po = result.scalars().first()
    if not po:
        return None
    if po.status == PurchaseOrderStatus.RECEIVED.value:
        raise ValueError("No se puede cancelar una orden ya recibida")
    po.status = PurchaseOrderStatus.CANCELLED.value
    await db.commit()
    await db.refresh(po)
    return po


async def _po_pdf_context(db: AsyncSession, po_id: int):
    """Gather a PO + its supplier, warehouse, resolved item names and the
    company profile, ready to feed the PDF builder. Returns None if the PO
    doesn't exist."""
    from app.modules.core_config import service as config_service

    result = await db.execute(
        select(PurchaseOrder).where(PurchaseOrder.id == po_id).options(selectinload(PurchaseOrder.items))
    )
    po = result.scalars().first()
    if not po:
        return None

    supplier = await db.get(Supplier, po.supplier_id)
    warehouse = await db.get(Warehouse, po.warehouse_id)
    company = await config_service.get_company_profile(db)

    item_rows = []
    for it in po.items:
        variant = await db.get(ProductVariant, it.variant_id)
        product = await db.get(Product, variant.product_id) if variant else None
        item_rows.append({
            "name": product.name if product else (variant.sku if variant else "—"),
            "sku": variant.sku if variant else "—",
            "quantity": it.quantity,
            "unit_cost": it.unit_cost,
            "subtotal": round(it.quantity * it.unit_cost, 2),
        })
    return po, supplier, (warehouse.name if warehouse else "—"), company, item_rows


async def generate_purchase_order_pdf(db: AsyncSession, po_id: int) -> Optional[bytes]:
    from app.modules.inventory import documents
    ctx = await _po_pdf_context(db, po_id)
    if not ctx:
        return None
    po, supplier, warehouse_name, company, item_rows = ctx
    return documents.build_purchase_order_pdf(company, supplier, warehouse_name, po, item_rows)


async def email_purchase_order(db: AsyncSession, po_id: int, to: Optional[str] = None) -> dict:
    """Render the PO as PDF and email it to the supplier (or an override
    address). Returns {sent: bool, to: str} so the UI can report the outcome
    honestly even when no SMTP integration is configured."""
    from app.core.email import send_email

    ctx = await _po_pdf_context(db, po_id)
    if not ctx:
        return {"sent": False, "to": "", "error": "not_found"}
    po, supplier, warehouse_name, company, item_rows = ctx
    recipient = to or (supplier.email if supplier else None)
    if not recipient:
        return {"sent": False, "to": "", "error": "no_recipient"}

    from app.modules.inventory import documents
    pdf = documents.build_purchase_order_pdf(company, supplier, warehouse_name, po, item_rows)
    folio = po.folio or f"OC-{po.id}"
    company_name = getattr(company, "legal_name", None) or "Kitchenette"
    body = (
        f"<p>Estimado proveedor,</p>"
        f"<p>Adjunto encontrará la orden de compra <b>{folio}</b> emitida por {company_name}.</p>"
        f"<p>Total: <b>${(po.total_amount or 0):,.2f}</b></p>"
        f"<p>Quedamos atentos. Saludos.</p>"
    )
    sent = await send_email(db, to=recipient, subject=f"Orden de compra {folio}", body_html=body,
                            attachments=[(f"{folio}.pdf", pdf, "pdf")])
    return {"sent": sent, "to": recipient}


async def generate_production_order_pdf(db: AsyncSession, prod_id: int) -> Optional[bytes]:
    from app.modules.inventory import documents
    from app.modules.core_config import service as config_service

    prod = await db.get(ProductionOrder, prod_id)
    if not prod:
        return None
    recipe = await get_recipe(db, prod.recipe_id)
    warehouse = await db.get(Warehouse, prod.warehouse_id)
    company = await config_service.get_company_profile(db)

    recipe_name = "—"
    item_rows = []
    if recipe:
        if getattr(recipe, "name", None):
            recipe_name = recipe.name
        else:
            out_variant = await db.get(ProductVariant, recipe.output_variant_id)
            out_product = await db.get(Product, out_variant.product_id) if out_variant else None
            recipe_name = out_product.name if out_product else (out_variant.sku if out_variant else "—")
        for it in recipe.items:
            variant = await db.get(ProductVariant, it.input_variant_id)
            product = await db.get(Product, variant.product_id) if variant else None
            item_rows.append({
                "name": product.name if product else (variant.sku if variant else "—"),
                "sku": variant.sku if variant else "—",
                "quantity": it.quantity * prod.runs,
            })
    return documents.build_production_order_pdf(company, prod, recipe_name, warehouse.name if warehouse else "—", item_rows)


async def pay_purchase_order(db: AsyncSession, po_id: int, pay_in: "schemas.SupplierPaymentCreate", user_id: Optional[int] = None) -> Optional[PurchaseOrder]:
    from app.modules.inventory.models import SupplierPayment
    result = await db.execute(select(PurchaseOrder).where(PurchaseOrder.id == po_id).options(selectinload(PurchaseOrder.items)))
    po = result.scalars().first()
    if not po:
        return None
    if po.status == PurchaseOrderStatus.CANCELLED.value:
        raise ValueError("No se puede pagar una orden cancelada")
    balance = round((po.total_amount or 0.0) - (po.paid_amount or 0.0), 2)
    if pay_in.amount > balance + 0.001:
        raise ValueError(f"El pago (${pay_in.amount:,.2f}) excede el saldo (${balance:,.2f})")
    db.add(SupplierPayment(
        purchase_order_id=po.id, amount=pay_in.amount, method=pay_in.method,
        reference=pay_in.reference, note=pay_in.note, user_id=user_id,
    ))
    po.paid_amount = round((po.paid_amount or 0.0) + pay_in.amount, 2)

    from app.modules.finance import models as fin
    db.add(fin.Transaction(
        type="expense", amount=round(pay_in.amount, 2), category="supplies",
        description=f"Pago a proveedor — OC {po.folio or '#' + str(po.id)}",
        reference=f"po:{po.id}",
    ))

    # ── Hook contable: póliza automática de pago a proveedor ──────────────
    # Cargo Proveedores / Abono Bancos + pase de IVA pendiente→pagado si aplica.
    # Idempotente por payment_id — pagos parciales generan pólizas separadas.
    try:
        from app.modules.accounting import service as acc
        # Fracción del pago que corresponde a IVA — asume 16% aplicado sobre
        # la parte de mercancía. Si el pago es proporcional al total de la OC,
        # el mismo porcentaje del pago corresponde a IVA.
        # Cuando en Fase 4B se desglose IVA por partida, esto se ajustará
        # con precisión — hoy usa un cálculo conservador.
        total_po = float(po.total_amount or 0.0)
        tax_portion = 0.0  # sin IVA calculado por defecto — el sistema aún no lo desglosa por OC
        # Necesitamos el id del SupplierPayment recién insertado para idempotencia.
        # Como aún no se hizo commit, buscamos el último (será el nuestro).
        await db.flush()
        last_payment = (await db.execute(
            select(SupplierPayment.id).where(SupplierPayment.purchase_order_id == po.id)
            .order_by(SupplierPayment.id.desc()).limit(1)
        )).scalar()
        await acc.record_supplier_payment(
            db, po_id=po.id, payment_id=last_payment or 0,
            amount=pay_in.amount, tax_portion=tax_portion,
            concept=f"Pago a proveedor — OC {po.folio or '#' + str(po.id)}",
            user_id=user_id,
        )

        # ── Hook 8: Diferencia cambiaria si la OC es en moneda extranjera ──
        # La OC se recibió a un TC (fx_rate); el pago es en MXN al TC del día.
        # Si el proveedor factura en USD/EUR, el importe pagado en MXN puede
        # diferir de lo originalmente registrado en Proveedores.
        po_currency = getattr(po, "currency", "MXN") or "MXN"
        if po_currency != "MXN":
            po_fx = float(getattr(po, "fx_rate", 1.0) or 1.0)
            # `pay_in.amount` viene en MXN ya (el usuario ingresa lo que sale del banco)
            # Fracción del pago sobre el total (para calcular la porción original)
            paid_mxn = float(pay_in.amount)
            frac = paid_mxn / total_po if total_po > 0 else 0.0
            # Lo que "debía" ser (a TC original): frac × total en divisa × po_fx
            # Como total_po ya está en MXN al TC original, la parte original = frac × total_po
            original_portion = round(frac * total_po, 2)
            if abs(paid_mxn - original_portion) >= 0.01:
                await acc.record_fx_difference(
                    db, source_ref=f"po:{po.id}:pay:{last_payment or 0}",
                    original_mxn=original_portion, paid_mxn=paid_mxn,
                    concept=f"Diferencia cambiaria — OC {po.folio or '#' + str(po.id)} ({po_currency})",
                    user_id=user_id,
                )
    except Exception as e:
        log.warning("hook contable pago proveedor falló", extra={"po_id": po.id, "error": str(e)}, exc_info=True)

    await db.commit()
    await db.refresh(po)
    return po


# --- BOM / Recipes -------------------------------------------------------------
async def create_recipe(db: AsyncSession, recipe_in: schemas.RecipeCreate) -> Recipe:
    recipe = Recipe(
        output_variant_id=recipe_in.output_variant_id, name=recipe_in.name,
        labor_cost=recipe_in.labor_cost, overhead_cost=recipe_in.overhead_cost,
        extra_costs=[c.model_dump() for c in recipe_in.extra_costs],
        yield_quantity=recipe_in.yield_quantity,
    )
    db.add(recipe)
    await db.flush()
    for item in recipe_in.items:
        db.add(RecipeItem(recipe_id=recipe.id, input_variant_id=item.input_variant_id, quantity=item.quantity))
    product = await db.get(Product, (await db.get(ProductVariant, recipe_in.output_variant_id)).product_id)
    if product:
        product.is_manufactured = True
    await db.commit()
    result = await db.execute(select(Recipe).where(Recipe.id == recipe.id).options(selectinload(Recipe.items)))
    return result.scalars().first()

async def get_recipes(db: AsyncSession) -> List[Recipe]:
    result = await db.execute(select(Recipe).options(selectinload(Recipe.items)).order_by(Recipe.id.desc()))
    return result.scalars().all()

async def get_recipe(db: AsyncSession, recipe_id: int) -> Optional[Recipe]:
    result = await db.execute(select(Recipe).where(Recipe.id == recipe_id).options(selectinload(Recipe.items)))
    return result.scalars().first()

async def update_recipe(db: AsyncSession, recipe_id: int, recipe_in: schemas.RecipeUpdate) -> Optional[Recipe]:
    recipe = await get_recipe(db, recipe_id)
    if not recipe:
        return None
    recipe.name = recipe_in.name
    recipe.labor_cost = recipe_in.labor_cost
    recipe.overhead_cost = recipe_in.overhead_cost
    recipe.extra_costs = [c.model_dump() for c in recipe_in.extra_costs]
    recipe.yield_quantity = recipe_in.yield_quantity
    recipe.is_active = recipe_in.is_active
    for old_item in list(recipe.items):
        await db.delete(old_item)
    await db.flush()
    for item in recipe_in.items:
        db.add(RecipeItem(recipe_id=recipe.id, input_variant_id=item.input_variant_id, quantity=item.quantity))
    await db.commit()
    return await get_recipe(db, recipe_id)

async def get_recipe_cost(db: AsyncSession, recipe_id: int) -> Optional[schemas.RecipeCostBreakdown]:
    recipe = await get_recipe(db, recipe_id)
    if not recipe:
        return None
    materials_cost = 0.0
    missing = []
    for item in recipe.items:
        variant = await db.get(ProductVariant, item.input_variant_id)
        if not variant or variant.cost_price is None:
            missing.append(variant.sku if variant else str(item.input_variant_id))
            continue
        materials_cost += item.quantity * variant.cost_price
    extra_costs_total = sum((c or {}).get("amount", 0) for c in (recipe.extra_costs or []))
    total = materials_cost + recipe.labor_cost + recipe.overhead_cost + extra_costs_total
    unit_cost = total / recipe.yield_quantity if recipe.yield_quantity else total
    return schemas.RecipeCostBreakdown(
        recipe_id=recipe.id, materials_cost=round(materials_cost, 4), labor_cost=recipe.labor_cost,
        overhead_cost=recipe.overhead_cost, extra_costs_total=round(extra_costs_total, 4),
        total_cost=round(total, 4), unit_cost=round(unit_cost, 4),
        missing_cost_inputs=missing,
    )


# --- Production Orders ----------------------------------------------------------
async def create_production_order(db: AsyncSession, po_in: schemas.ProductionOrderCreate, user_id: Optional[int] = None) -> ProductionOrder:
    recipe = await get_recipe(db, po_in.recipe_id)
    if not recipe:
        raise ValueError("Receta no encontrada")
    folio = await _next_folio(db, ProductionOrder, "PROD")
    prod = ProductionOrder(
        folio=folio, recipe_id=po_in.recipe_id, warehouse_id=po_in.warehouse_id,
        runs=po_in.runs, notes=po_in.notes, user_id=user_id,
    )
    db.add(prod)
    await db.commit()
    await db.refresh(prod)
    return prod

async def get_production_orders(db: AsyncSession) -> List[ProductionOrder]:
    result = await db.execute(select(ProductionOrder).order_by(ProductionOrder.created_at.desc()))
    return result.scalars().all()

async def update_production_order(db: AsyncSession, prod_id: int, po_in: schemas.ProductionOrderUpdate) -> Optional[ProductionOrder]:
    prod = await db.get(ProductionOrder, prod_id)
    if not prod:
        return None
    if prod.status != ProductionOrderStatus.DRAFT.value:
        raise ValueError("Solo se pueden editar órdenes en borrador")
    recipe = await get_recipe(db, po_in.recipe_id)
    if not recipe:
        raise ValueError("Receta no encontrada")
    prod.recipe_id = po_in.recipe_id
    prod.warehouse_id = po_in.warehouse_id
    prod.runs = po_in.runs
    prod.notes = po_in.notes
    await db.commit()
    await db.refresh(prod)
    return prod

async def complete_production_order(db: AsyncSession, prod_id: int, user_id: Optional[int] = None) -> ProductionOrder:
    prod = await db.get(ProductionOrder, prod_id)
    if not prod or prod.status != ProductionOrderStatus.DRAFT.value:
        return prod
    recipe = await get_recipe(db, prod.recipe_id)
    if not recipe:
        raise ValueError("Receta no encontrada")

    total_materials_cost = 0.0
    # 1. Consumir insumos (FIFO) del almacén de producción
    for item in recipe.items:
        qty_needed = item.quantity * prod.runs
        out_move = schemas.StockMovementCreate(
            variant_id=item.input_variant_id, warehouse_id=prod.warehouse_id, quantity=int(qty_needed),
            movement_type=StockMovementType.OUT.value, reference=prod.folio,
            notes=f"Consumo de insumo para orden de producción {prod.folio}",
        )
        movement = await adjust_stock(db, out_move, user_id=user_id)
        total_materials_cost += (movement.unit_cost or 0) * qty_needed

    # 2. Dar de alta el producto terminado al costo real calculado
    output_qty = recipe.yield_quantity * prod.runs
    extra_costs_total = sum((c or {}).get("amount", 0) for c in (recipe.extra_costs or []))
    total_cost = total_materials_cost + (recipe.labor_cost + recipe.overhead_cost + extra_costs_total) * prod.runs
    unit_cost_result = total_cost / output_qty if output_qty else 0.0

    in_move = schemas.StockMovementCreate(
        variant_id=recipe.output_variant_id, warehouse_id=prod.warehouse_id, quantity=int(output_qty),
        movement_type=StockMovementType.IN.value, unit_cost=round(unit_cost_result, 4), reference=prod.folio,
        notes=f"Producto terminado de orden de producción {prod.folio}",
    )
    await adjust_stock(db, in_move, user_id=user_id)

    prod.status = ProductionOrderStatus.COMPLETED.value
    prod.unit_cost_result = round(unit_cost_result, 4)
    prod.completed_at = datetime.now(timezone.utc)
    await db.commit()
    await db.refresh(prod)
    return prod


# --- Carga masiva (Excel/CSV) ---------------------------------------------------
PRODUCTS_TEMPLATE_COLUMNS = [
    "sku", "codigo_barras", "producto", "tipo", "categoria", "imagen_url", "fabricado_interno", "talla", "color", "material",
    "precio", "costo", "almacen", "stock_inicial", "punto_reorden", "stock_seguridad",
    "dias_entrega_proveedor",
]

RECIPES_TEMPLATE_COLUMNS = [
    "sku_producto_terminado", "nombre_receta", "sku_insumo", "cantidad_insumo",
    "costo_mano_obra_maquila", "costo_indirectos", "unidades_por_corrida",
]

def _build_xlsx(columns: List[str], example_rows: List[list], sheet_name: str) -> bytes:
    df = pd.DataFrame(example_rows, columns=columns)
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
    return buffer.getvalue()

def generate_products_template() -> bytes:
    # Una columna por cada entrada de PRODUCTS_TEMPLATE_COLUMNS (17). Orden:
    # sku, codigo_barras, producto, tipo, categoria, imagen_url, fabricado_interno,
    # talla, color, material, precio, costo, almacen, stock_inicial, punto_reorden,
    # stock_seguridad, dias_entrega_proveedor
    # tipo: "producto terminado" | "insumo" | "consumible" | "otro"
    example = [
        ["CAM-001-AZ-CH", "7501234567890", "Camisa de algodón", "producto terminado", "Ropa",
         "https://misitio.com/imagenes/camisa-azul.jpg", "no", "CH", "Azul", "Algodón",
         350.0, 180.0, "Almacén Principal", 50, 10, 5, 7],
        ["TEL-ALG-001", "", "Tela de algodón (insumo)", "insumo", "Materia prima",
         "", "no", "", "", "Algodón",
         0, 45.0, "Almacén Principal", 500, 100, 50, 15],
    ]
    return _build_xlsx(PRODUCTS_TEMPLATE_COLUMNS, example, "productos")

def generate_recipes_template() -> bytes:
    example = [
        ["CAM-001-AZ-CH", "Camisa de algodón - receta", "TEL-ALG-001", 1.5, 30.0, 10.0, 1],
    ]
    return _build_xlsx(RECIPES_TEMPLATE_COLUMNS, example, "recetas")

def _read_table(file_bytes: bytes, filename: str) -> pd.DataFrame:
    name = (filename or "").lower()
    if name.endswith(".csv"):
        df = pd.read_csv(BytesIO(file_bytes), dtype=str, keep_default_na=False)
    else:
        df = pd.read_excel(BytesIO(file_bytes), dtype=str, keep_default_na=False)
    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

def _to_float(v, default=None):
    v = (v or "").strip() if isinstance(v, str) else v
    if v in (None, ""):
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default

def _to_int(v, default=None):
    f = _to_float(v, None)
    return int(f) if f is not None else default

def _to_bool_si_no(v) -> bool:
    return str(v).strip().lower() in ("si", "sí", "true", "1", "yes")

# Mapea el valor de la columna "tipo" a ProductItemType. Acepta español/inglés;
# si viene vacío o desconocido, asume "producto terminado" (finished_good).
_ITEM_TYPE_ALIASES = {
    "producto terminado": "finished_good", "terminado": "finished_good", "producto": "finished_good",
    "finished_good": "finished_good", "finished good": "finished_good", "pt": "finished_good",
    "insumo": "raw_material", "materia prima": "raw_material", "raw_material": "raw_material",
    "raw material": "raw_material", "mp": "raw_material",
    "consumible": "consumable", "consumable": "consumable",
    "otro": "other", "other": "other",
}

def _parse_item_type(v) -> str:
    return _ITEM_TYPE_ALIASES.get(str(v or "").strip().lower(), "finished_good")

async def bulk_import_products(db: AsyncSession, file_bytes: bytes, filename: str, user_id: Optional[int] = None) -> schemas.BulkImportResult:
    df = _read_table(file_bytes, filename)
    missing_cols = [c for c in ("sku", "producto", "precio") if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Faltan columnas requeridas en el archivo: {', '.join(missing_cols)}")

    products_by_name: dict[str, Product] = {}
    result_warehouses: dict[str, Warehouse] = {}
    created, updated = 0, 0
    errors: List[schemas.BulkImportRowError] = []

    for idx, row in df.iterrows():
        row_num = idx + 2  # +1 por encabezado, +1 porque idx es 0-based
        try:
            sku = str(row.get("sku", "")).strip()
            product_name = str(row.get("producto", "")).strip()
            if not sku or not product_name:
                errors.append(schemas.BulkImportRowError(row=row_num, message="sku y producto son obligatorios"))
                continue
            price = _to_float(row.get("precio"))
            if price is None:
                errors.append(schemas.BulkImportRowError(row=row_num, message="precio es obligatorio y debe ser numérico"))
                continue

            product = products_by_name.get(product_name.lower())
            if not product:
                result = await db.execute(select(Product).where(Product.name == product_name))
                product = result.scalars().first()
                if not product:
                    product = Product(
                        name=product_name,
                        category=str(row.get("categoria", "")).strip() or None,
                        item_type=_parse_item_type(row.get("tipo")),
                        is_manufactured=_to_bool_si_no(row.get("fabricado_interno", "no")),
                        image_url=str(row.get("imagen_url", "")).strip() or None,
                    )
                    db.add(product)
                    await db.flush()
                products_by_name[product_name.lower()] = product
            img_url = str(row.get("imagen_url", "")).strip()
            if img_url and not product.image_url:
                product.image_url = img_url

            result = await db.execute(select(ProductVariant).where(ProductVariant.sku == sku))
            variant = result.scalars().first()
            cost_price = _to_float(row.get("costo"))
            if variant:
                variant.price = price
                if cost_price is not None:
                    variant.cost_price = cost_price
                variant.barcode = str(row.get("codigo_barras", "")).strip() or None
                variant.size = str(row.get("talla", "")).strip() or None
                variant.color = str(row.get("color", "")).strip() or None
                variant.material = str(row.get("material", "")).strip() or None
                variant.reorder_point = _to_int(row.get("punto_reorden"))
                variant.safety_stock = _to_int(row.get("stock_seguridad"))
                variant.lead_time_days = _to_int(row.get("dias_entrega_proveedor"))
                updated += 1
            else:
                variant = ProductVariant(
                    product_id=product.id, sku=sku, price=price, cost_price=cost_price,
                    barcode=str(row.get("codigo_barras", "")).strip() or None,
                    size=str(row.get("talla", "")).strip() or None,
                    color=str(row.get("color", "")).strip() or None,
                    material=str(row.get("material", "")).strip() or None,
                    reorder_point=_to_int(row.get("punto_reorden")),
                    safety_stock=_to_int(row.get("stock_seguridad")),
                    lead_time_days=_to_int(row.get("dias_entrega_proveedor")),
                )
                db.add(variant)
                await db.flush()
                created += 1

            stock_inicial = _to_int(row.get("stock_inicial"))
            almacen_name = str(row.get("almacen", "")).strip()
            if stock_inicial and stock_inicial > 0 and almacen_name:
                warehouse = result_warehouses.get(almacen_name.lower())
                if not warehouse:
                    result = await db.execute(select(Warehouse).where(Warehouse.name == almacen_name))
                    warehouse = result.scalars().first()
                    if not warehouse:
                        warehouse = Warehouse(name=almacen_name)
                        db.add(warehouse)
                        await db.flush()
                    result_warehouses[almacen_name.lower()] = warehouse
                movement_in = schemas.StockMovementCreate(
                    variant_id=variant.id, warehouse_id=warehouse.id, quantity=stock_inicial,
                    movement_type=StockMovementType.IN.value, unit_cost=cost_price,
                    reference="CARGA-MASIVA", notes="Stock inicial por carga masiva de inventario",
                )
                await adjust_stock(db, movement_in, user_id=user_id)
        except Exception as e:
            errors.append(schemas.BulkImportRowError(row=row_num, message=str(e)))

    await db.commit()
    return schemas.BulkImportResult(total_rows=len(df), created=created, updated=updated, errors=errors)

async def bulk_import_recipes(db: AsyncSession, file_bytes: bytes, filename: str) -> schemas.BulkImportResult:
    df = _read_table(file_bytes, filename)
    missing_cols = [c for c in ("sku_producto_terminado", "sku_insumo", "cantidad_insumo") if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Faltan columnas requeridas en el archivo: {', '.join(missing_cols)}")

    grouped: dict[str, dict] = {}
    errors: List[schemas.BulkImportRowError] = []

    for idx, row in df.iterrows():
        row_num = idx + 2
        try:
            output_sku = str(row.get("sku_producto_terminado", "")).strip()
            input_sku = str(row.get("sku_insumo", "")).strip()
            quantity = _to_float(row.get("cantidad_insumo"))
            if not output_sku or not input_sku or quantity is None:
                errors.append(schemas.BulkImportRowError(
                    row=row_num, message="sku_producto_terminado, sku_insumo y cantidad_insumo son obligatorios"))
                continue

            result = await db.execute(select(ProductVariant).where(ProductVariant.sku == output_sku))
            output_variant = result.scalars().first()
            if not output_variant:
                errors.append(schemas.BulkImportRowError(row=row_num, message=f"SKU de producto terminado no encontrado: {output_sku}"))
                continue
            result = await db.execute(select(ProductVariant).where(ProductVariant.sku == input_sku))
            input_variant = result.scalars().first()
            if not input_variant:
                errors.append(schemas.BulkImportRowError(row=row_num, message=f"SKU de insumo no encontrado: {input_sku}"))
                continue

            entry = grouped.setdefault(output_sku, {
                "output_variant_id": output_variant.id,
                "name": str(row.get("nombre_receta", "")).strip() or None,
                "labor_cost": _to_float(row.get("costo_mano_obra_maquila"), 0) or 0,
                "overhead_cost": _to_float(row.get("costo_indirectos"), 0) or 0,
                "yield_quantity": _to_int(row.get("unidades_por_corrida"), 1) or 1,
                "items": [],
            })
            entry["items"].append(schemas.RecipeItemCreate(input_variant_id=input_variant.id, quantity=quantity))
        except Exception as e:
            errors.append(schemas.BulkImportRowError(row=row_num, message=str(e)))

    created, updated = 0, 0
    for output_sku, entry in grouped.items():
        try:
            result = await db.execute(select(Recipe).where(Recipe.output_variant_id == entry["output_variant_id"]))
            existing = result.scalars().first()
            recipe_in = schemas.RecipeCreate(
                output_variant_id=entry["output_variant_id"], name=entry["name"],
                labor_cost=entry["labor_cost"], overhead_cost=entry["overhead_cost"],
                yield_quantity=entry["yield_quantity"], items=entry["items"],
            )
            if existing:
                await update_recipe(db, existing.id, schemas.RecipeUpdate(**recipe_in.model_dump(), is_active=True))
                updated += 1
            else:
                await create_recipe(db, recipe_in)
                created += 1
        except Exception as e:
            errors.append(schemas.BulkImportRowError(row=0, message=f"Receta para {output_sku}: {e}"))

    return schemas.BulkImportResult(total_rows=len(df), created=created, updated=updated, errors=errors)


# ═════════════════════════════════════════════════════════════════════════════
# Traspasos entre almacenes (Stock Transfer Orders)
#
# Flujo profesional de 6 estados:
#   draft → approved → in_preparation → shipped → received
#         ↘ cancelled (desde draft/approved/in_preparation, NO desde shipped)
#
# Contablemente NO afecta P&L — es solo movimiento físico de un almacén a otro.
# El kardex de ambos almacenes refleja el traspaso con la misma referencia
# (folio TR-XXXXXX), lo que permite conciliar visualmente que lo que salió
# del origen es lo mismo que entró al destino.
# ═════════════════════════════════════════════════════════════════════════════

from app.modules.inventory.models import (
    StockTransfer, StockTransferItem, StockTransferStatus,
)


async def _next_transfer_folio(db: AsyncSession) -> str:
    n = (await db.execute(select(func.count(StockTransfer.id)))).scalar() or 0
    return f"TR-{n + 1:06d}"


def _serialize_transfer(t: StockTransfer) -> dict:
    """Convierte el ORM a dict con nombres de almacén y de productos resueltos."""
    return {
        "id": t.id, "folio": t.folio, "status": t.status,
        "source_warehouse_id": t.source_warehouse_id,
        "destination_warehouse_id": t.destination_warehouse_id,
        "source_warehouse_name": t.source_warehouse.name if t.source_warehouse else None,
        "destination_warehouse_name": t.destination_warehouse.name if t.destination_warehouse else None,
        "notes": t.notes, "expected_delivery_date": t.expected_delivery_date,
        "created_at": t.created_at, "approved_at": t.approved_at,
        "shipped_at": t.shipped_at, "received_at": t.received_at,
        "cancelled_at": t.cancelled_at, "cancelled_reason": t.cancelled_reason,
        "items": [
            {
                "id": it.id, "variant_id": it.variant_id,
                "quantity_requested": it.quantity_requested,
                "quantity_shipped": it.quantity_shipped,
                "quantity_received": it.quantity_received,
                "unit_cost_snapshot": it.unit_cost_snapshot,
                "discrepancy_reason": it.discrepancy_reason,
                "product_name": (it.variant.product.name if it.variant and it.variant.product else None),
                "sku": (it.variant.sku if it.variant else None),
                "barcode": (it.variant.barcode if it.variant else None),
            } for it in (t.items or [])
        ],
    }


async def create_stock_transfer(db: AsyncSession, data: dict, user_id: Optional[int] = None) -> dict:
    folio = await _next_transfer_folio(db)
    transfer = StockTransfer(
        folio=folio,
        source_warehouse_id=data["source_warehouse_id"],
        destination_warehouse_id=data["destination_warehouse_id"],
        status=StockTransferStatus.DRAFT.value,
        notes=data.get("notes"),
        expected_delivery_date=data.get("expected_delivery_date"),
        created_by_id=user_id,
    )
    db.add(transfer)
    await db.flush()
    for item in data["items"]:
        db.add(StockTransferItem(
            transfer_id=transfer.id,
            variant_id=item["variant_id"],
            quantity_requested=int(item["quantity_requested"]),
        ))
    await db.commit()
    return await get_stock_transfer(db, transfer.id)


async def get_stock_transfer(db: AsyncSession, transfer_id: int) -> Optional[dict]:
    res = await db.execute(
        select(StockTransfer).where(StockTransfer.id == transfer_id).options(
            selectinload(StockTransfer.source_warehouse),
            selectinload(StockTransfer.destination_warehouse),
            selectinload(StockTransfer.items).selectinload(StockTransferItem.variant).selectinload(ProductVariant.product),
        )
    )
    t = res.scalars().first()
    return _serialize_transfer(t) if t else None


async def list_stock_transfers(db: AsyncSession, *, warehouse_id: Optional[int] = None,
                                status: Optional[str] = None,
                                limit: int = 100) -> list[dict]:
    """Lista traspasos. Filtro `warehouse_id` incluye tanto origen como destino
    (así el encargado de una tienda ve TODO lo que le compete: lo que envió y
    lo que va a recibir)."""
    stmt = select(StockTransfer).options(
        selectinload(StockTransfer.source_warehouse),
        selectinload(StockTransfer.destination_warehouse),
        selectinload(StockTransfer.items).selectinload(StockTransferItem.variant).selectinload(ProductVariant.product),
    ).order_by(StockTransfer.created_at.desc()).limit(limit)
    if status:
        stmt = stmt.where(StockTransfer.status == status)
    if warehouse_id is not None:
        stmt = stmt.where(
            or_(StockTransfer.source_warehouse_id == warehouse_id,
                StockTransfer.destination_warehouse_id == warehouse_id)
        )
    rows = (await db.execute(stmt)).scalars().all()
    return [_serialize_transfer(t) for t in rows]


async def approve_stock_transfer(db: AsyncSession, transfer_id: int,
                                  user_id: Optional[int] = None) -> Optional[dict]:
    t = await db.get(StockTransfer, transfer_id)
    if not t:
        return None
    if t.status != StockTransferStatus.DRAFT.value:
        raise ValueError(f"Solo se pueden aprobar traspasos en estado 'draft' (actual: {t.status})")
    t.status = StockTransferStatus.APPROVED.value
    t.approved_at = datetime.now(timezone.utc)
    t.approved_by_id = user_id
    await db.commit()
    return await get_stock_transfer(db, transfer_id)


async def start_preparation(db: AsyncSession, transfer_id: int,
                             user_id: Optional[int] = None) -> Optional[dict]:
    """Marca el traspaso como 'en preparación' (CEDIS empieza a armarlo).
    Estado intermedio para que el destino sepa que ya se está preparando."""
    t = await db.get(StockTransfer, transfer_id)
    if not t:
        return None
    if t.status != StockTransferStatus.APPROVED.value:
        raise ValueError(f"Solo se puede iniciar preparación desde 'approved' (actual: {t.status})")
    t.status = StockTransferStatus.IN_PREPARATION.value
    await db.commit()
    return await get_stock_transfer(db, transfer_id)


async def ship_stock_transfer(db: AsyncSession, transfer_id: int, items_shipped: list,
                               user_id: Optional[int] = None) -> Optional[dict]:
    """CEDIS confirma la salida. Consume stock del origen vía FIFO, snapshot el
    unit_cost y marca 'shipped'. items_shipped: [{item_id, quantity_shipped}]."""
    result = await db.execute(
        select(StockTransfer).where(StockTransfer.id == transfer_id).options(
            selectinload(StockTransfer.items),
        )
    )
    t = result.scalars().first()
    if not t:
        return None
    if t.status not in (StockTransferStatus.APPROVED.value, StockTransferStatus.IN_PREPARATION.value):
        raise ValueError(f"Solo se puede enviar desde 'approved' o 'in_preparation' (actual: {t.status})")

    from app.modules.inventory import fifo_service
    ship_map = {int(x["item_id"]): int(x["quantity_shipped"]) for x in items_shipped}

    for item in t.items:
        qty_ship = ship_map.get(item.id, 0)
        if qty_ship <= 0:
            continue
        try:
            r = await fifo_service.consume_stock(
                db, variant_id=item.variant_id, warehouse_id=t.source_warehouse_id,
                quantity=qty_ship, reference=t.folio,
                user_id=user_id, allow_negative=False, commit=False,
            )
            item.quantity_shipped = qty_ship
            item.unit_cost_snapshot = float(r.get("unit_cost_avg") or 0.0)
        except fifo_service.InsufficientStockError as e:
            raise ValueError(
                f"Stock insuficiente en origen para SKU (variant {item.variant_id}): "
                f"requerido {qty_ship}, disponible {e.available}"
            )

    t.status = StockTransferStatus.SHIPPED.value
    t.shipped_at = datetime.now(timezone.utc)
    t.shipped_by_id = user_id
    await db.commit()
    return await get_stock_transfer(db, transfer_id)


async def receive_stock_transfer(db: AsyncSession, transfer_id: int, items_received: list,
                                  user_id: Optional[int] = None) -> Optional[dict]:
    """Destino confirma la recepción. Ingresa stock al destino como lote nuevo
    con el costo snapshoteado en el envío (continuidad FIFO cruzando almacenes).
    items_received: [{item_id, quantity_received, discrepancy_reason}].
    Si received != shipped, se marca discrepancia con la razón dada."""
    result = await db.execute(
        select(StockTransfer).where(StockTransfer.id == transfer_id).options(
            selectinload(StockTransfer.items),
        )
    )
    t = result.scalars().first()
    if not t:
        return None
    if t.status != StockTransferStatus.SHIPPED.value:
        raise ValueError(f"Solo se puede recibir un traspaso 'shipped' (actual: {t.status})")

    from app.modules.inventory import fifo_service
    recv_map = {int(x["item_id"]): {"qty": int(x["quantity_received"]),
                                     "reason": x.get("discrepancy_reason")}
                for x in items_received}

    for item in t.items:
        rec = recv_map.get(item.id)
        if rec is None or rec["qty"] <= 0:
            continue
        qty_recv = rec["qty"]
        # Ingresar al destino como lote nuevo con el costo snapshoteado.
        await fifo_service.receive_stock(
            db, variant_id=item.variant_id, warehouse_id=t.destination_warehouse_id,
            quantity=qty_recv, unit_cost=item.unit_cost_snapshot,
            reference=t.folio, user_id=user_id, commit=False,
        )
        item.quantity_received = qty_recv
        if qty_recv != item.quantity_shipped:
            item.discrepancy_reason = rec.get("reason") or (
                f"Faltante: enviado {item.quantity_shipped}, recibido {qty_recv}"
                if qty_recv < item.quantity_shipped else
                f"Sobrante: enviado {item.quantity_shipped}, recibido {qty_recv}"
            )

    t.status = StockTransferStatus.RECEIVED.value
    t.received_at = datetime.now(timezone.utc)
    t.received_by_id = user_id
    await db.commit()
    return await get_stock_transfer(db, transfer_id)


async def cancel_stock_transfer(db: AsyncSession, transfer_id: int, reason: str,
                                 user_id: Optional[int] = None) -> Optional[dict]:
    t = await db.get(StockTransfer, transfer_id)
    if not t:
        return None
    if t.status == StockTransferStatus.SHIPPED.value:
        raise ValueError(
            "No se puede cancelar un traspaso ya enviado — el stock ya salió del origen. "
            "Marca como recibido en destino (aunque no lo llegara) y luego crea un ajuste "
            "manual con la razón de pérdida."
        )
    if t.status == StockTransferStatus.RECEIVED.value:
        raise ValueError("No se puede cancelar un traspaso ya recibido.")
    if t.status == StockTransferStatus.CANCELLED.value:
        return _serialize_transfer(t)  # idempotente
    t.status = StockTransferStatus.CANCELLED.value
    t.cancelled_at = datetime.now(timezone.utc)
    t.cancelled_by_id = user_id
    t.cancelled_reason = reason
    await db.commit()
    return await get_stock_transfer(db, transfer_id)


async def find_variant_by_code(db: AsyncSession, code: str) -> Optional[dict]:
    """Búsqueda unificada por SKU o barcode — usada por el escáner del módulo
    Traspasos para identificar productos rápidamente."""
    code = (code or "").strip()
    if not code:
        return None
    res = await db.execute(
        select(ProductVariant).where(
            or_(ProductVariant.sku == code, ProductVariant.barcode == code)
        ).options(selectinload(ProductVariant.product))
    )
    v = res.scalars().first()
    if not v:
        return None
    return {
        "variant_id": v.id,
        "sku": v.sku,
        "barcode": v.barcode,
        "product_name": v.product.name if v.product else None,
        "price": v.price,
    }


# ═════════════════════════════════════════════════════════════════════════════
# ALERTAS DE SOBREINVENTARIO
#
# Detecta variantes que tienen MÁS stock del necesario a la velocidad actual
# de ventas. El indicador estándar en retail es "Días de inventario":
#   días = stock_disponible / velocidad_diaria_de_venta
#
# Se calcula por variante × almacén. La velocidad se estima con las ventas
# del propio almacén en los últimos N días (default 60). Ignora variantes
# sin ventas en el período (no se puede estimar demanda de un producto
# nuevo o descontinuado — no es sobreinventario, es "por analizar").
#
# Umbrales estándar retail:
#   < 15 días  → bajo (comprar/traspasar)
#   15-90 días → sano
#   90-180 días → exceso (promocionar)
#   > 180 días → crítico (liquidar)
# ═════════════════════════════════════════════════════════════════════════════

from datetime import timedelta as _timedelta


async def get_overstock_alerts(
    db: AsyncSession,
    *, warehouse_ids: Optional[List[int]] = None,
    lookback_days: int = 60,
    days_threshold: int = 90,
) -> List[dict]:
    """Devuelve variantes con sobreinventario, por almacén, ordenadas por
    severidad (más días primero).

    Args:
      warehouse_ids: filtro; si es None, todos los almacenes visibles.
      lookback_days: ventana de venta para calcular la velocidad (default 60).
      days_threshold: umbral mínimo de días de inventario para incluir
                       en la lista (default 90 = solo exceso y crítico).
    """
    from app.modules.sales.models import Order, OrderItem
    from sqlalchemy import func as _f

    since = datetime.now(timezone.utc) - _timedelta(days=lookback_days)

    # Ventas por variante × almacén en la ventana (unidades)
    sales_stmt = (
        select(
            OrderItem.variant_id.label("vid"),
            Order.warehouse_id.label("wid"),
            _f.coalesce(_f.sum(OrderItem.quantity), 0).label("units_sold"),
        )
        .join(Order, OrderItem.order_id == Order.id)
        .where(
            Order.kind == "order",
            Order.status.notin_(["cancelled", "draft"]),
            Order.created_at >= since,
            OrderItem.variant_id.isnot(None),
            Order.warehouse_id.isnot(None),
        )
        .group_by(OrderItem.variant_id, Order.warehouse_id)
    )
    if warehouse_ids is not None:
        sales_stmt = sales_stmt.where(Order.warehouse_id.in_(warehouse_ids))
    sales_rows = (await db.execute(sales_stmt)).all()
    # Mapa (variant_id, warehouse_id) → unidades vendidas en la ventana
    sold_map: dict[tuple[int, int], int] = {(r.vid, r.wid): int(r.units_sold or 0) for r in sales_rows}

    if not sold_map:
        return []

    # Stock disponible actual por variante × almacén de los que tienen ventas
    variant_ids = {vid for (vid, _) in sold_map.keys()}
    wh_ids_used = {wid for (_, wid) in sold_map.keys()}
    stock_stmt = (
        select(StockLevel).where(
            StockLevel.variant_id.in_(variant_ids),
            StockLevel.warehouse_id.in_(wh_ids_used),
        ).options(
            selectinload(StockLevel.variant).selectinload(ProductVariant.product),
            selectinload(StockLevel.warehouse),
        )
    )
    stock_rows = (await db.execute(stock_stmt)).scalars().all()

    alerts = []
    for lvl in stock_rows:
        key = (lvl.variant_id, lvl.warehouse_id)
        sold_units = sold_map.get(key, 0)
        if sold_units <= 0:
            continue
        available = max(0, (lvl.quantity or 0) - (lvl.reserved_quantity or 0))
        if available <= 0:
            continue
        daily_velocity = sold_units / lookback_days
        if daily_velocity <= 0:
            continue
        days_of_stock = available / daily_velocity
        if days_of_stock < days_threshold:
            continue

        v = lvl.variant
        # Severidad y recomendación
        if days_of_stock >= 180:
            severity = "critical"
            recommendation = "Liquidar o promoción agresiva"
        elif days_of_stock >= 90:
            severity = "warning"
            recommendation = "Promocionar o traspasar a otro almacén"
        else:
            severity = "info"
            recommendation = "Monitorear"

        # Exceso ideal — asumiendo target de 60 días de stock, cuánto sobra
        target_units = int(daily_velocity * 60)
        excess_units = max(0, available - target_units)
        unit_cost = float(v.cost_price or 0.0) if v else 0.0
        excess_value = round(excess_units * unit_cost, 2)

        alerts.append({
            "variant_id": v.id,
            "sku": v.sku if v else "",
            "product_name": (v.product.name if v and v.product else "") if v else "",
            "warehouse_id": lvl.warehouse_id,
            "warehouse_name": lvl.warehouse.name if lvl.warehouse else "",
            "available": available,
            "units_sold_window": sold_units,
            "lookback_days": lookback_days,
            "daily_velocity": round(daily_velocity, 2),
            "days_of_stock": round(days_of_stock, 1),
            "severity": severity,
            "excess_units": excess_units,
            "excess_value": excess_value,
            "unit_cost": unit_cost,
            "recommendation": recommendation,
        })

    # Ordenar por severidad y días
    alerts.sort(key=lambda a: (-a["days_of_stock"], -a["excess_value"]))
    return alerts
