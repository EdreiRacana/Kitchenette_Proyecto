"""Presupuestos vs. Real — planeación anual por cuenta contable.

Módulo estándar de contabilidad ejecutiva: el contador captura el presupuesto
anual por cuenta (12 columnas mensuales) y el sistema lo compara mes a mes
contra los movimientos reales de las pólizas contabilizadas.

Fórmulas:
- Cuentas deudoras (activo, costo, gasto): real = SUM(debit) - SUM(credit)
- Cuentas acreedoras (pasivo, capital, ingreso): real = SUM(credit) - SUM(debit)
- Variación absoluta = real - presupuesto
- Variación % = (real - presupuesto) / |presupuesto| × 100

Rangos de cuenta típicamente presupuestables: ingresos (4), costos (5),
gastos (6). Se permite presupuestar cualquier cuenta postable para
flexibilidad (algunas empresas presupuestan liquidez o CxC objetivo).

Reglas del modelo:
- UniqueConstraint (period_year, account_id, branch_id) — un presupuesto
  por cuenta/año/sucursal.
- Sucursal null = presupuesto consolidado.
"""
from __future__ import annotations

from calendar import monthrange
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Optional

from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.accounting import models


MONTH_COLS = [f"m{i}" for i in range(1, 13)]


# ─── Helpers ────────────────────────────────────────────────────────────────


def _month_bounds(year: int, month: int) -> tuple[datetime, datetime]:
    first = datetime(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_exclusive = datetime(year, month, last_day) + timedelta(days=1)
    return first, end_exclusive


def _account_movements(nature: str, debit: float, credit: float) -> float:
    """Regresa el movimiento del mes en la naturaleza correcta."""
    return (debit - credit) if nature == "deudora" else (credit - debit)


# ─── CRUD ───────────────────────────────────────────────────────────────────


async def list_budgets(
    db: AsyncSession, year: int, branch_id: Optional[int] = None,
) -> List[dict]:
    """Devuelve todos los presupuestos del año enriquecidos con datos de cuenta."""
    q = (
        select(models.AccountBudget, models.Account)
        .join(models.Account, models.AccountBudget.account_id == models.Account.id)
        .where(models.AccountBudget.period_year == year)
        .order_by(models.Account.code)
    )
    if branch_id is not None:
        q = q.where(models.AccountBudget.branch_id == branch_id)
    rows = (await db.execute(q)).all()
    out = []
    for b, a in rows:
        item = {
            "id": b.id, "period_year": b.period_year, "branch_id": b.branch_id,
            "account_id": a.id, "account_code": a.code, "account_name": a.name,
            "account_type": a.account_type, "nature": a.nature,
            "notes": b.notes,
            "annual_total": round(sum(getattr(b, c) or 0.0 for c in MONTH_COLS), 2),
        }
        for c in MONTH_COLS:
            item[c] = float(getattr(b, c) or 0.0)
        out.append(item)
    return out


async def upsert_budget(
    db: AsyncSession, *, year: int, account_id: int, months: dict,
    branch_id: Optional[int] = None, notes: Optional[str] = None,
    user_id: Optional[int] = None,
) -> dict:
    """Crea o actualiza el presupuesto de una cuenta/año. months = {m1..m12: float}."""
    q = select(models.AccountBudget).where(
        models.AccountBudget.period_year == year,
        models.AccountBudget.account_id == account_id,
        models.AccountBudget.branch_id == branch_id,
    )
    b = (await db.execute(q)).scalars().first()
    if b is None:
        b = models.AccountBudget(
            period_year=year, account_id=account_id, branch_id=branch_id,
            created_by_id=user_id,
        )
        db.add(b)
    for c in MONTH_COLS:
        if c in months and months[c] is not None:
            setattr(b, c, float(months[c] or 0.0))
    if notes is not None:
        b.notes = notes
    b.updated_by_id = user_id
    await db.commit()
    await db.refresh(b)
    q2 = select(models.Account).where(models.Account.id == account_id)
    a = (await db.execute(q2)).scalars().first()
    return {
        "id": b.id, "period_year": b.period_year, "branch_id": b.branch_id,
        "account_id": a.id, "account_code": a.code, "account_name": a.name,
        "account_type": a.account_type, "nature": a.nature,
        "notes": b.notes,
        **{c: float(getattr(b, c) or 0.0) for c in MONTH_COLS},
        "annual_total": round(sum(getattr(b, c) or 0.0 for c in MONTH_COLS), 2),
    }


async def delete_budget(db: AsyncSession, budget_id: int) -> bool:
    q = select(models.AccountBudget).where(models.AccountBudget.id == budget_id)
    b = (await db.execute(q)).scalars().first()
    if not b:
        return False
    await db.delete(b)
    await db.commit()
    return True


async def copy_from_year(
    db: AsyncSession, source_year: int, target_year: int,
    factor: float = 1.0, branch_id: Optional[int] = None,
    user_id: Optional[int] = None,
) -> int:
    """Copia todos los presupuestos de source_year al target_year multiplicando
    por `factor` (útil para proyectar con inflación: factor=1.05 = +5%).
    No sobreescribe: si ya hay presupuesto en target_year para una cuenta,
    la respeta."""
    q_src = select(models.AccountBudget).where(
        models.AccountBudget.period_year == source_year,
    )
    if branch_id is not None:
        q_src = q_src.where(models.AccountBudget.branch_id == branch_id)
    sources = (await db.execute(q_src)).scalars().all()
    if not sources:
        return 0

    q_existing = select(models.AccountBudget.account_id).where(
        models.AccountBudget.period_year == target_year,
    )
    if branch_id is not None:
        q_existing = q_existing.where(models.AccountBudget.branch_id == branch_id)
    existing_ids = {r for r, in (await db.execute(q_existing)).all()}

    created = 0
    for src in sources:
        if src.account_id in existing_ids:
            continue
        new_b = models.AccountBudget(
            period_year=target_year,
            account_id=src.account_id,
            branch_id=src.branch_id,
            notes=(f"Copiado de {source_year}"
                    + (f" × {factor:.2%}" if factor != 1.0 else "")),
            created_by_id=user_id,
        )
        for c in MONTH_COLS:
            setattr(new_b, c, round(float(getattr(src, c) or 0.0) * factor, 2))
        db.add(new_b)
        created += 1
    await db.commit()
    return created


# ─── Variance (Presupuesto vs Real) ─────────────────────────────────────────


async def compute_variance(
    db: AsyncSession, year: int, branch_id: Optional[int] = None,
) -> dict:
    """Comparativo mes a mes entre presupuesto capturado y movimientos reales
    de las pólizas contabilizadas del año.
    """
    # 1) Trae presupuestos + cuenta
    budgets = await list_budgets(db, year, branch_id=branch_id)
    if not budgets:
        return {"period_year": year, "rows": [], "totals": _empty_totals()}

    # 2) Trae movimientos por cuenta y por mes de las pólizas del año
    real_by_acct = await _real_per_account_per_month(db, year)

    rows = []
    for b in budgets:
        acct_id = b["account_id"]
        nature = b["nature"]
        real_monthly = real_by_acct.get(acct_id, [0.0] * 12)

        budget_arr = [float(b[c] or 0.0) for c in MONTH_COLS]
        real_arr = [float(real_monthly[i] or 0.0) for i in range(12)]
        variance = [round(real_arr[i] - budget_arr[i], 2) for i in range(12)]
        variance_pct = [
            round(variance[i] / abs(budget_arr[i]) * 100, 1)
            if abs(budget_arr[i]) > 0.01 else None
            for i in range(12)
        ]

        rows.append({
            **b,
            "budget_monthly": budget_arr,
            "real_monthly": real_arr,
            "variance_monthly": variance,
            "variance_pct_monthly": variance_pct,
            "budget_ytd": round(sum(budget_arr), 2),
            "real_ytd": round(sum(real_arr), 2),
            "variance_ytd": round(sum(variance), 2),
            "variance_pct_ytd": (
                round(sum(variance) / abs(sum(budget_arr)) * 100, 1)
                if abs(sum(budget_arr)) > 0.01 else None
            ),
        })

    # 3) Totales generales (por mes)
    totals = _compute_totals(rows)
    return {"period_year": year, "rows": rows, "totals": totals}


def _empty_totals() -> dict:
    return {
        "budget_monthly": [0.0] * 12, "real_monthly": [0.0] * 12,
        "variance_monthly": [0.0] * 12,
        "budget_ytd": 0.0, "real_ytd": 0.0, "variance_ytd": 0.0,
    }


def _compute_totals(rows: List[dict]) -> dict:
    if not rows:
        return _empty_totals()
    bm = [round(sum(r["budget_monthly"][i] for r in rows), 2) for i in range(12)]
    rm = [round(sum(r["real_monthly"][i] for r in rows), 2) for i in range(12)]
    vm = [round(rm[i] - bm[i], 2) for i in range(12)]
    return {
        "budget_monthly": bm, "real_monthly": rm, "variance_monthly": vm,
        "budget_ytd": round(sum(bm), 2),
        "real_ytd": round(sum(rm), 2),
        "variance_ytd": round(sum(vm), 2),
    }


async def _real_per_account_per_month(
    db: AsyncSession, year: int,
) -> dict:
    """{account_id: [real_m1, real_m2, ..., real_m12]} respetando la
    naturaleza de la cuenta (deudora/acreedora)."""
    year_start = datetime(year, 1, 1)
    year_end = datetime(year + 1, 1, 1)

    # Trae account_id + mes + suma cargos + suma abonos
    month_expr = func.extract("month", models.JournalEntry.date)
    q = (
        select(
            models.JournalLine.account_id,
            month_expr.label("month"),
            func.coalesce(func.sum(models.JournalLine.debit), 0.0).label("debit"),
            func.coalesce(func.sum(models.JournalLine.credit), 0.0).label("credit"),
        )
        .join(models.JournalEntry, models.JournalLine.entry_id == models.JournalEntry.id)
        .where(
            models.JournalEntry.status == "posted",
            models.JournalEntry.date >= year_start,
            models.JournalEntry.date < year_end,
        )
        .group_by(models.JournalLine.account_id, month_expr)
    )

    # Necesitamos la naturaleza para cada cuenta
    q_accts = select(models.Account.id, models.Account.nature)
    natures = {aid: n for aid, n in (await db.execute(q_accts)).all()}

    out: dict[int, list[float]] = {}
    for aid, month, deb, cred in (await db.execute(q)).all():
        arr = out.setdefault(int(aid), [0.0] * 12)
        idx = int(month) - 1
        nat = natures.get(int(aid), "deudora")
        arr[idx] += _account_movements(nat, float(deb or 0.0), float(cred or 0.0))
    return out


# ─── XLSX export ────────────────────────────────────────────────────────────


def build_variance_xlsx(variance: dict) -> bytes:
    """XLSX profesional del comparativo presupuesto vs real anual."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    year = variance["period_year"]
    ws.title = f"Presupuesto {year}"

    brand = "33B2F5"
    thin = Side(border_style="thin", color="D0D0D0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=brand)
    section_fill = PatternFill("solid", fgColor="F0F5FA")
    header_font = Font(bold=True, color="FFFFFF")
    bold = Font(bold=True)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    ws["A1"] = f"Presupuesto vs Real — {year}"
    ws["A1"].font = Font(bold=True, size=14, color=brand)

    MESES = ["Ene", "Feb", "Mar", "Abr", "May", "Jun",
             "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

    # Header (Cuenta + 12 meses × 3 subcolumnas + YTD × 3)
    row0 = 3
    ws.cell(row=row0, column=1, value="Cuenta").fill = header_fill
    ws.cell(row=row0, column=1).font = header_font
    ws.cell(row=row0, column=2, value="Tipo").fill = header_fill
    ws.cell(row=row0, column=2).font = header_font

    col = 3
    for m in MESES:
        for sub in ("Presup.", "Real", "Var."):
            c = ws.cell(row=row0, column=col, value=f"{m} {sub}")
            c.fill = header_fill; c.font = header_font
            c.alignment = center
            col += 1
    # YTD
    for sub in ("Presup. YTD", "Real YTD", "Var. YTD", "Var. %"):
        c = ws.cell(row=row0, column=col, value=sub)
        c.fill = header_fill; c.font = header_font
        c.alignment = center
        col += 1

    # Rows
    r_idx = row0 + 1
    for r in variance["rows"]:
        ws.cell(row=r_idx, column=1, value=f"{r['account_code']} · {r['account_name']}")
        ws.cell(row=r_idx, column=2, value=r["account_type"])
        col = 3
        for i in range(12):
            for val in (r["budget_monthly"][i], r["real_monthly"][i], r["variance_monthly"][i]):
                c = ws.cell(row=r_idx, column=col, value=val)
                c.number_format = "#,##0.00"; c.alignment = right
                col += 1
        ws.cell(row=r_idx, column=col, value=r["budget_ytd"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col + 1, value=r["real_ytd"]).number_format = "#,##0.00"
        ws.cell(row=r_idx, column=col + 2, value=r["variance_ytd"]).number_format = "#,##0.00"
        vpct = r["variance_pct_ytd"]
        ws.cell(row=r_idx, column=col + 3, value=(f"{vpct:.1f}%" if vpct is not None else "—"))
        r_idx += 1

    # Totales generales
    totals = variance["totals"]
    r_idx += 1
    ws.cell(row=r_idx, column=1, value="TOTALES").font = Font(bold=True, size=12)
    ws.cell(row=r_idx, column=1).fill = section_fill
    ws.cell(row=r_idx, column=2).fill = section_fill
    col = 3
    for i in range(12):
        for val in (totals["budget_monthly"][i], totals["real_monthly"][i],
                    totals["variance_monthly"][i]):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.number_format = "#,##0.00"; c.alignment = right
            c.font = bold; c.fill = section_fill
            col += 1
    for val in (totals["budget_ytd"], totals["real_ytd"], totals["variance_ytd"]):
        c = ws.cell(row=r_idx, column=col, value=val)
        c.number_format = "#,##0.00"; c.font = bold; c.fill = section_fill
        col += 1

    # Anchos
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    for i in range(3, col + 1):
        letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[letter].width = 14

    ws.freeze_panes = "C4"

    buf = BytesIO(); wb.save(buf); return buf.getvalue()
