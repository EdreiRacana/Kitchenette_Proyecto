"""DIOT — Declaración Informativa de Operaciones con Terceros (SAT A29).

Genera el archivo `.txt` con layout oficial pipe-delimited (`|`) que se sube
al portal del SAT (Servicio de Declaraciones y Pagos → Declaraciones
Informativas → DIOT).

Fuente de datos: Facturas de proveedor (`SupplierBill`) del módulo Finance.
El DIOT se declara sobre **operaciones EFECTIVAMENTE PAGADAS** en el mes
(base de flujo de efectivo), no sobre facturas emitidas.

Un renglón por proveedor con:
1. tipoTercero (04 nacional | 05 extranjero | 15 global)
2. tipoOperacion (85 otros | 84 bienes | 03 serv. profesionales | 06 arrend.)
3. rfc (obligatorio si tipoTercero=04)
4. idFiscal, nombreExtranjero, pais, nacionalidad (solo extranjeros)
5. Valor + IVA por tasa: 16%, 8% frontera, 15% imp, 10% imp frontera, 0%,
   exento, no objeto
6. Retención IVA
7. IVA no acreditable

Referencia: portal SAT A29 y layout oficial DIOT vigente.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import List, Optional

from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.finance import models as fin_models
from app.modules.inventory import models as inv_models


# ── Códigos catálogos SAT (para validación y UI) ────────────────────────
THIRD_TYPES = {
    "04": "Proveedor nacional",
    "05": "Proveedor extranjero",
    "15": "Proveedor global",
}

OPERATION_TYPES = {
    "85": "Otros",
    "84": "Adquisición de bienes",
    "03": "Prestación de servicios profesionales",
    "06": "Arrendamiento de inmuebles",
}


@dataclass
class DiotRow:
    """Un renglón del DIOT — una fila = un proveedor en el periodo."""
    third_type: str = "04"
    operation_type: str = "85"
    rfc: Optional[str] = None
    foreign_tax_id: Optional[str] = None
    foreign_name: Optional[str] = None
    country_code: Optional[str] = None
    nationality: Optional[str] = None

    # Valores por tasa (en pesos, redondeados)
    value_16: int = 0
    iva_16: int = 0
    value_8_border: int = 0
    iva_8_border: int = 0
    value_15_import: int = 0
    iva_15_import: int = 0
    value_10_import_border: int = 0
    iva_10_import_border: int = 0
    value_0: int = 0
    value_exempt: int = 0
    value_no_object: int = 0
    iva_withheld: int = 0
    iva_non_deductible: int = 0

    # Meta para UI (no van al archivo)
    supplier_id: Optional[int] = None
    supplier_name: str = ""
    warnings: List[str] = field(default_factory=list)

    def total_paid(self) -> int:
        return (self.value_16 + self.value_8_border + self.value_15_import
                + self.value_10_import_border + self.value_0
                + self.value_exempt + self.value_no_object)

    def total_iva(self) -> int:
        return (self.iva_16 + self.iva_8_border + self.iva_15_import
                + self.iva_10_import_border)

    def to_pipe_row(self) -> str:
        """Serializa al layout SAT pipe-delimited. Enteros sin decimales.
        Campos vacíos si el valor es 0 o no aplica."""
        def _s(v) -> str:
            if v is None or v == 0 or v == "":
                return ""
            return str(v)
        cols = [
            self.third_type,
            self.operation_type,
            _s(self.rfc),
            _s(self.foreign_tax_id),
            _s(self.foreign_name),
            _s(self.country_code),
            _s(self.nationality),
            _s(self.value_16),
            _s(self.iva_16),
            _s(self.value_8_border),
            _s(self.iva_8_border),
            _s(self.value_15_import),
            _s(self.iva_15_import),
            _s(self.value_10_import_border),
            _s(self.iva_10_import_border),
            _s(self.value_0),
            _s(self.value_exempt),
            _s(self.value_no_object),
            _s(self.iva_withheld),
            _s(self.iva_non_deductible),
        ]
        return "|".join(cols)


def _month_range(year: int, month: int) -> tuple[date, date]:
    """Devuelve (primer_dia, primer_dia_mes_siguiente) para queries [start, end)."""
    from calendar import monthrange
    first = date(year, month, 1)
    last_day = monthrange(year, month)[1]
    end_exclusive = first + timedelta(days=last_day)
    return first, end_exclusive


async def generate_diot_rows(
    db: AsyncSession, year: int, month: int, branch_id: Optional[int] = None,
) -> List[DiotRow]:
    """Agrega SupplierBills pagadas en el mes, agrupadas por proveedor.

    Reglas:
    - Solo bills con `paid_amount > 0` cuyo `paid_at` cae en el mes.
    - Si `paid_at` es NULL pero `status = 'paid'` y `updated_at` cae en el
      mes, también cuenta (compatibilidad con capturas viejas).
    - El valor pagado se prorratea si la bill fue pagada parcialmente
      (paid_amount / total_amount aplicado sobre subtotal e IVA).
    - Por default asumimos IVA a 16% (el más común). Extensión futura:
      breakdown por tasa en cada bill.
    """
    start, end_excl = _month_range(year, month)

    # Bills con pago en el periodo (fecha del pago = paid_at o updated_at)
    q = (
        select(
            fin_models.SupplierBill,
            inv_models.Supplier,
        )
        .outerjoin(inv_models.Supplier,
                    fin_models.SupplierBill.supplier_id == inv_models.Supplier.id)
        .where(
            fin_models.SupplierBill.status.in_(("paid", "partial")),
            fin_models.SupplierBill.paid_amount > 0,
            # cae en el mes por paid_at O por updated_at (fallback)
            (
                (func.date(fin_models.SupplierBill.paid_at) >= start)
                & (func.date(fin_models.SupplierBill.paid_at) < end_excl)
            ) | (
                fin_models.SupplierBill.paid_at.is_(None)
                & (func.date(fin_models.SupplierBill.updated_at) >= start)
                & (func.date(fin_models.SupplierBill.updated_at) < end_excl)
            ),
        )
    )
    if branch_id is not None:
        q = q.where(fin_models.SupplierBill.branch_id == branch_id)

    rows = (await db.execute(q)).all()

    # Agrupamos por supplier_id (o -1 para "sin proveedor")
    by_supplier: dict[int, DiotRow] = {}
    for bill, sup in rows:
        key = sup.id if sup else -1
        if key not in by_supplier:
            drow = DiotRow(
                supplier_id=(sup.id if sup else None),
                supplier_name=(sup.name if sup else (bill.supplier_name or "Sin proveedor")),
                third_type=(sup.diot_third_type or "04") if sup else "04",
                operation_type=(sup.diot_operation_type or "85") if sup else "85",
                rfc=(sup.rfc if sup else None),
                foreign_tax_id=(sup.foreign_tax_id if sup else None),
                foreign_name=(sup.name if sup and sup.diot_third_type == "05" else None),
                country_code=(sup.country_code if sup else None),
                nationality=(sup.foreign_nationality if sup else None),
            )
            by_supplier[key] = drow
        drow = by_supplier[key]

        # Proporción del pago sobre el total (para bills pagadas parcialmente)
        total = float(bill.total_amount or 0.0)
        paid = float(bill.paid_amount or 0.0)
        if total <= 0:
            continue
        ratio = min(1.0, paid / total)
        sub_paid = round(float(bill.subtotal or 0.0) * ratio)
        tax_paid = round(float(bill.tax_amount or 0.0) * ratio)

        # Default: 16% para nacionales, 0% para extranjeros (importaciones fuera)
        if drow.third_type == "04":
            drow.value_16 += sub_paid
            drow.iva_16 += tax_paid
        elif drow.third_type == "05":
            drow.value_15_import += sub_paid
            drow.iva_15_import += tax_paid
        else:
            drow.value_16 += sub_paid
            drow.iva_16 += tax_paid

    # Warnings post-agregación
    for drow in by_supplier.values():
        if drow.third_type == "04" and not drow.rfc:
            drow.warnings.append("Falta RFC del proveedor nacional")
        if drow.third_type == "05" and not drow.country_code:
            drow.warnings.append("Falta país de residencia del extranjero")

    # Orden estable: primero con más pagado
    return sorted(by_supplier.values(), key=lambda r: -r.total_paid())


async def generate_diot_txt(
    db: AsyncSession, year: int, month: int, branch_id: Optional[int] = None,
) -> str:
    """Devuelve el string completo del archivo `.txt` para subir al SAT."""
    rows = await generate_diot_rows(db, year, month, branch_id=branch_id)
    lines = [r.to_pipe_row() for r in rows]
    return "\r\n".join(lines) + ("\r\n" if lines else "")


def generate_diot_xlsx(rows: List[DiotRow], year: int, month: int) -> bytes:
    """XLSX legible para el contador (revisión previa a subir el .txt)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = f"DIOT {year}-{month:02d}"

    brand = "33B2F5"
    thin = Side(border_style="thin", color="D0D0D0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=brand)
    header_font = Font(bold=True, color="FFFFFF")
    total_fill = PatternFill("solid", fgColor="F0F5FA")

    ws["A1"] = f"DIOT — {year}-{month:02d}"
    ws["A1"].font = Font(bold=True, size=14, color=brand)
    ws.merge_cells("A1:M1")

    headers = [
        "Proveedor", "Tipo tercero", "Tipo operación", "RFC",
        "País", "Nacionalidad",
        "Valor 16%", "IVA 16%",
        "Valor 0%", "Valor exento",
        "Retención IVA", "IVA no acreditable",
        "Advertencias",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=i, value=h)
        c.fill = header_fill; c.font = header_font
        c.alignment = Alignment(horizontal="center")
        c.border = box

    for i, r in enumerate(rows, start=4):
        cells = [
            r.supplier_name,
            f"{r.third_type} · {THIRD_TYPES.get(r.third_type, '')}",
            f"{r.operation_type} · {OPERATION_TYPES.get(r.operation_type, '')}",
            r.rfc or "",
            r.country_code or "",
            r.nationality or "",
            r.value_16, r.iva_16,
            r.value_0, r.value_exempt,
            r.iva_withheld, r.iva_non_deductible,
            " · ".join(r.warnings) if r.warnings else "",
        ]
        for j, v in enumerate(cells, start=1):
            c = ws.cell(row=i, column=j, value=v)
            c.border = box
            if j in (7, 8, 9, 10, 11, 12):
                c.number_format = "#,##0"
                c.alignment = Alignment(horizontal="right")
            if r.warnings:
                c.fill = PatternFill("solid", fgColor="FFF6E0")

    # Totales
    if rows:
        total_row = len(rows) + 4
        ws.cell(row=total_row, column=1, value="TOTALES").font = Font(bold=True)
        for col in (7, 8, 9, 10, 11, 12):
            letter = chr(ord("A") + col - 1)
            ws.cell(row=total_row, column=col,
                     value=f"=SUM({letter}4:{letter}{total_row-1})").number_format = "#,##0"
            ws.cell(row=total_row, column=col).font = Font(bold=True)
            ws.cell(row=total_row, column=col).fill = total_fill

    widths = [30, 22, 30, 15, 8, 15, 14, 14, 12, 14, 14, 16, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(ord("A") + i - 1)].width = w

    buf = BytesIO(); wb.save(buf); return buf.getvalue()
