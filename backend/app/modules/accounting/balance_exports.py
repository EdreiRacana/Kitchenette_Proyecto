"""Exports oficiales del Balance General.

Genera:
- PDF formato estado financiero: header con datos fiscales de la empresa,
  tabla con secciones (Activo/Pasivo/Capital), subtotales, totales,
  validación de la ecuación contable, razones financieras y firmas.
- XLSX profesional del balance con formato tabular, formulas, fills.
- XLSX histórico multi-hoja: una hoja por balance + hoja resumen de ratios.

Convención del diccionario `company` (viene de universal_service._get_company_dict):
  {commercial_name, legal_name, rfc, regimen_fiscal, address, brand_color,
   logo_bytes, logo_path, ...}

Convención del diccionario `balance` (viene de service._enrich_balance):
  {id, period_year, period_month, ...campos crudos...,
   totals: {...}, ratios: {...}, status, notes, created_at, updated_at}
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import List

from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image,
)


_MONTHS_ES = [
    "", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def _mxn(n) -> str:
    return "$" + f"{(n or 0):,.2f}"


def _period_label(year: int, month: int) -> str:
    return f"{_MONTHS_ES[month]} de {year}"


def _logo_source(company: dict):
    b = company.get("logo_bytes")
    if b:
        try:
            return BytesIO(b)
        except Exception:
            return None
    import os as _os
    p = company.get("logo_path")
    if p and _os.path.exists(p) and not p.lower().endswith(".svg"):
        return p
    return None


# ─────────────────────────────────────────────────────────────────────────────
# PDF oficial (LETTER portrait)
# ─────────────────────────────────────────────────────────────────────────────


def build_balance_pdf(company: dict, balance: dict) -> bytes:
    """PDF oficial del balance para firma. LETTER portrait."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=LETTER,
        leftMargin=18 * mm, rightMargin=18 * mm,
        topMargin=14 * mm, bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    brand = colors.HexColor(company.get("brand_color") or "#33B2F5")
    story = []

    # ── Header: logo + datos empresa + título ────────────────────
    logo_cell = ""
    src = _logo_source(company)
    if src is not None:
        try:
            logo_cell = Image(src, width=32 * mm, height=14 * mm, kind="proportional")
        except Exception:
            logo_cell = ""

    empresa_lines = []
    if company.get("commercial_name"):
        empresa_lines.append(f"<b>{company['commercial_name']}</b>")
    if company.get("legal_name") and company.get("legal_name") != company.get("commercial_name"):
        empresa_lines.append(company["legal_name"])
    if company.get("rfc"):
        empresa_lines.append(f"RFC: {company['rfc']}")
    if company.get("regimen_fiscal"):
        empresa_lines.append(f"Régimen: {company['regimen_fiscal']}")
    if company.get("address"):
        empresa_lines.append(company["address"])

    style_small = ParagraphStyle(
        "empresa", parent=styles["Normal"], fontSize=8.5, leading=11, alignment=TA_LEFT,
    )
    style_title = ParagraphStyle(
        "title", parent=styles["Heading1"], fontSize=15, alignment=TA_RIGHT,
        textColor=brand, leading=18, spaceAfter=0,
    )
    style_subtitle = ParagraphStyle(
        "sub", parent=styles["Normal"], fontSize=9.5, alignment=TA_RIGHT,
        textColor=colors.HexColor("#666666"), leading=12,
    )

    empresa_para = Paragraph("<br/>".join(empresa_lines) or "&nbsp;", style_small)
    period_lbl = _period_label(balance["period_year"], balance["period_month"])
    title_para = Paragraph("BALANCE GENERAL", style_title)
    sub_para = Paragraph(f"Al 31 de {period_lbl}", style_subtitle)

    header_tbl = Table(
        [[logo_cell, empresa_para, [title_para, sub_para]]],
        colWidths=[38 * mm, 82 * mm, 60 * mm],
    )
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 4 * mm))

    # Barra separadora con color de marca
    bar = Table([[""]], colWidths=[180 * mm], rowHeights=[1.4 * mm])
    bar.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), brand)]))
    story.append(bar)
    story.append(Spacer(1, 4 * mm))

    totals = balance.get("totals") or {}
    ratios = balance.get("ratios") or {}

    # ── Tabla del Balance ──────────────────────────────────────────
    # Dos columnas: Activo (izq) | Pasivo + Capital (der)
    header_row_style = ("BACKGROUND", brand)

    def _section(rows):
        return rows

    activo_rows = [
        ["ACTIVO", ""],
        ["Activo Circulante", ""],
        ["  Caja y equivalentes", _mxn(balance["cash_and_equivalents"])],
        ["  Cuentas por cobrar", _mxn(balance["accounts_receivable"])],
        ["  Inventarios", _mxn(balance["inventory"])],
        ["  Otros activos circulantes", _mxn(balance["other_current_assets"])],
        ["Suma Activo Circulante", _mxn(totals.get("current_assets_total"))],
        ["Activo No Circulante", ""],
        ["  Activo fijo (neto)", _mxn(balance["fixed_assets"])],
        ["  Activos intangibles", _mxn(balance["intangible_assets"])],
        ["  Inversiones largo plazo", _mxn(balance["long_term_investments"])],
        ["  Otros activos no circulantes", _mxn(balance["other_non_current_assets"])],
        ["Suma Activo No Circulante", _mxn(totals.get("non_current_assets_total"))],
        ["TOTAL ACTIVO", _mxn(totals.get("total_assets"))],
    ]

    pasivo_rows = [
        ["PASIVO Y CAPITAL", ""],
        ["Pasivo Corto Plazo", ""],
        ["  Proveedores", _mxn(balance["accounts_payable"])],
        ["  Deuda corto plazo", _mxn(balance["short_term_debt"])],
        ["  Impuestos por pagar", _mxn(balance["taxes_payable"])],
        ["  Otros pasivos circulantes", _mxn(balance["other_current_liabilities"])],
        ["Suma Pasivo Corto Plazo", _mxn(totals.get("current_liabilities_total"))],
        ["Pasivo Largo Plazo", ""],
        ["  Deuda largo plazo", _mxn(balance["long_term_debt"])],
        ["  Otros pasivos no circulantes", _mxn(balance["other_non_current_liabilities"])],
        ["Suma Pasivo Largo Plazo", _mxn(totals.get("non_current_liabilities_total"))],
        ["Capital Contable", ""],
        ["  Capital social", _mxn(balance["equity_capital"])],
        ["  Utilidades acumuladas", _mxn(balance["retained_earnings"])],
        ["  Resultado del ejercicio", _mxn(balance["period_result"])],
        ["Suma Capital Contable", _mxn(totals.get("total_equity"))],
        ["TOTAL PASIVO + CAPITAL", _mxn(totals.get("liabilities_plus_equity"))],
    ]

    # Empatar longitudes para render side-by-side
    max_len = max(len(activo_rows), len(pasivo_rows))
    while len(activo_rows) < max_len:
        activo_rows.append(["", ""])
    while len(pasivo_rows) < max_len:
        pasivo_rows.append(["", ""])

    combined_data = []
    for a, p in zip(activo_rows, pasivo_rows):
        combined_data.append([a[0], a[1], "", p[0], p[1]])

    tbl = Table(
        combined_data,
        colWidths=[52 * mm, 32 * mm, 4 * mm, 54 * mm, 32 * mm],
    )

    tbl_style = [
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("LINEABOVE", (0, 0), (1, 0), 0.6, brand),
        ("LINEABOVE", (3, 0), (4, 0), 0.6, brand),
        # Título ACTIVO / PASIVO
        ("BACKGROUND", (0, 0), (1, 0), brand),
        ("BACKGROUND", (3, 0), (4, 0), brand),
        ("TEXTCOLOR", (0, 0), (1, 0), colors.white),
        ("TEXTCOLOR", (3, 0), (4, 0), colors.white),
        ("FONTNAME", (0, 0), (1, 0), "Helvetica-Bold"),
        ("FONTNAME", (3, 0), (4, 0), "Helvetica-Bold"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
    ]

    # Bold en filas de subtotal y total
    def _mark_bold(rows, col_start):
        for i, r in enumerate(rows):
            label = r[0].strip()
            if label.startswith("Suma "):
                tbl_style.append(("FONTNAME", (col_start, i), (col_start + 1, i), "Helvetica-Bold"))
                tbl_style.append(("LINEABOVE", (col_start, i), (col_start + 1, i), 0.3, colors.HexColor("#999999")))
            if label.startswith("TOTAL"):
                tbl_style.append(("FONTNAME", (col_start, i), (col_start + 1, i), "Helvetica-Bold"))
                tbl_style.append(("FONTSIZE", (col_start, i), (col_start + 1, i), 10.5))
                tbl_style.append(("BACKGROUND", (col_start, i), (col_start + 1, i), colors.HexColor("#F0F5FA")))
                tbl_style.append(("LINEABOVE", (col_start, i), (col_start + 1, i), 0.8, brand))
            if label in ("Activo Circulante", "Activo No Circulante",
                          "Pasivo Corto Plazo", "Pasivo Largo Plazo",
                          "Capital Contable"):
                tbl_style.append(("FONTNAME", (col_start, i), (col_start, i), "Helvetica-Bold"))
                tbl_style.append(("TEXTCOLOR", (col_start, i), (col_start, i), brand))

    _mark_bold(activo_rows, 0)
    _mark_bold(pasivo_rows, 3)

    tbl.setStyle(TableStyle(tbl_style))
    story.append(tbl)
    story.append(Spacer(1, 6 * mm))

    # ── Validación ecuación contable ───────────────────────────────
    is_balanced = totals.get("is_balanced", False)
    check_color = colors.HexColor("#22c55e") if is_balanced else colors.HexColor("#ef4444")
    check_text = "✓ Ecuación contable validada" if is_balanced else "⚠ Ecuación NO cuadra"
    diff = (totals.get("total_assets") or 0) - (totals.get("liabilities_plus_equity") or 0)
    check_para = Paragraph(
        f"<font color='{check_color.hexval()}'><b>{check_text}</b></font>"
        f"  ·  Activo = {_mxn(totals.get('total_assets'))}"
        f"  ·  Pasivo + Capital = {_mxn(totals.get('liabilities_plus_equity'))}"
        + (f"  ·  Diferencia: {_mxn(diff)}" if not is_balanced else ""),
        ParagraphStyle("chk", parent=styles["Normal"], fontSize=8.5, alignment=TA_CENTER),
    )
    story.append(check_para)
    story.append(Spacer(1, 5 * mm))

    # ── Razones financieras ────────────────────────────────────────
    def _r(v, pct=False, decimals=2):
        if v is None:
            return "—"
        return f"{v:.{decimals}f}%" if pct else f"{v:.{decimals}f}"

    ratios_rows = [
        ["Liquidez Corriente", _r(ratios.get("current_ratio")),
         "Prueba Ácida", _r(ratios.get("quick_ratio"))],
        ["Endeudamiento", _r(ratios.get("debt_ratio")),
         "Apalancamiento", _r(ratios.get("debt_to_equity"))],
        ["Capital de Trabajo", _mxn(ratios.get("working_capital")),
         "ROE", _r(ratios.get("roe"), pct=True)],
        ["ROA", _r(ratios.get("roa"), pct=True), "", ""],
    ]
    ratios_tbl = Table(
        [["RAZONES FINANCIERAS", "", "", ""]] + ratios_rows,
        colWidths=[45 * mm, 40 * mm, 45 * mm, 40 * mm],
    )
    ratios_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F5FA")),
        ("TEXTCOLOR", (0, 0), (-1, 0), brand),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("SPAN", (0, 0), (-1, 0)),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#E0E0E0")),
    ]))
    story.append(ratios_tbl)
    story.append(Spacer(1, 8 * mm))

    # ── Notas ──────────────────────────────────────────────────────
    if balance.get("notes"):
        notes_style = ParagraphStyle(
            "notes", parent=styles["Normal"], fontSize=8.5, leading=11,
            leftIndent=4, rightIndent=4,
        )
        story.append(Paragraph(f"<b>Notas:</b> {balance['notes']}", notes_style))
        story.append(Spacer(1, 6 * mm))

    # ── Firmas ─────────────────────────────────────────────────────
    sign_line = "_" * 32
    sign_tbl = Table([
        [sign_line, sign_line, sign_line],
        ["Elaboró", "Revisó", "Autorizó"],
    ], colWidths=[58 * mm, 58 * mm, 58 * mm])
    sign_tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#999999")),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
    ]))
    story.append(sign_tbl)
    story.append(Spacer(1, 8 * mm))

    # ── Pie ────────────────────────────────────────────────────────
    footer_txt = (
        f"Documento generado el {datetime.utcnow().strftime('%d/%m/%Y %H:%M UTC')}"
        f" · Folio interno: BS-{balance['id']:06d}"
        f" · Estado: {balance.get('status', 'draft').upper()}"
    )
    story.append(Paragraph(
        f"<font color='#999999' size='7.5'>{footer_txt}</font>",
        ParagraphStyle("foot", parent=styles["Normal"], alignment=TA_CENTER),
    ))

    doc.build(story)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# XLSX profesional
# ─────────────────────────────────────────────────────────────────────────────


def _xlsx_write_balance_sheet(wb, ws, company: dict, balance: dict):
    """Escribe UN balance en la worksheet ya creada. Reutilizable para hoja
    única o multi-hoja histórico."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    brand_hex = (company.get("brand_color") or "#33B2F5").lstrip("#")

    brand_fill = PatternFill("solid", fgColor=brand_hex)
    subtotal_fill = PatternFill("solid", fgColor="F0F5FA")
    total_fill = PatternFill("solid", fgColor="E6EEF5")
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    small = Font(size=9)
    brand_font = Font(bold=True, color=brand_hex, size=11)
    right = Alignment(horizontal="right")
    center = Alignment(horizontal="center")

    thin = Side(border_style="thin", color="D0D0D0")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header con datos de la empresa
    ws["A1"] = company.get("commercial_name") or "STHENOVA"
    ws["A1"].font = Font(bold=True, size=14, color=brand_hex)
    ws["A2"] = company.get("legal_name") or ""
    ws["A3"] = f"RFC: {company.get('rfc', '')}"
    ws["A4"] = company.get("regimen_fiscal") or ""
    for r in range(1, 5):
        ws[f"A{r}"].alignment = Alignment(horizontal="left")

    ws["D1"] = "BALANCE GENERAL"
    ws["D1"].font = Font(bold=True, size=13, color=brand_hex)
    ws["D1"].alignment = center
    ws.merge_cells("D1:E1")

    period_lbl = _period_label(balance["period_year"], balance["period_month"])
    ws["D2"] = f"Al 31 de {period_lbl}"
    ws["D2"].alignment = center
    ws.merge_cells("D2:E2")

    ws["D3"] = f"Estado: {balance.get('status', 'draft').upper()}"
    ws["D3"].alignment = center
    ws["D3"].font = small
    ws.merge_cells("D3:E3")

    start_row = 6

    # Cabecera de tabla
    ws.cell(row=start_row, column=1, value="ACTIVO").fill = brand_fill
    ws.cell(row=start_row, column=1).font = white_bold
    ws.cell(row=start_row, column=2, value="Monto").fill = brand_fill
    ws.cell(row=start_row, column=2).font = white_bold
    ws.cell(row=start_row, column=2).alignment = right

    ws.cell(row=start_row, column=4, value="PASIVO Y CAPITAL").fill = brand_fill
    ws.cell(row=start_row, column=4).font = white_bold
    ws.cell(row=start_row, column=5, value="Monto").fill = brand_fill
    ws.cell(row=start_row, column=5).font = white_bold
    ws.cell(row=start_row, column=5).alignment = right

    totals = balance.get("totals") or {}
    ratios = balance.get("ratios") or {}

    activo_rows = [
        ("Activo Circulante", None, "section"),
        ("  Caja y equivalentes", balance["cash_and_equivalents"], "line"),
        ("  Cuentas por cobrar", balance["accounts_receivable"], "line"),
        ("  Inventarios", balance["inventory"], "line"),
        ("  Otros activos circulantes", balance["other_current_assets"], "line"),
        ("Suma Activo Circulante", totals.get("current_assets_total"), "subtotal"),
        ("Activo No Circulante", None, "section"),
        ("  Activo fijo (neto)", balance["fixed_assets"], "line"),
        ("  Activos intangibles", balance["intangible_assets"], "line"),
        ("  Inversiones largo plazo", balance["long_term_investments"], "line"),
        ("  Otros activos no circulantes", balance["other_non_current_assets"], "line"),
        ("Suma Activo No Circulante", totals.get("non_current_assets_total"), "subtotal"),
        ("TOTAL ACTIVO", totals.get("total_assets"), "total"),
    ]
    pasivo_rows = [
        ("Pasivo Corto Plazo", None, "section"),
        ("  Proveedores", balance["accounts_payable"], "line"),
        ("  Deuda corto plazo", balance["short_term_debt"], "line"),
        ("  Impuestos por pagar", balance["taxes_payable"], "line"),
        ("  Otros pasivos circulantes", balance["other_current_liabilities"], "line"),
        ("Suma Pasivo Corto Plazo", totals.get("current_liabilities_total"), "subtotal"),
        ("Pasivo Largo Plazo", None, "section"),
        ("  Deuda largo plazo", balance["long_term_debt"], "line"),
        ("  Otros pasivos no circulantes", balance["other_non_current_liabilities"], "line"),
        ("Suma Pasivo Largo Plazo", totals.get("non_current_liabilities_total"), "subtotal"),
        ("Capital Contable", None, "section"),
        ("  Capital social", balance["equity_capital"], "line"),
        ("  Utilidades acumuladas", balance["retained_earnings"], "line"),
        ("  Resultado del ejercicio", balance["period_result"], "line"),
        ("Suma Capital Contable", totals.get("total_equity"), "subtotal"),
        ("TOTAL PASIVO + CAPITAL", totals.get("liabilities_plus_equity"), "total"),
    ]

    def _write_col(rows, r0, col_label, col_val):
        for i, (label, val, kind) in enumerate(rows):
            r = r0 + i
            c_l = ws.cell(row=r, column=col_label, value=label)
            c_v = ws.cell(row=r, column=col_val, value=val)
            c_v.number_format = '"$"#,##0.00'
            c_v.alignment = right
            c_l.border = box
            c_v.border = box
            if kind == "section":
                c_l.font = brand_font
            elif kind == "subtotal":
                c_l.font = bold
                c_v.font = bold
                c_l.fill = subtotal_fill
                c_v.fill = subtotal_fill
            elif kind == "total":
                c_l.font = Font(bold=True, size=11)
                c_v.font = Font(bold=True, size=11)
                c_l.fill = total_fill
                c_v.fill = total_fill

    _write_col(activo_rows, start_row + 1, 1, 2)
    _write_col(pasivo_rows, start_row + 1, 4, 5)

    last_row = start_row + max(len(activo_rows), len(pasivo_rows)) + 2

    # Validación
    is_balanced = totals.get("is_balanced", False)
    check_row = last_row
    ws.cell(row=check_row, column=1,
             value=("Ecuación contable validada" if is_balanced
                    else "Ecuación contable NO cuadra")).font = Font(
        bold=True, color=("22C55E" if is_balanced else "EF4444"),
    )
    ws.merge_cells(start_row=check_row, start_column=1, end_row=check_row, end_column=5)

    # Ratios
    rr = check_row + 2
    ws.cell(row=rr, column=1, value="RAZONES FINANCIERAS").font = bold
    ws.cell(row=rr, column=1).fill = subtotal_fill
    ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=5)

    def _r(v, pct=False, decimals=2):
        if v is None:
            return "—"
        return f"{v:.{decimals}f}{'%' if pct else ''}"

    ratio_lines = [
        ("Liquidez Corriente", _r(ratios.get("current_ratio")),
         "Prueba Ácida", _r(ratios.get("quick_ratio"))),
        ("Endeudamiento", _r(ratios.get("debt_ratio")),
         "Apalancamiento", _r(ratios.get("debt_to_equity"))),
        ("Capital de Trabajo", f"${ratios.get('working_capital', 0):,.2f}",
         "ROE", _r(ratios.get("roe"), pct=True)),
        ("ROA", _r(ratios.get("roa"), pct=True), "", ""),
    ]
    for i, (l1, v1, l2, v2) in enumerate(ratio_lines):
        r = rr + 1 + i
        ws.cell(row=r, column=1, value=l1)
        ws.cell(row=r, column=2, value=v1).alignment = right
        ws.cell(row=r, column=4, value=l2)
        ws.cell(row=r, column=5, value=v2).alignment = right

    # Notas
    if balance.get("notes"):
        r_notes = rr + len(ratio_lines) + 2
        ws.cell(row=r_notes, column=1, value=f"Notas: {balance['notes']}")
        ws.merge_cells(start_row=r_notes, start_column=1, end_row=r_notes, end_column=5)

    # Anchos de columna
    for col_letter, w in [("A", 34), ("B", 18), ("C", 3), ("D", 34), ("E", 18)]:
        ws.column_dimensions[col_letter].width = w


def build_balance_xlsx(company: dict, balance: dict) -> bytes:
    """XLSX con un solo balance."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    period_lbl = _period_label(balance["period_year"], balance["period_month"])
    ws.title = f"Balance {balance['period_year']}-{balance['period_month']:02d}"
    _xlsx_write_balance_sheet(wb, ws, company, balance)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_balance_history_xlsx(company: dict, balances: List[dict]) -> bytes:
    """XLSX multi-hoja: una hoja por balance + resumen de razones."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()

    # Hoja resumen primero
    ws_sum = wb.active
    ws_sum.title = "Resumen"
    brand_hex = (company.get("brand_color") or "#33B2F5").lstrip("#")
    brand_fill = PatternFill("solid", fgColor=brand_hex)
    white_bold = Font(bold=True, color="FFFFFF")

    ws_sum["A1"] = f"Histórico de Balances — {company.get('commercial_name', '')}"
    ws_sum["A1"].font = Font(bold=True, size=13, color=brand_hex)
    ws_sum.merge_cells("A1:H1")

    headers = [
        "Periodo", "Total Activo", "Total Pasivo", "Total Capital",
        "Liquidez", "Endeudamiento", "ROE", "ROA",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws_sum.cell(row=3, column=i, value=h)
        c.fill = brand_fill
        c.font = white_bold
        c.alignment = Alignment(horizontal="center")

    for i, b in enumerate(balances):
        r = 4 + i
        t = b.get("totals") or {}
        ra = b.get("ratios") or {}
        ws_sum.cell(row=r, column=1, value=f"{b['period_year']}-{b['period_month']:02d}")
        ws_sum.cell(row=r, column=2, value=t.get("total_assets"))
        ws_sum.cell(row=r, column=3, value=t.get("total_liabilities"))
        ws_sum.cell(row=r, column=4, value=t.get("total_equity"))
        ws_sum.cell(row=r, column=5, value=ra.get("current_ratio"))
        ws_sum.cell(row=r, column=6, value=ra.get("debt_ratio"))
        ws_sum.cell(row=r, column=7, value=ra.get("roe"))
        ws_sum.cell(row=r, column=8, value=ra.get("roa"))
        for col in (2, 3, 4):
            ws_sum.cell(row=r, column=col).number_format = '"$"#,##0.00'

    for col_letter, w in [("A", 12), ("B", 18), ("C", 18), ("D", 18),
                            ("E", 12), ("F", 14), ("G", 10), ("H", 10)]:
        ws_sum.column_dimensions[col_letter].width = w

    # Una hoja por balance (más reciente primero, ya vienen ordenados)
    for b in balances:
        sheet_name = f"{b['period_year']}-{b['period_month']:02d}"
        ws = wb.create_sheet(title=sheet_name)
        _xlsx_write_balance_sheet(wb, ws, company, b)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
